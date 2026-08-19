import base64
import hashlib
import io
import json
import os
import pickle
import re
import unicodedata
import uuid
import zipfile
from html import escape
from datetime import date, datetime
from difflib import SequenceMatcher
from pathlib import Path
from urllib.parse import parse_qs, quote, urlencode, urlparse, urlunparse
from urllib.request import Request, urlopen

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, Side
from datablix_scanner_panel import render_website_scanner_panel

try:
    from supabase import Client, create_client
except ImportError:  # Cloud persistence remains optional until dependencies are installed.
    Client = object
    create_client = None

st.set_page_config(page_title="Datablix", page_icon="✅", layout="wide")

DATABLIX_BUILD = "SaaS Dashboard UI V1 + Deliverables Generator 2026.08.18-v96"

SOURCE_BASELINE_PARSE_VERSION = 2

# Project-wide municipal boundary. A company's marketing label (for example,
# "Ottawa Region" or "National Capital Region") is never sufficient evidence.
# A property is in scope only when its physical location is within the municipal
# boundaries of the City of Ottawa, Ontario, Canada.
PROJECT_CITY = "Ottawa"
PROJECT_PROVINCE = "Ontario"
PROJECT_COUNTRY = "Canada"
PROJECT_REGION_NAME = "City of Ottawa"
PROJECT_GEOGRAPHIC_SCOPE = (
    "Current residential rental properties physically located within the municipal "
    "boundaries of the City of Ottawa, Ontario, Canada, including apartment "
    "buildings or units, condominium rentals, townhomes, duplexes, and garden homes"
)

# Property-form scope follows the categories present in the project Starting Data.
# A current townhome, duplex, or garden home must not be excluded merely because
# it is not a conventional apartment building. Detached single-family homes are
# retained for human scope review unless the project owner gives a company-specific
# instruction to include or exclude them.
PROJECT_PROPERTY_SCOPE = (
    "Include current apartment buildings and units, condominium rentals, townhomes, "
    "duplexes, and garden homes. Retain current detached single-family homes for "
    "human scope review instead of silently excluding them."
)

# Height-based building classification used by the project.
# Classification is derived only when a reliable storey count (or multiple
# non-conflicting counts that remain in the same band) supports the result.
BUILDING_CLASSIFICATION_BANDS = (
    ("Low-rise", 1, 4),
    ("Mid-rise", 5, 11),
    ("High-rise", 12, None),
)

# Locality labels physically within the City of Ottawa. These labels are useful
# for text normalization, but exact coordinates plus a municipal-boundary check
# are stronger evidence whenever geocoding is available.
OTTAWA_MUNICIPAL_LOCALITIES = {
    "ottawa", "kanata", "nepean", "orleans", "orléans", "gloucester",
    "barrhaven", "stittsville", "vanier", "rockcliffe park", "manotick",
    "carp", "cumberland", "greely", "metcalfe", "osgoode", "richmond",
    "north gower", "navan", "vars", "constance bay", "dunrobin",
    "fitzroy harbour", "munster", "sarsfield", "kinburn",
}

# Nearby independent municipalities are explicitly out of scope even when a
# company groups them under an Ottawa-area marketing page.
OUT_OF_SCOPE_NEARBY_LOCALITIES = {
    "carleton place", "smiths falls", "renfrew", "arnprior", "almonte",
    "mississippi mills", "perth", "kemptville", "north grenville",
    "rockland", "clarence-rockland", "casselman", "embrun", "russell",
    "winchester", "north dundas", "alexandria", "north glengarry",
    "hawkesbury", "cornwall", "gatineau",
}

OTTAWA_LOCALITY_LABELS = OTTAWA_MUNICIPAL_LOCALITIES

# =========================================================
# Configuration
# =========================================================

INTERNAL_COLUMNS = [
    "Record ID", "Company ID", "Building Name", "Management/Owner", "Street Address",
    "Address Line 2", "City", "Province", "Postal Code", "Country",
    "Mailing Address", "PO Box", "PO Box City", "PO Box Province",
    "PO Box Postal Code", "PO Box Search Status", "PO Box Source URL",
    "PO Box Evidence", "PO Box Confidence",
    "Latitude", "Longitude", "Geocoded Municipality",
    "Geographic Scope Status", "Geographic Evidence", "Geographic Confidence",
    "Phone", "Primary Email", "Secondary Email", "Website",
    "Property Website", "Company Website", "Number of Apartments",
    "Apartment Count Search Status", "Apartment Count Source URL",
    "Apartment Count Evidence", "Apartment Count Confidence",
    "Number of Storeys", "Storey Count Search Status", "Storey Count Source URL",
    "Storey Count Evidence", "Storey Count Confidence",
    "Rental Rate Range", "Suite Types", "Amenities",
    "Parking", "Laundry", "Utilities", "Elevator", "Accessibility",
    "Pet Policy", "Smoke-Free", "Property Type", "Building Classification",
    "Current Inventory Status", "Inventory Evidence",
    "Rental Availability Status", "Rental Availability Detail",
    "Rental Availability Evidence",
    "Found on City/Portfolio Page", "Found on HTML Sitemap",
    "Found on XML Sitemap", "Inventory Exclusion Reason",
    "Directory Discovery Status", "Discovery Status Source", "Directory Entry Status", "Source URL", "Date Researched", "Researcher", "Research Status",
    "Source Status", "Verification Status", "Missing Information",
    "Reviewer Notes", "Record Decision",
]

LISTING_COLUMNS = [
    "Apartment Building Name",
    "Street Address",
    "City and Postal Code",
    "Building Classification",
    "Storeys",
    "Number of Apartments",
    "Apartment Building Management/Owner",
    "Phone Number",
    "Email Contact",
    "WebSite",
]

# Required listing fields are kept in the exact order shown in the sample.
# Other useful findings are placed below the main listing instead of widening it.
LISTING_FIELD_MAP = [
    ("Apartment Building Name", "Building Name"),
    ("Street Address", "Street Address"),
    ("City and Postal Code", None),
    ("Building Classification", "Building Classification"),
    ("Storeys", "Number of Storeys"),
    ("Number of Apartments", "Number of Apartments"),
    ("Apartment Building Management/Owner", "Management/Owner"),
    ("Phone Number", "Phone"),
    ("Email Contact", "Primary Email"),
    ("WebSite", "Website"),
]

LISTING_ADDITIONAL_FIELD_MAP = [
    ("Address Line 2", "Address Line 2"),
    ("Mailing Address", "Mailing Address"),
    ("PO Box", "PO Box"),
    ("PO Box City", "PO Box City"),
    ("PO Box Province", "PO Box Province"),
    ("PO Box Postal Code", "PO Box Postal Code"),
    ("PO Box Search Status", "PO Box Search Status"),
    ("PO Box Source URL", "PO Box Source URL"),
    ("PO Box Evidence", "PO Box Evidence"),
    ("PO Box Confidence", "PO Box Confidence"),
    ("Latitude", "Latitude"),
    ("Longitude", "Longitude"),
    ("Geocoded Municipality", "Geocoded Municipality"),
    ("Geographic Scope Status", "Geographic Scope Status"),
    ("Geographic Evidence", "Geographic Evidence"),
    ("Geographic Confidence", "Geographic Confidence"),
    ("Secondary Email", "Secondary Email"),
    ("Property Website", "Property Website"),
    ("Company Website", "Company Website"),
    ("Apartment Count Search Status", "Apartment Count Search Status"),
    ("Apartment Count Source URL", "Apartment Count Source URL"),
    ("Apartment Count Evidence", "Apartment Count Evidence"),
    ("Apartment Count Confidence", "Apartment Count Confidence"),
    ("Storey Count Search Status", "Storey Count Search Status"),
    ("Storey Count Source URL", "Storey Count Source URL"),
    ("Storey Count Evidence", "Storey Count Evidence"),
    ("Storey Count Confidence", "Storey Count Confidence"),
    ("Rental Rate Range", "Rental Rate Range"),
    ("Suite Types", "Suite Types"),
    ("Amenities", "Amenities"),
    ("Parking", "Parking"),
    ("Laundry", "Laundry"),
    ("Utilities", "Utilities"),
    ("Elevator", "Elevator"),
    ("Accessibility", "Accessibility"),
    ("Pet Policy", "Pet Policy"),
    ("Smoke-Free", "Smoke-Free"),
    ("Property Type", "Property Type"),
    ("Current Inventory Status", "Current Inventory Status"),
    ("Inventory Evidence", "Inventory Evidence"),
    ("Rental Availability Status", "Rental Availability Status"),
    ("Rental Availability Detail", "Rental Availability Detail"),
    ("Rental Availability Evidence", "Rental Availability Evidence"),
    ("Found on City/Portfolio Page", "Found on City/Portfolio Page"),
    ("Found on HTML Sitemap", "Found on HTML Sitemap"),
    ("Found on XML Sitemap", "Found on XML Sitemap"),
    ("Inventory Exclusion Reason", "Inventory Exclusion Reason"),
    ("Directory Discovery Status", "Directory Discovery Status"),
    ("Directory Entry Status", "Directory Entry Status"),
    ("Country", "Country"),
    ("Official Source URL", "Source URL"),
    ("Date Researched", "Date Researched"),
    ("Researcher", "Researcher"),
    ("Verification Status", "Verification Status"),
    ("Missing Information", "Missing Information"),
    ("Reviewer Notes", "Reviewer Notes"),
]

TEMPLATE_COLUMNS = LISTING_COLUMNS + [
    "Source URL", "Date Researched", "Researcher", "Research Status",
    "Source Status", "Verification Status", "Missing Information",
    "Reviewer Notes", "Record Decision",
]

ALIASES = {
    "Record ID": ["Record ID", "ID", "Directory ID"],
    "Company ID": ["Company ID", "Organization ID", "Owner ID", "Management ID"],
    "Building Name": [
        "Building Name", "Apartment Building Name",
        "Apartment Building Name (Draft - Check)", "Property Name",
    ],
    "Management/Owner": [
        "Management/Owner", "Apartment Building Management/Owner",
        "Apartment Building Management / Owner", "Assigned Company",
        "Management Company", "Owner", "Original Owner Name", "Company",
    ],
    "Street Address": ["Street Address", "Address (Street Address)", "Address"],
    "Address Line 2": ["Address Line 2", "Address (Address Line 2)", "Suite / Unit"],
    "City": ["City", "Address (City)"],
    "Province": ["Province", "State / Province", "Address (State / Province)"],
    "Postal Code": ["Postal Code", "ZIP / Postal Code", "Address (ZIP / Postal Code)"],
    "Country": ["Country", "Address (Country)"],
    "Mailing Address": ["Mailing Address", "Corporate Mailing Address", "Postal Address"],
    "PO Box": ["PO Box", "P.O. Box", "Postal Box", "Case Postale", "CP", "Box Number"],
    "PO Box City": ["PO Box City", "Mailing City"],
    "PO Box Province": ["PO Box Province", "Mailing Province", "Mailing State / Province"],
    "PO Box Postal Code": ["PO Box Postal Code", "Mailing Postal Code", "Mailing ZIP / Postal Code"],
    "PO Box Search Status": ["PO Box Search Status", "PO Box Status", "Mailing Address Search Status"],
    "PO Box Source URL": ["PO Box Source URL", "Mailing Address Source URL", "PO Box Source"],
    "PO Box Evidence": ["PO Box Evidence", "Mailing Address Evidence"],
    "PO Box Confidence": ["PO Box Confidence", "Mailing Address Confidence"],
    "Latitude": ["Latitude", "Lat"],
    "Longitude": ["Longitude", "Lng", "Lon", "Long"],
    "Geocoded Municipality": ["Geocoded Municipality", "Detected Municipality", "Mapped Municipality"],
    "Geographic Scope Status": ["Geographic Scope Status", "Ottawa Boundary Status", "Municipal Scope Status"],
    "Geographic Evidence": ["Geographic Evidence", "Geocoding Evidence", "Boundary Evidence"],
    "Geographic Confidence": ["Geographic Confidence", "Geocoding Confidence", "Boundary Confidence"],
    "Phone": ["Phone", "Phone Number", "Primary Phone"],
    "Primary Email": ["Primary Email", "Primary Email (Enter Email)", "Email", "Email Contact"],
    "Secondary Email": ["Secondary Email", "Alternate Email"],
    "Website": ["Website", "WebSite", "Website / Source URL", "Official Website"],
    "Property Website": ["Property Website", "Official Property Website", "Building Website"],
    "Company Website": ["Company Website", "Management Company Website", "Corporate Website"],
    "Number of Apartments": [
        "Number of Apartments", "Apartment Count", "No. of Units", "Number of Units",
        "Unit Count", "Total Units", "Units", "Residential Units", "Rental Units",
        "Dwelling Units", "Number of Suites", "Suite Count", "Total Suites",
        "Suites", "Number of Residences", "Residences", "Doors",
    ],
    "Apartment Count Search Status": [
        "Apartment Count Search Status", "Unit Count Search Status",
        "Apartment Count Status", "Unit Count Status",
    ],
    "Apartment Count Source URL": [
        "Apartment Count Source URL", "Unit Count Source URL",
        "Apartment Count Source", "Unit Count Source",
    ],
    "Apartment Count Evidence": [
        "Apartment Count Evidence", "Unit Count Evidence",
        "Apartment Count Supporting Evidence", "Unit Count Supporting Evidence",
    ],
    "Apartment Count Confidence": [
        "Apartment Count Confidence", "Unit Count Confidence",
    ],
    "Number of Storeys": [
        "Number of Storeys", "Number of Floors", "Number of Stories", "Number of Levels",
        "Storey", "Storeys", "Story", "Stories", "Floor", "Floors", "Level", "Levels",
    ],
    "Storey Count Search Status": [
        "Storey Count Search Status", "Storey Search Status", "Storeys Search Status",
        "Storey Count Status", "Storeys Status",
    ],
    "Storey Count Source URL": [
        "Storey Count Source URL", "Storey Source URL", "Storeys Source URL",
        "Storey Count Source", "Storey Source",
    ],
    "Storey Count Evidence": [
        "Storey Count Evidence", "Storey Evidence", "Storeys Evidence",
        "Storey Count Supporting Evidence",
    ],
    "Storey Count Confidence": [
        "Storey Count Confidence", "Storey Confidence", "Storeys Confidence",
    ],
    "Rental Rate Range": ["Rental Rate Range", "Rental Rates", "Rent Range", "Rent"],
    "Suite Types": ["Suite Types", "Unit Types", "Bedroom Types", "Floor Plan Types"],
    "Amenities": ["Amenities", "Detected Amenities", "Features"],
    "Parking": ["Parking", "Parking Details"],
    "Laundry": ["Laundry", "Laundry Details"],
    "Utilities": ["Utilities", "Utilities Included"],
    "Elevator": ["Elevator", "Elevator Available"],
    "Accessibility": ["Accessibility", "Accessible Features"],
    "Pet Policy": ["Pet Policy", "Pets"],
    "Smoke-Free": ["Smoke-Free", "Smoke Free", "Non-Smoking"],
    "Property Type": [
        "Property Type", "Property Form", "Housing Type", "Residential Property Type",
        "Building Type",
    ],
    "Building Classification": ["Building Classification", "Verified Building Classification"],
    "Current Inventory Status": ["Current Inventory Status", "Inventory Status", "Portfolio Status"],
    "Inventory Evidence": ["Inventory Evidence", "Current Inventory Evidence"],
    "Rental Availability Status": [
        "Rental Availability Status", "Rental Status", "Current Rental Status",
        "Availability Status", "Leasing Status",
    ],
    "Rental Availability Detail": [
        "Rental Availability Detail", "Availability Detail", "Availability Date",
        "Available From", "Rental Availability Date",
    ],
    "Rental Availability Evidence": [
        "Rental Availability Evidence", "Rental Status Evidence",
        "Availability Evidence", "Leasing Status Evidence",
    ],
    "Found on City/Portfolio Page": ["Found on City/Portfolio Page", "On City Page", "On Portfolio Page"],
    "Found on HTML Sitemap": ["Found on HTML Sitemap", "On HTML Sitemap"],
    "Found on XML Sitemap": ["Found on XML Sitemap", "On XML Sitemap"],
    "Inventory Exclusion Reason": ["Inventory Exclusion Reason", "Exclusion Reason"],
    "Directory Discovery Status": ["Directory Discovery Status", "Discovery Status", "Record Origin", "Directory Origin"],
    "Discovery Status Source": ["Discovery Status Source", "Discovery Classification Source", "Discovery Override Source"],
    "Directory Entry Status": ["Directory Entry Status", "Entry Status", "Directory Submission Status"],
    "Source URL": ["Source URL", "Official Source URL", "Research Source", "Website / Source URL"],
    "Date Researched": ["Date Researched", "Date Verified", "Verification Date", "Research Date"],
    "Researcher": ["Researcher", "Assigned To"],
    "Research Status": ["Research Status"],
    "Source Status": ["Source Status"],
    "Verification Status": ["Verification Status", "Review Status"],
    "Missing Information": ["Missing Information", "Information Missing"],
    "Reviewer Notes": ["Reviewer Notes", "Research Notes", "Notes"],
    "Record Decision": ["Record Decision", "Decision"],
}

COMBINED_LOCATION_ALIASES = [
    "City and Postal Code", "City & Postal Code",
    "City, Province and Postal Code", "City Province Postal Code",
]

PROPERTY_TYPE_SOURCE_COLUMNS = [
    "Apartment", "Condominium", "Condo", "Townhome", "Duplex", "Garden Home",
    "Semi-detached", "Semi Detached", "Detached", "Detached Home",
]
PROPERTY_TYPE_LABELS = {
    "Apartment": "Apartment",
    "Condominium": "Condominium",
    "Condo": "Condominium",
    "Townhome": "Townhome",
    "Duplex": "Duplex",
    "Garden Home": "Garden Home",
    "Semi-detached": "Semi-detached",
    "Semi Detached": "Semi-detached",
    "Detached": "Detached",
    "Detached Home": "Detached",
}

CORE_FIELDS = ["Management/Owner", "Street Address", "City"]
TARGET_FIELDS = [
    "Building Name", "Province", "Postal Code", "Phone", "Primary Email",
    "Website", "Number of Apartments", "Number of Storeys",
    "Property Type", "Building Classification",
]
ALL_RESEARCH_FIELDS = CORE_FIELDS + TARGET_FIELDS

RESEARCH_STATUSES = [
    "Imported - Needs Review", "Not Started", "In Progress", "Needs Follow-up",
    "Ready for Review", "Completed",
]
SOURCE_STATUSES = ["Not Checked", "Active", "Needs Follow-up", "Unavailable"]
VERIFICATION_STATUSES = ["Not Reviewed", "Needs Review", "Verified"]
RECORD_DECISIONS = ["Undecided", "Keep", "Update", "Possible Duplicate", "Remove"]
DISCOVERY_STATUSES = [
    "Needs Classification",
    "Existing Source Record",
    "Newly Discovered",
    "Possible Duplicate",
    "Excluded / Not Current",
]
DISCOVERY_STATUS_SOURCES = ["Automatic", "Manual"]
DIRECTORY_ENTRY_STATUSES = ["Not Entered", "Entered", "Needs Correction"]
PO_BOX_SEARCH_STATUSES = [
    "Not Checked", "Found", "Not Found after Search", "Not Applicable", "Needs Review",
]
EVIDENCE_CONFIDENCE_LEVELS = ["Not Checked", "High", "Medium", "Low"]
COUNT_SEARCH_STATUSES = [
    "Not Checked", "Official page", "Official document", "Public record",
    "Reputable property source", "Not Found after Search", "Conflict — Needs Review",
]
GEOGRAPHIC_SCOPE_STATUSES = [
    "Not Checked", "Inside City of Ottawa", "Outside City of Ottawa",
    "Needs Geographic Review",
]
RENTAL_AVAILABILITY_STATUSES = [
    "Not Checked",
    "Available Now",
    "Available Future",
    "Rented",
    "Status Unclear — Needs Review",
]
RENTAL_AVAILABILITY_ALIASES = {
    "available": "Available Now",
    "available now": "Available Now",
    "for rent": "Available Now",
    "for-rent": "Available Now",
    "vacant": "Available Now",
    "available future": "Available Future",
    "future availability": "Available Future",
    "upcoming": "Available Future",
    "rented": "Rented",
    "leased": "Rented",
    "taken": "Rented",
    "no longer available": "Rented",
    "not available": "Rented",
    "review": "Status Unclear — Needs Review",
    "unclear": "Status Unclear — Needs Review",
    "status unclear": "Status Unclear — Needs Review",
    "needs review": "Status Unclear — Needs Review",
}

COMPANY_STATUSES = [
    "Not started", "Researching", "Needs follow-up", "Ready for QA",
    "Complete", "Complete with limitations",
]
COMPANY_SCOPE_TYPES = ["Initial assignment", "Added later", "Imported"]
COMPANY_COLUMNS = [
    "Company ID", "Management/Owner", "Main Website",
    "Related Official Links", "Special Website Notes", "Scope Type",
    "Date Assigned", "Company Status", "Notes",
    "Prompt Scope", "Prompt Source Policy", "Prompt Priority Notes",
    "Prompt Output Notes", "Research Prompt", "Prompt Updated", "AI Tool Used",
]

UNRESOLVED = {
    "", "n/a", "na", "n.a.", "unknown", "not known", "not available",
    "not found", "not provided", "not researched", "tbd", "-", "--",
    "none", "null",
}
YES_VALUES = {"yes", "y", "true", "1"}
NO_VALUES = {"no", "n", "false", "0"}

STATUS_ALIASES = {
    "Research Status": {
        "imported": "Imported - Needs Review", "complete": "Completed",
        "completed": "Completed", "ready": "Ready for Review",
        "follow up": "Needs Follow-up", "follow-up": "Needs Follow-up",
    },
    "Source Status": {
        "verified": "Active", "working": "Active", "broken": "Unavailable",
        "follow up": "Needs Follow-up", "follow-up": "Needs Follow-up",
    },
    "Verification Status": {
        "complete": "Verified", "completed": "Verified", "reviewed": "Verified",
        "not verified": "Not Reviewed",
    },
    "Record Decision": {"duplicate": "Possible Duplicate", "delete": "Remove"},
}

PROVINCES = {
    "ab": ("Alberta", "AB"), "alberta": ("Alberta", "AB"),
    "bc": ("British Columbia", "BC"), "british columbia": ("British Columbia", "BC"),
    "mb": ("Manitoba", "MB"), "manitoba": ("Manitoba", "MB"),
    "nb": ("New Brunswick", "NB"), "new brunswick": ("New Brunswick", "NB"),
    "nl": ("Newfoundland and Labrador", "NL"),
    "newfoundland and labrador": ("Newfoundland and Labrador", "NL"),
    "ns": ("Nova Scotia", "NS"), "nova scotia": ("Nova Scotia", "NS"),
    "nt": ("Northwest Territories", "NT"),
    "northwest territories": ("Northwest Territories", "NT"),
    "nu": ("Nunavut", "NU"), "nunavut": ("Nunavut", "NU"),
    "on": ("Ontario", "ON"), "ontario": ("Ontario", "ON"),
    "pe": ("Prince Edward Island", "PE"),
    "prince edward island": ("Prince Edward Island", "PE"),
    "qc": ("Quebec", "QC"), "quebec": ("Quebec", "QC"), "québec": ("Quebec", "QC"),
    "sk": ("Saskatchewan", "SK"), "saskatchewan": ("Saskatchewan", "SK"),
    "yt": ("Yukon", "YT"), "yukon": ("Yukon", "YT"),
}

FRESHNESS_DAYS = 180

S_FILE = "db_file_signature"
S_ORIGINAL = "db_original"
S_WORKING = "db_working"
S_NAME = "db_name"
S_SHEET = "db_sheet"
S_MAPPING = "db_mapping"
S_FLASH = "db_flash"
S_SOURCE_TYPE = "db_source_type"
S_SOURCE_REF = "db_source_ref"
S_SELECTOR = "db_selector"
S_EDIT_COUNT = "db_edit_count"
S_PROJECT_NAME = "db_project_name"
S_COMPANIES = "db_company_registry"
S_ACTIVE_COMPANY = "db_active_company_id"
S_PROJECT_LOADED = "db_project_loaded"
S_SCAN_HISTORY = "db_scan_history"
S_SCAN_CANDIDATES = "db_scan_candidates_history"
S_SCAN_PAGES = "db_scan_pages_history"
S_MANUAL_ENTRY_OPEN = "db_manual_entry_open"
S_PENDING_ACTIVE_COMPANY = "db_pending_active_company"
S_CLOUD_PROJECT_ID = "db_cloud_project_id"
S_AUTH_USER_ID = "db_auth_user_id"
S_AUTH_EMAIL = "db_auth_email"
S_AUTH_ACCESS_TOKEN = "db_auth_access_token"
S_AUTH_REFRESH_TOKEN = "db_auth_refresh_token"
S_PROJECT_ROLE = "db_project_role"
S_CLOUD_STATE_HASH = "db_cloud_state_hash"
S_SKIP_CLOUD_RESTORE = "db_skip_cloud_restore"
S_DEMO_MODE = "db_demo_mode"
S_SHOW_AUTH = "db_show_auth"
S_SOURCE_BASELINE_META = "db_source_baseline_meta"
S_SOURCE_VERSIONS = "db_source_versions"
S_CLASSIFICATION_RULES = "db_classification_rules"
S_COMPANY_LINK_REPAIR_VERSION = "db_company_link_repair_version"
S_REPORTING_SNAPSHOT = "db_reporting_snapshot"

AUTOSAVE_DIRECTORY = Path(
    os.environ.get("DATABLIX_AUTOSAVE_DIRECTORY", "/tmp/datablix_autosave")
)
AUTOSAVE_FILE = AUTOSAVE_DIRECTORY / "current_project.pkl"

def _autosave_file() -> Path:
    """Use a separate local fallback file for each signed-in account."""
    email = str(st.session_state.get(S_AUTH_EMAIL, "anonymous")).strip().lower() or "anonymous"
    identity = hashlib.sha256(email.encode("utf-8")).hexdigest()[:20]
    return AUTOSAVE_DIRECTORY / f"current_project_{identity}.pkl"
AUTOSAVE_STATE_KEYS = [
    S_FILE, S_ORIGINAL, S_WORKING, S_NAME, S_SHEET, S_MAPPING,
    S_SOURCE_TYPE, S_SOURCE_REF, S_SELECTOR, S_EDIT_COUNT,
    S_PROJECT_NAME, S_COMPANIES, S_ACTIVE_COMPANY,
    S_PROJECT_LOADED, S_SCAN_HISTORY, S_SCAN_CANDIDATES, S_SCAN_PAGES,
    S_SOURCE_BASELINE_META, S_SOURCE_VERSIONS, S_CLASSIFICATION_RULES,
    S_COMPANY_LINK_REPAIR_VERSION, S_REPORTING_SNAPSHOT,
    S_CLOUD_PROJECT_ID, "db_section",
]


def _secret_value(name: str, default: str = "") -> str:
    """Read a Streamlit secret or environment variable without exposing it."""
    try:
        value = st.secrets.get(name, default)
    except Exception:
        value = default
    return str(value or os.environ.get(name, default)).strip()


@st.cache_resource(show_spinner=False)
def get_supabase_client():
    """Create the server-side Supabase client when cloud saving is configured."""
    if create_client is None:
        return None
    url = _secret_value("SUPABASE_URL")
    key = _secret_value("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not key:
        return None
    try:
        return create_client(url, key)
    except Exception:
        return None


def cloud_persistence_available() -> bool:
    return get_supabase_client() is not None


def get_supabase_auth_client():
    """Create a session-local Supabase client for email-and-password authentication."""
    if create_client is None:
        return None
    url = _secret_value("SUPABASE_URL")
    key = _secret_value("SUPABASE_PUBLISHABLE_KEY")
    if not url or not key:
        return None
    try:
        return create_client(url, key)
    except Exception:
        return None


def current_user_email() -> str:
    return str(st.session_state.get(S_AUTH_EMAIL, "")).strip().lower()


def current_user_id() -> str:
    return str(st.session_state.get(S_AUTH_USER_ID, "")).strip()


def user_is_authenticated() -> bool:
    return bool(current_user_email() and current_user_id())


def _remember_auth_response(response) -> bool:
    user = getattr(response, "user", None)
    session = getattr(response, "session", None)
    if user is None:
        return False
    if session is None:
        return False
    st.session_state[S_AUTH_USER_ID] = str(getattr(user, "id", "") or "")
    st.session_state[S_AUTH_EMAIL] = str(getattr(user, "email", "") or "").strip().lower()
    st.session_state[S_AUTH_ACCESS_TOKEN] = str(getattr(session, "access_token", "") or "")
    st.session_state[S_AUTH_REFRESH_TOKEN] = str(getattr(session, "refresh_token", "") or "")
    return user_is_authenticated()


def sign_out_datablix() -> None:
    client = get_supabase_auth_client()
    if client is not None:
        try:
            client.auth.sign_out()
        except Exception:
            pass
    for key in list(st.session_state.keys()):
        if str(key).startswith(("db_", "website_scan", "full_scan")):
            st.session_state.pop(key, None)


def _valid_email(value: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", str(value or "").strip()))


def sign_in_with_password(email: str, password: str) -> tuple[bool, str]:
    """Sign an approved Datablix user into a private workspace."""
    clean_email = str(email or "").strip().lower()
    if not _valid_email(clean_email):
        return False, "Enter a valid email address."
    if not str(password or ""):
        return False, "Enter your password."
    client = get_supabase_auth_client()
    if client is None:
        return False, "Workspace sign-in is not configured."
    try:
        response = client.auth.sign_in_with_password(
            {"email": clean_email, "password": password}
        )
        if _remember_auth_response(response):
            st.session_state.pop(S_SHOW_AUTH, None)
            return True, "Signed in."
        return False, "The workspace could not be opened."
    except Exception:
        return False, "Sign-in failed. Check your email and password, then try again."


def render_public_entry_gate() -> None:
    """Show a short public landing page before demo access or private sign-in."""
    if user_is_authenticated() or st.session_state.get(S_DEMO_MODE):
        return
    if st.session_state.get(S_SHOW_AUTH):
        return

    render_brand_header()
    st.markdown("### Choose how you’d like to begin")

    demo_col, access_col = st.columns(2)
    with demo_col:
        with st.container(border=True):
            st.markdown("#### Explore Your Demo")
            st.write("Explore your rental property research workspace using realistic sample data.")
            if st.button("Explore Demo", type="primary", width="stretch", key="db_public_demo"):
                start_demo_workspace()
                st.rerun()
            st.caption("No account required.")

    with access_col:
        with st.container(border=True):
            st.markdown("#### Access Your Workspace")
            st.write("Sign in to open your saved projects and continue your work with your team.")
            if st.button("Continue", width="stretch", key="db_public_continue"):
                st.session_state[S_SHOW_AUTH] = True
                st.rerun()
            st.caption("Authorized users only.")

    if not st.session_state.get(S_SHOW_AUTH):
        st.stop()


def render_auth_gate() -> None:
    """Require email-and-password authentication for private Datablix workspaces."""
    if user_is_authenticated() or st.session_state.get(S_DEMO_MODE):
        return

    render_brand_header()
    if st.button("Back", key="db_auth_back"):
        st.session_state.pop(S_SHOW_AUTH, None)
        st.rerun()

    st.markdown("### Access Your Workspace")
    st.write("Sign in with the email and password assigned to your Datablix account.")
    if get_supabase_auth_client() is None:
        st.error("Authentication is not configured. Add SUPABASE_PUBLISHABLE_KEY to Streamlit Secrets.")
        st.stop()

    with st.form("db_sign_in_form"):
        email = st.text_input("Email address", placeholder="name@example.com")
        password = st.text_input("Password", type="password")
        submitted = st.form_submit_button("Access Workspace", type="primary", use_container_width=True)

    if submitted:
        ok, message = sign_in_with_password(email, password)
        if ok:
            st.rerun()
        st.error(message)

    st.caption("Only authorized users with an existing account can sign in.")
    st.stop()


def project_access_role(project_id: str) -> str:
    """Return owner, editor, viewer, or an empty string for no access."""
    email = current_user_email()
    if not email or not project_id:
        return ""
    client = get_supabase_client()
    if client is None:
        return ""
    try:
        project = (
            client.table("datablix_project_state")
            .select("owner_email")
            .eq("project_id", project_id)
            .limit(1)
            .execute()
        )
        rows = list(project.data or [])
        if rows and str(rows[0].get("owner_email", "")).strip().lower() == email:
            return "owner"
        membership = (
            client.table("datablix_project_members")
            .select("role")
            .eq("project_id", project_id)
            .eq("member_email", email)
            .limit(1)
            .execute()
        )
        members = list(membership.data or [])
        return str(members[0].get("role", "")) if members else ""
    except Exception:
        return ""


def user_can_edit_project(project_id: str | None = None) -> bool:
    if st.session_state.get(S_DEMO_MODE):
        return True
    target = str(project_id or st.session_state.get(S_CLOUD_PROJECT_ID, "")).strip()
    role = str(st.session_state.get(S_PROJECT_ROLE, "") or project_access_role(target)).lower()
    return role in {"owner", "editor"}


def list_project_members(project_id: str) -> list[dict]:
    if not project_id or st.session_state.get(S_PROJECT_ROLE) != "owner":
        return []
    try:
        response = (
            get_supabase_client().table("datablix_project_members")
            .select("member_email,role,added_at")
            .eq("project_id", project_id)
            .order("added_at")
            .execute()
        )
        return list(response.data or [])
    except Exception:
        return []


def add_project_member(project_id: str, email: str, role: str) -> tuple[bool, str]:
    if st.session_state.get(S_PROJECT_ROLE) != "owner":
        return False, "Only the project owner can manage access."
    clean_email = str(email).strip().lower()
    if not re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", clean_email):
        return False, "Enter a valid email address."
    if clean_email == current_user_email():
        return False, "You already own this project."
    try:
        get_supabase_client().table("datablix_project_members").upsert(
            {
                "project_id": project_id,
                "member_email": clean_email,
                "role": role if role in {"editor", "viewer"} else "editor",
                "added_by": current_user_email(),
                "added_at": datetime.now().astimezone().isoformat(),
            },
            on_conflict="project_id,member_email",
        ).execute()
        return True, f"Access saved for {clean_email}."
    except Exception:
        return False, "Access could not be saved. Run the updated Supabase SQL first."


def remove_project_member(project_id: str, email: str) -> bool:
    if st.session_state.get(S_PROJECT_ROLE) != "owner":
        return False
    try:
        get_supabase_client().table("datablix_project_members").delete().eq(
            "project_id", project_id
        ).eq("member_email", str(email).strip().lower()).execute()
        return True
    except Exception:
        return False


def delete_cloud_project(project_id: str) -> tuple[bool, str]:
    """Permanently delete one cloud project and its membership rows.

    Only the authenticated project owner may perform this action.
    Other Datablix projects are not touched.
    """
    project_id = str(project_id or "").strip()
    if not project_id:
        return False, "No saved project is selected."

    if st.session_state.get(S_DEMO_MODE):
        return False, "Demo workspaces cannot be permanently deleted."

    client = get_supabase_client()
    if client is None:
        return False, "Cloud storage is not configured."

    email = current_user_email()
    if not email:
        return False, "Sign in before deleting a project."

    # Verify ownership from cloud state immediately before deletion instead of
    # trusting only the role cached in the Streamlit session.
    workspace_key = _secret_value("DATABLIX_WORKSPACE_KEY", "default")
    try:
        response = (
            client.table("datablix_project_state")
            .select("project_id,project_name,owner_email")
            .eq("workspace_key", workspace_key)
            .eq("project_id", project_id)
            .limit(1)
            .execute()
        )
        rows = list(response.data or [])
    except Exception:
        return False, "Datablix could not verify project ownership."

    if not rows:
        return False, "The saved project could not be found."

    owner_email = str(rows[0].get("owner_email", "") or "").strip().lower()
    if owner_email != email:
        return False, "Only the project owner can permanently delete this project."

    try:
        # Delete access rows first so this works whether or not the database
        # relationship is configured with cascading deletes.
        (
            client.table("datablix_project_members")
            .delete()
            .eq("project_id", project_id)
            .execute()
        )
        (
            client.table("datablix_project_state")
            .delete()
            .eq("workspace_key", workspace_key)
            .eq("project_id", project_id)
            .eq("owner_email", email)
            .execute()
        )
    except Exception:
        return False, "The project could not be deleted from cloud storage."

    return True, "Project permanently deleted."


def _json_safe(value):
    """Convert Streamlit and pandas state into JSON-safe structures."""
    if isinstance(value, pd.DataFrame):
        return {
            "__datablix_type__": "dataframe",
            "value": json.loads(value.to_json(orient="split", date_format="iso")),
        }
    if isinstance(value, pd.Series):
        return {
            "__datablix_type__": "series",
            "value": json.loads(value.to_json(date_format="iso")),
        }
    if isinstance(value, (datetime, date, pd.Timestamp)):
        return {"__datablix_type__": "datetime", "value": value.isoformat()}
    if isinstance(value, Path):
        return {"__datablix_type__": "path", "value": str(value)}
    if isinstance(value, (bytes, bytearray)):
        return {
            "__datablix_type__": "bytes",
            "value": base64.b64encode(bytes(value)).decode("ascii"),
        }
    if isinstance(value, tuple):
        return {"__datablix_type__": "tuple", "value": [_json_safe(v) for v in value]}
    if isinstance(value, set):
        return {"__datablix_type__": "set", "value": [_json_safe(v) for v in value]}
    if isinstance(value, dict):
        return {str(k): _json_safe(v) for k, v in value.items()}
    if isinstance(value, list):
        return [_json_safe(v) for v in value]
    if value is pd.NA:
        return None
    try:
        if pd.isna(value):
            return None
    except (TypeError, ValueError):
        pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            pass
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def _from_json_safe(value):
    if isinstance(value, list):
        return [_from_json_safe(v) for v in value]
    if not isinstance(value, dict):
        return value
    marker = value.get("__datablix_type__")
    if marker == "dataframe":
        payload = value.get("value", {})
        return pd.DataFrame(
            data=payload.get("data", []),
            columns=payload.get("columns", []),
            index=payload.get("index", None),
        )
    if marker == "series":
        return pd.Series(value.get("value", {}))
    if marker == "datetime":
        raw = value.get("value", "")
        try:
            return datetime.fromisoformat(raw)
        except ValueError:
            return raw
    if marker == "path":
        return Path(value.get("value", ""))
    if marker == "bytes":
        try:
            return base64.b64decode(value.get("value", ""))
        except Exception:
            return b""
    if marker == "tuple":
        return tuple(_from_json_safe(v) for v in value.get("value", []))
    if marker == "set":
        return set(_from_json_safe(v) for v in value.get("value", []))
    return {k: _from_json_safe(v) for k, v in value.items()}


def _current_state_payload() -> dict:
    state = {
        key: st.session_state[key]
        for key in AUTOSAVE_STATE_KEYS
        if key in st.session_state
    }
    return {
        "schema_version": 1,
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "state": _json_safe(state),
    }


def _state_hash(payload: dict) -> str:
    stable = json.dumps(payload.get("state", {}), sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(stable.encode("utf-8")).hexdigest()


def claim_legacy_projects() -> None:
    """Attach projects created before multi-user support to the first signed-in owner."""
    client = get_supabase_client()
    email = current_user_email()
    if client is None or not email or st.session_state.get("db_legacy_claim_checked"):
        return
    st.session_state["db_legacy_claim_checked"] = True
    workspace_key = _secret_value("DATABLIX_WORKSPACE_KEY", "default")
    try:
        client.table("datablix_project_state").update(
            {"owner_email": email}
        ).eq("workspace_key", workspace_key).is_("owner_email", "null").execute()
    except Exception:
        pass


def list_cloud_projects() -> list[dict]:
    client = get_supabase_client()
    email = current_user_email()
    if client is None or not email:
        return []
    workspace_key = _secret_value("DATABLIX_WORKSPACE_KEY", "default")
    claim_legacy_projects()
    try:
        owned_response = (
            client.table("datablix_project_state")
            .select("project_id,project_name,updated_at,owner_email")
            .eq("workspace_key", workspace_key)
            .eq("owner_email", email)
            .order("updated_at", desc=True)
            .execute()
        )
        membership_response = (
            client.table("datablix_project_members")
            .select("project_id,role")
            .eq("member_email", email)
            .execute()
        )
        membership_roles = {
            str(row.get("project_id", "")): str(row.get("role", "viewer"))
            for row in list(membership_response.data or [])
        }
        shared_rows = []
        if membership_roles:
            shared_response = (
                client.table("datablix_project_state")
                .select("project_id,project_name,updated_at,owner_email")
                .eq("workspace_key", workspace_key)
                .in_("project_id", list(membership_roles))
                .execute()
            )
            shared_rows = list(shared_response.data or [])
        combined = {}
        for row in list(owned_response.data or []):
            row["role"] = "owner"
            combined[str(row.get("project_id", ""))] = row
        for row in shared_rows:
            pid = str(row.get("project_id", ""))
            row["role"] = membership_roles.get(pid, "viewer")
            combined[pid] = row
        return sorted(combined.values(), key=lambda r: str(r.get("updated_at", "")), reverse=True)
    except Exception:
        return []

def restore_cloud_project(project_id: str | None = None) -> bool:
    """Restore a selected cloud project, or the most recently updated one."""
    if S_WORKING in st.session_state or st.session_state.get(S_SKIP_CLOUD_RESTORE):
        return False
    client = get_supabase_client()
    if client is None:
        return False
    workspace_key = _secret_value("DATABLIX_WORKSPACE_KEY", "default")
    email = current_user_email()
    try:
        accessible = {str(row.get("project_id", "")): str(row.get("role", "")) for row in list_cloud_projects()}
        if project_id and str(project_id) not in accessible:
            return False
        if not project_id and not accessible:
            return False
        query = (
            client.table("datablix_project_state")
            .select("project_id,project_name,state_json,state_hash,updated_at,owner_email")
            .eq("workspace_key", workspace_key)
            .in_("project_id", list(accessible))
        )
        if project_id:
            query = query.eq("project_id", project_id).limit(1)
        else:
            query = query.order("updated_at", desc=True).limit(1)
        response = query.execute()
        rows = list(response.data or [])
        if not rows:
            return False
        row = rows[0]
        payload = row.get("state_json") or {}
        state = _from_json_safe(payload.get("state", {}))
        if not isinstance(state, dict) or S_WORKING not in state:
            return False
        for key, value in state.items():
            if key in AUTOSAVE_STATE_KEYS:
                st.session_state[key] = value
        st.session_state[S_CLOUD_PROJECT_ID] = str(row.get("project_id", ""))
        st.session_state[S_PROJECT_ROLE] = accessible.get(str(row.get("project_id", "")), "viewer")
        st.session_state[S_CLOUD_STATE_HASH] = str(row.get("state_hash", ""))
        st.session_state[S_FLASH] = "Your project was restored from permanent cloud storage."
        return True
    except Exception:
        return False


def save_cloud_project() -> bool:
    """Upsert the active project into Supabase only when its state changed."""
    if S_WORKING not in st.session_state:
        return False
    client = get_supabase_client()
    if client is None or not user_is_authenticated():
        return False
    project_id = str(st.session_state.get(S_CLOUD_PROJECT_ID, "")).strip()
    if not project_id:
        project_id = str(uuid.uuid4())
        st.session_state[S_CLOUD_PROJECT_ID] = project_id
        st.session_state[S_PROJECT_ROLE] = "owner"
    elif not user_can_edit_project(project_id):
        return False
    payload = _current_state_payload()
    fingerprint = _state_hash(payload)
    if fingerprint == st.session_state.get(S_CLOUD_STATE_HASH):
        return True
    workspace_key = _secret_value("DATABLIX_WORKSPACE_KEY", "default")
    project_name = str(st.session_state.get(S_PROJECT_NAME, "Datablix project")).strip() or "Datablix project"
    owner_email = current_user_email()
    if st.session_state.get(S_PROJECT_ROLE) != "owner":
        try:
            existing = client.table("datablix_project_state").select("owner_email").eq("project_id", project_id).limit(1).execute()
            existing_rows = list(existing.data or [])
            if existing_rows:
                owner_email = str(existing_rows[0].get("owner_email", owner_email))
        except Exception:
            pass
    row = {
        "workspace_key": workspace_key,
        "project_id": project_id,
        "owner_email": owner_email,
        "project_name": project_name,
        "state_json": payload,
        "state_hash": fingerprint,
        "updated_at": datetime.now().astimezone().isoformat(),
    }
    try:
        (
            client.table("datablix_project_state")
            .upsert(row, on_conflict="workspace_key,project_id")
            .execute()
        )
        st.session_state[S_CLOUD_STATE_HASH] = fingerprint
        return True
    except Exception:
        return False


def restore_autosaved_project() -> bool:
    """Restore cloud state first, then use the local refresh fallback."""
    if st.session_state.get(S_DEMO_MODE) or not user_is_authenticated():
        return False
    if restore_cloud_project():
        return True
    if S_WORKING in st.session_state or not _autosave_file().exists():
        return False
    try:
        payload = pickle.loads(_autosave_file().read_bytes())
        if not isinstance(payload, dict):
            return False
        state = payload.get("state", {})
        if not isinstance(state, dict) or S_WORKING not in state:
            return False
        for key, value in state.items():
            if key in AUTOSAVE_STATE_KEYS:
                st.session_state[key] = value
        st.session_state[S_FLASH] = "Your last local project was restored automatically."
        return True
    except (OSError, pickle.PickleError, EOFError, AttributeError, ValueError, TypeError):
        return False


def autosave_current_project() -> bool:
    """Save to permanent cloud storage and retain a local refresh fallback."""
    if st.session_state.get(S_DEMO_MODE) or not user_is_authenticated():
        return False
    if S_WORKING not in st.session_state:
        return False
    cloud_saved = save_cloud_project()
    state = {
        key: st.session_state[key]
        for key in AUTOSAVE_STATE_KEYS
        if key in st.session_state
    }
    payload = {
        "saved_at": datetime.now().isoformat(timespec="seconds"),
        "state": state,
    }
    local_saved = False
    try:
        AUTOSAVE_DIRECTORY.mkdir(parents=True, exist_ok=True)
        temporary = _autosave_file().with_suffix(".tmp")
        temporary.write_bytes(pickle.dumps(payload, protocol=pickle.HIGHEST_PROTOCOL))
        os.replace(temporary, _autosave_file())
        local_saved = True
    except (OSError, pickle.PickleError, TypeError, AttributeError):
        pass
    return cloud_saved or local_saved


def clear_autosaved_project() -> None:
    """Clear only this user's temporary local copy; cloud projects remain available."""
    try:
        _autosave_file().unlink(missing_ok=True)
    except OSError:
        pass


# =========================================================
# Helpers
# =========================================================

def render_brand_header():
    """Render the complete horizontal Datablix logo without clipping."""
    logo_path = Path("datablix_logo.png")

    if logo_path.exists():
        encoded_logo = base64.b64encode(logo_path.read_bytes()).decode("utf-8")

        st.html(f"""
        <div style="
            width:100%;
            display:flex;
            justify-content:center;
            align-items:center;
            overflow:visible;
            padding:0.35rem 0 0.75rem;
        ">
            <img
                src="data:image/png;base64,{encoded_logo}"
                alt="Datablix logo"
                style="
                    display:block;
                    width:420px;
                    max-width:100%;
                    height:auto !important;
                    max-height:none !important;
                    object-fit:contain !important;
                    object-position:center center;
                    margin:0 auto;
                    padding:0;
                    border:0;
                    clip-path:none !important;
                    overflow:visible;
                "
            >
        </div>
        """)
    else:
        st.html("""
        <div class="db-brand-name">Datablix</div>
        """)

    st.html("""
    <div class="db-brand">
        <div class="db-tag">Turn your rental property research into structured, review-ready listings.</div>
        <div class="db-subtag">Collect public information, verify key details, preserve additional findings, and prepare consistent records for review or export.</div>
    </div>
    """)


def safe_text(value, default=""):
    """Convert a scalar value to text without evaluating pd.NA as a boolean."""
    if isinstance(value, (pd.Series, pd.DataFrame)):
        return default
    if value is None or value is pd.NA:
        return default
    try:
        if pd.api.types.is_scalar(value) and pd.isna(value):
            return default
    except Exception:
        pass
    return str(value).strip()


def norm_header(value):
    return re.sub(r"[^a-z0-9]+", "", safe_text(value).lower())


def norm_scalar(value):
    return safe_text(value).lower()


def is_unresolved(value):
    return norm_scalar(value) in UNRESOLVED


def coalesce_duplicate_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Return one column per label, keeping the first resolved value across duplicates.

    Some saved/imported workbooks can contain duplicate labels after schema mapping.
    Pandas returns a DataFrame rather than a Series for df["column"] in that case,
    which breaks string operations such as .str.strip(). This helper repairs that
    condition before normalization.
    """
    if not isinstance(df, pd.DataFrame):
        return df
    out = df.copy()
    if not out.columns.duplicated().any():
        return out

    result = pd.DataFrame(index=out.index)
    seen = []
    for column in out.columns:
        if column not in seen:
            seen.append(column)

    for column in seen:
        positions = [i for i, label in enumerate(out.columns) if label == column]
        combined = out.iloc[:, positions[0]].copy()

        for pos in positions[1:]:
            candidate = out.iloc[:, pos]
            blank_mask = (
                combined.isna()
                | combined.astype("string").fillna("").str.strip().eq("")
            )
            combined = combined.where(~blank_mask, candidate)

        result[column] = combined

    return result


def unresolved_mask(series):
    text = series.astype("string").fillna("").str.strip().str.lower()
    return series.isna() | text.isin(UNRESOLVED)


def resolved(series):
    out = series.copy()
    out.loc[unresolved_mask(out)] = pd.NA
    return out


def prepare_data(df):
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]
    out = coalesce_duplicate_columns(out)
    return out.replace(r"^\s*$", pd.NA, regex=True)


def display_values(series, blank="Blank"):
    return series.astype("string").fillna(blank).str.strip().replace("", blank)


def csv_bytes(df):
    return df.to_csv(index=False).encode("utf-8-sig")


def dataframe_content_fingerprint(df: pd.DataFrame) -> str:
    """Return a deterministic fingerprint for a DataFrame's current contents.

    Reporting snapshots use this to detect when the active Starting Data baseline
    has changed.  A snapshot may intentionally preserve older research values, but
    it must never preserve counts from a different source workbook.
    """
    if not isinstance(df, pd.DataFrame):
        return ""
    if df.empty:
        return "empty"

    stable = coalesce_duplicate_columns(df.copy())
    stable.columns = [str(column).strip() for column in stable.columns]
    stable = stable.reindex(sorted(stable.columns), axis=1)
    for column in stable.columns:
        stable[column] = stable[column].astype("string").fillna("").str.strip()
    payload = stable.to_csv(index=False, lineterminator="\n")
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def safe_filename(name):
    stem = name.rsplit(".", 1)[0].strip()
    return "".join(c if c.isalnum() or c in "-_" else "_" for c in stem) or "datablix"


def _excel_display_value(value):
    return "" if is_unresolved(value) else str(value).strip()


def _write_listing_blocks_sheet(ws, listings):
    """Write one apartment building at a time in the supplied two-column layout."""
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.merge_cells("A1:B1")
    ws["A1"] = "Create a listing for each Apartment Building as per sample below"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[1].height = 30

    row_number = 3
    required = LISTING_COLUMNS
    additional = [label for label, _ in LISTING_ADDITIONAL_FIELD_MAP]

    for listing_number, (_, record) in enumerate(listings.iterrows(), start=1):
        name = _excel_display_value(record.get("Apartment Building Name")) or f"Listing {listing_number}"
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=2)
        title_cell = ws.cell(
            row=row_number,
            column=1,
            value=f"Apartment Building {listing_number}: {name}",
        )
        title_cell.font = Font(bold=True, size=12)
        title_cell.alignment = Alignment(wrap_text=True, vertical="top")
        row_number += 1

        for field_name in required:
            value = _excel_display_value(record.get(field_name))
            field_cell = ws.cell(row=row_number, column=1, value=field_name)
            value_cell = ws.cell(row=row_number, column=2, value=value)
            field_cell.font = Font(bold=True)
            field_cell.border = border
            value_cell.border = border
            field_cell.alignment = Alignment(wrap_text=True, vertical="top")
            value_cell.alignment = Alignment(wrap_text=True, vertical="top")
            if field_name == "Email Contact" and value:
                value_cell.hyperlink = f"mailto:{value}"
                value_cell.style = "Hyperlink"
            elif field_name == "WebSite" and value.startswith(("http://", "https://")):
                value_cell.hyperlink = value
                value_cell.style = "Hyperlink"
            row_number += 1

        populated_additional = [
            field_name
            for field_name in additional
            if _excel_display_value(record.get(field_name))
        ]
        if populated_additional:
            ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=2)
            section_cell = ws.cell(
                row=row_number,
                column=1,
                value="Additional information and research reference",
            )
            section_cell.font = Font(bold=True)
            section_cell.alignment = Alignment(wrap_text=True, vertical="top")
            row_number += 1
            for field_name in populated_additional:
                value = _excel_display_value(record.get(field_name))
                field_cell = ws.cell(row=row_number, column=1, value=field_name)
                value_cell = ws.cell(row=row_number, column=2, value=value)
                field_cell.font = Font(bold=True)
                field_cell.border = border
                value_cell.border = border
                field_cell.alignment = Alignment(wrap_text=True, vertical="top")
                value_cell.alignment = Alignment(wrap_text=True, vertical="top")
                if field_name == "Official Source URL" and value.startswith(("http://", "https://")):
                    value_cell.hyperlink = value
                    value_cell.style = "Hyperlink"
                elif field_name == "Secondary Email" and value:
                    value_cell.hyperlink = f"mailto:{value}"
                    value_cell.style = "Hyperlink"
                row_number += 1

        row_number += 2

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 75
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def excel_bytes(sheets):
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        used = set()
        for requested, df in sheets.items():
            name = re.sub(r"[:\\/?*\[\]]", " ", str(requested))
            name = re.sub(r"\s+", " ", name).strip()[:31] or "Sheet"
            base, n = name, 2
            while name in used:
                suffix = f" {n}"
                name = f"{base[:31-len(suffix)]}{suffix}"
                n += 1
            used.add(name)

            if requested == "Building Listings":
                ws = writer.book.create_sheet(title=name)
                _write_listing_blocks_sheet(ws, df)
                continue

            df.to_excel(writer, sheet_name=name, index=False)
            ws = writer.book[name]
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for cells in ws.columns:
                lengths = [len(str(c.value)) for c in cells[:101] if c.value is not None]
                ws.column_dimensions[cells[0].column_letter].width = min(
                    max(lengths + [12]) + 2,
                    42,
                )
    output.seek(0)
    return output.getvalue()


def canonical_province(value):
    if is_unresolved(value):
        return pd.NA
    text = re.sub(r"\s+", " ", str(value)).strip()
    return PROVINCES.get(text.lower(), (text, text))[0]


def province_code(value):
    if is_unresolved(value):
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    return PROVINCES.get(text.lower(), (text, text.upper() if len(text) == 2 else text))[1]


def postal_code(value):
    if is_unresolved(value):
        return pd.NA
    text = re.sub(r"\s+", "", str(value)).upper()
    return f"{text[:3]} {text[3:]}" if re.fullmatch(r"[A-Z]\d[A-Z]\d[A-Z]\d", text) else str(value).strip().upper()


def parse_combined_location(value):
    if is_unresolved(value):
        return pd.NA, pd.NA, pd.NA
    text = re.sub(r"\s+", " ", str(value)).strip(" ,")
    match = re.search(r"\b([ABCEGHJ-NPRSTVXY]\d[ABCEGHJ-NPRSTV-Z][ -]?\d[ABCEGHJ-NPRSTV-Z]\d)\b", text, re.I)
    pc = pd.NA
    if match:
        pc = postal_code(match.group(1))
        text = (text[:match.start()] + text[match.end():]).strip(" ,")
    province_pattern = "|".join(re.escape(k) for k in sorted(PROVINCES, key=len, reverse=True))
    pm = re.search(rf"(?:,\s*|\s+)({province_pattern})$", text, re.I)
    province = pd.NA
    city = text
    if pm:
        province = canonical_province(pm.group(1))
        city = text[:pm.start()].strip(" ,")
    return city or pd.NA, province, pc


def formatted_location(row):
    city = "" if is_unresolved(row.get("City")) else str(row.get("City")).strip()
    province = province_code(row.get("Province"))
    pc = "" if is_unresolved(row.get("Postal Code")) else str(postal_code(row.get("Postal Code")))
    tail = " ".join(v for v in [province, pc] if v)
    return f"{city}, {tail}" if city and tail else city or tail or pd.NA




def _address_for_geocoding(row) -> str:
    """Build a physical-address query without mixing in PO Box information."""
    parts = []
    for field in ["Street Address", "City", "Province", "Postal Code", "Country"]:
        value = safe_text(row.get(field, ""))
        if value and not is_unresolved(value):
            parts.append(value)
    return ", ".join(parts)


def _component_value(result: dict, *wanted_types: str) -> str:
    wanted = set(wanted_types)
    for component in result.get("address_components", []) or []:
        if wanted.intersection(component.get("types", []) or []):
            return safe_text(component.get("long_name", ""))
    return ""


def _point_in_ring(longitude: float, latitude: float, ring) -> bool:
    """Ray-casting point-in-polygon test for one GeoJSON linear ring."""
    inside = False
    if not isinstance(ring, list) or len(ring) < 3:
        return False
    j = len(ring) - 1
    for i in range(len(ring)):
        try:
            xi, yi = float(ring[i][0]), float(ring[i][1])
            xj, yj = float(ring[j][0]), float(ring[j][1])
        except (TypeError, ValueError, IndexError):
            j = i
            continue
        crosses = ((yi > latitude) != (yj > latitude)) and (
            longitude < (xj - xi) * (latitude - yi) / ((yj - yi) or 1e-15) + xi
        )
        if crosses:
            inside = not inside
        j = i
    return inside


def _point_in_polygon(longitude: float, latitude: float, polygon) -> bool:
    if not polygon or not _point_in_ring(longitude, latitude, polygon[0]):
        return False
    # A point in a hole is outside the polygon.
    return not any(_point_in_ring(longitude, latitude, hole) for hole in polygon[1:])


def _point_in_geojson(longitude: float, latitude: float, payload) -> bool:
    """Test a point against Polygon/MultiPolygon features in a GeoJSON object."""
    if not isinstance(payload, dict):
        return False
    object_type = safe_text(payload.get("type", ""))
    if object_type == "FeatureCollection":
        return any(
            _point_in_geojson(longitude, latitude, feature)
            for feature in payload.get("features", []) or []
        )
    if object_type == "Feature":
        return _point_in_geojson(longitude, latitude, payload.get("geometry") or {})
    coordinates = payload.get("coordinates") or []
    if object_type == "Polygon":
        return _point_in_polygon(longitude, latitude, coordinates)
    if object_type == "MultiPolygon":
        return any(_point_in_polygon(longitude, latitude, polygon) for polygon in coordinates)
    if object_type == "GeometryCollection":
        return any(
            _point_in_geojson(longitude, latitude, geometry)
            for geometry in payload.get("geometries", []) or []
        )
    return False


@st.cache_data(ttl=86400, show_spinner=False)
def _load_boundary_geojson(url: str) -> dict:
    """Load a configured City of Ottawa municipal-boundary GeoJSON file."""
    if not safe_text(url):
        return {}
    request = Request(
        safe_text(url),
        headers={"User-Agent": "Datablix/1.0 municipal-scope-validator"},
    )
    try:
        with urlopen(request, timeout=25) as response:
            payload = json.loads(response.read().decode("utf-8"))
        return payload if isinstance(payload, dict) else {}
    except Exception:
        return {}


@st.cache_data(ttl=2592000, show_spinner=False)
def _google_geocode_address(address: str) -> dict:
    """Geocode one physical address when GOOGLE_MAPS_API_KEY is configured."""
    api_key = _secret_value("GOOGLE_MAPS_API_KEY")
    if not api_key or not safe_text(address):
        return {}
    query = urlencode(
        {
            "address": address,
            "components": "country:CA|administrative_area:ON",
            "key": api_key,
        }
    )
    request = Request(
        f"https://maps.googleapis.com/maps/api/geocode/json?{query}",
        headers={"User-Agent": "Datablix/1.0 address-validator"},
    )
    try:
        with urlopen(request, timeout=25) as response:
            payload = json.loads(response.read().decode("utf-8"))
    except Exception:
        return {}
    if payload.get("status") != "OK" or not payload.get("results"):
        return {}
    result = payload["results"][0]
    return result if isinstance(result, dict) else {}


def _geographic_result_fields(result: dict) -> dict:
    """Convert a Google geocode result into auditable Ottawa-scope fields."""
    geometry = result.get("geometry") or {}
    location = geometry.get("location") or {}
    try:
        latitude = float(location.get("lat"))
        longitude = float(location.get("lng"))
    except (TypeError, ValueError):
        return {}

    locality = _component_value(result, "locality", "postal_town")
    admin3 = _component_value(result, "administrative_area_level_3")
    admin2 = _component_value(result, "administrative_area_level_2")
    province = _component_value(result, "administrative_area_level_1")
    country = _component_value(result, "country")
    municipality = locality or admin3 or admin2
    normalized_components = {
        norm_header(value)
        for value in [locality, admin3, admin2]
        if safe_text(value)
    }

    boundary_url = _secret_value("OTTAWA_BOUNDARY_GEOJSON_URL")
    boundary = _load_boundary_geojson(boundary_url) if boundary_url else {}
    method = "Google address components"
    if boundary:
        inside = _point_in_geojson(longitude, latitude, boundary)
        status = "Inside City of Ottawa" if inside else "Outside City of Ottawa"
        method = "Configured City of Ottawa boundary point-in-polygon check"
    elif normalized_components & {norm_header(v) for v in OTTAWA_MUNICIPAL_LOCALITIES}:
        status = "Inside City of Ottawa"
    elif normalized_components & {norm_header(v) for v in OUT_OF_SCOPE_NEARBY_LOCALITIES}:
        status = "Outside City of Ottawa"
    elif province and norm_header(province) not in {"ontario", "on"}:
        status = "Outside City of Ottawa"
    elif country and norm_header(country) not in {"canada", "ca"}:
        status = "Outside City of Ottawa"
    elif locality or admin3:
        # An explicit non-Ottawa municipality is normally outside the municipal
        # boundary, but remains reviewable when Google returns unusual components.
        status = "Outside City of Ottawa"
    else:
        status = "Needs Geographic Review"

    partial_match = bool(result.get("partial_match"))
    location_type = safe_text(geometry.get("location_type", "")) or "Unknown"
    if partial_match:
        confidence = "Low"
    elif location_type in {"ROOFTOP", "RANGE_INTERPOLATED"} and status != "Needs Geographic Review":
        confidence = "High"
    elif status != "Needs Geographic Review":
        confidence = "Medium"
    else:
        confidence = "Low"

    evidence_parts = [
        f"Provider: Google Maps Geocoding API",
        f"Formatted address: {safe_text(result.get('formatted_address', ''))}",
        f"Location type: {location_type}",
        f"Partial match: {'Yes' if partial_match else 'No'}",
        f"Municipality components: {', '.join(v for v in [locality, admin3, admin2] if v) or 'Not returned'}",
        f"Scope method: {method}",
    ]
    place_id = safe_text(result.get("place_id", ""))
    if place_id:
        evidence_parts.append(f"Place ID: {place_id}")

    return {
        "Latitude": latitude,
        "Longitude": longitude,
        "Geocoded Municipality": municipality,
        "Geographic Scope Status": status,
        "Geographic Evidence": "; ".join(evidence_parts),
        "Geographic Confidence": confidence,
    }


def enrich_geographic_scope(df: pd.DataFrame, *, max_requests: int = 100) -> pd.DataFrame:
    """Geocode unresolved physical addresses and classify City of Ottawa scope.

    The function is intentionally opt-in through GOOGLE_MAPS_API_KEY. It never
    geocodes PO Box or mailing-address fields and never overwrites the physical
    street address supplied by the source.
    """
    out = normalize_workflow(df)
    if not _secret_value("GOOGLE_MAPS_API_KEY") or out.empty:
        return out

    requests_used = 0
    for index, row in out.iterrows():
        current_status = safe_text(row.get("Geographic Scope Status", ""))
        if current_status not in {"", "Not Checked", "Needs Geographic Review"}:
            continue
        address = _address_for_geocoding(row)
        if not address or is_unresolved(row.get("Street Address", "")):
            continue
        if requests_used >= max(1, int(max_requests)):
            break
        result = _google_geocode_address(address)
        requests_used += 1
        fields = _geographic_result_fields(result) if result else {}
        if not fields:
            out.at[index, "Geographic Scope Status"] = "Needs Geographic Review"
            out.at[index, "Geographic Confidence"] = "Low"
            out.at[index, "Geographic Evidence"] = _append_note(
                out.at[index, "Geographic Evidence"],
                f"Google Maps geocoding did not return a usable exact result for: {address}",
            )
            continue
        for field, value in fields.items():
            out.at[index, field] = value
        if fields["Geographic Scope Status"] == "Outside City of Ottawa":
            out.at[index, "Verification Status"] = "Needs Review"
            out.at[index, "Reviewer Notes"] = _append_note(
                out.at[index, "Reviewer Notes"],
                "Geographic validation indicates that the physical property is outside the City of Ottawa municipal boundary. Review before export.",
            )
    return normalize_workflow(out)


def normalize_choice(series, choices, default, aliases=None):
    mapping = {v.lower(): v for v in choices}
    mapping.update(aliases or {})
    return series.astype("string").fillna("").str.strip().str.lower().map(mapping).fillna(default)


def synchronize_missing_information(df):
    """Keep Missing Information aligned with current gaps using column masks.

    This avoids repeated scalar ``.at`` access for every row/field combination,
    which was noticeable on larger project datasets during every Streamlit rerun.
    """
    out = df.copy()
    if out.empty:
        out["Missing Information"] = pd.Series(index=out.index, dtype="object")
        return out

    # Build each field's unresolved mask once, then assemble the row labels.
    gap_masks = {field: unresolved_mask(out[field]).to_numpy() for field in TARGET_FIELDS}
    missing_values = []
    for position in range(len(out)):
        missing_values.append(
            ", ".join(field for field in TARGET_FIELDS if gap_masks[field][position])
        )
    out["Missing Information"] = pd.Series(missing_values, index=out.index, dtype="object")
    return out


def normalize_workflow(df):
    out = coalesce_duplicate_columns(df.copy())
    for c in INTERNAL_COLUMNS:
        if c not in out.columns:
            out[c] = pd.NA

    # Backward-compatible migration for projects saved before Property Type was
    # separated from Building Classification. Recover only recognized property-form
    # labels from the old mixed field, then make Building Classification height-only.
    legacy_property_type = out["Building Classification"].apply(
        property_type_from_legacy_classification
    )
    migrated_property_type = pd.Series(pd.NA, index=out.index, dtype="object")
    for idx in out.index:
        migrated_property_type.loc[idx] = merge_property_type_values(
            out.at[idx, "Property Type"],
            legacy_property_type.loc[idx],
        )
    out["Property Type"] = migrated_property_type
    out["Building Classification"] = derive_height_classification_from_storeys(
        out["Number of Storeys"]
    )

    out["Research Status"] = normalize_choice(out["Research Status"], RESEARCH_STATUSES, "Not Started", STATUS_ALIASES["Research Status"])
    out["Source Status"] = normalize_choice(out["Source Status"], SOURCE_STATUSES, "Not Checked", STATUS_ALIASES["Source Status"])
    out["Verification Status"] = normalize_choice(out["Verification Status"], VERIFICATION_STATUSES, "Not Reviewed", STATUS_ALIASES["Verification Status"])
    out["Record Decision"] = normalize_choice(out["Record Decision"], RECORD_DECISIONS, "Undecided", STATUS_ALIASES["Record Decision"])
    # Migrate the earlier wording without losing saved project choices.
    out["Directory Discovery Status"] = out["Directory Discovery Status"].replace(
        {"Existing Client Record": "Existing Source Record"}
    )
    out["Directory Discovery Status"] = normalize_choice(
        out["Directory Discovery Status"],
        DISCOVERY_STATUSES,
        "Needs Classification",
    )
    # Track whether discovery classification came from Datablix or a human reviewer.
    # Legacy/imported projects default to Automatic until a reviewer explicitly
    # changes Directory Discovery Status in the review table.
    out["Discovery Status Source"] = normalize_choice(
        out["Discovery Status Source"],
        DISCOVERY_STATUS_SOURCES,
        "Automatic",
    )
    out["Directory Entry Status"] = normalize_choice(
        out["Directory Entry Status"],
        DIRECTORY_ENTRY_STATUSES,
        "Not Entered",
    )
    out["PO Box Search Status"] = normalize_choice(
        out["PO Box Search Status"],
        PO_BOX_SEARCH_STATUSES,
        "Not Checked",
    )
    out["Apartment Count Search Status"] = normalize_choice(
        out["Apartment Count Search Status"],
        COUNT_SEARCH_STATUSES,
        "Not Checked",
    )
    out["Apartment Count Confidence"] = normalize_choice(
        out["Apartment Count Confidence"],
        EVIDENCE_CONFIDENCE_LEVELS,
        "Not Checked",
    )
    out["Storey Count Search Status"] = normalize_choice(
        out["Storey Count Search Status"],
        COUNT_SEARCH_STATUSES,
        "Not Checked",
    )
    out["Storey Count Confidence"] = normalize_choice(
        out["Storey Count Confidence"],
        EVIDENCE_CONFIDENCE_LEVELS,
        "Not Checked",
    )
    out["PO Box Confidence"] = normalize_choice(
        out["PO Box Confidence"],
        EVIDENCE_CONFIDENCE_LEVELS,
        "Not Checked",
    )
    out["Geographic Scope Status"] = normalize_choice(
        out["Geographic Scope Status"],
        GEOGRAPHIC_SCOPE_STATUSES,
        "Not Checked",
    )
    out["Geographic Confidence"] = normalize_choice(
        out["Geographic Confidence"],
        EVIDENCE_CONFIDENCE_LEVELS,
        "Not Checked",
    )
    out["Rental Availability Status"] = normalize_choice(
        out["Rental Availability Status"],
        RENTAL_AVAILABILITY_STATUSES,
        "Not Checked",
        RENTAL_AVAILABILITY_ALIASES,
    )
    out["PO Box Province"] = out["PO Box Province"].apply(canonical_province)
    out["PO Box Postal Code"] = out["PO Box Postal Code"].apply(postal_code)
    out = synchronize_missing_information(out)
    for c in ["Researcher", "Missing Information", "Reviewer Notes"]:
        out[c] = out[c].fillna("").astype(str)
    return out


# -------------------------------------------------------------------------
# Robust Starting Data identity matching
# -------------------------------------------------------------------------
# The research CSV and the project Starting Data often describe the same regional
# property using slightly different address text.  Discovery status therefore
# must never depend on one literal string key.  The helpers below normalize
# civic numbers, compound addresses, street types, directions, Ottawa municipal
# locality labels, other municipality names, postal codes and URLs before Datablix
# decides that a property is new.

_STREET_TYPE_ALIASES = {
    "street": "st", "st": "st",
    "avenue": "ave", "ave": "ave",
    "road": "rd", "rd": "rd",
    "boulevard": "blvd", "blvd": "blvd",
    "drive": "dr", "dr": "dr",
    "lane": "ln", "ln": "ln",
    "court": "ct", "ct": "ct",
    "place": "pl", "pl": "pl",
    "crescent": "cres", "cres": "cres",
    "terrace": "terr", "terr": "terr",
    "parkway": "pkwy", "pkwy": "pkwy",
    "way": "way", "trail": "trl", "trl": "trl",
    "circle": "cir", "cir": "cir",
    "highway": "hwy", "hwy": "hwy",
}

# Correct the accidental comma in the static mapping above defensively below.
_STREET_TYPE_ALIASES["pkwy"] = "pkwy"
_STREET_TYPE_ALIASES["parkway"] = "pkwy"

_DIRECTION_ALIASES = {
    "north": "n", "n": "n",
    "south": "s", "s": "s",
    "east": "e", "e": "e",
    "west": "w", "w": "w",
}

_OTTawa_LOCALITY_KEYS = {
    norm_header(value) for value in OTTAWA_LOCALITY_LABELS
} | {"ottawa"}


def _fold_identity_text(value) -> str:
    """Lowercase and simplify punctuation for identity comparison only."""
    text = safe_text(value).lower()
    text = (
        text.replace("–", "-")
        .replace("—", "-")
        .replace("−", "-")
        .replace("’", "'")
    )
    text = re.sub(r"[.]", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _canonical_postal(value) -> str:
    return re.sub(r"[^A-Z0-9]", "", safe_text(value).upper())


def _canonical_city(value) -> str:
    key = norm_header(value)
    if key in _OTTawa_LOCALITY_KEYS:
        return "ottawa"
    return key


def _canonical_phone(value) -> str:
    digits = re.sub(r"\D+", "", safe_text(value))
    if len(digits) > 10:
        digits = digits[-10:]
    return digits


def _canonical_url(value) -> str:
    raw = safe_text(value)
    if not raw:
        return ""
    try:
        parsed = urlparse(raw if "://" in raw else f"https://{raw}")
        host = (parsed.netloc or "").lower().removeprefix("www.")
        path = re.sub(r"/+", "/", parsed.path or "/").rstrip("/")
        if not host:
            return norm_header(raw)
        return f"{host}{path}"
    except Exception:
        return norm_header(raw)


def _url_hostname(value) -> str:
    """Return a normalized hostname without treating a subdomain as a company."""
    raw = safe_text(value)
    if not raw:
        return ""
    try:
        parsed = urlparse(raw if "://" in raw else f"https://{raw}")
        return (parsed.hostname or "").lower().strip(".").removeprefix("www.")
    except Exception:
        return ""


def _is_official_property_subdomain(value, company_website) -> bool:
    """Return True when a URL is an official subdomain of the selected company.

    Example: ``wildwood.milyservice.com`` is a property microsite belonging to
    the company rooted at ``milyservice.com``. It is a property source, not a
    second company.
    """
    candidate_host = _url_hostname(value)
    company_host = _url_hostname(company_website)
    candidate_root = _company_domain_key(value)
    company_root = _company_domain_key(company_website)
    return bool(
        candidate_host
        and company_host
        and candidate_root
        and company_root
        and candidate_root == company_root
        and candidate_host != company_host
    )


def normalize_official_website_roles(
    df: pd.DataFrame,
    company_website: str,
) -> pd.DataFrame:
    """Keep corporate and property URLs in their correct Datablix fields.

    The selected company's registered website remains ``Company Website``. An
    official subdomain under the same registrable root domain is retained as
    ``Property Website`` and never creates a separate company identity. The
    exact page used as evidence remains available in ``Source URL``.
    """
    if not isinstance(df, pd.DataFrame) or df.empty:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()

    out = df.copy()
    for column in [
        "Company Website", "Property Website", "Website", "Source URL",
        "Reviewer Notes",
    ]:
        if column not in out.columns:
            out[column] = pd.NA

    canonical_company_website = safe_text(company_website)
    company_root = _company_domain_key(canonical_company_website)

    for index, row in out.iterrows():
        property_website = safe_text(row.get("Property Website", ""))
        detected_microsite = ""

        # AI tools sometimes put a property microsite in Company Website. Move
        # that URL to the property field before restoring the selected company's
        # canonical corporate website.
        current_company_url = safe_text(row.get("Company Website", ""))
        if (
            not property_website
            and _is_official_property_subdomain(
                current_company_url, canonical_company_website
            )
        ):
            property_website = current_company_url
            detected_microsite = current_company_url

        if not property_website:
            for column in ["Website", "Source URL"]:
                candidate = safe_text(row.get(column, ""))
                if _is_official_property_subdomain(
                    candidate, canonical_company_website
                ):
                    property_website = candidate
                    detected_microsite = candidate
                    break

        if canonical_company_website:
            out.at[index, "Company Website"] = canonical_company_website
        if property_website:
            out.at[index, "Property Website"] = property_website

        if is_unresolved(out.at[index, "Website"]):
            out.at[index, "Website"] = (
                property_website or canonical_company_website or pd.NA
            )

        if detected_microsite:
            note = (
                "Official property microsite grouped under the parent company "
                f"root domain {company_root or 'the selected company domain'}; "
                "retained as Property Website rather than creating a separate company."
            )
            out.at[index, "Reviewer Notes"] = _append_note(
                out.at[index, "Reviewer Notes"], note
            )

    return out


def _address_signature(value) -> dict:
    """Return civic-number and street signatures from a Street Address value.

    Compound values such as ``1161-1171 Wellington Street`` and
    ``1140/1150 Fisher Avenue`` expose every explicitly written civic number.
    Datablix intentionally does NOT invent intermediate numbers in a range.
    """
    raw = _fold_identity_text(value)
    if not raw:
        return {"numbers": tuple(), "street": "", "street_lenient": "", "full": ""}

    # A research tool occasionally puts city/province/postal text into the Street
    # Address cell.  Comparison uses only the street-address part before a comma.
    street_part = raw.split(",", 1)[0].strip()
    street_part = re.sub(
        r"\b[abceghj-nprstvxy]\d[abceghj-nprstv-z][ -]?\d[abceghj-nprstv-z]\d\b",
        " ",
        street_part,
        flags=re.I,
    )
    street_part = re.sub(r"\s+", " ", street_part).strip()

    # Capture an explicitly listed civic-number sequence at the beginning.
    prefix = re.match(
        r"^\s*((?:\d+[a-z]?)(?:\s*(?:-|/|&|\band\b)\s*\d+[a-z]?)*)(?:\s+)(.+)$",
        street_part,
        flags=re.I,
    )
    if prefix:
        number_text = prefix.group(1)
        street_text = prefix.group(2)
        numbers = tuple(dict.fromkeys(re.findall(r"\d+[a-z]?", number_text, flags=re.I)))
    else:
        one_number = re.match(r"^\s*(\d+[a-z]?)\s+(.+)$", street_part, flags=re.I)
        if one_number:
            numbers = (one_number.group(1),)
            street_text = one_number.group(2)
        else:
            numbers = tuple()
            street_text = street_part

    tokens = re.findall(r"[a-z0-9]+", street_text.lower())
    normalized_tokens = []
    for token in tokens:
        token = _STREET_TYPE_ALIASES.get(token, token)
        token = _DIRECTION_ALIASES.get(token, token)
        normalized_tokens.append(token)

    street = " ".join(normalized_tokens).strip()
    lenient_tokens = [
        token for token in normalized_tokens if token not in {"n", "s", "e", "w"}
    ]
    street_lenient = " ".join(lenient_tokens).strip()
    number_key = "/".join(sorted(set(numbers)))
    full = f"{number_key}|{street}" if number_key or street else ""
    return {
        "numbers": tuple(sorted(set(numbers))),
        "street": street,
        "street_lenient": street_lenient,
        "full": full,
    }


def _building_name_key(value) -> str:
    key = norm_header(value)
    generic = {
        "apartments", "apartment", "property", "properties", "building",
        "residences", "residence", "community", "communities",
    }
    return "" if key in generic else key


def _row_identity(row) -> dict:
    address = _address_signature(row.get("Street Address", ""))
    property_url = _canonical_url(row.get("Property Website", ""))
    website = _canonical_url(row.get("Website", ""))
    source_url = _canonical_url(row.get("Source URL", ""))
    return {
        "address": address,
        "postal": _canonical_postal(row.get("Postal Code", "")),
        "city": _canonical_city(row.get("City", "")),
        "name": _building_name_key(row.get("Building Name", "")),
        "property_url": property_url,
        "website": website,
        "source_url": source_url,
        "phone": _canonical_phone(row.get("Phone", "")),
        "email": safe_text(row.get("Primary Email", "")).lower(),
    }


def _discovery_keys_for_row(row) -> set[str]:
    """Return normalized identity keys for diagnostics/backward compatibility."""
    identity = _row_identity(row)
    address = identity["address"]
    keys = set()
    if address["full"]:
        keys.add(f"address:{address['full']}")
    if address["street_lenient"] and address["numbers"]:
        for number in address["numbers"]:
            keys.add(f"civic_street:{number}|{address['street_lenient']}")
    if identity["postal"] and address["street_lenient"]:
        keys.add(f"street_postal:{address['street_lenient']}|{identity['postal']}")
    if identity["name"] and identity["postal"]:
        keys.add(f"name_postal:{identity['name']}|{identity['postal']}")
    return keys


def _source_match_score_from_identities(research: dict, source: dict) -> tuple[int, str]:
    """Score two pre-normalized property identities conservatively."""
    ra = research["address"]
    sa = source["address"]

    research_numbers = set(ra["numbers"])
    source_numbers = set(sa["numbers"])
    numbers_overlap = bool(research_numbers & source_numbers)
    research_numbers_are_component = bool(
        research_numbers
        and source_numbers
        and research_numbers.issubset(source_numbers)
    )
    street_exact = bool(ra["street"] and ra["street"] == sa["street"])
    street_lenient = bool(
        ra["street_lenient"]
        and ra["street_lenient"] == sa["street_lenient"]
    )
    postal_same = bool(research["postal"] and research["postal"] == source["postal"])
    city_same = bool(research["city"] and research["city"] == source["city"])
    name_same = bool(research["name"] and research["name"] == source["name"])

    # Strongest signal: same civic address after standardizing street wording.
    if numbers_overlap and street_exact:
        if postal_same:
            return 100, "same civic address and postal code"
        if city_same:
            return 97, "same civic address and Ottawa locality"
        return 93, "same civic address"

    # Directional text is often omitted from either the research or source file.
    if numbers_overlap and street_lenient:
        if postal_same:
            return 99, "same civic address after street-direction normalization"
        if city_same:
            return 95, "same civic address after Ottawa/street normalization"
        return 90, "same civic address after street normalization"

    # A single AI row may represent one component of a combined source address,
    # e.g. 1171 Wellington versus 1161-1171 Wellington.
    if research_numbers_are_component and street_lenient:
        if postal_same:
            return 99, "research address is a component of a combined source address"
        if city_same:
            return 95, "research address is a component of a combined Ottawa source address"
        return 90, "research address is a component of a combined source address"

    # Property-specific URLs can rescue a formatting-heavy address mismatch.
    url_pairs = [
        (research["property_url"], source["property_url"]),
        (research["property_url"], source["website"]),
        (research["website"], source["property_url"]),
    ]
    if any(a and b and a == b for a, b in url_pairs):
        if postal_same or city_same:
            return 91, "same property website with compatible location"
        return 82, "same property website"

    # Building name is supporting evidence, not enough by itself to force a match.
    if name_same and postal_same:
        return 88, "same building name and postal code"
    if name_same and city_same and street_lenient:
        return 84, "same building name, Ottawa locality and street"
    if street_lenient and postal_same:
        return 80, "same street and postal code but civic number needs review"
    if name_same and city_same:
        return 72, "same building name and Ottawa locality; address needs review"
    return 0, "no credible source match"


def _source_match_score(research_row, source_row) -> tuple[int, str]:
    """Backward-compatible row scorer used by diagnostics and tests."""
    return _source_match_score_from_identities(
        _row_identity(research_row),
        _row_identity(source_row),
    )


def _build_discovery_source_index(source_frame: pd.DataFrame) -> tuple[list[dict], dict]:
    """Pre-normalize Starting Data and index every signal that can score > 0.

    Earlier versions reparsed every source address/URL for every research row.
    The index keeps the exact same scoring rules while limiting comparisons to
    source rows that share at least one possible matching signal.
    """
    identities = []
    indexes = {
        "civic_exact": {},
        "civic_lenient": {},
        "property_url": {},
        "website": {},
        "name_postal": {},
        "name_city_street": {},
        "street_postal": {},
        "name_city": {},
    }

    def add(index_name: str, key, position: int) -> None:
        if key is None or key == "" or key == ():
            return
        indexes[index_name].setdefault(key, []).append(position)

    for position, source_row in enumerate(source_frame.to_dict(orient="records")):
        identity = _row_identity(source_row)
        identities.append(identity)
        address = identity["address"]

        for number in address["numbers"]:
            if address["street"]:
                add("civic_exact", (number, address["street"]), position)
            if address["street_lenient"]:
                add("civic_lenient", (number, address["street_lenient"]), position)

        if identity["property_url"]:
            add("property_url", identity["property_url"], position)
        if identity["website"]:
            add("website", identity["website"], position)
        if identity["name"] and identity["postal"]:
            add("name_postal", (identity["name"], identity["postal"]), position)
        if identity["name"] and identity["city"] and address["street_lenient"]:
            add(
                "name_city_street",
                (identity["name"], identity["city"], address["street_lenient"]),
                position,
            )
        if address["street_lenient"] and identity["postal"]:
            add("street_postal", (address["street_lenient"], identity["postal"]), position)
        if identity["name"] and identity["city"]:
            add("name_city", (identity["name"], identity["city"]), position)

    return identities, indexes


def _candidate_source_positions(research: dict, indexes: dict) -> set[int]:
    """Return only source rows capable of receiving a non-zero match score."""
    positions = set()
    address = research["address"]

    for number in address["numbers"]:
        if address["street"]:
            positions.update(indexes["civic_exact"].get((number, address["street"]), ()))
        if address["street_lenient"]:
            positions.update(
                indexes["civic_lenient"].get((number, address["street_lenient"]), ())
            )

    # Preserve the asymmetric URL comparisons from _source_match_score.
    if research["property_url"]:
        positions.update(indexes["property_url"].get(research["property_url"], ()))
        positions.update(indexes["website"].get(research["property_url"], ()))
    if research["website"]:
        positions.update(indexes["property_url"].get(research["website"], ()))

    if research["name"] and research["postal"]:
        positions.update(
            indexes["name_postal"].get((research["name"], research["postal"]), ())
        )
    if research["name"] and research["city"] and address["street_lenient"]:
        positions.update(
            indexes["name_city_street"].get(
                (research["name"], research["city"], address["street_lenient"]),
                (),
            )
        )
    if address["street_lenient"] and research["postal"]:
        positions.update(
            indexes["street_postal"].get(
                (address["street_lenient"], research["postal"]),
                (),
            )
        )
    if research["name"] and research["city"]:
        positions.update(
            indexes["name_city"].get((research["name"], research["city"]), ())
        )

    return positions


def current_starting_source_records() -> pd.DataFrame:
    """Return the active structured Starting Data with a safe legacy fallback.

    Saved projects may contain a structured baseline produced by an older Datablix
    parser.  When the original workbook bytes are still preserved, v94 rebuilds that
    structured baseline once before returning it.  This prevents an old cached subset
    (for example, Lepine=4) from surviving after the counting/parser rules were fixed.
    """
    try:
        _repair_active_source_baseline_from_raw_if_needed()
    except Exception:
        # Never make an existing project unusable because a legacy source cannot be
        # reparsed. The existing baseline remains available as a safe fallback.
        pass

    versions = st.session_state.get(S_SOURCE_VERSIONS, [])
    if isinstance(versions, list):
        for version in reversed(versions):
            if not isinstance(version, dict) or not bool(version.get("is_active")):
                continue
            records = version.get("records")
            if isinstance(records, pd.DataFrame) and not records.empty:
                return records.copy()
        for version in reversed(versions):
            if not isinstance(version, dict):
                continue
            records = version.get("records")
            if isinstance(records, pd.DataFrame) and not records.empty:
                return records.copy()

    fallback = st.session_state.get(S_ORIGINAL)
    if isinstance(fallback, pd.DataFrame) and not fallback.empty:
        return fallback.copy()
    return pd.DataFrame()


def classify_discovery_status(df, original=None):
    """Classify website-research rows against the current project Starting Data.

    The status is deliberately conservative:
    * strong normalized source match -> Existing Source Record;
    * plausible/ambiguous source candidate -> Needs Classification;
    * Newly Discovered -> only after every Starting Data row fails the stronger
      identity checks and the research row has current-property evidence.

    Performance note: Starting Data identities are normalized once and indexed by
    the same address/name/URL signals used by the scorer. This preserves the v48
    thresholds while avoiding an all-against-all comparison on every rerun.
    """
    out = normalize_workflow(df)

    source_frame = pd.DataFrame()
    source_identities = []
    source_indexes = {}
    if isinstance(original, pd.DataFrame) and not original.empty:
        # Discovery comparison only needs property identity fields. Avoid running
        # the full workflow normalizer over Starting Data on every rerun.
        source_frame = coalesce_duplicate_columns(original.copy())
        source_identities, source_indexes = _build_discovery_source_index(source_frame)

    for idx, row in out.iterrows():
        decision = safe_text(row.get("Record Decision", ""))
        inventory_status = safe_text(row.get("Current Inventory Status", "")).lower()
        rental_status = safe_text(row.get("Rental Availability Status", "")).lower()
        verification_status = safe_text(row.get("Verification Status", ""))
        discovery_source = safe_text(row.get("Discovery Status Source", "Automatic"))
        current_discovery_status = safe_text(row.get("Directory Discovery Status", ""))

        # Record-level decisions still take precedence over discovery classification.
        if decision == "Possible Duplicate":
            out.at[idx, "Directory Discovery Status"] = "Possible Duplicate"
            out.at[idx, "Discovery Status Source"] = "Automatic"
            continue
        if (
            decision == "Remove"
            or inventory_status.startswith("excluded")
            or rental_status == "rented"
        ):
            out.at[idx, "Directory Discovery Status"] = "Excluded / Not Current"
            out.at[idx, "Discovery Status Source"] = "Automatic"
            continue

        # A human-reviewed classification is final unless the reviewer changes it
        # again. This prevents the app-wide rerun from undoing a saved choice.
        if discovery_source == "Manual" and current_discovery_status in DISCOVERY_STATUSES:
            continue

        if source_frame.empty:
            out.at[idx, "Directory Discovery Status"] = "Needs Classification"
            out.at[idx, "Discovery Status Source"] = "Automatic"
            continue

        research_identity = _row_identity(row)
        candidate_positions = _candidate_source_positions(
            research_identity, source_indexes
        )

        best_score = 0
        for position in candidate_positions:
            score, _reason = _source_match_score_from_identities(
                research_identity, source_identities[position]
            )
            if score > best_score:
                best_score = score
                if best_score >= 100:
                    break

        if best_score >= 88:
            out.at[idx, "Directory Discovery Status"] = "Existing Source Record"
            out.at[idx, "Discovery Status Source"] = "Automatic"
            continue

        # A plausible source candidate blocks an automatic "new" claim. Human
        # review is safer than overstating discovery when address evidence conflicts.
        if best_score >= 72:
            out.at[idx, "Directory Discovery Status"] = "Needs Classification"
            out.at[idx, "Discovery Status Source"] = "Automatic"
            continue

        address = research_identity["address"]
        has_property_identity = bool(
            address["numbers"] and address["street_lenient"]
        ) or bool(research_identity["property_url"]) or bool(
            research_identity["name"] and research_identity["postal"]
        )
        active_rental_status = rental_status in {"available now", "available future"}
        legacy_status_not_checked = rental_status in {"", "not checked"}
        current_evidence = (
            inventory_status.startswith("current")
            and (active_rental_status or legacy_status_not_checked)
        ) or (
            verification_status == "Verified"
            and decision in {"Keep", "Update"}
            and rental_status != "rented"
        )
        out.at[idx, "Directory Discovery Status"] = (
            "Newly Discovered"
            if has_property_identity and current_evidence
            else "Needs Classification"
        )
        out.at[idx, "Discovery Status Source"] = "Automatic"

    return out


def empty_company_registry():
    return pd.DataFrame(columns=COMPANY_COLUMNS)


def _normalize_official_url(value: str) -> str:
    """Return a clean HTTP(S) URL for company-level official entry points."""
    raw = safe_text(value)
    if not raw:
        return ""
    candidate = raw.strip().strip("<>")
    if not re.match(r"^[a-z][a-z0-9+.-]*://", candidate, flags=re.I):
        candidate = f"https://{candidate}"
    parsed = urlparse(candidate)
    if parsed.scheme.lower() not in {"http", "https"} or not parsed.netloc:
        return ""
    host = parsed.netloc.strip().lower()
    path = re.sub(r"/{2,}", "/", parsed.path or "/")
    if path != "/":
        path = path.rstrip("/") + "/"
    return urlunparse((parsed.scheme.lower(), host, path, "", parsed.query, ""))


def _official_url_identity(value: str) -> str:
    normalized = _normalize_official_url(value)
    if not normalized:
        return ""
    parsed = urlparse(normalized)
    path = (parsed.path or "/").rstrip("/") or "/"
    query = f"?{parsed.query}" if parsed.query else ""
    return f"{parsed.netloc.lower()}{path.lower()}{query}"


def _parse_related_official_links(value, main_website: str = "") -> tuple[list[str], list[str]]:
    """Parse one official URL per line, deduplicate, and report invalid entries."""
    raw = safe_text(value)
    if not raw:
        return [], []
    candidates = [
        item.strip()
        for item in re.split(r"[\r\n]+", raw)
        if item.strip()
    ]
    main_identity = _official_url_identity(main_website)
    links: list[str] = []
    invalid: list[str] = []
    seen: set[str] = set()
    for candidate in candidates:
        normalized = _normalize_official_url(candidate)
        identity = _official_url_identity(normalized)
        if not normalized or not identity:
            invalid.append(candidate)
            continue
        if identity == main_identity or identity in seen:
            continue
        seen.add(identity)
        links.append(normalized)
    return links, invalid


def _normalize_related_official_links(value, main_website: str = "") -> str:
    links, _ = _parse_related_official_links(value, main_website)
    return "\n".join(links)


def _merge_related_official_links(*values, main_website: str = "") -> str:
    links: list[str] = []
    seen: set[str] = set()
    main_identity = _official_url_identity(main_website)
    for value in values:
        parsed_links, _ = _parse_related_official_links(value, main_website)
        for link in parsed_links:
            identity = _official_url_identity(link)
            if not identity or identity == main_identity or identity in seen:
                continue
            seen.add(identity)
            links.append(link)
    return "\n".join(links)


def _contains_legacy_regional_scope(value) -> bool:
    """Identify saved prompt text that would wrongly restore regional scope."""
    text = safe_text(value).lower()
    legacy_markers = (
        "greater ottawa area",
        "surrounding eastern ontario",
        "ottawa-region portfolio",
        "nearby eastern ontario communities",
        "include city of ottawa properties plus nearby",
        "carleton place",
        "smiths falls",
        "renfrew",
        "current apartment listings",
        "only physical apartment properties",
        "apartment properties physically located",
    )
    return any(marker in text for marker in legacy_markers)


def normalize_company_registry(registry):
    if not isinstance(registry, pd.DataFrame):
        registry = empty_company_registry()
    out = coalesce_duplicate_columns(registry.copy())
    for column in COMPANY_COLUMNS:
        if column not in out.columns:
            out[column] = pd.NA
    out = out[COMPANY_COLUMNS].copy()
    out["Management/Owner"] = out["Management/Owner"].fillna("").astype(str).str.strip()
    out["Company ID"] = out["Company ID"].fillna("").astype(str).str.strip()
    out["Main Website"] = out["Main Website"].fillna("").astype(str).str.strip()
    out["Related Official Links"] = out.apply(
        lambda row: _normalize_related_official_links(
            row.get("Related Official Links", ""),
            row.get("Main Website", ""),
        ),
        axis=1,
    )
    out["Special Website Notes"] = (
        out["Special Website Notes"].fillna("").astype(str).str.strip()
    )
    out["Scope Type"] = normalize_choice(
        out["Scope Type"], COMPANY_SCOPE_TYPES, "Imported"
    )
    out["Company Status"] = normalize_choice(
        out["Company Status"], COMPANY_STATUSES, "Not started"
    )
    out["Date Assigned"] = out["Date Assigned"].fillna("").astype(str).str.strip()
    out["Notes"] = out["Notes"].fillna("").astype(str)
    for prompt_column in [
        "Prompt Scope", "Prompt Source Policy", "Prompt Priority Notes",
        "Prompt Output Notes", "Research Prompt",
    ]:
        out[prompt_column] = out[prompt_column].fillna("").astype(str)

    # Project scope is fixed globally. Remove previously saved regional-scope
    # instructions so old company notes cannot reintroduce nearby municipalities.
    out["Prompt Scope"] = PROJECT_GEOGRAPHIC_SCOPE
    legacy_priority_mask = out["Prompt Priority Notes"].apply(
        _contains_legacy_regional_scope
    )
    out.loc[legacy_priority_mask, "Prompt Priority Notes"] = ""
    legacy_source_mask = out["Prompt Source Policy"].apply(
        _contains_legacy_regional_scope
    )
    out.loc[legacy_source_mask, "Prompt Source Policy"] = ""

    out["Prompt Updated"] = out["Prompt Updated"].fillna("").astype(str).str.strip()
    out["AI Tool Used"] = out["AI Tool Used"].fillna("").astype(str).str.strip()
    out = out.loc[out["Management/Owner"].ne("") | out["Company ID"].ne("")].copy()

    used_ids = set(out.loc[out["Company ID"].ne(""), "Company ID"].astype(str))
    next_number = 1
    for index in out.index[out["Company ID"].eq("")]:
        while f"CMP-{next_number:03d}" in used_ids:
            next_number += 1
        company_id = f"CMP-{next_number:03d}"
        out.at[index, "Company ID"] = company_id
        used_ids.add(company_id)
        next_number += 1

    return out.drop_duplicates(subset=["Company ID"], keep="last").reset_index(drop=True)


def next_company_id(registry):
    existing = set(
        normalize_company_registry(registry)["Company ID"].astype(str).str.strip()
    )
    number = 1
    while f"CMP-{number:03d}" in existing:
        number += 1
    return f"CMP-{number:03d}"


def company_name_key(value):
    """Return a conservative company-name key for matching aliases.

    Legal suffixes are ignored, but meaningful words such as ``properties`` or
    ``management`` are retained. URL-like values are handled through the
    company-domain matcher instead of being treated as company names.
    """
    raw = safe_text(value)
    if not raw or _looks_like_company_url_label(raw):
        return ""
    tokens = re.findall(r"[a-z0-9]+", raw.lower())
    if tokens and tokens[0] == "the":
        tokens = tokens[1:]
    legal_suffixes = {
        "inc", "incorporated", "ltd", "limited", "corp", "corporation",
        "co", "company", "llc", "lp", "llp", "ulc",
    }
    while tokens and tokens[-1] in legal_suffixes:
        tokens.pop()
    return "".join(tokens)


_GENERIC_COMPANY_DOMAINS = {
    "rentcafe.com", "propertyvista.com", "buildingstack.com", "appfolio.com",
    "entrata.com", "realpage.com", "rentmanager.com", "yardi.com",
    "google.com", "facebook.com", "instagram.com", "linkedin.com",
}


def _looks_like_company_url_label(value) -> bool:
    raw = safe_text(value).strip().lower()
    if not raw:
        return False
    return bool(
        "://" in raw
        or raw.startswith("www.")
        or raw.startswith("/")
        or re.match(r"^(?:ca|com|org|net|io|co)/", raw)
        or re.search(r"\b[a-z0-9-]+\.(?:ca|com|org|net|io|co)(?:/|$)", raw)
    )


def _company_domain_key(value) -> str:
    """Return a registrable-looking website domain for company reconciliation."""
    raw = safe_text(value).strip()
    if not raw:
        return ""

    # Email fields occasionally appear in imported website columns.
    if "@" in raw and "://" not in raw and "/" not in raw:
        raw = raw.rsplit("@", 1)[-1]

    if "://" not in raw:
        if not re.search(r"[a-z0-9-]+\.[a-z]{2,}", raw, flags=re.I):
            return ""
        raw = f"https://{raw}"

    try:
        host = (urlparse(raw).hostname or "").lower().strip(".")
    except Exception:
        return ""
    if not host:
        return ""

    for prefix in ("www.", "m."):
        if host.startswith(prefix):
            host = host[len(prefix):]

    labels = [label for label in host.split(".") if label]
    if len(labels) < 2:
        return ""

    country_second_levels = {
        "co.uk", "org.uk", "com.au", "net.au", "co.nz", "co.za",
    }
    tail_two = ".".join(labels[-2:])
    domain = ".".join(labels[-3:]) if tail_two in country_second_levels and len(labels) >= 3 else tail_two
    return "" if domain in _GENERIC_COMPANY_DOMAINS else domain


def _company_domain_brand_key(value) -> str:
    """Return the normalized brand portion of a company website domain."""
    domain = _company_domain_key(value)
    if not domain:
        return ""
    return norm_header(domain.split(".", 1)[0])


def _company_brand_keys(row) -> set[str]:
    """Return conservative name/domain brand keys for one registry row."""
    keys = set()
    name_key = company_name_key(row.get("Management/Owner", ""))
    if name_key:
        keys.add(name_key)
    for value in [row.get("Main Website", ""), row.get("Management/Owner", "")]:
        brand = _company_domain_brand_key(value)
        if brand:
            keys.add(brand)
    return keys


def _company_brand_keys_compatible(left: set[str], right: set[str]) -> bool:
    """Match a human company name to its official-domain brand safely.

    This specifically repairs cases such as ``Hazelview Properties`` versus
    ``hazelviewproperties.ca/residential`` even when the assigned company row
    has no Main Website saved. Partial matching is allowed only for a meaningful
    brand fragment of at least five characters.
    """
    for left_key in left:
        for right_key in right:
            if not left_key or not right_key:
                continue
            if left_key == right_key:
                return True
            shorter, longer = sorted((left_key, right_key), key=len)
            if len(shorter) >= 5 and longer.startswith(shorter):
                return True
    return False


def _record_company_brand_keys(row) -> set[str]:
    """Return company brand keys found in one building/research record."""
    keys = set()
    owner_key = company_name_key(row.get("Management/Owner", ""))
    if owner_key:
        keys.add(owner_key)
    for column in [
        "Company Website", "Website", "Property Website", "Source URL",
        "Management/Owner",
    ]:
        brand = _company_domain_brand_key(row.get(column, ""))
        if brand:
            keys.add(brand)
    return keys


def _unique_company_brand_map(registry: pd.DataFrame) -> dict[str, str]:
    grouped: dict[str, set[str]] = {}
    for _, row in registry.iterrows():
        company_id = safe_text(row.get("Company ID"))
        if not company_id:
            continue
        for key in _company_brand_keys(row):
            grouped.setdefault(key, set()).add(company_id)
    return {
        key: next(iter(company_ids))
        for key, company_ids in grouped.items()
        if len(company_ids) == 1
    }


def _company_scope_rank(value) -> int:
    return {
        "Initial assignment": 0,
        "Added later": 1,
        "Imported": 2,
    }.get(safe_text(value), 3)


def _company_status_rank(value) -> int:
    return {
        "Not started": 0,
        "Researching": 1,
        "Needs follow-up": 2,
        "Ready for QA": 3,
        "Complete with limitations": 4,
        "Complete": 5,
    }.get(safe_text(value), 0)


def _company_row_domain(row) -> str:
    return (
        _company_domain_key(row.get("Main Website", ""))
        or _company_domain_key(row.get("Management/Owner", ""))
    )


def _merge_company_registry_duplicates(registry: pd.DataFrame) -> tuple[pd.DataFrame, dict[str, str]]:
    """Merge legacy duplicate company rows and return old-to-canonical ID mapping.

    Exact normalized company names are always considered aliases. A shared
    official domain is considered an alias only when at least one row was
    auto-imported or has a URL-like company label. This keeps assigned companies
    distinct while repairing the legacy rows that caused completed Hazelview
    records to appear under separate website-name companies.
    """
    source = normalize_company_registry(registry)
    if source.empty:
        return source, {}

    source = source.copy()
    source["__position"] = range(len(source))
    source["__scope_rank"] = source["Scope Type"].map(_company_scope_rank)
    source["__label_penalty"] = source["Management/Owner"].map(
        lambda value: 1 if (not safe_text(value) or _looks_like_company_url_label(value)) else 0
    )
    source = source.sort_values(
        ["__scope_rank", "__label_penalty", "__position"],
        kind="stable",
    )

    canonical_rows: list[dict] = []
    id_remap: dict[str, str] = {}

    for _, series in source.iterrows():
        row = {column: series.get(column, "") for column in COMPANY_COLUMNS}
        row_id = safe_text(row.get("Company ID"))
        name_key = company_name_key(row.get("Management/Owner"))
        domain_key = _company_row_domain(row)
        row_is_imported = safe_text(row.get("Scope Type")) == "Imported"
        row_is_suspicious = _looks_like_company_url_label(row.get("Management/Owner")) or not safe_text(row.get("Management/Owner"))
        row_brand_keys = _company_brand_keys(row)

        match_index = None
        for position, candidate in enumerate(canonical_rows):
            candidate_name = company_name_key(candidate.get("Management/Owner"))
            candidate_domain = _company_row_domain(candidate)
            exact_name_match = bool(name_key and candidate_name and name_key == candidate_name)
            domain_match = bool(domain_key and candidate_domain and domain_key == candidate_domain)
            candidate_is_imported = safe_text(candidate.get("Scope Type")) == "Imported"
            candidate_is_suspicious = _looks_like_company_url_label(candidate.get("Management/Owner")) or not safe_text(candidate.get("Management/Owner"))
            brand_match = _company_brand_keys_compatible(
                row_brand_keys, _company_brand_keys(candidate)
            )
            alias_context = (
                row_is_imported
                or candidate_is_imported
                or row_is_suspicious
                or candidate_is_suspicious
            )

            if exact_name_match or ((domain_match or brand_match) and alias_context):
                match_index = position
                break

        if match_index is None:
            canonical_rows.append(row)
            id_remap[row_id] = row_id
            continue

        canonical = canonical_rows[match_index]
        canonical_id = safe_text(canonical.get("Company ID"))
        id_remap[row_id] = canonical_id

        # Keep the human-readable assigned name and scope, while preserving any
        # useful prompt, website, notes, or status stored on the duplicate row.
        if (
            (not safe_text(canonical.get("Management/Owner")) or _looks_like_company_url_label(canonical.get("Management/Owner")))
            and safe_text(row.get("Management/Owner"))
            and not _looks_like_company_url_label(row.get("Management/Owner"))
        ):
            canonical["Management/Owner"] = safe_text(row.get("Management/Owner"))

        fill_if_blank = [
            "Main Website", "Date Assigned", "Prompt Scope", "Prompt Source Policy",
            "Prompt Priority Notes", "Prompt Output Notes", "Research Prompt",
            "Prompt Updated", "AI Tool Used",
        ]
        for column in fill_if_blank:
            if not safe_text(canonical.get(column)) and safe_text(row.get(column)):
                canonical[column] = safe_text(row.get(column))

        canonical["Related Official Links"] = _merge_related_official_links(
            canonical.get("Related Official Links", ""),
            row.get("Related Official Links", ""),
            main_website=canonical.get("Main Website", "") or row.get("Main Website", ""),
        )
        incoming_special_note = safe_text(row.get("Special Website Notes"))
        existing_special_note = safe_text(canonical.get("Special Website Notes"))
        if incoming_special_note and incoming_special_note not in existing_special_note:
            canonical["Special Website Notes"] = (
                f"{existing_special_note} | {incoming_special_note}".strip(" |")
            )

        incoming_status = safe_text(row.get("Company Status"))
        canonical_status = safe_text(canonical.get("Company Status"))
        # Do not transfer a legacy duplicate's manual Complete flag blindly.
        # Once its records are relinked, the normal QA calculation decides
        # whether the canonical company is truly complete.
        if (
            incoming_status in {"Researching", "Needs follow-up", "Ready for QA"}
            and _company_status_rank(incoming_status) > _company_status_rank(canonical_status)
        ):
            canonical["Company Status"] = incoming_status

        incoming_note = safe_text(row.get("Notes"))
        existing_note = safe_text(canonical.get("Notes"))
        if incoming_note and incoming_note not in existing_note:
            canonical["Notes"] = f"{existing_note} | {incoming_note}".strip(" |")

    merged = pd.DataFrame(canonical_rows, columns=COMPANY_COLUMNS)
    return normalize_company_registry(merged), id_remap


def _unique_company_alias_map(registry: pd.DataFrame, key_function) -> dict[str, str]:
    grouped: dict[str, set[str]] = {}
    for _, row in registry.iterrows():
        key = key_function(row)
        company_id = safe_text(row.get("Company ID"))
        if key and company_id:
            grouped.setdefault(key, set()).add(company_id)
    return {
        key: next(iter(company_ids))
        for key, company_ids in grouped.items()
        if len(company_ids) == 1
    }


def _record_company_domains(row) -> list[str]:
    domains = []
    for column in [
        "Company Website", "Website", "Property Website", "Source URL",
        "Management/Owner",
    ]:
        domain = _company_domain_key(row.get(column, ""))
        if domain and domain not in domains:
            domains.append(domain)
    return domains


def synchronize_company_registry(records, registry=None):
    """Keep building records linked to one canonical company registry row.

    This includes a migration for projects saved by older Datablix builds where
    research rows could remain under imported URL/company aliases. The selected
    assigned company now wins, legacy duplicate IDs are remapped, and every
    linked record receives the canonical company name.
    """
    data = normalize_workflow(prepare_data(records.copy()))
    registry, id_remap = _merge_company_registry_duplicates(registry)

    # Apply duplicate-ID migration before any progress calculation.
    if not data.empty:
        data["Company ID"] = data["Company ID"].apply(
            lambda value: id_remap.get(safe_text(value), safe_text(value) or pd.NA)
        )

    active_id = safe_text(st.session_state.get(S_ACTIVE_COMPANY, ""))
    if active_id and active_id in id_remap:
        st.session_state[S_ACTIVE_COMPANY] = id_remap[active_id]

    def rebuild_maps():
        name_map = _unique_company_alias_map(
            registry,
            lambda row: company_name_key(row.get("Management/Owner", "")),
        )
        domain_map = _unique_company_alias_map(registry, _company_row_domain)
        brand_map = _unique_company_brand_map(registry)
        rows_by_id = {
            safe_text(row.get("Company ID")): row
            for _, row in registry.iterrows()
            if safe_text(row.get("Company ID"))
        }
        return name_map, domain_map, brand_map, rows_by_id

    name_to_id, domain_to_id, brand_to_id, registry_by_id = rebuild_maps()

    # Repair rows whose old Company ID points to an imported alias, or whose ID
    # is missing/unknown but the canonical assigned name or official domain is clear.
    for index, row in data.iterrows():
        current_id = safe_text(row.get("Company ID"))
        current_company = registry_by_id.get(current_id)
        owner_candidate = name_to_id.get(company_name_key(row.get("Management/Owner", "")))
        domain_candidates = [
            domain_to_id[domain]
            for domain in _record_company_domains(row)
            if domain in domain_to_id
        ]
        domain_candidate = domain_candidates[0] if len(set(domain_candidates)) == 1 and domain_candidates else ""
        brand_candidates = {
            brand_to_id[key]
            for key in _record_company_brand_keys(row)
            if key in brand_to_id
        }
        brand_candidate = next(iter(brand_candidates)) if len(brand_candidates) == 1 else ""
        candidate_id = owner_candidate or domain_candidate or brand_candidate

        if current_company is None:
            chosen_id = candidate_id
        elif candidate_id and candidate_id != current_id:
            current_is_imported = safe_text(current_company.get("Scope Type")) == "Imported"
            current_is_alias = current_is_imported or _looks_like_company_url_label(
                current_company.get("Management/Owner", "")
            )
            candidate_company = registry_by_id.get(candidate_id)
            candidate_is_human_named = (
                candidate_company is not None
                and safe_text(candidate_company.get("Management/Owner", ""))
                and not _looks_like_company_url_label(
                    candidate_company.get("Management/Owner", "")
                )
            )
            candidate_is_stronger = (
                candidate_company is not None
                and _company_scope_rank(candidate_company.get("Scope Type"))
                <= _company_scope_rank(current_company.get("Scope Type"))
            )
            chosen_id = (
                candidate_id
                if current_is_alias and (candidate_is_human_named or candidate_is_stronger)
                else current_id
            )
        else:
            chosen_id = current_id

        if chosen_id:
            data.at[index, "Company ID"] = chosen_id

    # Register genuinely new owners only after existing aliases have had a chance
    # to resolve to an assigned company.
    used_ids = set(registry["Company ID"].astype(str))
    for index, row in data.iterrows():
        current_id = safe_text(row.get("Company ID"))
        if current_id in used_ids:
            continue

        owner = safe_text(row.get("Management/Owner"))
        owner_key = company_name_key(owner)
        domains = _record_company_domains(row)
        existing_id = name_to_id.get(owner_key, "") if owner_key else ""
        if not existing_id:
            matching_domains = {domain_to_id[d] for d in domains if d in domain_to_id}
            if len(matching_domains) == 1:
                existing_id = next(iter(matching_domains))
        if not existing_id:
            matching_brands = {
                brand_to_id[key]
                for key in _record_company_brand_keys(row)
                if key in brand_to_id
            }
            if len(matching_brands) == 1:
                existing_id = next(iter(matching_brands))

        if existing_id:
            data.at[index, "Company ID"] = existing_id
            continue

        domain = domains[0] if domains else ""
        if not owner and not domain:
            continue

        candidate_id = current_id if current_id and current_id not in used_ids else next_company_id(registry)
        while candidate_id in used_ids:
            candidate_id = next_company_id(registry)

        if owner and not _looks_like_company_url_label(owner):
            company_name = owner
        elif domain:
            company_name = domain.split(".", 1)[0].replace("-", " ").title()
        else:
            company_name = f"Imported company {candidate_id}"

        website = ""
        for column in ["Company Website", "Website", "Property Website", "Source URL"]:
            value = safe_text(row.get(column, ""))
            if _company_domain_key(value):
                website = value
                break

        registry = pd.concat([
            registry,
            pd.DataFrame([{
                "Company ID": candidate_id,
                "Management/Owner": company_name,
                "Main Website": website,
                "Scope Type": "Imported",
                "Date Assigned": "",
                "Company Status": "Researching",
                "Notes": "",
            }]),
        ], ignore_index=True)
        registry = normalize_company_registry(registry)
        used_ids.add(candidate_id)
        data.at[index, "Company ID"] = candidate_id
        name_to_id, domain_to_id, brand_to_id, registry_by_id = rebuild_maps()

    registry = normalize_company_registry(registry)
    registry_by_id = {
        safe_text(row.get("Company ID")): row
        for _, row in registry.iterrows()
        if safe_text(row.get("Company ID"))
    }

    # The registry is the source of truth for the displayed owner name. This
    # prevents a URL or AI alias in one imported row from creating another
    # company on the next Streamlit rerun.
    for index, company_id in data["Company ID"].items():
        canonical = registry_by_id.get(safe_text(company_id))
        if canonical is not None and safe_text(canonical.get("Management/Owner")):
            data.at[index, "Management/Owner"] = safe_text(canonical.get("Management/Owner"))

    # Remove orphaned URL/blank imported rows left behind after the migration.
    linked_ids = set(data["Company ID"].fillna("").astype(str).str.strip())
    orphan_mask = (
        ~registry["Company ID"].isin(linked_ids)
        & registry["Management/Owner"].map(
            lambda value: not safe_text(value) or _looks_like_company_url_label(value)
        )
        & registry["Research Prompt"].fillna("").astype(str).str.strip().eq("")
    )
    registry = normalize_company_registry(registry.loc[~orphan_mask].copy())
    # Keep the registry status synchronized with actual linked work. Explicit
    # Complete statuses are preserved. A company with linked records can never
    # remain Not started, even in projects created by older builds.
    for reg_index, company in registry.iterrows():
        cid = safe_text(company.get("Company ID"))
        cname = safe_text(company.get("Management/Owner"))
        csite = safe_text(company.get("Main Website"))
        linked = _company_records_for_progress(
            data, company_id=cid, company_name=cname, company_website=csite
        )
        if linked.empty:
            continue
        current_status = safe_text(company.get("Company Status"), "Not started")
        if current_status == "Not started":
            registry.at[reg_index, "Company Status"] = "Researching"

        # Research activity and directory eligibility are different concepts.
        # Every linked row proves that company research started, even when the row
        # is intentionally excluded from the final directory. Only active rows are
        # used for export-readiness and verification completion.
        linked_qa = qa_checks(linked)
        active = linked_qa.loc[
            ~linked_qa["Record Readiness"].eq("Excluded from Listings")
        ].copy()
        excluded_count = int(
            linked_qa["Record Readiness"].eq("Excluded from Listings").sum()
        )

        linked_reviewed = (
            linked_qa["Research Status"].fillna("").astype(str).eq("Completed")
            | linked_qa["Record Decision"].fillna("").astype(str).isin(
                ["Keep", "Update", "Possible Duplicate", "Remove"]
            )
            | linked_qa["Record Readiness"].eq("Excluded from Listings")
        ).all()

        if active.empty:
            # A company with researched-but-excluded records is not "Not started".
            # When all linked rows have been reviewed, the honest status is
            # Complete with limitations because there are no eligible listings.
            registry.at[reg_index, "Company Status"] = (
                "Complete with limitations" if linked_reviewed and excluded_count else "Researching"
            )
        else:
            all_verified = active["Verification Status"].fillna("").astype(str).eq("Verified").all()
            all_reviewed = (
                active["Research Status"].fillna("").astype(str).eq("Completed")
                | active["Record Decision"].fillna("").astype(str).isin(
                    ["Keep", "Update", "Possible Duplicate"]
                )
            ).all()
            unresolved_discovery = active["Directory Discovery Status"].fillna("").astype(str).isin(
                ["Needs Classification", "Possible Duplicate"]
            ).any()
            if all_verified and all_reviewed and not unresolved_discovery:
                registry.at[reg_index, "Company Status"] = "Complete"

    registry = normalize_company_registry(registry)
    st.session_state[S_COMPANY_LINK_REPAIR_VERSION] = "v61"

    return data, registry

def active_company_row():
    registry = normalize_company_registry(st.session_state.get(S_COMPANIES))
    active_id = str(st.session_state.get(S_ACTIVE_COMPANY, "")).strip()
    match = registry.loc[registry["Company ID"].eq(active_id)]
    return None if match.empty else match.iloc[0]


def company_label(row):
    name = str(row.get("Management/Owner", "")).strip() or "Unnamed company"
    company_id = str(row.get("Company ID", "")).strip()
    return f"{name} · {company_id}" if company_id else name


def assign_active_company(records, row_mask):
    out = records.copy()
    active = active_company_row()
    if active is None:
        return out
    out.loc[row_mask, "Company ID"] = active["Company ID"]
    blank_owner = row_mask & unresolved_mask(out["Management/Owner"])
    out.loc[blank_owner, "Management/Owner"] = active["Management/Owner"]
    return out


def add_company_to_project(
    name,
    website="",
    scope_type="Added later",
    notes="",
    related_official_links="",
    special_website_notes="",
):
    clean_name = re.sub(r"\s+", " ", str(name or "")).strip()
    if not clean_name:
        raise ValueError("Enter the management company or owner name.")
    registry = normalize_company_registry(st.session_state.get(S_COMPANIES))
    key = company_name_key(clean_name)
    existing = registry.loc[
        registry["Management/Owner"].apply(company_name_key).eq(key)
    ]
    if not existing.empty:
        company_id = existing.iloc[0]["Company ID"]
        st.session_state[S_ACTIVE_COMPANY] = company_id
        st.session_state[S_PENDING_ACTIVE_COMPANY] = company_id
        return company_id, False
    company_id = next_company_id(registry)
    new_row = {
        "Company ID": company_id,
        "Management/Owner": clean_name,
        "Main Website": str(website or "").strip(),
        "Related Official Links": _normalize_related_official_links(
            related_official_links,
            website,
        ),
        "Special Website Notes": str(special_website_notes or "").strip(),
        "Scope Type": scope_type if scope_type in COMPANY_SCOPE_TYPES else "Added later",
        "Date Assigned": date.today().isoformat(),
        "Company Status": "Not started",
        "Notes": str(notes or "").strip(),
    }
    st.session_state[S_COMPANIES] = normalize_company_registry(
        pd.concat([registry, pd.DataFrame([new_row])], ignore_index=True)
    )
    st.session_state[S_ACTIVE_COMPANY] = company_id
    st.session_state[S_PENDING_ACTIVE_COMPANY] = company_id
    return company_id, True


def _company_row_mask(
    frame: pd.DataFrame,
    *,
    company_id: str,
    company_name: str,
) -> pd.Series:
    """Return rows that belong to one company across Datablix data/history frames."""
    if not isinstance(frame, pd.DataFrame):
        return pd.Series(dtype="bool")
    if frame.empty:
        return pd.Series(False, index=frame.index, dtype="bool")

    mask = pd.Series(False, index=frame.index, dtype="bool")
    id_fields = ["Company ID", "company_id"]
    name_fields = [
        "Management/Owner",
        "management_owner",
        "Assigned Company",
        "assigned_company",
        "Company",
        "company",
    ]

    clean_company_id = str(company_id or "").strip()
    clean_company_name = re.sub(r"\s+", " ", str(company_name or "")).strip()
    normalized_name = company_name_key(clean_company_name)

    id_evidence_found = False
    if clean_company_id:
        for field in id_fields:
            if field in frame.columns:
                values = frame[field].astype("string").fillna("").str.strip()
                field_match = values.eq(clean_company_id)
                if field_match.any():
                    id_evidence_found = True
                    mask |= field_match

    # Use the organization name as a fallback for older imported/history rows that
    # predate Company ID assignment. We only use this fallback when no ID match was
    # available in the frame, reducing the chance of deleting similarly named records.
    if not id_evidence_found and normalized_name:
        for field in name_fields:
            if field in frame.columns:
                values = frame[field].astype("string").fillna("").apply(company_name_key)
                mask |= values.eq(normalized_name)

    return mask


def delete_company_from_project(company_id: str) -> tuple[bool, str, dict]:
    """Delete one company and its company-scoped project data.

    The project itself and every other company remain untouched. For cloud/shared
    projects, permanent company removal is restricted to the project owner.
    """
    clean_company_id = str(company_id or "").strip()
    registry = normalize_company_registry(st.session_state.get(S_COMPANIES))

    match = registry.loc[registry["Company ID"].astype(str).eq(clean_company_id)]
    if match.empty:
        return False, "The selected company could not be found.", {}

    company_name = str(match.iloc[0].get("Management/Owner", "") or "").strip()
    project_id = str(st.session_state.get(S_CLOUD_PROJECT_ID, "") or "").strip()
    project_role = str(
        st.session_state.get(S_PROJECT_ROLE, "")
        or (project_access_role(project_id) if project_id else "owner")
    ).strip().lower()

    if (
        project_id
        and not st.session_state.get(S_DEMO_MODE)
        and project_role != "owner"
    ):
        return (
            False,
            "Only the project owner can permanently delete a company from this project.",
            {},
        )

    stats = {
        "company": company_name,
        "records_removed": 0,
        "scan_history_rows_removed": 0,
        "scan_candidate_rows_removed": 0,
        "scan_page_rows_removed": 0,
    }

    # Remove company-owned rows from the working project and from the imported
    # original snapshot so a later reset cannot silently restore the deleted company.
    for state_key, stat_key in [
        (S_WORKING, "records_removed"),
        (S_ORIGINAL, None),
        (S_SCAN_HISTORY, "scan_history_rows_removed"),
        (S_SCAN_CANDIDATES, "scan_candidate_rows_removed"),
        (S_SCAN_PAGES, "scan_page_rows_removed"),
    ]:
        frame = st.session_state.get(state_key)
        if not isinstance(frame, pd.DataFrame):
            continue

        mask = _company_row_mask(
            frame,
            company_id=clean_company_id,
            company_name=company_name,
        )
        removed = int(mask.sum()) if len(mask) else 0
        st.session_state[state_key] = frame.loc[~mask].reset_index(drop=True).copy()
        if stat_key:
            stats[stat_key] = removed


    # Remove the organization from the project registry itself.
    registry = registry.loc[
        ~registry["Company ID"].astype(str).eq(clean_company_id)
    ].reset_index(drop=True)
    st.session_state[S_COMPANIES] = normalize_company_registry(registry)

    # Remove scanner UI/state cached specifically for this company.
    scanner_store_key = "_db_company_scan_states"
    scanner_active_key = "_db_active_scan_company"
    scanner_store = st.session_state.get(scanner_store_key)
    if isinstance(scanner_store, dict):
        scanner_store = dict(scanner_store)
        scanner_store.pop(clean_company_id, None)
        st.session_state[scanner_store_key] = scanner_store

    if str(st.session_state.get(scanner_active_key, "")).strip() == clean_company_id:
        st.session_state.pop(scanner_active_key, None)
        for session_key in list(st.session_state.keys()):
            if str(session_key).startswith(("website_scan_", "full_scan_")):
                st.session_state.pop(session_key, None)

    # Clear prompt/editor widgets belonging to the deleted company.
    company_widget_prefixes = (
        "db_prompt_",
        "db_master_prompt_",
        "db_save_company_prompt_",
        "db_project_next_",
        "db_project_alternate_",
        "db_main_company_details_",
    )
    for session_key in list(st.session_state.keys()):
        session_text = str(session_key)
        if (
            clean_company_id in session_text
            and session_text.startswith(company_widget_prefixes)
        ):
            st.session_state.pop(session_key, None)

    # The main company selectbox may still hold the deleted ID. Clearing its widget
    # state prevents Streamlit from trying to render a value that no longer exists.
    for session_key in list(st.session_state.keys()):
        if str(session_key).startswith("db_main_active_company_"):
            st.session_state.pop(session_key, None)

    # Move safely to another company if one remains.
    updated_registry = normalize_company_registry(st.session_state.get(S_COMPANIES))
    if not updated_registry.empty:
        next_company_id = str(updated_registry.iloc[0]["Company ID"]).strip()
        st.session_state[S_ACTIVE_COMPANY] = next_company_id
        st.session_state[S_PENDING_ACTIVE_COMPANY] = next_company_id
    else:
        st.session_state.pop(S_ACTIVE_COMPANY, None)
        st.session_state.pop(S_PENDING_ACTIVE_COMPANY, None)

    # Rebuild missing-information fields after company rows are removed.
    working = st.session_state.get(S_WORKING)
    if isinstance(working, pd.DataFrame):
        st.session_state[S_WORKING] = normalize_workflow(working)

    # Persist the company deletion into the active project immediately.
    autosave_current_project()

    return True, f'{company_name or clean_company_id} was removed from this project.', stats


# =========================================================
# Reading and mapping
# =========================================================

def source_columns(df, aliases):
    lookup = {}
    for c in df.columns:
        lookup.setdefault(norm_header(c), []).append(c)
    matches = []
    for alias in aliases:
        for c in lookup.get(norm_header(alias), []):
            if c not in matches:
                matches.append(c)
    return matches


def combine_columns(df, columns):
    out = pd.Series(pd.NA, index=df.index, dtype="object")
    for c in columns:
        candidate = resolved(df[c])
        mask = unresolved_mask(out) & ~unresolved_mask(candidate)
        out.loc[mask] = candidate.loc[mask]
    return out


def derive_property_type(df):
    """Derive property-form labels from legacy source flag columns only."""
    available = [c for c in PROPERTY_TYPE_SOURCE_COLUMNS if c in df.columns]
    if not available:
        return pd.Series(pd.NA, index=df.index, dtype="object")

    def derive(row):
        values = []
        for c in available:
            value = row[c]
            n = norm_scalar(value)
            if n in UNRESOLVED or n in NO_VALUES:
                continue
            label = (
                PROPERTY_TYPE_LABELS.get(c, c)
                if n in YES_VALUES or norm_header(value) == norm_header(c)
                else str(value).strip()
            )
            if label and norm_header(label) not in {norm_header(v) for v in values}:
                values.append(label)
        return " | ".join(values) if values else pd.NA

    return df[available].apply(derive, axis=1)


def classification_from_storey_value(value):
    """Return Low-rise, Mid-rise, or High-rise when storey evidence supports one band.

    Examples:
    - 4 -> Low-rise
    - 11 -> Mid-rise
    - 12 -> High-rise
    - "14 or 15" -> High-rise because both reported counts remain in one band
    - "11 or 12" -> unresolved because the conflicting counts cross a band boundary
    """
    if is_unresolved(value):
        return pd.NA

    numbers = [
        int(number)
        for number in re.findall(r"\b\d+\b", str(value))
        if int(number) > 0
    ]
    if not numbers:
        return pd.NA

    labels = set()
    for storeys in numbers:
        if 1 <= storeys <= 4:
            labels.add("Low-rise")
        elif 5 <= storeys <= 11:
            labels.add("Mid-rise")
        elif storeys >= 12:
            labels.add("High-rise")

    return next(iter(labels)) if len(labels) == 1 else pd.NA


def derive_height_classification_from_storeys(series: pd.Series) -> pd.Series:
    """Derive the project's controlled height classification from storey evidence."""
    return series.apply(classification_from_storey_value).astype("object")


def normalize_height_classification(value):
    """Normalize common wording to the project's three controlled labels."""
    if is_unresolved(value):
        return pd.NA
    key = norm_header(value)
    mapping = {
        "lowrise": "Low-rise",
        "midrise": "Mid-rise",
        "highrise": "High-rise",
        "hirise": "High-rise",
    }
    return mapping.get(key, str(value).strip())


def _property_type_parts(value) -> list[str]:
    """Split and deduplicate property-form labels without treating them as height bands."""
    if is_unresolved(value):
        return []
    parts = re.split(r"\s*(?:\||;|,)\s*", str(value).strip())
    normalized = []
    seen = set()
    for part in parts:
        clean = str(part).strip()
        key = norm_header(clean)
        if clean and key and key not in seen:
            seen.add(key)
            normalized.append(clean)
    return normalized


def merge_property_type_values(*values):
    """Combine property-form labels only, for example Apartment | Townhome."""
    parts = []
    seen = set()
    for value in values:
        for part in _property_type_parts(value):
            key = norm_header(part)
            if key and key not in seen:
                seen.add(key)
                parts.append(part)
    return " | ".join(parts) if parts else pd.NA


def property_type_from_legacy_classification(value):
    """Recover only recognized property-form labels from older mixed classifications."""
    recognized = {
        "apartment": "Apartment",
        "apartmentbuilding": "Apartment",
        "apartmentunit": "Apartment",
        "condominium": "Condominium",
        "condo": "Condominium",
        "condominiumrental": "Condominium",
        "townhome": "Townhome",
        "townhouse": "Townhome",
        "duplex": "Duplex",
        "gardenhome": "Garden Home",
        "semidetached": "Semi-detached",
        "semidetachedhome": "Semi-detached",
        "detached": "Detached",
        "detachedhome": "Detached",
        "detachedsinglefamilyhome": "Detached",
    }
    recovered = []
    seen = set()
    for part in _property_type_parts(value):
        key = norm_header(part)
        label = recognized.get(key)
        if label and norm_header(label) not in seen:
            seen.add(norm_header(label))
            recovered.append(label)
    return " | ".join(recovered) if recovered else pd.NA


def ensure_ids(df):
    """Ensure every Datablix row has one unique stable Record ID."""
    out = df.copy()
    supplied_ids = [
        str(value).strip()
        for value in out["Record ID"]
        if not is_unresolved(value)
    ]
    reserved = set(supplied_ids)
    used = set()
    result = []
    counter = 1

    for value in out["Record ID"]:
        clean = "" if is_unresolved(value) else str(value).strip()
        if clean and clean not in used:
            result.append(clean)
            used.add(clean)
            continue

        while f"DB-{counter:04d}" in reserved or f"DB-{counter:04d}" in used:
            counter += 1
        candidate = f"DB-{counter:04d}"
        result.append(candidate)
        used.add(candidate)
        counter += 1

    out["Record ID"] = result
    return out


def map_schema(df):
    imported = prepare_data(df)
    mapped = imported.copy()
    rows = []
    for target in INTERNAL_COLUMNS:
        matches = source_columns(imported, ALIASES.get(target, [target]))
        if matches:
            mapped[target] = combine_columns(imported, matches)
            rows.append({"Datablix Field": target, "Imported Column(s)": ", ".join(matches), "Mapping Status": "Mapped"})
        else:
            mapped[target] = pd.NA
            rows.append({"Datablix Field": target, "Imported Column(s)": "None", "Mapping Status": "Not found"})

    combined = source_columns(imported, COMBINED_LOCATION_ALIASES)
    if combined:
        parsed = pd.DataFrame(
            combine_columns(imported, combined).apply(parse_combined_location).tolist(),
            columns=["City", "Province", "Postal Code"], index=imported.index,
        )
        for field in ["City", "Province", "Postal Code"]:
            current, derived = resolved(mapped[field]), resolved(parsed[field])
            mask = unresolved_mask(current) & ~unresolved_mask(derived)
            current.loc[mask] = derived.loc[mask]
            mapped[field] = current
            if mask.any():
                for row in rows:
                    if row["Datablix Field"] == field and row["Mapping Status"] == "Not found":
                        row["Imported Column(s)"] = ", ".join(combined)
                        row["Mapping Status"] = "Derived"

    # Keep property form and height classification as separate dimensions.
    # Property Type preserves supported labels such as Apartment, Townhome, Duplex,
    # Garden Home, or Semi-detached. Building Classification is height-only and is
    # derived exclusively from verified Number of Storeys.
    supplied_property_type = resolved(mapped["Property Type"])
    legacy_property_type = derive_property_type(imported)
    combined_property_type = pd.Series(pd.NA, index=mapped.index, dtype="object")
    for idx in mapped.index:
        combined_property_type.loc[idx] = merge_property_type_values(
            supplied_property_type.loc[idx],
            legacy_property_type.loc[idx],
        )
    mapped["Property Type"] = combined_property_type

    mapped["Building Classification"] = derive_height_classification_from_storeys(
        mapped["Number of Storeys"]
    )

    source = resolved(mapped["Source URL"])
    website = resolved(mapped["Website"])
    mask = unresolved_mask(source) & ~unresolved_mask(website)
    source.loc[mask] = website.loc[mask]
    mapped["Source URL"] = source

    mapped["Province"] = mapped["Province"].apply(canonical_province)
    mapped["Postal Code"] = mapped["Postal Code"].apply(postal_code)
    mapped["Management/Owner"] = mapped["Management/Owner"].apply(
        lambda v: pd.NA if is_unresolved(v) else re.sub(r"\s+", " ", str(v)).strip()
    )
    mapped = ensure_ids(mapped)

    imported_mask = pd.Series(False, index=mapped.index)
    for c in ["Building Name", "Management/Owner", "Street Address", "City", "Website", "Phone"]:
        imported_mask |= ~unresolved_mask(mapped[c])
    mapped.loc[imported_mask & unresolved_mask(mapped["Research Status"]), "Research Status"] = "Imported - Needs Review"

    canonical = [c for c in INTERNAL_COLUMNS if c in mapped.columns]
    originals = [c for c in imported.columns if c not in canonical]
    return normalize_workflow(mapped[canonical + originals]), pd.DataFrame(rows)


def validate_input(df):
    groups = [ALIASES["Building Name"], ALIASES["Management/Owner"], ALIASES["Street Address"], ALIASES["City"], COMBINED_LOCATION_ALIASES, ALIASES["Website"], ALIASES["Phone"]]
    if sum(bool(source_columns(df, g)) for g in groups) < 2:
        raise ValueError(
            "Datablix could not find rental property columns in this worksheet. "
            "Pick the tab where the first row contains headings such as building name, "
            "address, or owner, and each row below is one building."
        )


def excel_sheet_names(uploaded):
    with pd.ExcelFile(io.BytesIO(uploaded.getvalue()), engine="openpyxl") as workbook:
        return workbook.sheet_names


def preferred_sheet(names):
    keywords = ["working", "research", "apartmentbuildings", "buildings", "directory", "listing"]
    normalized = [norm_header(n) for n in names]
    for keyword in keywords:
        for i, name in enumerate(normalized):
            if keyword in name:
                return i
    return 0


def read_upload(uploaded, sheet=None):
    data = uploaded.getvalue()
    extension = uploaded.name.rsplit(".", 1)[-1].lower()
    df = pd.read_csv(io.BytesIO(data)) if extension == "csv" else pd.read_excel(io.BytesIO(data), sheet_name=sheet, engine="openpyxl")
    return prepare_data(df), data


def sheet_id(url):
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9_-]+)", str(url))
    return match.group(1) if match else None


def sheet_gid(url):
    parsed = urlparse(str(url))
    for text in [parsed.query, parsed.fragment]:
        values = parse_qs(text).get("gid", [])
        if values and str(values[0]).isdigit():
            return str(values[0])
    match = re.search(r"(?:[?#&]gid=)(\d+)", str(url))
    return match.group(1) if match else None


def sheet_csv_url(url, selector=""):
    clean, selector = str(url).strip(), str(selector).strip()
    if not clean:
        raise ValueError("Paste a Google Sheets link first.")
    parsed = urlparse(clean)
    if "docs.google.com" in parsed.netloc.lower() and "/spreadsheets/d/e/" in parsed.path:
        parts = urlparse(clean.replace("/pubhtml", "/pub"))
        query = parse_qs(parts.query)
        query["output"] = ["csv"]
        if selector.isdigit():
            query["gid"] = [selector]
        return urlunparse(parts._replace(query=urlencode({k: v[-1] for k, v in query.items() if v})))
    if clean.lower().endswith(".csv") or "output=csv" in clean.lower():
        return clean
    sid = sheet_id(clean)
    if not sid:
        raise ValueError(
            "This link does not look like a Google Sheets sharing link. "
            "Copy it from Share > Copy link inside the Sheet."
        )
    if selector and not selector.isdigit():
        return f"https://docs.google.com/spreadsheets/d/{sid}/gviz/tq?tqx=out:csv&sheet={quote(selector)}"
    gid = selector if selector.isdigit() else sheet_gid(clean)
    return f"https://docs.google.com/spreadsheets/d/{sid}/export?format=csv" + (f"&gid={gid}" if gid else "")


def read_google_sheet(url, selector=""):
    request = Request(sheet_csv_url(url, selector), headers={"User-Agent": "Mozilla/5.0 (compatible; Datablix/1.0)"})
    try:
        with urlopen(request, timeout=30) as response:
            data = response.read()
            content_type = response.headers.get("Content-Type", "").lower()
    except Exception as error:
        raise ValueError(
            "Datablix could not open this Google Sheet. Check the link, then set the Sheet's "
            "General access to 'Anyone with the link' with the Viewer role and try again."
        ) from error
    preview = data[:500].decode("utf-8", errors="ignore").lower()
    if "text/html" in content_type or "<html" in preview:
        raise ValueError(
            "Google returned a webpage instead of spreadsheet data. Set the Sheet's General "
            "access to 'Anyone with the link', or paste a published CSV link instead."
        )
    try:
        df = pd.read_csv(io.BytesIO(data))
    except Exception as error:
        raise ValueError(
            "The Sheet opened, but its first row could not be read as column headings. "
            "Make sure row 1 contains the column names."
        ) from error
    sid = sheet_id(url)
    return prepare_data(df), data, f"google_sheet_{sid[:10]}.csv" if sid else "google_sheet.csv", selector or sheet_gid(url) or "linked worksheet"




AI_RESEARCH_DELIVERABLE_COLUMNS = [
    "Building Name", "Street Address", "Address Line 2", "City", "Province",
    "Postal Code", "Country", "Mailing Address", "PO Box", "PO Box City",
    "PO Box Province", "PO Box Postal Code", "PO Box Search Status",
    "PO Box Source URL", "PO Box Evidence", "PO Box Confidence",
    "Latitude", "Longitude", "Geocoded Municipality",
    "Geographic Scope Status", "Geographic Evidence", "Geographic Confidence",
    "Management/Owner", "Phone", "Primary Email",
    "Secondary Email", "Property Website", "Company Website", "Source URL",
    "Number of Apartments", "Apartment Count Search Status",
    "Apartment Count Source URL", "Apartment Count Evidence",
    "Apartment Count Confidence", "Property Type", "Number of Storeys",
    "Storey Count Search Status", "Storey Count Source URL",
    "Storey Count Evidence", "Storey Count Confidence", "Rental Rate Range",
    "Suite Types", "Amenities", "Parking",
    "Laundry", "Utilities", "Elevator", "Accessibility", "Pet Policy",
    "Smoke-Free",
    "Current Inventory Status", "Inventory Evidence",
    "Rental Availability Status", "Rental Availability Detail",
    "Rental Availability Evidence",
    "Found on City/Portfolio Page", "Found on HTML Sitemap",
    "Found on XML Sitemap", "Inventory Exclusion Reason",
    "Supporting Evidence", "Confidence", "Missing Information",
    "Reviewer Notes",
]


def ai_research_template(company_name: str = "", company_website: str = "") -> pd.DataFrame:
    """Return a blank spreadsheet structure for external AI research deliverables."""
    row = {column: "" for column in AI_RESEARCH_DELIVERABLE_COLUMNS}
    row["Management/Owner"] = str(company_name or "").strip()
    row["Company Website"] = str(company_website or "").strip()
    return pd.DataFrame([row])


def company_source_records_for_research(
    df: pd.DataFrame,
    company_id: str = "",
    company_name: str = "",
) -> pd.DataFrame:
    """Return *all* current Starting Data rows for one company.

    Source Records means rows that physically exist in the current Starting Data.
    Geographic scope is assessed separately and must never silently remove a source
    row from this count.  Company ID and canonical company-name matching are UNIONED
    so a stale/partial ID assignment cannot hide valid Starting Data rows.
    """
    if not isinstance(df, pd.DataFrame) or df.empty:
        return pd.DataFrame()

    rows = df.copy()
    clean_company_id = safe_text(company_id)
    clean_company_name = safe_text(company_name)
    match_mask = pd.Series(False, index=rows.index)

    # Keep compatible rows carrying the canonical Company ID.
    if clean_company_id and "Company ID" in rows.columns:
        id_mask = rows["Company ID"].astype("string").fillna("").str.strip().eq(clean_company_id)
        if clean_company_name and "Management/Owner" in rows.columns:
            owner_compatible = rows["Management/Owner"].apply(
                lambda value: (
                    not safe_text(value)
                    or _company_core_matches(safe_text(value), clean_company_name)
                )
            )
            id_mask &= owner_compatible
        match_mask |= id_mask

    # UNION canonical/alias-aware owner-name matches with ID matches.  This is the
    # important difference from v92: name matches are not merely a fallback when
    # some ID rows already exist.  Therefore Lepine's 3 out-of-scope rows still
    # remain part of its original Starting Data count.
    if clean_company_name and "Management/Owner" in rows.columns:
        name_mask = rows["Management/Owner"].apply(
            lambda value: _company_core_matches(
                safe_text(value),
                clean_company_name,
            )
        )
        match_mask |= name_mask

    matched = rows.loc[match_mask].copy()
    if matched.empty:
        return matched

    preferred_columns = [
        "Record ID",
        "Building Name",
        "Street Address",
        "Address Line 2",
        "City",
        "Province",
        "Postal Code",
        "Country",
        "Geographic Scope Status",
        "Geographic Evidence",
        "Geographic Confidence",
        "Property Type",
        "Building Classification",
        "Number of Storeys",
        "Number of Apartments",
        "Management/Owner",
        "Phone",
        "Primary Email",
        "Website",
        "Property Website",
        "Source URL",
    ]
    available = [column for column in preferred_columns if column in matched.columns]
    if available:
        matched = matched[available].copy()

    # Starting Data can legitimately contain blanks. Export them as blank cells.
    return matched.replace({pd.NA: ""}).fillna("")


def _starting_source_scope_status(row) -> str:
    """Classify source geography without changing or filtering the source baseline.

    Prefer an explicit Geographic Scope Status.  When source rows have not yet been
    geocoded, use only the project's explicit municipality/province/country labels.
    Uncertain locations remain Needs Geographic Review rather than being guessed.
    """
    explicit = safe_text(row.get("Geographic Scope Status", ""))
    if explicit in {"Inside City of Ottawa", "Outside City of Ottawa", "Needs Geographic Review"}:
        return explicit

    city = safe_text(row.get("City", "")).lower()
    province = safe_text(row.get("Province", "")).lower()
    country = safe_text(row.get("Country", "")).lower()

    if country and country not in {"canada", "ca"}:
        return "Outside City of Ottawa"
    if province and province not in {"ontario", "on"}:
        return "Outside City of Ottawa"
    if city in OUT_OF_SCOPE_NEARBY_LOCALITIES:
        return "Outside City of Ottawa"
    if city in OTTAWA_MUNICIPAL_LOCALITIES:
        return "Inside City of Ottawa"
    return "Needs Geographic Review"


def starting_data_geographic_scope_table(
    baseline: pd.DataFrame,
    registry=None,
) -> pd.DataFrame:
    """Return one row per Starting Data record with a separate scope assessment."""
    registry = normalize_company_registry(registry)
    columns = [
        "Company", "Building Name", "Street Address", "City", "Province",
        "Postal Code", "Scope Status", "Scope Basis",
    ]
    if not isinstance(baseline, pd.DataFrame) or baseline.empty or registry.empty:
        return pd.DataFrame(columns=columns)

    output_rows = []
    seen = set()
    for _, company in registry.iterrows():
        company_id = safe_text(company.get("Company ID", ""))
        company_name = safe_text(company.get("Management/Owner", "")) or company_id or "Unassigned"
        source = company_source_records_for_research(
            baseline,
            company_id=company_id,
            company_name=company_name,
        )
        for source_index, row in source.iterrows():
            record_id = safe_text(row.get("Record ID", ""))
            identity = record_id or "|".join([
                company_name,
                safe_text(row.get("Building Name", "")),
                safe_text(row.get("Street Address", "")),
                safe_text(row.get("City", "")),
                safe_text(row.get("Postal Code", "")),
            ])
            if identity in seen:
                continue
            seen.add(identity)

            status = _starting_source_scope_status(row)
            explicit = safe_text(row.get("Geographic Scope Status", ""))
            if explicit in {"Inside City of Ottawa", "Outside City of Ottawa", "Needs Geographic Review"}:
                basis = "Recorded geographic scope status"
            else:
                basis = "Starting Data municipality/province/country label"

            output_rows.append({
                "Company": company_name,
                "Building Name": safe_text(row.get("Building Name", "")),
                "Street Address": safe_text(row.get("Street Address", "")),
                "City": safe_text(row.get("City", "")),
                "Province": safe_text(row.get("Province", "")),
                "Postal Code": safe_text(row.get("Postal Code", "")),
                "Scope Status": status,
                "Scope Basis": basis,
            })

    return pd.DataFrame(output_rows, columns=columns)


def company_source_presence_reconciliation(
    source_records: pd.DataFrame,
    research_records: pd.DataFrame,
) -> pd.DataFrame:
    """Compare one company's Starting Data with every returned research row.

    This is a presence/reconciliation view for human review, not an automatic
    inventory decision.  Strong matches use the same >=88 threshold as Datablix
    discovery classification; scores from 72-87 remain Possible match unless a
    reviewer has already manually accepted that research row as an Existing Source
    Record.  Removed/excluded research rows still count as rediscovered because the
    purpose here is to show whether a Starting Data record appeared anywhere in the
    returned research, not whether it remains active inventory.
    """
    columns = [
        "Starting Source Position",
        "Matched Research Position",
        "Starting Record ID",
        "Starting Record",
        "Street Address",
        "City",
        "Postal Code",
        "Starting Source URL",
        "Reconciliation",
        "Best Match Score",
        "Match Reason",
        "Matched Research Record",
        "Research Result State",
    ]
    if not isinstance(source_records, pd.DataFrame) or source_records.empty:
        return pd.DataFrame(columns=columns)

    source = coalesce_duplicate_columns(source_records.copy()).reset_index(drop=True)
    research = (
        coalesce_duplicate_columns(research_records.copy()).reset_index(drop=True)
        if isinstance(research_records, pd.DataFrame)
        else pd.DataFrame()
    )

    source_identities, source_indexes = _build_discovery_source_index(source)
    best_matches = {
        position: {"score": 0, "reason": "no credible source match", "research_position": None}
        for position in range(len(source))
    }
    research_rows = research.to_dict(orient="records") if not research.empty else []

    for research_position, research_row in enumerate(research_rows):
        research_identity = _row_identity(research_row)
        candidate_positions = _candidate_source_positions(
            research_identity, source_indexes
        )
        for source_position in candidate_positions:
            score, reason = _source_match_score_from_identities(
                research_identity, source_identities[source_position]
            )
            if score > best_matches[source_position]["score"]:
                best_matches[source_position] = {
                    "score": int(score),
                    "reason": reason,
                    "research_position": research_position,
                }

    rows = []
    for source_position, source_row in enumerate(source.to_dict(orient="records")):
        match = best_matches[source_position]
        matched_row = (
            research_rows[match["research_position"]]
            if match["research_position"] is not None
            else {}
        )
        score = int(match["score"] or 0)
        manual_existing = bool(
            matched_row
            and safe_text(matched_row.get("Discovery Status Source", "")) == "Manual"
            and safe_text(matched_row.get("Directory Discovery Status", ""))
            == "Existing Source Record"
        )

        matched_rental_status = safe_text(
            matched_row.get("Rental Availability Status", "")
        ).lower()
        excluded_match = bool(
            matched_row
            and (
                safe_text(matched_row.get("Record Decision", "")) == "Remove"
                or safe_text(matched_row.get("Directory Discovery Status", ""))
                == "Excluded / Not Current"
                or safe_text(matched_row.get("Current Inventory Status", "")).lower().startswith("excluded")
                or matched_rental_status == "rented"
            )
        )

        if score >= 88 or (manual_existing and score >= 72):
            if matched_rental_status == "rented":
                reconciliation = "Rediscovered — rented"
            elif excluded_match:
                reconciliation = "Rediscovered — excluded/not current"
            else:
                reconciliation = "Rediscovered"
        elif score >= 72:
            reconciliation = "Possible match"
        else:
            reconciliation = "Not rediscovered"

        source_label = (
            safe_text(source_row.get("Building Name", ""))
            or safe_text(source_row.get("Street Address", ""))
            or safe_text(source_row.get("Record ID", ""))
            or "Unlabelled source record"
        )
        matched_label = (
            safe_text(matched_row.get("Working Record Label", ""))
            or safe_text(matched_row.get("Building Name", ""))
            or safe_text(matched_row.get("Street Address", ""))
            or safe_text(matched_row.get("Record ID", ""))
        )
        starting_url = (
            safe_text(source_row.get("Property Website", ""))
            or safe_text(source_row.get("Website", ""))
            or safe_text(source_row.get("Source URL", ""))
        )
        result_state_parts = [
            safe_text(matched_row.get("Current Inventory Status", "")),
            safe_text(matched_row.get("Rental Availability Status", "")),
            safe_text(matched_row.get("Rental Availability Detail", "")),
            safe_text(matched_row.get("Directory Discovery Status", "")),
            safe_text(matched_row.get("Record Decision", "")),
        ]
        result_state = " · ".join(
            dict.fromkeys(part for part in result_state_parts if part)
        )

        rows.append({
            "Starting Source Position": source_position,
            "Matched Research Position": match["research_position"] if match["research_position"] is not None else pd.NA,
            "Starting Record ID": safe_text(source_row.get("Record ID", "")),
            "Starting Record": source_label,
            "Street Address": safe_text(source_row.get("Street Address", "")),
            "City": safe_text(source_row.get("City", "")),
            "Postal Code": safe_text(source_row.get("Postal Code", "")),
            "Starting Source URL": starting_url,
            "Reconciliation": reconciliation,
            "Best Match Score": score,
            "Match Reason": match["reason"],
            "Matched Research Record": matched_label,
            "Research Result State": result_state,
        })

    return pd.DataFrame(rows, columns=columns)


def _comparison_ascii_key(value) -> str:
    """Return a punctuation-insensitive ASCII key for review-only comparisons."""
    text = unicodedata.normalize("NFKD", safe_text(value))
    text = "".join(character for character in text if not unicodedata.combining(character))
    return re.sub(r"[^a-z0-9]+", "", text.lower())


def _comparison_number(value) -> str:
    """Normalize ordinary integer-like counts without inventing a value."""
    text = safe_text(value)
    if not text:
        return ""
    compact = text.replace(",", "").strip()
    if re.fullmatch(r"\d+(?:\.0+)?", compact):
        return str(int(float(compact)))
    numbers = re.findall(r"(?<!\d)(\d{1,6})(?!\d)", compact)
    return numbers[0] if len(numbers) == 1 else ""


def _comparison_province(value) -> str:
    key = _comparison_ascii_key(value)
    aliases = {
        "on": "ontario",
        "ont": "ontario",
        "ontario": "ontario",
        "qc": "quebec",
        "pq": "quebec",
        "quebec": "quebec",
    }
    return aliases.get(key, key)


def _comparison_classification(value) -> str:
    """Normalize the height-only Building Classification for source comparison.

    Building Classification is a single controlled height band derived from
    Number of Storeys (Low-rise, Mid-rise, or High-rise). Property-form labels
    belong in Property Type, so this comparison must not use the older
    multi-part classification parser.
    """
    normalized = normalize_height_classification(value)
    if is_unresolved(normalized):
        return ""
    return _comparison_ascii_key(normalized)


def _comparison_field_key(field: str, value) -> str:
    """Normalize a field only enough to identify formatting-equivalent values."""
    text = safe_text(value)
    if not text:
        return ""
    if field == "Street Address":
        signature = _address_signature(text)
        if signature["full"]:
            return signature["full"]
        return _comparison_ascii_key(text)
    if field == "City":
        return _canonical_city(text)
    if field == "Province":
        return _comparison_province(text)
    if field == "Postal Code":
        return _canonical_postal(text)
    if field == "Phone":
        return _canonical_phone(text)
    if field == "Primary Email":
        return text.lower().replace(" ", "")
    if field in {"Website", "Property Website", "Source URL"}:
        return _canonical_url(text)
    if field in {"Number of Storeys", "Number of Apartments"}:
        return _comparison_number(text)
    if field == "Building Classification":
        return _comparison_classification(text)
    return _comparison_ascii_key(text)


def _comparison_text_similarity(first, second) -> float:
    first_key = _comparison_ascii_key(first)
    second_key = _comparison_ascii_key(second)
    if not first_key or not second_key:
        return 0.0
    return SequenceMatcher(None, first_key, second_key).ratio()


def _classify_field_difference(field: str, source_value, research_value) -> tuple[str, str]:
    """Classify one Starting Data vs research value without choosing a winner."""
    source_text = safe_text(source_value)
    research_text = safe_text(research_value)

    if not source_text and not research_text:
        return "Same", "Both blank"
    if not source_text and research_text:
        return "Added in research", "Research filled a field that was blank in Starting Data"
    if source_text and not research_text:
        return "Missing from research", "Starting Data had a value but the matched research row is blank"
    if source_text == research_text:
        return "Same", "Exact match"

    source_key = _comparison_field_key(field, source_text)
    research_key = _comparison_field_key(field, research_text)
    if source_key and research_key and source_key == research_key:
        return "Formatting only", "Same normalized value; spelling, punctuation, abbreviation, accents, or formatting differ"

    if field in {"Building Name", "Street Address", "Management/Owner", "City"}:
        similarity = _comparison_text_similarity(source_text, research_text)
        if similarity >= 0.87:
            return "Possible typo / minor text change", f"Very similar text ({similarity:.0%}); verify the preferred wording"

    return "Changed — verify", "Values differ materially; Datablix does not choose which one is correct"


def company_source_field_comparison(
    source_records: pd.DataFrame,
    research_records: pd.DataFrame,
    reconciliation: pd.DataFrame | None = None,
) -> pd.DataFrame:
    """Return field-level differences for strongly rediscovered Starting Data rows.

    The comparison is retroactive and read-only: it operates on whatever research
    records are already saved in the project and never mutates Starting Data or the
    working research dataframe.
    """
    columns = [
        "Starting Source Position",
        "Matched Research Position",
        "Matched Research Record ID",
        "Starting Record",
        "Matched Research Record",
        "Field",
        "Starting Data",
        "Research Result",
        "Comparison",
        "Comparison Note",
        "Best Match Score",
        "Match Reason",
    ]
    if not isinstance(source_records, pd.DataFrame) or source_records.empty:
        return pd.DataFrame(columns=columns)
    if not isinstance(research_records, pd.DataFrame) or research_records.empty:
        return pd.DataFrame(columns=columns)

    source = coalesce_duplicate_columns(source_records.copy()).reset_index(drop=True)
    research = coalesce_duplicate_columns(research_records.copy()).reset_index(drop=True)
    if reconciliation is None:
        reconciliation = company_source_presence_reconciliation(source, research)
    if not isinstance(reconciliation, pd.DataFrame) or reconciliation.empty:
        return pd.DataFrame(columns=columns)

    compare_fields = [
        "Building Name",
        "Street Address",
        "Address Line 2",
        "City",
        "Province",
        "Postal Code",
        "Property Type",
        "Building Classification",
        "Number of Storeys",
        "Number of Apartments",
        "Management/Owner",
        "Phone",
        "Primary Email",
        "Website",
        "Property Website",
        "Source URL",
    ]

    rows = []
    strong = reconciliation["Reconciliation"].astype(str).str.startswith("Rediscovered")
    for _, link in reconciliation.loc[strong].iterrows():
        try:
            source_position = int(link["Starting Source Position"])
            research_position = int(link["Matched Research Position"])
        except (TypeError, ValueError):
            continue
        if source_position not in source.index or research_position not in research.index:
            continue

        source_row = source.loc[source_position]
        research_row = research.loc[research_position]
        source_label = safe_text(link.get("Starting Record", ""))
        research_label = safe_text(link.get("Matched Research Record", ""))

        for field in compare_fields:
            source_value = source_row.get(field, "")
            research_value = research_row.get(field, "")
            comparison, note = _classify_field_difference(field, source_value, research_value)
            if comparison == "Same":
                continue
            rows.append({
                "Starting Source Position": source_position,
                "Matched Research Position": research_position,
                "Matched Research Record ID": safe_text(research_row.get("Record ID", "")),
                "Starting Record": source_label,
                "Matched Research Record": research_label,
                "Field": field,
                "Starting Data": safe_text(source_value),
                "Research Result": safe_text(research_value),
                "Comparison": comparison,
                "Comparison Note": note,
                "Best Match Score": int(link.get("Best Match Score", 0) or 0),
                "Match Reason": safe_text(link.get("Match Reason", "")),
            })

    return pd.DataFrame(rows, columns=columns)


def _render_company_source_field_comparison(
    company_source: pd.DataFrame,
    research_records: pd.DataFrame,
    reconciliation: pd.DataFrame,
    differences: pd.DataFrame | None = None,
) -> None:
    """Render Changes to review as a compact matrix plus complete change details.

    The matrix is the quick scan. Every meaningful source-vs-research difference
    remains visible below it, grouped by property, and reviewers can save notes to
    the matched record's existing Reviewer Notes field. Harmless formatting-only
    differences remain collapsed separately.
    """
    if differences is None:
        differences = company_source_field_comparison(
            company_source,
            research_records,
            reconciliation=reconciliation,
        )

    if not isinstance(differences, pd.DataFrame) or differences.empty:
        return

    material_labels = {
        "Changed — verify",
        "Possible typo / minor text change",
        "Added in research",
        "Missing from research",
    }
    material = differences.loc[differences["Comparison"].isin(material_labels)].copy()
    formatting = differences.loc[differences["Comparison"].eq("Formatting only")].copy()

    if not material.empty:
        priority_order = {
            "Changed — verify": 0,
            "Possible typo / minor text change": 1,
            "Missing from research": 2,
            "Added in research": 3,
        }
        simple_labels = {
            "Changed — verify": "Changed",
            "Possible typo / minor text change": "Possible typo",
            "Missing from research": "Missing in new research",
            "Added in research": "New information",
        }
        material["_priority"] = material["Comparison"].map(priority_order).fillna(9)
        material = material.sort_values(
            ["Starting Record", "_priority", "Field"], kind="stable"
        ).drop(columns=["_priority"])
        material["Change"] = material["Comparison"].map(simple_labels).fillna(
            material["Comparison"]
        )

        def compact_value(value, limit=24):
            value = safe_text(value)
            if len(value) <= limit:
                return value
            return value[: max(0, limit - 1)].rstrip() + "…"

        def matrix_cell(group: pd.DataFrame, field: str) -> str:
            rows = group.loc[group["Field"].eq(field)]
            if rows.empty:
                return "✓"
            row = rows.iloc[0]
            before = safe_text(row.get("Starting Data", ""))
            now = safe_text(row.get("Research Result", ""))
            comparison = safe_text(row.get("Comparison", ""))
            if comparison == "Added in research":
                return f"New: {compact_value(now)}" if now else "New"
            if comparison == "Missing from research":
                return "Missing"
            if comparison == "Possible typo / minor text change":
                return "Check wording"
            if field in {"Number of Apartments", "Number of Storeys"}:
                return f"{compact_value(before, 12)} → {compact_value(now, 12)}"
            if field == "Phone" and before and now and len(before) <= 18 and len(now) <= 18:
                return f"{before} → {now}"
            if field in {"Property Type", "Building Classification"} and before and now and len(before) + len(now) <= 50:
                return f"{before} → {now}"
            return "Updated"

        matrix_fields = [
            ("Apartments", "Number of Apartments"),
            ("Storeys", "Number of Storeys"),
            ("Phone", "Phone"),
            ("Email", "Primary Email"),
            ("Website", "Website"),
            ("Property Type", "Property Type"),
            ("Classification", "Building Classification"),
        ]
        matrix_rows = []
        for property_name, group in material.groupby("Starting Record", sort=False):
            row = {"Property": property_name}
            for display_name, field_name in matrix_fields:
                row[display_name] = matrix_cell(group, field_name)

            changed_fields = len(group)
            review_count = int(
                group["Comparison"].isin({
                    "Changed — verify",
                    "Possible typo / minor text change",
                    "Missing from research",
                }).sum()
            )
            new_count = int(group["Comparison"].eq("Added in research").sum())
            if review_count:
                row["Overall"] = f"⚠ {changed_fields} change" + ("" if changed_fields == 1 else "s")
            elif new_count:
                row["Overall"] = f"ℹ {new_count} new"
            else:
                row["Overall"] = f"{changed_fields} change" + ("" if changed_fields == 1 else "s")
            matrix_rows.append(row)

        matrix = pd.DataFrame(matrix_rows)
        changed_records = len(matrix)

        st.markdown("#### Changes to review")
        st.caption(
            f"{changed_records:,} properties · {len(material):,} meaningful differences. "
            "The matrix is shown directly so it is always visible."
        )
        st.caption(
            "One row per property. ✓ means no meaningful difference was detected for that field. "
            "Less common changed fields are not added as extra matrix columns; every change is shown below."
        )
        st.dataframe(
            matrix,
            width="stretch",
            hide_index=True,
            column_config={
                "Property": st.column_config.TextColumn("Property", width="medium"),
                "Apartments": st.column_config.TextColumn("Apartments", width="small"),
                "Storeys": st.column_config.TextColumn("Storeys", width="small"),
                "Phone": st.column_config.TextColumn("Phone", width="medium"),
                "Email": st.column_config.TextColumn("Email", width="medium"),
                "Website": st.column_config.TextColumn("Website", width="small"),
                "Classification": st.column_config.TextColumn("Classification", width="medium"),
                "Overall": st.column_config.TextColumn("Overall", width="small"),
            },
        )

        st.markdown("#### Review change details")
        st.caption(
            "The matrix above shows every property at a glance. Choose one property only when you want "
            "to inspect its exact Before → New Research differences or add research notes."
        )

        research_reset = (
            coalesce_duplicate_columns(research_records.copy()).reset_index(drop=True)
            if isinstance(research_records, pd.DataFrame)
            else pd.DataFrame()
        )

        property_options = list(dict.fromkeys(material["Starting Record"].dropna().astype(str)))
        selected_property = st.selectbox(
            "Property to review",
            property_options,
            index=None,
            placeholder="Select a property to view its changes",
            key="db_source_change_property_selector",
        )

        if selected_property:
            group = material.loc[
                material["Starting Record"].astype(str).eq(str(selected_property))
            ].copy()
            detail = group.rename(
                columns={
                    "Starting Data": "Before",
                    "Research Result": "Now",
                }
            )[["Field", "Before", "Now", "Change"]]
            property_change_count = len(detail)

            st.markdown(
                f"**{selected_property}** · {property_change_count} change"
                f"{'' if property_change_count == 1 else 's'}"
            )
            st.dataframe(
                detail,
                width="stretch",
                hide_index=True,
                column_config={
                    "Field": st.column_config.TextColumn("Field", width="medium"),
                    "Before": st.column_config.TextColumn("Starting Data", width="medium"),
                    "Now": st.column_config.TextColumn("New Research", width="medium"),
                    "Change": st.column_config.TextColumn("Change", width="small"),
                },
            )

            first = group.iloc[0]
            research_position = pd.to_numeric(
                first.get("Matched Research Position", pd.NA), errors="coerce"
            )
            research_record_id = safe_text(first.get("Matched Research Record ID", ""))
            current_note = ""
            if pd.notna(research_position):
                pos = int(research_position)
                if pos in research_reset.index:
                    current_note = safe_text(
                        research_reset.loc[pos].get("Reviewer Notes", "")
                    )
                    if not research_record_id:
                        research_record_id = safe_text(
                            research_reset.loc[pos].get("Record ID", "")
                        )

            note_identity = research_record_id or (
                f"{selected_property}|{int(research_position) if pd.notna(research_position) else 'x'}"
            )
            note_hash = hashlib.sha256(note_identity.encode("utf-8")).hexdigest()[:12]
            with st.form(f"db_change_research_notes_{note_hash}"):
                note_text = st.text_area(
                    "Research notes about these changes",
                    value=current_note,
                    height=110,
                    placeholder=(
                        "Example: Verified the new unit count in a planning document; "
                        "phone number still needs confirmation."
                    ),
                    help=(
                        "Saved to this property's Reviewer Notes field and retained with the project."
                    ),
                )
                save_note = st.form_submit_button(
                    "Save research notes",
                    type="primary",
                    width="stretch",
                )

            if save_note:
                working = st.session_state.get(S_WORKING)
                if not isinstance(working, pd.DataFrame) or working.empty:
                    st.error("The working project records are not available to save this note.")
                else:
                    updated = working.copy()
                    if "Reviewer Notes" not in updated.columns:
                        updated["Reviewer Notes"] = ""
                    matched_mask = pd.Series(False, index=updated.index)
                    if research_record_id and "Record ID" in updated.columns:
                        matched_mask = (
                            updated["Record ID"].astype("string").fillna("").str.strip()
                            .eq(research_record_id)
                        )
                    if not matched_mask.any() and pd.notna(research_position):
                        # Safe fallback for older records without a reliable Record ID.
                        pos = int(research_position)
                        if pos in research_reset.index:
                            target_identity = _row_identity(research_reset.loc[pos].to_dict())
                            for working_index, working_row in updated.iterrows():
                                score, _ = _source_match_score_from_identities(
                                    _row_identity(working_row.to_dict()),
                                    target_identity,
                                )
                                if score >= 88:
                                    matched_mask.at[working_index] = True
                                    break
                    if matched_mask.any():
                        updated.loc[matched_mask, "Reviewer Notes"] = note_text.strip()
                        st.session_state[S_WORKING] = updated
                        st.session_state[S_EDIT_COUNT] = (
                            st.session_state.get(S_EDIT_COUNT, 0) + 1
                        )
                        autosave_current_project()
                        st.success("Research notes saved for this property.")
                    else:
                        st.error(
                            "Datablix could not safely link this note to the saved research record."
                        )

        st.info(
            "Datablix never overwrites Starting Data or research values from this comparison. "
            "Verify factual changes against the supporting source before deciding which value to keep."
        )

    if not formatting.empty:
        with smart_expander(
            "Technical details: formatting differences",
            count=len(formatting),
            status="usually no factual conflict",
            expanded=False,
        ):
            st.caption(
                "Hidden by default because these values normalize to the same information—for example "
                "St vs Street, accents, phone formatting, capitalization, or equivalent URL formatting."
            )
            formatting_view = formatting.rename(
                columns={
                    "Starting Record": "Property",
                    "Starting Data": "Before",
                    "Research Result": "Now",
                }
            )[["Property", "Field", "Before", "Now"]]
            st.dataframe(formatting_view, width="stretch", hide_index=True)


def render_company_source_presence_reconciliation(
    company_id: str,
    research_records: pd.DataFrame,
) -> None:
    """Render a low-cognitive-load Starting Data vs research comparison."""
    company_id = safe_text(company_id)
    if not company_id:
        return

    registry = normalize_company_registry(st.session_state.get(S_COMPANIES))
    company_name = ""
    if not registry.empty:
        exact = registry.loc[registry["Company ID"].astype(str).eq(company_id)]
        if not exact.empty:
            company_name = safe_text(exact.iloc[0].get("Management/Owner", ""))
    if not company_name and isinstance(research_records, pd.DataFrame) and not research_records.empty:
        owner_values = research_records.get("Management/Owner", pd.Series(dtype="object"))
        owner_values = owner_values.apply(safe_text)
        owner_values = owner_values.loc[owner_values.ne("")]
        if not owner_values.empty:
            company_name = owner_values.iloc[0]

    starting_data = current_starting_source_records()
    company_source = company_source_records_for_research(
        starting_data,
        company_id=company_id,
        company_name=company_name,
    )
    if company_source.empty:
        return

    reconciliation = company_source_presence_reconciliation(
        company_source, research_records
    )
    if reconciliation.empty:
        return

    differences = company_source_field_comparison(
        company_source,
        research_records,
        reconciliation=reconciliation,
    )
    material_labels = {
        "Changed — verify",
        "Possible typo / minor text change",
        "Added in research",
        "Missing from research",
    }
    material = (
        differences.loc[differences["Comparison"].isin(material_labels)].copy()
        if isinstance(differences, pd.DataFrame) and not differences.empty
        else pd.DataFrame()
    )

    rediscovered_mask = reconciliation["Reconciliation"].str.startswith("Rediscovered")
    possible_mask = reconciliation["Reconciliation"].eq("Possible match")
    missing_mask = reconciliation["Reconciliation"].eq("Not rediscovered")
    excluded_mask = reconciliation["Reconciliation"].eq("Rediscovered — excluded/not current")

    found_count = int(rediscovered_mask.sum())
    possible_count = int(possible_mask.sum())
    missing_count = int(missing_mask.sum())
    excluded_count = int(excluded_mask.sum())

    review_labels = set()
    if not material.empty:
        review_labels.update(material["Starting Record"].dropna().astype(str))
    review_labels.update(
        reconciliation.loc[possible_mask, "Starting Record"].dropna().astype(str)
    )
    review_labels.update(
        reconciliation.loc[excluded_mask, "Starting Record"].dropna().astype(str)
    )
    need_review_count = len(review_labels)

    changed_labels = set(material["Starting Record"].dropna().astype(str)) if not material.empty else set()
    excluded_labels = set(
        reconciliation.loc[excluded_mask, "Starting Record"].dropna().astype(str)
    )
    clean_found_count = max(0, found_count - len(changed_labels | excluded_labels))

    st.divider()
    st.markdown("### Source comparison")
    st.caption(
        "A simple check of this company's Starting Data against the research already saved in Datablix. Nothing is overwritten automatically."
    )
    metric_columns = st.columns(4)
    metric_columns[0].metric("Starting records", f"{len(reconciliation):,}")
    metric_columns[1].metric("Found again", f"{found_count:,}")
    metric_columns[2].metric("Need review", f"{need_review_count:,}")
    metric_columns[3].metric("Not found", f"{missing_count:,}")

    if clean_found_count:
        st.success(
            f"{clean_found_count:,} Starting Data record(s) were found again with no important field changes to review."
        )
    if not need_review_count and not missing_count:
        st.success("The Starting Data and saved research are aligned for this company.")

    _render_company_source_field_comparison(
        company_source,
        research_records,
        reconciliation,
        differences=differences,
    )

    if missing_count:
        with smart_expander(
            "Not found in new research",
            count=missing_count,
            status="manual check recommended",
            expanded=True,
        ):
            st.warning(
                "Not found does not mean the property is inactive or no longer managed. It only means Datablix could not find a credible matching research row."
            )
            missing = reconciliation.loc[missing_mask, [
                "Starting Record ID", "Starting Record", "Street Address",
                "City", "Postal Code", "Starting Source URL",
            ]].copy().rename(
                columns={
                    "Starting Record ID": "Source ID",
                    "Starting Record": "Property",
                    "Starting Source URL": "Source",
                }
            )
            st.dataframe(
                missing,
                width="stretch",
                hide_index=True,
                column_config={
                    "Source": st.column_config.LinkColumn(
                        "Source", display_text="Open source"
                    ),
                },
            )

    if possible_count:
        with smart_expander(
            "Possible same property",
            count=possible_count,
            status="confirm the match",
            expanded=False,
        ):
            st.caption(
                "Datablix found similarities, but not enough evidence for a strong automatic match. Compare the two records before confirming they are the same property."
            )
            possible = reconciliation.loc[possible_mask, [
                "Starting Record", "Street Address", "Postal Code",
                "Matched Research Record",
            ]].copy().rename(
                columns={
                    "Starting Record": "Starting Data property",
                    "Matched Research Record": "Possible research match",
                }
            )
            st.dataframe(possible, width="stretch", hide_index=True)

    if excluded_count:
        with smart_expander(
            "Found, but marked no longer current",
            count=excluded_count,
            status="review status",
            expanded=False,
        ):
            st.caption(
                "These properties were found in the research, so they are not missing. The saved research currently marks them excluded, removed, or not current."
            )
            excluded = reconciliation.loc[excluded_mask, [
                "Starting Record", "Street Address", "Matched Research Record",
                "Research Result State",
            ]].copy().rename(
                columns={
                    "Starting Record": "Starting Data property",
                    "Matched Research Record": "Research property",
                    "Research Result State": "Current research status",
                }
            )
            st.dataframe(excluded, width="stretch", hide_index=True)



def build_research_package_bytes(
    company_name: str,
    prompt_text: str,
    project_source_records: pd.DataFrame,
    company_source_records: pd.DataFrame,
    research_template: pd.DataFrame,
    source_meta: dict | None = None,
    raw_source_bytes: bytes = b"",
    raw_source_filename: str = "",
) -> bytes:
    """Create a package containing the actual current project source file."""
    company_stem = safe_filename(company_name)
    prompt_name = f"{company_stem}_website_research_prompt.txt"
    meta = source_meta if isinstance(source_meta, dict) else {}
    source_version = safe_text(meta.get("version_label", "")) or f"v{_safe_int(meta.get('version_number', 1), 1)}"

    original_source_name = safe_text(raw_source_filename or meta.get("workbook_name", ""))
    original_source_name = Path(original_source_name).name if original_source_name else f"project_starting_source_{source_version}.xlsx"
    structured_source_name = f"project_starting_source_records_{safe_filename(source_version)}.csv"
    company_source_name = f"{company_stem}_source_matches_{safe_filename(source_version)}.csv"
    template_name = f"{company_stem}_research_template.csv"

    source_mode = safe_text(meta.get("source_mode", "")) or (
        "Structured records + original file"
        if isinstance(project_source_records, pd.DataFrame) and not project_source_records.empty
        else "Original project source file"
    )
    structured_count = len(project_source_records) if isinstance(project_source_records, pd.DataFrame) else 0
    company_match_count = len(company_source_records) if isinstance(company_source_records, pd.DataFrame) else 0

    readme_lines = [
        "DATABLIX RESEARCH PACKAGE",
        "",
        "ACTIVE COMPANY",
        company_name,
        "",
        "CURRENT PROJECT SOURCE",
        f"Source version: {source_version}",
        f"Source mode: {source_mode}",
        f"Original project source file: {original_source_name}",
        f"Structured source rows detected by Datablix: {structured_count:,}",
        f"Company-specific source matches detected: {company_match_count:,}",
        "",
        "HOW TO USE THIS PACKAGE",
        "1. Upload the research prompt to your AI research tool. Use the blank research template only if the tool needs a schema reference.",
        "2. Do NOT give the AI Starting Data for company comparison. Website research must be independent; Datablix performs source reconciliation after import.",
        f"3. For {company_name}, establish the current official website/listing inventory from authoritative company pages.",
        "4. For every current listing, separately verify rental availability. A property can remain on a For Rent page after it has been rented.",
        "5. Record explicit statuses such as FOR RENT, AVAILABLE SEP/OCT 2026, RENTED, LEASED, or NO LONGER AVAILABLE in Rental Availability Status/Detail/Evidence.",
        "6. Do not treat page existence, a displayed rent, or placement on a For Rent page as proof that a listing is currently rentable.",
        "7. Return exactly ONE completed research CSV using the required headings.",
        "8. Keep explicit Rented rows when they are still identifiable on a current official inventory page so Datablix can reconcile and remove them from the active directory.",
        "9. Do not create separate active, rented, legacy, duplicate, or reconciliation CSV files.",
        "10. Import that one consolidated CSV back into Datablix for source comparison and human review.",
        "",
        "FILES",
        f"- {prompt_name}: company-specific research instructions.",
        f"- {original_source_name}: the actual current project-wide Starting Data file.",
    ]
    if structured_count:
        readme_lines.append(f"- {structured_source_name}: normalized project source rows Datablix could parse from the original file.")
    if company_match_count:
        readme_lines.append(f"- {company_source_name}: convenience subset Datablix matched to {company_name}.")
    readme_lines += [
        f"- {template_name}: required structure for the returned research CSV.",
        "",
        "IMPORTANT",
        "The original project source file is the authoritative starting reference for this research cycle. Structured CSV extracts are conveniences only and may not capture every sheet, note, alias, or layout element in the original source file.",
        "The required research OUTPUT is exactly one consolidated company research CSV. The ZIP may contain multiple INPUT/reference files, but the AI research result must be one CSV only.",
    ]
    readme = "\n".join(readme_lines)

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, "w", compression=zipfile.ZIP_DEFLATED) as package:
        package.writestr(prompt_name, prompt_text.encode("utf-8"))
        if isinstance(raw_source_bytes, (bytes, bytearray)) and raw_source_bytes:
            package.writestr(original_source_name, bytes(raw_source_bytes))
        if isinstance(project_source_records, pd.DataFrame) and not project_source_records.empty:
            package.writestr(structured_source_name, project_source_records.to_csv(index=False).encode("utf-8-sig"))
        if isinstance(company_source_records, pd.DataFrame) and not company_source_records.empty:
            package.writestr(company_source_name, company_source_records.to_csv(index=False).encode("utf-8-sig"))
        package.writestr(template_name, research_template.to_csv(index=False).encode("utf-8-sig"))
        package.writestr("README.txt", readme.encode("utf-8"))
    return buffer.getvalue()


def prompt_record_identity_lines(
    df: pd.DataFrame,
    company_id: str = "",
    company_name: str = "",
    limit: int = 150,
) -> list[str]:
    """Return unique property identity lines for one company."""
    if not isinstance(df, pd.DataFrame) or df.empty:
        return []

    matched = company_source_records_for_research(
        df,
        company_id=company_id,
        company_name=company_name,
    )

    if matched.empty:
        return []

    lines = []
    seen = set()
    for _, row in matched.iterrows():
        parts = []
        for field in ["Building Name", "Street Address", "City", "Postal Code"]:
            value = safe_text(row.get(field, ""))
            if value and not is_unresolved(value):
                parts.append(value)

        label = " · ".join(parts)
        normalized = norm_header(label)
        if label and normalized not in seen:
            seen.add(normalized)
            lines.append(f"- {label}")

        if len(lines) >= max(1, int(limit)):
            break

    return lines


def build_company_website_research_prompt(
    *,
    company_name: str,
    company_website: str,
    related_official_links: str,
    special_website_notes: str,
    geographic_scope: str,
    priority_notes: str,
    source_policy: str,
    output_notes: str,
) -> str:
    """Create a City-of-Ottawa website-research prompt for Datablix import."""
    related_links, _ = _parse_related_official_links(
        related_official_links,
        company_website,
    )
    if related_links:
        related_links_block = "\n".join(f"- {link}" for link in related_links)
        official_entry_points_section = f"""## Reviewer-supplied official website entry points
Main company website:
- {company_website or '[enter official website]'}

Additional official links that must be inspected:
{related_links_block}

Special multi-link handling:
- Treat every supplied URL as an official research entry point for the selected company, subject to verification of branding, management statements, navigation, or cross-links.
- Do not create separate company records for subdomains, microsites, or property paths.
- Research every supplied link even when it is not discoverable from the main navigation or sitemap.
- Use the main website as Company Website.
- Use the most specific property page or microsite as Property Website and Source URL.
- Consolidate duplicate references to the same property using normalized physical address and municipality.
- Retain out-of-Ottawa findings only as research/audit evidence; do not include them in the final Ottawa CSV.
"""
    else:
        official_entry_points_section = """## Official website entry points
Only the main company website was supplied. Follow its current official navigation, portfolio pages, sitemaps, and confirmed related property pages or subdomains. Do not assume that additional official links exist.
"""

    special_notes_section = ""
    if safe_text(special_website_notes):
        special_notes_section = f"""## Reviewer notes about this website structure
{safe_text(special_website_notes)}

Use these notes as research context only. They cannot broaden the City of Ottawa geographic scope, weaken source requirements, or override evidence.
"""

    return f"""# Datablix City of Ottawa Company Research Prompt

You are acting as a careful public-source rental-property research analyst. Research the selected company and produce exactly ONE downloadable consolidated CSV file for import into Datablix.

IMPORTANT WORKFLOW BOUNDARY:
- Research the company's CURRENT public website/inventory independently.
- Do NOT compare the findings with Datablix Starting Data, prior research, or known-record lists.
- Do NOT label a row as Existing Source Record, Newly Discovered, or Possible Duplicate relative to the project. Datablix performs that comparison after import.

## Company context
- Company or management owner: {company_name or '[enter company name]'}
- Official company website: {company_website or '[enter official website]'}
- Geographic scope: {geographic_scope or PROJECT_GEOGRAPHIC_SCOPE}

{official_entry_points_section}
{special_notes_section}
## Non-negotiable property-type scope
Research the company's current official residential listing inventory—not only conventional apartment towers. Website inventory presence and rental availability are separate facts and must be recorded separately.

- Include current apartment buildings and apartment units, condominium rentals, townhomes, duplexes, and garden homes.
- Do not exclude a current property solely because it is a townhome, duplex, garden home, condominium unit, or another low-density rental form recognized by the project Starting Data.
- A current detached single-family home must not be silently discarded. Retain it in the CSV, keep Current Inventory Status based on the official inventory evidence, and add `Detached single-family home — Needs Scope Review` to Reviewer Notes unless company-specific instructions clearly resolve its scope.
- Do not relabel an ordinary detached home as a Garden Home without explicit official or source-file evidence.
- Record the supported property-form label in `Property Type`, for example `Apartment`, `Townhome`, `Duplex`, `Garden Home`, `Semi-detached`, or a supported combination such as `Apartment | Townhome | Semi-detached`. Never place property-form labels in Building Classification.
- Inactive, archived, or no-longer-supported listings may still be marked Excluded because of inventory status; townhome, duplex, or garden-home form alone is never an exclusion reason.

## Non-negotiable City of Ottawa municipal boundary
Return only in-scope residential rental properties whose PHYSICAL LOCATION is within the municipal boundaries of the City of Ottawa, Ontario, Canada.

- Website labels such as “Ottawa,” “Ottawa Region,” “Greater Ottawa,” “National Capital Region,” or “Eastern Ontario” are discovery hints, not geographic proof.
- Ottawa localities such as Kanata, Nepean, Orléans, Gloucester, Barrhaven, Stittsville, Vanier, Manotick, Carp, Cumberland, Richmond, and other communities physically inside the amalgamated City of Ottawa remain in scope.
- Carleton Place, Smiths Falls, Renfrew, Arnprior, Almonte/Mississippi Mills, Perth, Kemptville/North Grenville, Rockland/Clarence-Rockland, Casselman, Embrun/Russell, Cornwall, Gatineau, and every other independent municipality are out of scope.
- Determine scope from the exact physical street address, municipality, postal code, and—when needed—geocoded latitude/longitude or map position. Do not rely on the company's portfolio grouping alone.
- When exact municipal-boundary evidence remains uncertain, set Geographic Scope Status to Needs Geographic Review. Do not call the property inside Ottawa without evidence.
- Properties confirmed outside the City of Ottawa must be omitted from the final CSV. Do not keep them merely because they appear on an Ottawa-area portfolio page.

## Core inventory principle
Do not begin by collecting every URL that exists. First establish the company's CURRENT City of Ottawa website/listing inventory from its strongest official inventory/navigation evidence. A dedicated property URL that loads is not, by itself, proof that the property is current.

CRITICAL AVAILABILITY DISTINCTION:
- `Current Inventory Status` answers: Is this property/listing currently supported as part of the company's official website or managed listing inventory?
- `Rental Availability Status` answers: Can a renter currently pursue this listing, is it becoming available later, or has it already been rented?
- A property may legitimately have `Current Inventory Status = Current` and `Rental Availability Status = Rented`.
- A page title, navigation label, or collection called `For Rent` is NOT proof that every listing displayed there is available.
- Page existence, a displayed rental price, a contact form, or a working property URL is NOT proof of current rental availability.

## Official company, subdomain, and property-page hierarchy
Treat the organization and its official property pages as a hierarchy, not as separate companies.

- The main/root website is the corporate or management-company source.
- A property page on the main domain and an official property microsite on a subdomain can both belong to that same company.
- URLs sharing the same registrable root domain—such as `milyservice.com`, `wildwood.milyservice.com`, and `hillpark.milyservice.com`—must remain under one company when branding, management statements, navigation, or cross-links confirm the relationship.
- Do not create a new company merely because the hostname or subdomain is different.
- Do not create one property row per URL. When a portfolio page, property page, and microsite describe the same physical property or leasing community, consolidate them into one property record.
- Store the selected root/corporate site in Company Website, the exact property page or microsite in Property Website, and the strongest exact supporting page in Source URL. Preserve additional official URLs in Supporting Evidence.
- A separate branded subdomain is not sufficient by itself: confirm that it is owned, managed, linked, or explicitly identified by the selected company.
- Community names and physical buildings are not automatically equivalent. Keep one row for a single named complex with one leasing identity; split into multiple building rows only when official evidence confirms distinct physical buildings or separately leased addresses.

### Phase 1 — Establish current official inventory
Prioritize:
1. current city/location pages;
2. current property-search or portfolio pages;
3. current human-readable HTML sitemaps;
4. current building/community index pages;
5. official property pages linked from those sources; and
6. confirmed official property subdomains or microsites under the same registrable company domain.

Use XML sitemaps only as discovery evidence. They may contain stale, orphaned, archived, or legacy URLs.

Use these inventory values:
- Current — supported as part of the company's current official website/listing inventory, regardless of whether it is presently rentable.
- Review — current website/inventory status is uncertain because evidence is incomplete, conflicting, blocked, or unavailable.
- Excluded — an identifiable website property is no longer supported as part of the current official inventory.

Do not create a property row for an orphan, empty, generic, redirected, placeholder, or template-only page with no meaningful property evidence. A sparse property page must still be retained when current official inventory evidence confirms the property.

### Phase 1B — Mandatory rental-availability validation
For EVERY identifiable Current or Review property, inspect the most current official property card, listing page, property page, leasing page, and any visible status badge/label before deciding whether it is actively rentable.

Normalize Rental Availability Status to exactly one of:
- `Available Now` — explicitly shown as FOR RENT, AVAILABLE, AVAILABLE NOW, or equivalent with no future availability period.
- `Available Future` — explicitly available from a stated future date or period, such as `AVAILABLE SEP 2026`, `AVAILABLE OCT 2026`, or `Available September 1, 2026`.
- `Rented` — explicitly marked RENTED, LEASED, TAKEN, NO LONGER AVAILABLE, or equivalent.
- `Status Unclear — Needs Review` — the property is identifiable/current on the website but reliable current rental availability cannot be established or current official signals conflict.

Rental Availability Detail:
- Preserve the site's useful human-readable detail, for example `AVAILABLE OCT 2026`, `Available September 1, 2026`, or `RENTED`.
- Do not invent an exact day when the website gives only a month or general future period.

Rental Availability Evidence:
- Record the exact official status wording/context and the page/card URL where it was observed.
- Prefer the most current property-specific status evidence.
- If a current portfolio/card and exact property page conflict and the conflict cannot be resolved by a clearly newer date/status, set `Status Unclear — Needs Review` and preserve both signals.

NON-NEGOTIABLE STATUS RULES:
- A listing on a page named `For Rent` can still be `Rented`.
- An explicit `Rented`/`Leased`/`No Longer Available` status overrides any assumption based only on page location, a displayed price, or a still-working URL.
- Do not infer `Available Now` merely because a rent amount or application/showing button is present.
- Do not infer `Rented` merely because no current unit is visible; use `Status Unclear — Needs Review` when there is no explicit reliable status.
- Keep an explicitly Rented row in the research CSV when it is still an identifiable listing on a current official inventory/For Rent page. Datablix needs that row for source reconciliation, but it must not count as active rental inventory or as a Newly Discovered active property.
- Historical sitemap-only/orphaned rows that are not supported by current official inventory evidence remain Excluded/omitted under the existing inventory rules; do not revive them merely to populate Rented records.

### Phase 2 — Research every Current or Review Ottawa property deeply
Inspect the complete relevant official content, including property overview, physical location, contact information, floor plans, rates, amenities, parking, laundry, utilities, accessibility, elevator, policies, official PDFs, brochures, leasing pages, JavaScript-rendered content, footer details, linked official property websites, and rental-availability evidence.

For Number of Apartments, actively search the official property/company material for equivalent total-inventory wording such as apartments, units, residential units, rental units, dwelling units, suites, residences, homes, doors, unit count, suite count, and total units. Do not stop merely because the exact phrase `Number of Apartments` is absent.

### Phase 3 — Exhaustive PO Box and mailing-address search
PO Box research is separate from the rental property's physical address.

Search the company's official web presence first, including:
- property contact and leasing pages;
- company Contact, About, Corporate, Legal, Privacy, Terms, Accessibility, Careers, Media, footer, and tenant-document pages;
- official PDFs, forms, notices, brochures, application documents, rent-payment instructions, and corporate filings hosted on the official domain.

Recognize formats including PO Box, P.O. Box, Post Office Box, Postal Box, Box, CP, C.P., Case postale, RR, Rural Route, and General Delivery/GD when the context clearly shows a mailing address.

When the official website does not provide a PO Box or complete mailing address, a controlled Google/web lookup is allowed:
- search the exact company name plus terms such as PO Box, P.O. Box, mailing address, postal address, case postale, corporate address, rent payment address, and tenant correspondence;
- use Google Search or Google Maps to locate the underlying source, but do not treat a search-result snippet as evidence;
- prefer official company documents, government/public records, reliable corporate filings, municipal documents, and clearly attributable public records;
- never copy a PO Box from a similarly named company or another branch without evidence linking it to the selected company/property;
- record the full source URL, evidence text/context, and confidence.

CRITICAL ADDRESS SEPARATION:
- Street Address is the rental property's physical civic address.
- Mailing Address is the complete correspondence address when found.
- PO Box contains only the box identifier/number.
- Never replace Street Address with a corporate office, management office, PO Box, mailing address, leasing office, or payment address.
- A PO Box does not prove the location of a rental property and must never be used for geographic-scope determination.
- When no PO Box is found after the required search, leave PO Box blank and set PO Box Search Status to Not Found after Search. Do not invent one.

### Phase 4 — Controlled geographic-position verification
After a property is identified from the official company website, verify whether its physical coordinates fall within the City of Ottawa.

Use the verified physical street address—not a PO Box or corporate mailing address—to search Google Maps or another reliable geocoder. Capture:
- Latitude;
- Longitude;
- Geocoded Municipality;
- Geographic Scope Status;
- Geographic Evidence; and
- Geographic Confidence.

Rules:
- Inside City of Ottawa requires a clear exact-address match and municipal evidence supporting Ottawa.
- Outside City of Ottawa applies when the exact address/geographic position is in another municipality.
- Needs Geographic Review applies to partial matches, conflicting municipalities, ambiguous rural addressing, missing coordinates, or uncertain boundary placement.
- Record formatted address, municipality/address components, coordinates, map/geocoder source, partial-match warning, and boundary method in Geographic Evidence.
- Geocoding may verify or exclude an official-site candidate geographically, but it must not be used to discover additional rental properties.

### Phase 5 — Controlled external field recovery
Only after an official-site property candidate is established may outside evidence be used for these limited purposes:

1. Postal Code recovery — use the exact verified physical address; accept only an exact civic-address match. Never infer from a neighbour, neighbourhood, FSA, or partial code.
2. PO Box/mailing-address recovery — follow the exhaustive method above and keep it separate from the property address.
3. Geographic-position verification — exact-address geocoding and municipal-boundary evidence only.
4. Number of Storeys — use the mandatory storey-verification sequence below. The AI researches and evidences the storey count; Datablix derives Building Classification automatically after import from a verified Number of Storeys.
5. Number of Apartments / Total Unit Count — use the mandatory apartment-count verification sequence below. The value must represent total residential inventory for the exact property scope represented by the row.

External research for apartment/storey counts is a RECOVERY step for an already-established property, not a discovery method for inventing additional properties.

## Number of Apartments — mandatory verification rule
Research `Number of Apartments` for EVERY Current or Review property.

`Number of Apartments` means the TOTAL residential inventory associated with the exact property represented by the row. It does NOT mean currently available, vacant, advertised, remaining, or search-result units.

Recognize equivalent total-inventory wording including:
- apartments;
- units;
- residential units;
- rental units;
- dwelling units;
- suites;
- residences;
- rental homes;
- doors;
- unit count;
- suite count;
- number of units;
- number of suites;
- total units; and
- total suites.

### Mandatory apartment-count search sequence
Do NOT leave Number of Apartments blank after checking only the main property/leasing page.

For every unresolved property, complete this sequence in order:
1. Exact official property page or official property microsite.
2. Other exact-property pages on the official company website.
3. Official company/property PDFs, brochures, reports, filings, acquisition pages, development pages, planning material, or tenant/property documents.
4. Search the exact civic address in quotation marks together with several total-count terms: `units`, `apartments`, `suites`, `residential units`, `dwelling units`, `unit count`, `number of units`, and `total units`.
5. Repeat with the exact building/property name plus exact address when a name is known.
6. Municipal, planning, development-application, assessment, CMHC/government, permit, or comparable reliable public records tied to the exact property.
7. Only after stronger tiers are silent, use a reputable property/database or industry source that clearly matches the exact civic address/property.

Open the underlying page/document. A search-result snippet, AI summary, generated answer, or search preview alone is NOT evidence.

### Apartment-count property identity and scope test
Before accepting any apartment/unit total, verify that the number refers to the SAME property scope represented by the CSV row.

Determine whether the candidate number refers to:
- the exact building;
- the exact multi-address complex represented by the row;
- a larger community containing several buildings;
- a phase within a larger development;
- the management company's entire portfolio; or
- only currently available/vacant units.

Accept the number only when its scope matches the row.

Do NOT:
- use available units, vacant units, units remaining, or availability-search counts;
- use the number of floor plans, bedroom types, listings, or advertised suites;
- assign a whole-community or portfolio total to one building unless the source explicitly allocates that total to the row's exact building/property;
- transfer a count from a nearby, similarly named, or same-street property;
- infer a count from storeys, windows, balconies, maps, photos, elevators, unit numbers, or visual appearance.

For an external exact-address match, the civic number, street name, street type, and directional component when present must refer to the same property. A matching civic number alone is insufficient.

### Apartment-count source hierarchy and confidence
Use source quality, exact-property match, context, and scope—not repetition alone—to judge confidence.

- `High`: clear explicit total tied to the exact property from an official property/company source, official document, or strong authoritative public record.
- `Medium`: clear exact-property total from a reputable secondary property/industry source after stronger tiers were checked and silent.
- `Low`: usable but weak/ambiguous evidence, unresolved scope concerns, or a material source conflict.
- Two weak sources agreeing do NOT automatically become High confidence.

Apartment Count Search Status values:
- `Official page` — confirmed on the exact official property/company web page.
- `Official document` — confirmed in an official company PDF, brochure, report, filing, acquisition/development document, or equivalent official document.
- `Public record` — confirmed in a municipal/planning/government/CMHC or comparable reliable public record tied to the exact property.
- `Reputable property source` — confirmed only after stronger source tiers were silent, using a reputable exact-address property/database source.
- `Not Found after Search` — the mandatory recovery sequence was completed but no reliable total could be confirmed.
- `Conflict — Needs Review` — reliable exact-property sources report materially different totals and the conflict cannot be safely resolved.

### Conflicting apartment counts
If reliable exact-property sources report different totals:
1. Do not silently choose the value that appears most often.
2. Compare source authority, exact-address match, date, context, and property scope.
3. Prefer one value only when a clearly stronger and more property-specific source reasonably resolves the discrepancy.
4. Otherwise leave Number of Apartments blank, set Apartment Count Search Status to `Conflict — Needs Review`, record both values and URLs in Apartment Count Evidence and Reviewer Notes, and use Low confidence.

### Mandatory apartment-count closure pass
Before producing the final CSV, revisit EVERY Current or Review row where Number of Apartments is blank.

A completed property must NEVER have both Number of Apartments blank AND Apartment Count Search Status blank.

If no reliable total can be confirmed after the full sequence:
- leave Number of Apartments blank;
- set Apartment Count Search Status to `Not Found after Search`;
- set Apartment Count Confidence to `Low`;
- use Apartment Count Evidence to briefly state which source tiers/searches were completed; and
- do not invent or estimate a number merely to avoid a blank.

For EVERY populated Number of Apartments, Apartment Count Evidence must make clear what the source says and why the number represents TOTAL residential inventory for the exact property—not availability.

## Number of Storeys — mandatory evidence-backed verification rule
Research `Number of Storeys` for EVERY Current or Review property. Do NOT research or return Building Classification; Datablix calculates that field after import.

Number of Storeys means the verified TOTAL building storey count. It does not mean the floor on which a particular apartment/unit is located.

Recognize equivalent wording including `storey/storeys`, `story/stories`, `floor/floors`, and `level/levels` ONLY when context clearly states the TOTAL building height.

Accept examples such as:
- `8-storey building` → Number of Storeys = 8;
- `8-story apartment building` → Number of Storeys = 8;
- `the building consists of 8 floors` → Number of Storeys = 8.

Reject examples such as:
- `apartment located on the 8th floor` → NOT proof of 8 storeys;
- `unit on level 4` → NOT proof of 4 storeys;
- `Suite 1204` → NOT proof of 12 storeys.

### Mandatory storey-count search sequence
Do NOT leave Number of Storeys unresolved after checking only the official property/leasing page.

For every unresolved property, complete this sequence in order:
1. Exact official property page or official property microsite.
2. Other exact-property pages on the official company website.
3. Official brochures, PDFs, reports, acquisition/development pages, architectural/planning material, or other official documents.
4. Search the exact civic address in quotation marks with several height terms: `storeys`, `stories`, `floors`, `levels`, `building height`, and `storey apartment building`.
5. Repeat with the exact building/property name plus exact address when a name is known.
6. Municipal planning/development/building/permit or comparable authoritative public records tied to the exact address.
7. Only after stronger tiers are silent, use a reputable exact-address property/real-estate/database or industry source.

Open the underlying source. Search-result snippets, AI summaries, generated answers, and previews alone are NOT evidence.

### Storey-count property identity and scope test
Before accepting a storey count, verify that it belongs to the exact building/property represented by the row. Match the full civic address and, where available, the property/building name or official property URL.

For multi-building communities or combined-address properties:
- do not assign one building's height to another building;
- do not assign one storey count to an entire mixed-height community when buildings differ;
- if the row genuinely represents a multi-building complex and no single storey count applies, leave Number of Storeys blank, set Storey Count Search Status to `Conflict — Needs Review`, set Storey Count Confidence to `Low`, and explain the mixed-height condition in Storey Count Evidence and Reviewer Notes;
- if a source gives a range such as `2–3 storeys` for a mixed community, do not convert the range to a single integer.

### What must NOT be counted or inferred
Do not independently add basements, underground parking levels, mezzanines, podium/mechanical levels, or rooftop structures. Include one only when the source itself explicitly includes it in the stated storey total.

Never determine Number of Storeys by visually counting photographs, Google Street View, satellite imagery, windows, balconies, entrances, or rooflines. Do not infer storeys from elevators, unit numbers, apartment floor locations, the property name, the word `Tower`, `High-Rise`, `Luxury`, or other marketing language.

### Storey-count source hierarchy and confidence
Use source quality, exact-property match, context, and scope—not repetition alone—to judge confidence.

- `High`: clear explicit total storey count tied to the exact property from an official property/company source, official document, or strong authoritative public record.
- `Medium`: clear exact-property total storey count from a reputable secondary property/industry source after stronger tiers were checked and silent.
- `Low`: weak/ambiguous evidence, unresolved scope concerns, mixed-height scope, or a material source conflict.
- Two weak sources agreeing do NOT automatically become High confidence.

Storey Count Search Status values:
- `Official page` — confirmed on the exact official property/company web page.
- `Official document` — confirmed in an official company PDF, brochure, report, filing, development/planning document, or equivalent official document.
- `Public record` — confirmed in a municipal/planning/building/government or comparable authoritative public record tied to the exact property.
- `Reputable property source` — confirmed only after stronger source tiers were silent, using a reputable exact-address property/database source.
- `Not Found after Search` — the mandatory recovery sequence was completed but no reliable total could be confirmed.
- `Conflict — Needs Review` — reliable sources materially disagree or the row represents a mixed-height property for which no single storey count safely applies.

### Conflicting storey counts
If reliable exact-property sources report different storey totals:
1. Compare exact-address match, source authority, date, context, and whether each source describes the same building/phase.
2. Do not average the values.
3. Do not choose the most frequently repeated value merely because it appears more often.
4. If one clearly stronger exact-property source resolves the discrepancy, use it, set the status/confidence according to that source, and preserve the conflicting source in Storey Count Evidence and Reviewer Notes.
5. Otherwise leave Number of Storeys blank, set Storey Count Search Status to `Conflict — Needs Review`, set Storey Count Confidence to `Low`, and record both values, contexts, and URLs in Storey Count Evidence.

### Mandatory storey-count closure pass
Before producing the final CSV, revisit EVERY Current or Review row where Number of Storeys is blank.

A completed property must NEVER have both Number of Storeys blank AND Storey Count Search Status blank/Not Checked.

When no reliable storey count can be verified after the full sequence:
- leave Number of Storeys blank;
- set Storey Count Search Status to `Not Found after Search`;
- set Storey Count Confidence to `Low`;
- use Storey Count Evidence to briefly state which source tiers/searches were completed;
- leave Storey Count Source URL blank when no source supports a count; and
- do not guess.

For EVERY populated Number of Storeys:
- Storey Count Search Status must identify the source tier;
- Storey Count Source URL must contain the exact page/document supporting the count;
- Storey Count Evidence must state the wording/context proving TOTAL building height; and
- Storey Count Confidence must be High, Medium, or Low according to the rules above.

## Property Type and Datablix-derived Building Classification
Identify the official residential property form independently and record it ONLY in `Property Type`: Apartment, Condominium, Townhome, Duplex, Garden Home, Semi-detached, Detached, or another clearly stated residential form. If an official community genuinely contains more than one supported form, separate the Property Type labels with ` | `. Do not infer property form from appearance.

Do NOT include `Building Classification` in the research CSV. Datablix derives it automatically from verified Number of Storeys after import:
- 1–4 storeys → Low-rise
- 5–11 storeys → Mid-rise
- 12+ storeys → High-rise

If Number of Storeys is unresolved, Datablix leaves Building Classification unresolved. Property Type never substitutes for the height classification.

## Final count verification pass — mandatory before CSV output
Before producing the final CSV, audit EVERY Current or Review property row:

1. If Number of Apartments is blank, confirm the mandatory apartment-count recovery sequence was completed and Apartment Count Search Status documents the outcome.
2. If Number of Storeys is blank, confirm the mandatory storey-count recovery sequence was completed and Storey Count Search Status/Evidence document the outcome.
3. For every populated Number of Apartments, confirm the number belongs to the exact row/property scope and represents TOTAL residential inventory rather than available/vacant inventory.
4. For every populated Number of Storeys, confirm Storey Count Source URL/Evidence prove the TOTAL building height rather than an individual unit's floor.
5. Recheck multi-building communities so a complex-wide unit count or one building's storey count is not incorrectly assigned to another building/row.
6. Recheck conflicting numbers using source authority, exact-property identity, date, context, and scope. Do not silently select a convenient value.
7. Do not finish a row merely because the official leasing page omitted these fields. Complete the permitted external recovery hierarchy first.
8. Accuracy is more important than completeness: a documented unresolved value is preferable to an unsupported or estimated number.

External evidence must never invent a property, override the company's official inventory status, or fill unrelated ordinary fields such as amenities, rates, unit types, leasing contacts, or policies.

## Source policy
{source_policy}

## Exact CSV columns
Use these headings in this exact order:

{', '.join(AI_RESEARCH_DELIVERABLE_COLUMNS)}

Field requirements:
- Company Website: the selected company's root or canonical corporate/management website, not a property subdomain.
- Property Website: the exact official building/community page or official property microsite.
- Source URL: the strongest exact page supporting the row; place additional official URLs in Supporting Evidence.
- Current Inventory Status: Current, Review, or Excluded — not in current website inventory. This field describes current website/inventory presence, NOT whether the property is presently rentable. Property form alone does not determine this status.
- Inventory Evidence: official website evidence supporting Current/Review/Excluded inventory presence.
- Rental Availability Status: required for every Current or Review completed row; use Available Now, Available Future, Rented, or Status Unclear — Needs Review. Do not equate presence on a For Rent page with availability.
- Rental Availability Detail: preserve the useful official status/date wording, such as `AVAILABLE OCT 2026`, `Available September 1, 2026`, or `RENTED`; blank only when the status has no additional detail.
- Rental Availability Evidence: concise exact official wording/context plus the property/card URL supporting the rental status. Preserve conflicts rather than guessing.
- Number of Apartments: total residential inventory only. Recognize units/suites/residences and equivalent total-count wording; never substitute available/vacant listings. Use the dedicated exact-address recovery hierarchy when the official property page is silent.
- Apartment Count Search Status: required for every completed row; use Official page, Official document, Public record, Reputable property source, Not Found after Search, or Conflict — Needs Review.
- Apartment Count Source URL: the exact page/document supporting the total; blank only when no count was confirmed.
- Apartment Count Evidence: concise exact-property wording/context supporting the count, or a short note describing the completed unsuccessful/conflicting recovery search.
- Apartment Count Confidence: High, Medium, Low, or Not Checked. Use Not Checked only before the mandatory apartment-count closure pass has been completed. Once the closure pass is complete, do not leave confidence as Not Checked; use High only for clear exact-property evidence from strong sources, Medium for usable but less authoritative exact-property evidence, and Low for unresolved conflicts or weak/ambiguous evidence.
- Property Type: record only the officially supported residential property form(s), such as Apartment, Condominium, Townhome, Duplex, Garden Home, Semi-detached, or Detached. Use ` | ` only when the same official property/community genuinely contains multiple supported forms.
- Number of Storeys: verified TOTAL building storey count only; never an individual unit floor or visually inferred count.
- Storey Count Search Status: required for every completed row; use Official page, Official document, Public record, Reputable property source, Not Found after Search, or Conflict — Needs Review.
- Storey Count Source URL: the exact page/document supporting the storey count; blank only when no count was confirmed.
- Storey Count Evidence: concise exact-property wording/context supporting the total building height, or a short note describing the completed unsuccessful/conflicting recovery search.
- Storey Count Confidence: High, Medium, Low, or Not Checked. Once the mandatory closure pass is complete, do not leave confidence as Not Checked.
- Building Classification is NOT an AI output column. Datablix derives it automatically from verified Number of Storeys after import.
- PO Box Search Status: Not Checked, Found, Not Found after Search, Not Applicable, or Needs Review.
- PO Box Confidence: High, Medium, Low, or Not Checked.
- Geographic Scope Status: Inside City of Ottawa, Outside City of Ottawa, Needs Geographic Review, or Not Checked.
- Geographic Confidence: High, Medium, Low, or Not Checked.
- Supporting Evidence: concise field-level notes and URLs, clearly labeling secondary evidence.
- Missing Information: only fields that were actively checked and could not be confirmed.
- Reviewer Notes: conflicts, limitations, assumptions, and required follow-up.

## Mandatory quality rules
1. Never invent, estimate, or populate a field merely for completeness.
2. Leave unresolved values blank and document the research gap.
3. Use No only when a source explicitly says a feature is unavailable, prohibited, or not offered.
4. Separate property, company, mailing, corporate, leasing, and PO Box addresses.
5. Do not use a PO Box, corporate office, or leasing office to determine the building's municipality.
6. Search snippets are navigation aids, not evidence; open and cite the underlying page/document.
7. Preserve conflicting values rather than silently choosing one.
8. Deduplicate only within this company-research result.
9. Treat the main domain and confirmed official property subdomains as one company; never create a company per hostname.
10. Consolidate multiple official URLs for the same property into one row and preserve the property microsite in Property Website.
11. Keep multiple civic addresses in one row only when official evidence shows one named property/complex with one leasing identity.
12. Omit geographically out-of-scope properties from the final CSV.
13. Treat AI findings as preliminary and subject to Datablix validation and human approval.
14. Prefer transparent blanks over unsupported completeness.
15. Never exclude a current townhome, duplex, or garden home merely because it is not a conventional apartment building.
16. Keep detached single-family homes visible for human scope review unless a project or company rule explicitly resolves them.
17. Never use the title `For Rent`, page existence, displayed rent, or a working URL as a substitute for explicit rental-availability verification.
18. A Current website listing marked Rented is not active rental inventory. Preserve `Current Inventory Status = Current` when supported, set `Rental Availability Status = Rented`, and retain the row for Datablix reconciliation when it appears in current official inventory.
19. Available Now and Available Future are the only statuses that count as active/upcoming rental inventory. Status Unclear requires human review and Rented must never be counted as an active/new rental discovery.

## Priority or company-specific instructions
The City of Ottawa municipal boundary, residential property-type scope, mandatory rental-availability validation, physical-vs-mailing address separation, exhaustive PO Box search, exact-address geographic verification, postal-code recovery, and evidence-backed storey rules are project-wide. Company notes may refine priorities but must not broaden the project to nearby municipalities, treat `For Rent` page placement as availability proof, or weaken these rules.

{priority_notes or 'No additional priorities were provided.'}

## Required deliverable — EXACTLY ONE CSV file
Create exactly one downloadable CSV file named clearly, for example:
`company_name_ottawa_website_research_results.csv`

The file must contain one unique researched City of Ottawa property/listing per row, use the exact headings above, preserve evidence and blanks, and remain directly importable into Datablix. Keep Current and Review website inventory records, including listings explicitly marked Rented when they are still identifiable on a current official inventory/For Rent page, so Datablix can reconcile them against Starting Data after import. Do not count Rented rows as active inventory or newly discovered active rentals. Do not bulk-import stale sitemap-only/orphaned legacy rows that lack current official inventory support. Use Current Inventory Status and Rental Availability Status together to distinguish website presence from actual rental availability. Do not return Excel, Google Sheets, JSON, PDF, Word, Markdown tables, ZIP files, or multiple research files.

When the platform cannot create a downloadable file, return that one CSV as raw RFC-style CSV text in a fenced csv code block with only a one-line limitation notice.

Additional output instructions:
{output_notes or 'Return exactly one clean, evidence-based City of Ottawa research CSV ready for direct Datablix import.'}
"""


def _same_resolved_value(a, b, normalizer=lambda value: safe_text(value).lower()) -> bool:
    left = normalizer(a)
    right = normalizer(b)
    return bool(left and right and left == right)


def _strong_shared_leasing_identity(left, right) -> bool:
    """Return True only when two AI rows clearly represent one leased property.

    Same-company contact information alone is NOT enough.  Datablix requires the
    same property/building name, the same street, and shared property/leasing
    evidence before it combines civic addresses.
    """
    left_id = _row_identity(left)
    right_id = _row_identity(right)
    la = left_id["address"]
    ra = right_id["address"]

    if not left_id["name"] or left_id["name"] != right_id["name"]:
        return False
    if not la["street_lenient"] or la["street_lenient"] != ra["street_lenient"]:
        return False

    property_url_same = bool(
        left_id["property_url"]
        and left_id["property_url"] == right_id["property_url"]
    )
    source_url_same = bool(
        left_id["source_url"]
        and left_id["source_url"] == right_id["source_url"]
        and left_id["source_url"].count("/") >= 1
    )
    phone_same = bool(left_id["phone"] and left_id["phone"] == right_id["phone"])
    email_same = bool(left_id["email"] and left_id["email"] == right_id["email"])

    # Postal conflict is a warning sign.  Do not auto-merge it unless a property
    # URL itself is shared, which is strong evidence that the company leases it
    # as one record.
    postal_conflict = bool(
        left_id["postal"] and right_id["postal"] and left_id["postal"] != right_id["postal"]
    )
    if postal_conflict and not property_url_same:
        return False

    return property_url_same or source_url_same or (phone_same and email_same) or phone_same


def _street_tail_for_display(value) -> str:
    text = safe_text(value).strip()
    text = text.replace("–", "-").replace("—", "-")
    match = re.match(
        r"^\s*(?:\d+[A-Za-z]?)(?:\s*(?:-|/|&|\band\b)\s*\d+[A-Za-z]?)*\s+(.+)$",
        text,
        flags=re.I,
    )
    return match.group(1).strip() if match else text


def _combined_address_for_group(group: pd.DataFrame) -> str:
    signatures = [_address_signature(value) for value in group["Street Address"]]
    streets = [signature["street_lenient"] for signature in signatures if signature["street_lenient"]]
    if not streets or len(set(streets)) != 1:
        return safe_text(group.iloc[0].get("Street Address", ""))

    civic_numbers = []
    for signature in signatures:
        for number in signature["numbers"]:
            if number not in civic_numbers:
                civic_numbers.append(number)
    if not civic_numbers:
        return safe_text(group.iloc[0].get("Street Address", ""))

    def number_sort_key(value):
        match = re.match(r"(\d+)([a-z]?)", str(value), flags=re.I)
        return (int(match.group(1)) if match else 10**9, (match.group(2) if match else ""))

    civic_numbers = sorted(civic_numbers, key=number_sort_key)
    tail = _street_tail_for_display(group.iloc[0].get("Street Address", ""))
    if len(civic_numbers) == 1:
        return f"{civic_numbers[0]} {tail}".strip()
    if len(civic_numbers) == 2:
        number_text = f"{civic_numbers[0]}-{civic_numbers[1]}"
    else:
        number_text = "/".join(civic_numbers)
    return f"{number_text} {tail}".strip()


def _append_note(existing, note) -> str:
    current = safe_text(existing)
    note = safe_text(note)
    if not note:
        return current
    if not current:
        return note
    if note.lower() in current.lower():
        return current
    return f"{current}; {note}"


def consolidate_shared_leasing_rows(df: pd.DataFrame) -> pd.DataFrame:
    """Consolidate AI rows the company clearly leases as one property record.

    This is intentionally conservative.  It protects against research tools
    splitting a multi-address complex into one row per civic number when the
    company itself uses one property name and one leasing identity.
    """
    if not isinstance(df, pd.DataFrame) or len(df) < 2:
        return df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()

    work = df.reset_index(drop=True).copy()
    parent = list(range(len(work)))

    def find(i):
        while parent[i] != i:
            parent[i] = parent[parent[i]]
            i = parent[i]
        return i

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for left_index in range(len(work)):
        for right_index in range(left_index + 1, len(work)):
            if _strong_shared_leasing_identity(work.iloc[left_index], work.iloc[right_index]):
                union(left_index, right_index)

    groups = {}
    for index in range(len(work)):
        groups.setdefault(find(index), []).append(index)

    rows = []
    list_like_fields = {
        "Amenities", "Suite Types", "Rental Rate Range", "Property Type", "Supporting Evidence",
        "PO Box Evidence", "Geographic Evidence", "Apartment Count Evidence",
        "Storey Count Evidence", "Missing Information", "Reviewer Notes",
    }
    protected_conflict_fields = {
        "Number of Apartments", "Apartment Count Search Status",
        "Apartment Count Source URL", "Apartment Count Confidence",
        "Number of Storeys", "Storey Count Search Status",
        "Storey Count Source URL", "Storey Count Confidence",
        "Postal Code", "Phone",
        "Primary Email", "Secondary Email", "Building Classification", "PO Box",
        "PO Box Postal Code", "Latitude", "Longitude", "Geographic Scope Status",
    }

    for member_indexes in groups.values():
        group = work.iloc[member_indexes].copy()
        if len(group) == 1:
            rows.append(group.iloc[0].to_dict())
            continue

        merged = group.iloc[0].to_dict()
        merged["Street Address"] = _combined_address_for_group(group)
        conflicts = []

        for column in group.columns:
            if column in {"Street Address", "Record ID"}:
                continue
            values = []
            for value in group[column]:
                if is_unresolved(value):
                    continue
                clean = safe_text(value)
                if clean and clean not in values:
                    values.append(clean)
            if not values:
                continue
            if len(values) == 1:
                merged[column] = values[0]
            elif column == "Property Type":
                merged[column] = merge_property_type_values(*values)
            elif column in list_like_fields:
                merged[column] = "; ".join(values)
            elif column == "Current Inventory Status":
                lowered = [value.lower() for value in values]
                merged[column] = "Current" if any(value.startswith("current") for value in lowered) else values[0]
            elif column == "Confidence":
                rank = {"low": 0, "medium": 1, "high": 2}
                merged[column] = min(values, key=lambda value: rank.get(value.lower(), 1))
            elif column in protected_conflict_fields:
                # Preserve the first value but force the disagreement into human review.
                merged[column] = values[0]
                conflicts.append(f"{column} differs across consolidated AI rows: {' | '.join(values)}")
            else:
                merged[column] = values[0]

        merged["Reviewer Notes"] = _append_note(
            merged.get("Reviewer Notes", ""),
            f"Datablix consolidated {len(group)} AI rows because they shared the same property name, street and leasing identity; civic addresses were kept as one property record.",
        )
        for conflict in conflicts:
            merged["Reviewer Notes"] = _append_note(merged.get("Reviewer Notes", ""), conflict)
        rows.append(merged)

    return pd.DataFrame(rows, columns=work.columns)


def append_external_research_results(
    imported: pd.DataFrame,
    *,
    company_id: str,
    company_name: str,
    company_website: str,
) -> int:
    """Map an external AI spreadsheet into the current human-review workflow."""
    validate_input(imported)
    mapped, _mapping = map_schema(imported)
    if mapped.empty:
        return 0

    for column in INTERNAL_COLUMNS:
        if column not in mapped.columns:
            mapped[column] = pd.NA

    mapped["Company ID"] = company_id
    owner_blank = unresolved_mask(mapped["Management/Owner"])
    mapped.loc[owner_blank, "Management/Owner"] = company_name
    if "Company Website" in mapped.columns:
        company_site_blank = unresolved_mask(mapped["Company Website"])
        mapped.loc[company_site_blank, "Company Website"] = company_website
    website_blank = unresolved_mask(mapped["Website"])
    if "Property Website" in mapped.columns:
        mapped.loc[website_blank, "Website"] = mapped.loc[website_blank, "Property Website"]

    # Enforce the company -> property website hierarchy before matching or
    # consolidation. Official subdomains remain property sources under the
    # selected company instead of becoming separate company identities.
    mapped = normalize_official_website_roles(mapped, company_website)

    mapped["Research Status"] = "Imported - Needs Review"
    mapped["Verification Status"] = "Needs Review"
    mapped["Record Decision"] = "Undecided"
    mapped["Directory Entry Status"] = "Not Entered"

    # When Google Maps geocoding is configured, validate physical coordinates
    # before consolidation. Mailing/PO Box fields are never used for geocoding.
    mapped = enrich_geographic_scope(mapped, max_requests=100)

    # Protect the review queue from AI over-splitting.  When the company clearly
    # leases multiple civic addresses as one named property, retain one row.
    mapped = consolidate_shared_leasing_rows(mapped)

    current = st.session_state.get(S_WORKING, pd.DataFrame()).copy()
    combined = pd.concat([current, mapped], ignore_index=True, sort=False)
    combined = ensure_ids(normalize_workflow(prepare_data(combined)))

    # Project-level source comparison belongs in Datablix, not in the AI prompt.
    # Always prefer the active source-version records, with S_ORIGINAL only as a
    # fallback.  This prevents a valid source row such as 165 Chapel Street from
    # being called new because a stale/empty session copy was consulted.
    active_source_records = current_starting_source_records()
    combined = classify_discovery_status(
        combined,
        active_source_records if not active_source_records.empty else None,
    )

    st.session_state[S_WORKING] = combined

    # Importing research is an explicit start signal. Persist it in the company
    # registry so every page agrees that the company is no longer Not started.
    registry = normalize_company_registry(st.session_state.get(S_COMPANIES))
    company_mask = registry["Company ID"].astype(str).str.strip().eq(str(company_id).strip())
    registry.loc[company_mask, "Company Status"] = registry.loc[
        company_mask, "Company Status"
    ].replace("Not started", "Researching")
    st.session_state[S_COMPANIES] = normalize_company_registry(registry)
    autosave_current_project()
    return len(mapped)


# =========================================================
# Quality checks and output views
# =========================================================

@st.cache_data(show_spinner=False, ttl=300, max_entries=16)
def qa_checks(df):
    out = normalize_workflow(df)
    issues = pd.Series([[] for _ in range(len(out))], index=out.index, dtype="object")
    core_gaps = pd.Series([[] for _ in range(len(out))], index=out.index, dtype="object")
    research_gaps = pd.Series([[] for _ in range(len(out))], index=out.index, dtype="object")

    def flag(mask, severity, message):
        for idx in out.index[mask.fillna(False)]:
            issues.at[idx].append((severity, message))

    for field in CORE_FIELDS:
        mask = unresolved_mask(out[field])
        flag(mask, "Critical", f"Missing {field}")
        for idx in out.index[mask]:
            core_gaps.at[idx].append(field)
    for field in TARGET_FIELDS:
        mask = unresolved_mask(out[field])
        for idx in out.index[mask]:
            research_gaps.at[idx].append(field)

    ids = out["Record ID"].astype("string").fillna("").str.lower().str.replace(r"[^a-z0-9]", "", regex=True)
    flag(ids.ne("") & ids.duplicated(False), "Critical", "Duplicate Record ID")

    address_key = (
        out["Street Address"].astype("string").fillna("").str.lower().str.replace(r"[^a-z0-9]", "", regex=True)
        + "|" + out["City"].astype("string").fillna("").str.lower().str.replace(r"[^a-z0-9]", "", regex=True)
        + "|" + out["Postal Code"].astype("string").fillna("").str.lower().str.replace(r"[^a-z0-9]", "", regex=True)
    )
    flag(address_key.str.split("|").str[0].ne("") & address_key.duplicated(False), "Warning", "Possible duplicate address")

    units = pd.to_numeric(out["Number of Apartments"].astype("string").str.replace(",", "", regex=False).str.extract(r"(\d+(?:\.\d+)?)", expand=False), errors="coerce")
    flag(~unresolved_mask(out["Number of Apartments"]) & (units.isna() | units.le(0)), "Warning", "Invalid number of apartments")

    # Evidence-backed count QA. A populated researched count should carry a
    # source tier, exact supporting URL, evidence, and confidence. A completed
    # blank count should document that the mandatory closure pass was performed.
    research_completed = out["Research Status"].astype("string").fillna("").str.strip().eq("Completed")

    apartment_count_present = ~unresolved_mask(out["Number of Apartments"])
    apartment_status = out["Apartment Count Search Status"].astype("string").fillna("").str.strip()
    apartment_confidence = out["Apartment Count Confidence"].astype("string").fillna("").str.strip()
    flag(
        apartment_count_present & apartment_status.isin({"", "Not Checked"}),
        "Warning",
        "Apartment count has no verified search-status/source tier",
    )
    flag(
        apartment_count_present & unresolved_mask(out["Apartment Count Source URL"]),
        "Warning",
        "Apartment count has no exact supporting source URL",
    )
    flag(
        apartment_count_present & unresolved_mask(out["Apartment Count Evidence"]),
        "Warning",
        "Apartment count has no supporting evidence",
    )
    flag(
        apartment_count_present & apartment_confidence.isin({"", "Not Checked"}),
        "Warning",
        "Apartment count confidence has not been assessed",
    )
    flag(
        research_completed
        & ~apartment_count_present
        & apartment_status.isin({"", "Not Checked"}),
        "Warning",
        "Apartment count closure pass is not documented",
    )
    flag(
        apartment_count_present
        & apartment_status.isin({"Not Found after Search", "Conflict — Needs Review"}),
        "Warning",
        "Apartment count is populated despite an unresolved/conflict search status",
    )
    flag(
        apartment_status.eq("Reputable property source") & apartment_confidence.eq("High"),
        "Warning",
        "Apartment count confidence is High although only a secondary property source is recorded",
    )

    storey_text = out["Number of Storeys"].astype("string").fillna("").str.strip()
    storey_numbers = storey_text.str.findall(r"\b\d+\b")
    storey_count_present = ~unresolved_mask(out["Number of Storeys"])
    storey_status = out["Storey Count Search Status"].astype("string").fillna("").str.strip()
    storey_confidence = out["Storey Count Confidence"].astype("string").fillna("").str.strip()
    flag(
        storey_count_present & storey_numbers.apply(lambda values: len(values) != 1),
        "Warning",
        "Number of Storeys should be one verified whole-building count",
    )
    flag(
        storey_count_present & storey_status.isin({"", "Not Checked"}),
        "Warning",
        "Storey count has no verified search-status/source tier",
    )
    flag(
        storey_count_present & unresolved_mask(out["Storey Count Source URL"]),
        "Warning",
        "Storey count has no exact supporting source URL",
    )
    flag(
        storey_count_present & unresolved_mask(out["Storey Count Evidence"]),
        "Warning",
        "Storey count has no supporting evidence",
    )
    flag(
        storey_count_present & storey_confidence.isin({"", "Not Checked"}),
        "Warning",
        "Storey count confidence has not been assessed",
    )
    flag(
        research_completed
        & ~storey_count_present
        & storey_status.isin({"", "Not Checked"}),
        "Warning",
        "Storey count closure pass is not documented",
    )
    flag(
        storey_count_present
        & storey_status.isin({"Not Found after Search", "Conflict — Needs Review"}),
        "Warning",
        "Storey count is populated despite an unresolved/conflict search status",
    )
    flag(
        storey_status.eq("Reputable property source") & storey_confidence.eq("High"),
        "Warning",
        "Storey count confidence is High although only a secondary property source is recorded",
    )

    for count_source_field in ["Apartment Count Source URL", "Storey Count Source URL"]:
        count_urls = out[count_source_field].astype("string").fillna("").str.lower().str.strip()
        flag(
            ~unresolved_mask(out[count_source_field])
            & ~count_urls.str.startswith(("http://", "https://"), na=False),
            "Warning",
            f"Invalid {count_source_field}",
        )

    email = out["Primary Email"].astype("string").fillna("").str.strip()
    flag(~unresolved_mask(out["Primary Email"]) & ~email.str.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", na=False), "Warning", "Invalid email format")
    phone = out["Phone"].astype("string").fillna("").str.replace(r"\D", "", regex=True)
    flag(~unresolved_mask(out["Phone"]) & ~phone.str.len().isin([10, 11]), "Warning", "Phone number does not contain 10 or 11 digits")
    pc = out["Postal Code"].astype("string").fillna("").str.upper().str.strip()
    flag(~unresolved_mask(out["Postal Code"]) & ~pc.str.match(r"^[A-Z]\d[A-Z][ -]?\d[A-Z]\d$", na=False), "Warning", "Invalid Canadian postal code format")

    rental_status = out["Rental Availability Status"].astype("string").fillna("").str.strip()
    flag(
        rental_status.eq("Rented"),
        "Warning",
        "Rental listing is explicitly marked Rented; keep for reconciliation but exclude from active rental inventory",
    )
    flag(
        rental_status.eq("Status Unclear — Needs Review"),
        "Warning",
        "Rental availability is unclear and requires human review before active-directory inclusion",
    )

    # City of Ottawa project checks. Text labels are useful, but geocoded
    # coordinates and a configured municipal-boundary polygon are stronger.
    province_text = out["Province"].astype("string").fillna("").str.strip().str.lower()
    country_text = out["Country"].astype("string").fillna("").str.strip().str.lower()
    city_text = out["City"].astype("string").fillna("").str.strip().str.lower()
    geo_status = out["Geographic Scope Status"].astype("string").fillna("").str.strip()

    flag(~unresolved_mask(out["Province"]) & ~province_text.isin({"ontario", "on"}), "Critical", "Outside City of Ottawa project province")
    flag(~unresolved_mask(out["Country"]) & ~country_text.isin({"canada", "ca"}), "Critical", "Outside City of Ottawa project country")
    flag(
        ~unresolved_mask(out["City"])
        & city_text.isin(OUT_OF_SCOPE_NEARBY_LOCALITIES),
        "Critical",
        "Physical municipality is outside the City of Ottawa",
    )
    flag(
        ~unresolved_mask(out["City"])
        & ~city_text.isin(OTTAWA_MUNICIPAL_LOCALITIES)
        & ~city_text.isin(OUT_OF_SCOPE_NEARBY_LOCALITIES),
        "Warning",
        "Verify City of Ottawa municipal scope",
    )
    flag(geo_status.eq("Outside City of Ottawa"), "Critical", "Geographic coordinates are outside the City of Ottawa boundary")
    flag(geo_status.eq("Needs Geographic Review"), "Warning", "Geographic position needs manual review")
    flag(geo_status.eq("Not Checked"), "Warning", "City of Ottawa geographic position has not been checked")

    latitude = pd.to_numeric(out["Latitude"], errors="coerce")
    longitude = pd.to_numeric(out["Longitude"], errors="coerce")
    has_coordinate_text = ~unresolved_mask(out["Latitude"]) | ~unresolved_mask(out["Longitude"])
    invalid_coordinates = has_coordinate_text & (
        latitude.isna() | longitude.isna()
        | latitude.lt(-90) | latitude.gt(90)
        | longitude.lt(-180) | longitude.gt(180)
    )
    flag(invalid_coordinates, "Warning", "Invalid latitude/longitude")
    flag(
        geo_status.eq("Inside City of Ottawa")
        & city_text.isin(OUT_OF_SCOPE_NEARBY_LOCALITIES),
        "Critical",
        "City text conflicts with geographic Ottawa-boundary result",
    )

    # PO Box and mailing-address QA. A PO Box must remain separate from the
    # apartment building's physical Street Address.
    street_text = out["Street Address"].astype("string").fillna("").str.strip()
    po_box_text = out["PO Box"].astype("string").fillna("").str.strip()
    po_box_pattern = r"(?i)\b(?:p\.?\s*o\.?\s*box|post(?:al| office)?\s+box|case\s+postale|c\.?p\.?|box)\s*[#:-]?\s*[a-z0-9-]+"
    flag(street_text.str.contains(po_box_pattern, regex=True, na=False), "Critical", "PO Box or mailing address placed in physical Street Address")
    flag(
        ~unresolved_mask(out["PO Box"])
        & ~po_box_text.str.contains(po_box_pattern, regex=True, na=False),
        "Warning",
        "PO Box format needs review",
    )
    flag(
        ~unresolved_mask(out["PO Box"])
        & unresolved_mask(out["PO Box Source URL"])
        & unresolved_mask(out["PO Box Evidence"]),
        "Warning",
        "PO Box has no supporting source or evidence",
    )
    po_pc = out["PO Box Postal Code"].astype("string").fillna("").str.upper().str.strip()
    flag(
        ~unresolved_mask(out["PO Box Postal Code"])
        & ~po_pc.str.match(r"^[A-Z]\d[A-Z][ -]?\d[A-Z]\d$", na=False),
        "Warning",
        "Invalid PO Box postal code format",
    )

    # Classification consistency: a reliable storey count determines the height
    # band. Conflicting or non-standard classifications must be reviewed.
    #
    # IMPORTANT: pandas 3.x can raise ``TypeError: boolean value of NA is
    # ambiguous`` when two object Series containing pd.NA are compared directly
    # with ``Series.ne``. Build NA-free comparison Series first, then apply the
    # validity masks. This keeps manual classifications and unresolved values
    # working without crashing the demo, sidebar, or QA view.
    derived_height = derive_height_classification_from_storeys(out["Number of Storeys"])
    normalized_classification = out["Building Classification"].apply(normalize_height_classification)
    controlled_labels = {"Low-rise", "Mid-rise", "High-rise"}

    derived_height_text = derived_height.astype("string").fillna("").str.strip()
    normalized_classification_text = (
        normalized_classification.astype("string").fillna("").str.strip()
    )

    classification_is_controlled = normalized_classification_text.isin(controlled_labels)
    classification_conflicts = (
        ~unresolved_mask(derived_height)
        & ~unresolved_mask(normalized_classification)
        & classification_is_controlled
        & normalized_classification_text.ne(derived_height_text)
    )

    flag(
        ~unresolved_mask(out["Building Classification"])
        & ~classification_is_controlled,
        "Warning",
        "Building classification is not Low-rise, Mid-rise, or High-rise",
    )
    flag(
        classification_conflicts,
        "Warning",
        "Building classification conflicts with storey count",
    )
    flag(
        ~unresolved_mask(normalized_classification)
        & classification_is_controlled
        & unresolved_mask(out["Number of Storeys"]),
        "Warning",
        "Building classification lacks supporting storey evidence",
    )
    for field in ["Website", "Source URL"]:
        urls = out[field].astype("string").fillna("").str.lower().str.strip()
        flag(~unresolved_mask(out[field]) & ~urls.str.startswith(("http://", "https://"), na=False), "Warning", f"Invalid {field}")

    dates = pd.to_datetime(out["Date Researched"], errors="coerce")
    today = pd.Timestamp.today().normalize()
    flag(~unresolved_mask(out["Date Researched"]) & dates.isna(), "Warning", "Invalid Date Researched")
    flag(dates.notna() & dates.gt(today), "Warning", "Date Researched is in the future")

    out["Working Record Label"] = resolved(out["Building Name"]).combine_first(resolved(out["Street Address"])).combine_first(resolved(out["Record ID"])).fillna("Unlabelled record")
    out["Core Gap Count"] = core_gaps.apply(len)
    out["Core Gaps"] = core_gaps.apply(lambda v: ", ".join(v) if v else "None")
    out["Research Gap Count"] = research_gaps.apply(len)
    out["Research Gaps"] = research_gaps.apply(lambda v: ", ".join(v) if v else "None")
    out["Critical Issue Count"] = issues.apply(lambda v: sum(s == "Critical" for s, _ in v))
    out["Warning Count"] = issues.apply(lambda v: sum(s == "Warning" for s, _ in v))
    out["QA Flag Count"] = issues.apply(len)
    out["QA Flags"] = issues.apply(lambda v: "; ".join(f"{s}: {m}" for s, m in v) if v else "No rental property data issues found")
    out["QA Status"] = out.apply(lambda r: "Critical" if r["Critical Issue Count"] else "Review" if r["Warning Count"] else "Pass", axis=1)
    out["Core Completeness %"] = ((len(CORE_FIELDS) - out["Core Gap Count"]) / len(CORE_FIELDS) * 100).round(1)
    out["Target Coverage %"] = ((len(TARGET_FIELDS) - out["Research Gap Count"]) / len(TARGET_FIELDS) * 100).round(1)

    workflow = pd.Series([[] for _ in range(len(out))], index=out.index, dtype="object")
    def gap(mask, message):
        for idx in out.index[mask.fillna(False)]:
            workflow.at[idx].append(message)
    gap(unresolved_mask(out["Source URL"]), "Research source not recorded")
    gap(unresolved_mask(out["Date Researched"]), "Research date not recorded")
    gap(unresolved_mask(out["Researcher"]), "Researcher not recorded")
    gap(out["Research Status"].isin(["Imported - Needs Review", "Not Started", "In Progress"]), "Research is not complete")
    gap(out["Research Status"].eq("Needs Follow-up"), "Research requires follow-up")
    gap(out["Source Status"].eq("Not Checked"), "Source not checked")
    gap(out["Source Status"].isin(["Needs Follow-up", "Unavailable"]), "Source requires documentation or follow-up")
    gap(out["Verification Status"].ne("Verified"), "Human verification not completed")
    gap(out["Record Decision"].eq("Undecided"), "Record decision not made")
    gap(out["Record Decision"].isin(["Update", "Possible Duplicate"]), "Record decision requires action")
    gap(out["Research Status"].eq("Completed") & out["Research Gap Count"].gt(0) & unresolved_mask(out["Missing Information"]), "Document unavailable information")
    out["Workflow Gap Count"] = workflow.apply(len)
    out["Workflow Gaps"] = workflow.apply(lambda v: "; ".join(v) if v else "No workflow gaps")

    def readiness(row):
        # Human approval is authoritative for non-critical QA warnings. Critical
        # identity/data failures still block a record, and the research trail must
        # still be complete before a record can be treated as ready.
        if row["Record Decision"] == "Remove": return "Excluded from Listings"
        if row["Record Decision"] == "Possible Duplicate": return "Duplicate Review"
        if row["Critical Issue Count"]: return "Fix Critical Data"
        if row["Research Status"] in ["Imported - Needs Review", "Not Started", "In Progress"]: return "Needs Research"
        if row["Research Status"] == "Needs Follow-up": return "Needs Follow-up"
        if row["Research Status"] != "Completed": return "Needs Review"
        if row["Verification Status"] != "Verified": return "Needs Verification"
        if row["Record Decision"] == "Update": return "Needs Update"
        if row["Record Decision"] != "Keep": return "Needs Decision"
        if is_unresolved(row["Source URL"]) and row["Source Status"] != "Unavailable": return "Record Research Source"
        if is_unresolved(row["Date Researched"]) or is_unresolved(row["Researcher"]): return "Complete Research Trail"
        if row["Research Gap Count"] and is_unresolved(row["Missing Information"]): return "Document Research Gaps"
        if row["Warning Count"]: return "Ready with Reviewed Warnings"
        if row["Research Gap Count"]: return "Ready with Documented Gaps"
        return "Ready to Use"
    out["Record Readiness"] = out.apply(readiness, axis=1)

    # One clear business-facing export decision. A human-approved record can be
    # exported when review is complete, verification is complete, the reviewer
    # decided to keep it, and no critical identity/data issue remains. Warnings
    # and documented research gaps stay visible for audit purposes but do not
    # silently remove an otherwise approved record from export.
    approved_for_export = (
        out["Research Status"].eq("Completed")
        & out["Verification Status"].eq("Verified")
        & out["Record Decision"].eq("Keep")
        & out["Critical Issue Count"].eq(0)
    )
    out["Export Status"] = "Still in Review"
    out.loc[approved_for_export, "Export Status"] = "Approved for Export"
    out.loc[out["Record Decision"].eq("Remove"), "Export Status"] = "Excluded"

    out["Follow-up Priority"] = out.apply(
        lambda r: "None" if r["Record Readiness"] in ["Ready to Use", "Ready with Documented Gaps", "Ready with Reviewed Warnings", "Excluded from Listings"]
        else "High" if r["Critical Issue Count"] or r["Record Readiness"] in ["Duplicate Review", "Needs Follow-up"]
        else "Medium" if r["Warning Count"] or r["Research Gap Count"] else "Low", axis=1
    )
    age = (today - dates).dt.days
    out["Source Age (Days)"] = age.where(dates.notna() & dates.le(today)).astype("Int64")
    out["Freshness Status"] = "Current"
    out.loc[unresolved_mask(out["Date Researched"]), "Freshness Status"] = "Missing date"
    out.loc[~unresolved_mask(out["Date Researched"]) & dates.isna(), "Freshness Status"] = "Invalid date"
    out.loc[dates.gt(today), "Freshness Status"] = "Future date"
    out.loc[dates.notna() & dates.le(today) & age.gt(FRESHNESS_DAYS), "Freshness Status"] = "Stale"
    return out


def listing_export(df):
    """Return a flat export with the required sample fields first."""
    listing = pd.DataFrame(index=df.index)
    for label, source_field in LISTING_FIELD_MAP:
        listing[label] = (
            df.apply(formatted_location, axis=1)
            if source_field is None
            else df[source_field]
        )
    for label, source_field in LISTING_ADDITIONAL_FIELD_MAP:
        listing[label] = df[source_field] if source_field in df.columns else pd.NA
    columns = LISTING_COLUMNS + [label for label, _ in LISTING_ADDITIONAL_FIELD_MAP]
    return listing[columns]


def listing_block_dataframe(row, include_additional=True):
    """Turn one record into the same vertical field/value order as the sample."""
    rows = []
    for label, source_field in LISTING_FIELD_MAP:
        value = formatted_location(row) if source_field is None else row.get(source_field)
        rows.append({
            "Listing Field": label,
            "Listing Value": _excel_display_value(value),
        })

    if include_additional:
        additional_rows = []
        for label, source_field in LISTING_ADDITIONAL_FIELD_MAP:
            value = row.get(source_field)
            if not is_unresolved(value):
                additional_rows.append({
                    "Listing Field": label,
                    "Listing Value": _excel_display_value(value),
                })
        if additional_rows:
            rows.append({
                "Listing Field": "Additional information and research reference",
                "Listing Value": "",
            })
            rows.extend(additional_rows)
    return pd.DataFrame(rows)


def render_listing_preview(df, limit=5):
    """Show sample-style listing blocks without turning the page into a wide table."""
    if df.empty:
        st.info("No residential rental-property records are available to preview yet.")
        return

    visible = df.head(limit)
    for listing_number, (_, row) in enumerate(visible.iterrows(), start=1):
        name = _excel_display_value(row.get("Building Name")) or "Unnamed rental property"
        with st.expander(
            f"Apartment Building {listing_number}: {name}",
            expanded=listing_number == 1,
        ):
            source_url = _excel_display_value(row.get("Source URL"))
            if source_url.startswith(("http://", "https://")):
                st.link_button(
                    "Open supporting source",
                    source_url,
                    type="secondary",
                    use_container_width=False,
                )
            else:
                st.caption("No supporting source link has been recorded for this listing.")

            st.dataframe(
                listing_block_dataframe(row),
                width="stretch",
                hide_index=True,
                column_config={
                    "Listing Field": st.column_config.TextColumn(
                        "Listing Field",
                        width="medium",
                    ),
                    "Listing Value": st.column_config.TextColumn(
                        "Listing Value",
                        width="large",
                    ),
                },
            )

    if len(df) > limit:
        st.caption(
            f"Showing {limit:,} of {len(df):,} listings. "
            "The workbook download includes every building."
        )


def approved_for_export_mask(df):
    """Return records a reviewer has explicitly approved for CSV export."""
    if "Export Status" in df.columns:
        return df["Export Status"].eq("Approved for Export")
    return (
        df["Research Status"].eq("Completed")
        & df["Verification Status"].eq("Verified")
        & df["Record Decision"].eq("Keep")
        & df.get("Critical Issue Count", pd.Series(0, index=df.index)).eq(0)
    )


def ready_mask(df):
    """Backward-compatible alias for the user-facing approved-for-export state."""
    return approved_for_export_mask(df)


def research_log(df):
    columns = [
        "Record ID", "Working Record Label", "Building Name", "Management/Owner",
        "Street Address", "City", "Province", "Postal Code", "Source URL",
        "Date Researched", "Source Age (Days)", "Freshness Status", "Researcher",
        "Research Status", "Source Status", "Verification Status",
        "Research Gap Count", "Research Gaps", "Missing Information",
        "Reviewer Notes", "Record Decision", "Directory Entry Status", "Follow-up Priority",
        "Workflow Gap Count", "Workflow Gaps", "Record Readiness", "Export Status",
    ]
    return df[[c for c in columns if c in df.columns]].copy()


def owner_summary(df):
    working = df.assign(_owner=display_values(df["Management/Owner"], "Unassigned"))
    rows = []
    for owner, group in working.groupby("_owner", dropna=False):
        rows.append({
            "Management/Owner": owner,
            "Building Records": len(group),
            "Named Buildings": int((~unresolved_mask(group["Building Name"])).sum()),
            "Cities": ", ".join(sorted(set(resolved(group["City"]).dropna().astype(str).str.strip()))),
            "Records with Website": int((~unresolved_mask(group["Website"])).sum()),
            "Records with Apartment Count": int((~unresolved_mask(group["Number of Apartments"])).sum()),
            "Verified Records": int(group["Verification Status"].eq("Verified").sum()),
            "Approved for Export": int(approved_for_export_mask(group).sum()),
            "Still in Review": int((~approved_for_export_mask(group) & ~group["Record Decision"].eq("Remove")).sum()),
        })
    return pd.DataFrame(rows).sort_values(["Still in Review", "Building Records"], ascending=[False, False]).reset_index(drop=True) if rows else pd.DataFrame()


def draft_profiles(df):
    rows = []
    for _, row in df.iterrows():
        if row["Record Decision"] == "Remove":
            continue
        label = str(row["Working Record Label"]).strip() or "Rental property"
        sentences = [f"{label} is located at {row['Street Address']}, {formatted_location(row)}."]
        if not is_unresolved(row["Management/Owner"]): sentences.append(f"The recorded management or owner is {row['Management/Owner']}.")
        if not is_unresolved(row["Property Type"]): sentences.append(f"The recorded property type is {row['Property Type']}.")
        if not is_unresolved(row["Building Classification"]): sentences.append(f"The current building classification is {row['Building Classification']}.")
        if not is_unresolved(row["Number of Apartments"]): sentences.append(f"The source records approximately {row['Number of Apartments']} apartments.")
        contact = []
        for label_text, field in [("phone", "Phone"), ("email", "Primary Email"), ("website", "Website")]:
            if not is_unresolved(row[field]): contact.append(f"{label_text}: {row[field]}")
        if contact: sentences.append("Contact information: " + "; ".join(contact) + ".")
        rows.append({
            "Record ID": row["Record ID"], "Profile Heading": label,
            "Management/Owner": row["Management/Owner"], "Draft Profile": " ".join(sentences),
            "Research Gaps": row["Research Gaps"], "Source URL": row["Source URL"],
            "Verification Status": row["Verification Status"],
            "Profile Status": "Ready for editorial review" if ready_mask(pd.DataFrame([row])).iloc[0] else "Needs research or verification",
            "Editorial Note": "Confirm the facts and refine the wording before use.",
        })
    return pd.DataFrame(rows)


def field_coverage(df):
    rows = []
    for field in ALL_RESEARCH_FIELDS:
        missing = int(unresolved_mask(df[field]).sum())
        rows.append({
            "Field": field,
            "Field Group": "Core field" if field in CORE_FIELDS else "Useful detail",
            "Missing Records": missing,
            "Populated Records": len(df) - missing,
            "Coverage": f"{((len(df)-missing)/len(df)*100 if len(df) else 0):.1f}%",
            "How Datablix treats a blank": "Prevents the record from being treated as complete" if field in CORE_FIELDS else "Keeps the detail visible as an open gap rather than an error",
        })
    return pd.DataFrame(rows)



# =========================================================
# Project deliverables analytics
# =========================================================

# These are candidate listing and search fields that Datablix can evaluate
# from the evidence actually collected during research.  The matrix is deliberately
# evidence-first: coverage is calculated before any recommendation is shown.
CLOSEOUT_FIELD_SPECS = [
    {
        "Field": "Location",
        "Source Fields": ["Street Address", "City", "Postal Code"],
        "Mode": "all",
        "Directory Use": "Profile + Filter",
        "Decision Value": "High",
        "Base Recommendation": "Priority",
        "Rationale": "Location is fundamental when households compare areas and proximity to services, transportation, family support, and familiar neighbourhoods.",
    },
    {
        "Field": "Rental Rate",
        "Source Fields": ["Rental Rate Range"],
        "Mode": "all",
        "Directory Use": "Filter",
        "Decision Value": "High",
        "Base Recommendation": "Priority",
        "Rationale": "Housing cost is a primary decision factor, so reliable rent information provides direct comparison value.",
    },
    {
        "Field": "Suite Types",
        "Source Fields": ["Suite Types"],
        "Mode": "all",
        "Directory Use": "Filter",
        "Decision Value": "High",
        "Base Recommendation": "Priority",
        "Rationale": "Studio, one-bedroom, two-bedroom, and other suite configurations help users match housing to household size and space needs.",
    },
    {
        "Field": "Property Type",
        "Source Fields": ["Property Type"],
        "Mode": "all",
        "Directory Use": "Category + Filter",
        "Decision Value": "High",
        "Base Recommendation": "Priority",
        "Rationale": "Property form helps users distinguish apartments, townhomes, duplexes, garden homes, semi-detached homes, and other supported rental formats.",
    },
    {
        "Field": "Building Classification",
        "Source Fields": ["Building Classification"],
        "Mode": "all",
        "Directory Use": "Filter",
        "Decision Value": "High",
        "Base Recommendation": "Priority",
        "Rationale": "A separate storey-based height band lets users distinguish low-rise, mid-rise, and high-rise buildings without mixing height with property form.",
    },
    {
        "Field": "Rental Availability",
        "Source Fields": ["Rental Availability Status"],
        "Mode": "availability_status",
        "Directory Use": "Filter + Freshness Date",
        "Decision Value": "High",
        "Base Recommendation": "Use with Last Verified date",
        "Rationale": "Availability is useful but time-sensitive. A current property page can remain online after a unit is rented, so the status should be paired with a verification date and clear wording.",
    },
    {
        "Field": "Elevator",
        "Source Fields": ["Elevator"],
        "Mode": "all",
        "Directory Use": "Filter",
        "Decision Value": "Very High",
        "Base Recommendation": "Recommended",
        "Rationale": "Elevator access can materially affect usability and suitability for people with different mobility needs.",
    },
    {
        "Field": "Accessibility",
        "Source Fields": ["Accessibility"],
        "Mode": "all",
        "Directory Use": "Future Filter",
        "Decision Value": "Very High",
        "Base Recommendation": "Recommended",
        "Rationale": "Accessibility has high decision value, but a reliable filter requires sufficiently complete and consistently defined data.",
    },
    {
        "Field": "Parking",
        "Source Fields": ["Parking"],
        "Mode": "all",
        "Directory Use": "Filter",
        "Decision Value": "High",
        "Base Recommendation": "Recommended",
        "Rationale": "Parking can matter to residents, caregivers, family visitors, and service providers, making it a practical comparison field.",
    },
    {
        "Field": "Laundry",
        "Source Fields": ["Laundry"],
        "Mode": "all",
        "Directory Use": "Field + Filter",
        "Decision Value": "High",
        "Base Recommendation": "Recommended",
        "Rationale": "In-suite versus shared laundry can affect convenience and day-to-day independence, so variation across properties is useful to surface.",
    },
    {
        "Field": "Utilities",
        "Source Fields": ["Utilities"],
        "Mode": "all",
        "Directory Use": "Field + Filter",
        "Decision Value": "High",
        "Base Recommendation": "Recommended",
        "Rationale": "Utilities affect the real monthly housing cost, so they add affordability context beyond the advertised rent alone.",
    },
    {
        "Field": "Pet Policy",
        "Source Fields": ["Pet Policy"],
        "Mode": "all",
        "Directory Use": "Optional Filter",
        "Decision Value": "Medium",
        "Base Recommendation": "Useful",
        "Rationale": "Pet ownership can determine whether a property is suitable for some households, although it is less universal than cost, location, or accessibility.",
    },
    {
        "Field": "Smoke-Free",
        "Source Fields": ["Smoke-Free"],
        "Mode": "all",
        "Directory Use": "Optional Filter",
        "Decision Value": "High",
        "Base Recommendation": "Useful",
        "Rationale": "Smoke-free policies can influence comfort and health-related housing preferences and may be useful when the information is consistently available.",
    },
    {
        "Field": "Number of Apartments",
        "Source Fields": ["Number of Apartments"],
        "Mode": "all",
        "Directory Use": "Profile Field",
        "Decision Value": "Medium",
        "Base Recommendation": "Profile only",
        "Rationale": "Unit count helps describe the scale of a rental community but is less likely to be a primary search criterion than location, price, or accessibility.",
    },
    {
        "Field": "Phone",
        "Source Fields": ["Phone"],
        "Mode": "all",
        "Directory Use": "Profile Field",
        "Decision Value": "High",
        "Base Recommendation": "Profile only",
        "Rationale": "A direct contact number supports follow-up and verification but is better presented as profile information than as a search filter.",
    },
    {
        "Field": "Website",
        "Source Fields": ["Property Website", "Website", "Company Website"],
        "Mode": "any",
        "Directory Use": "Profile Link",
        "Decision Value": "High",
        "Base Recommendation": "Profile only",
        "Rationale": "A verified official webpage gives users a direct path to current property details and contact information, but it is not itself a meaningful filter.",
    },
    {
        "Field": "Last Verified / Research Date",
        "Source Fields": ["Date Researched"],
        "Mode": "all",
        "Directory Use": "Freshness / Trust Field",
        "Decision Value": "High",
        "Base Recommendation": "Recommended",
        "Rationale": "Public rental information changes over time, so a visible or internally maintained verification date helps users assess and manage data freshness.",
    },
]


def closeout_research_population(qa_frame: pd.DataFrame) -> pd.DataFrame:
    """Return directory-relevant researched rows for field-availability analysis.

    Excluded, removed, and confirmed out-of-Ottawa records remain part of the audit
    trail, but they should not lower coverage for fields intended for the public
    directory.
    """
    if not isinstance(qa_frame, pd.DataFrame) or qa_frame.empty:
        return qa_frame.iloc[0:0].copy() if isinstance(qa_frame, pd.DataFrame) else pd.DataFrame()
    out = qa_frame.copy()
    mask = pd.Series(True, index=out.index)
    if "Record Decision" in out.columns:
        mask &= ~out["Record Decision"].astype(str).eq("Remove")
    if "Directory Discovery Status" in out.columns:
        mask &= ~out["Directory Discovery Status"].astype(str).eq("Excluded / Not Current")
    if "Current Inventory Status" in out.columns:
        inventory = out["Current Inventory Status"].astype("string").fillna("").str.strip().str.lower()
        mask &= ~inventory.str.startswith("excluded")
    if "Geographic Scope Status" in out.columns:
        mask &= ~out["Geographic Scope Status"].astype(str).eq("Outside City of Ottawa")
    return out.loc[mask].copy()


def deliverable_research_population(qa_frame: pd.DataFrame) -> pd.DataFrame:
    """Return the full valid/current researched inventory used by project deliverables.

    Deliverable inclusion is intentionally broader than Approved for Export.
    Existing researched records must not disappear simply because they have not
    been marked Keep for an external-entry workflow.
    """
    population = closeout_research_population(qa_frame)
    if population.empty:
        return population.copy()
    sort_fields = [
        field for field in ["Management/Owner", "Building Name", "Street Address"]
        if field in population.columns
    ]
    if sort_fields:
        population = population.sort_values(sort_fields, kind="stable")
    return population.copy()


def deliverable_needs_review_mask(qa_frame: pd.DataFrame) -> pd.Series:
    """Return genuine unresolved review items without using export approval as a proxy.

    An existing record can be fully researched and verified while still lacking
    Record Decision = Keep. That is an export-workflow state, not automatically
    a research-quality problem.
    """
    if not isinstance(qa_frame, pd.DataFrame) or qa_frame.empty:
        return pd.Series(False, index=getattr(qa_frame, "index", pd.Index([])), dtype=bool)

    mask = pd.Series(False, index=qa_frame.index, dtype=bool)

    if "Verification Status" in qa_frame.columns:
        mask |= ~qa_frame["Verification Status"].astype(str).eq("Verified")

    if "QA Status" in qa_frame.columns:
        mask |= qa_frame["QA Status"].astype(str).eq("Critical")

    if "Research Status" in qa_frame.columns:
        mask |= qa_frame["Research Status"].astype(str).isin({
            "Imported - Needs Review", "Not Started", "In Progress",
            "Needs Follow-up", "Ready for Review",
        })

    if "Directory Discovery Status" in qa_frame.columns:
        mask |= qa_frame["Directory Discovery Status"].astype(str).isin({
            "Needs Classification", "Possible Duplicate",
        })

    if "Geographic Scope Status" in qa_frame.columns:
        mask |= qa_frame["Geographic Scope Status"].astype(str).eq("Needs Geographic Review")

    # Removed / excluded / confirmed out-of-scope records are audit records,
    # not unresolved active directory records.
    active_indices = deliverable_research_population(qa_frame).index
    return mask & qa_frame.index.to_series().isin(active_indices)


def full_research_status_summary(qa_frame: pd.DataFrame) -> pd.DataFrame:
    """Summarize every researched row without applying final-directory filters."""
    if not isinstance(qa_frame, pd.DataFrame) or qa_frame.empty:
        return pd.DataFrame(columns=["Status Area", "Status", "Records", "Share of Full Research"])

    total = len(qa_frame)
    rows = []

    def add_area(area: str, column: str):
        if column not in qa_frame.columns:
            return
        values = (
            qa_frame[column]
            .astype("string")
            .fillna("")
            .str.strip()
            .replace("", "Blank / Not Recorded")
        )
        counts = values.value_counts(dropna=False)
        for status, count in counts.items():
            rows.append({
                "Status Area": area,
                "Status": str(status),
                "Records": int(count),
                "Share of Full Research": f"{(int(count) / total * 100):.1f}%" if total else "0.0%",
            })

    add_area("Discovery", "Directory Discovery Status")
    add_area("Inventory", "Current Inventory Status")
    add_area("Rental Availability", "Rental Availability Status")
    add_area("Research", "Research Status")
    add_area("Verification", "Verification Status")
    add_area("Record Decision", "Record Decision")
    add_area("Geographic Scope", "Geographic Scope Status")
    return pd.DataFrame(rows)


def full_company_research_summary(qa_frame: pd.DataFrame, registry=None) -> pd.DataFrame:
    """Summarize all researched records by organization, including excluded/review rows."""
    registry = normalize_company_registry(registry)
    if not isinstance(qa_frame, pd.DataFrame) or qa_frame.empty:
        return pd.DataFrame(columns=[
            "Company", "All Research Records", "Existing Source Records",
            "Newly Discovered", "Needs Classification", "Possible Duplicates",
            "Excluded / Not Current", "Completed", "Verified", "Rented",
        ])

    rows = []
    represented = set()

    def summarize(company_name: str, group: pd.DataFrame):
        discovery = group.get("Directory Discovery Status", pd.Series("", index=group.index)).astype(str)
        rental = group.get("Rental Availability Status", pd.Series("", index=group.index)).astype(str)
        research = group.get("Research Status", pd.Series("", index=group.index)).astype(str)
        verification = group.get("Verification Status", pd.Series("", index=group.index)).astype(str)
        return {
            "Company": company_name,
            "All Research Records": len(group),
            "Existing Source Records": int(discovery.eq("Existing Source Record").sum()),
            "Newly Discovered": int(discovery.eq("Newly Discovered").sum()),
            "Needs Classification": int(discovery.eq("Needs Classification").sum()),
            "Possible Duplicates": int(discovery.eq("Possible Duplicate").sum()),
            "Excluded / Not Current": int(discovery.eq("Excluded / Not Current").sum()),
            "Completed": int(research.eq("Completed").sum()),
            "Verified": int(verification.eq("Verified").sum()),
            "Rented": int(rental.eq("Rented").sum()),
        }

    for _, company in registry.iterrows():
        company_id = safe_text(company.get("Company ID", ""))
        company_name = safe_text(company.get("Management/Owner", "")) or company_id or "Unassigned"
        represented.add(company_id)
        group = (
            qa_frame.loc[qa_frame["Company ID"].astype(str).eq(company_id)].copy()
            if company_id and "Company ID" in qa_frame.columns
            else qa_frame.iloc[0:0].copy()
        )
        rows.append(summarize(company_name, group))

    if "Company ID" in qa_frame.columns:
        remaining = qa_frame.loc[~qa_frame["Company ID"].astype(str).isin(represented)].copy()
    else:
        remaining = qa_frame.copy()

    if not remaining.empty:
        owners = display_values(
            remaining.get("Management/Owner", pd.Series("", index=remaining.index)),
            "Unassigned",
        )
        for owner, group in remaining.assign(_owner=owners).groupby("_owner", dropna=False):
            rows.append(summarize(str(owner), group))

    return pd.DataFrame(rows)


def _closeout_field_available_mask(frame: pd.DataFrame, spec: dict) -> pd.Series:
    if not isinstance(frame, pd.DataFrame) or frame.empty:
        return pd.Series(False, index=getattr(frame, "index", pd.Index([])), dtype=bool)
    fields = [field for field in spec.get("Source Fields", []) if field in frame.columns]
    if not fields:
        return pd.Series(False, index=frame.index, dtype=bool)

    mode = spec.get("Mode", "all")
    if mode == "availability_status":
        status = frame[fields[0]].astype("string").fillna("").str.strip()
        return status.isin({"Available Now", "Available Future", "Rented"})
    if mode == "any":
        mask = pd.Series(False, index=frame.index)
        for field in fields:
            mask |= ~unresolved_mask(frame[field])
        return mask

    mask = pd.Series(True, index=frame.index)
    for field in fields:
        mask &= ~unresolved_mask(frame[field])
    return mask


def _coverage_assessment(coverage: float) -> str:
    if coverage >= 90:
        return "Strong"
    if coverage >= 70:
        return "Good"
    if coverage >= 40:
        return "Partial"
    if coverage > 0:
        return "Weak"
    return "Unavailable"


def _coverage_recommendation(spec: dict, coverage: float) -> str:
    field = spec.get("Field", "")
    base = spec.get("Base Recommendation", "Recommended")
    if base == "Profile only":
        return "Profile only"
    if field == "Rental Availability":
        return "Use with Last Verified date" if coverage > 0 else "Improve data collection first"
    if field == "Accessibility" and coverage < 70:
        return "Improve data collection first"
    if coverage < 40:
        return "Future field — improve data"
    if coverage < 70 and base in {"Priority", "Recommended"}:
        return "Pilot / optional until coverage improves"
    return base


def field_availability_summary(qa_frame: pd.DataFrame, include_all: bool = False) -> pd.DataFrame:
    """Summarize field availability across full research or directory candidates."""
    population = qa_frame.copy() if include_all else closeout_research_population(qa_frame)
    total = len(population)
    rows = []
    for spec in CLOSEOUT_FIELD_SPECS:
        available = _closeout_field_available_mask(population, spec)
        found = int(available.sum()) if total else 0
        missing = max(total - found, 0)
        coverage = (found / total * 100) if total else 0.0
        assessment = _coverage_assessment(coverage)
        recommendation = _coverage_recommendation(spec, coverage)
        if assessment in {"Strong", "Good"}:
            finding = f"{assessment} public-data availability across the researched directory candidates."
        elif assessment == "Partial":
            finding = "Information was available for a meaningful share of properties, but coverage is not yet consistent."
        elif assessment == "Weak":
            finding = "The information was rarely confirmed from public sources and represents a material data gap."
        else:
            finding = "No usable information was confirmed for this field in the current research scope."

        rows.append({
            "Field": spec["Field"],
            "Found": found,
            "Missing": missing,
            "Coverage %": round(coverage, 1),
            "Availability Assessment": assessment,
            "Research Finding": finding,
            "Decision Value": spec["Decision Value"],
            "Suggested Directory Use": spec["Directory Use"],
            "Recommendation": recommendation,
            "Rationale": spec["Rationale"],
        })
    return pd.DataFrame(rows)


def field_availability_matrix(qa_frame: pd.DataFrame, registry=None, include_all: bool = False) -> pd.DataFrame:
    """Return a company x field matrix across full research or directory candidates."""
    registry = normalize_company_registry(registry)
    population = qa_frame.copy() if include_all else closeout_research_population(qa_frame)
    rows = []
    represented = set()

    for _, company in registry.iterrows():
        company_id = safe_text(company.get("Company ID", ""))
        company_name = safe_text(company.get("Management/Owner", "")) or company_id or "Unassigned"
        group = population.loc[population["Company ID"].astype(str).eq(company_id)].copy() if company_id and "Company ID" in population.columns else population.iloc[0:0].copy()
        represented.add(company_id)
        row = {"Company": company_name, "Records": len(group)}
        for spec in CLOSEOUT_FIELD_SPECS:
            if group.empty:
                row[spec["Field"]] = "—"
            else:
                found = int(_closeout_field_available_mask(group, spec).sum())
                row[spec["Field"]] = f"{found / len(group) * 100:.0f}%"
        rows.append(row)

    if not population.empty and "Company ID" in population.columns:
        unregistered = population.loc[~population["Company ID"].astype(str).isin(represented)].copy()
        for owner, group in unregistered.assign(
            _owner=display_values(unregistered["Management/Owner"], "Unassigned")
        ).groupby("_owner", dropna=False):
            row = {"Company": str(owner), "Records": len(group)}
            for spec in CLOSEOUT_FIELD_SPECS:
                found = int(_closeout_field_available_mask(group, spec).sum())
                row[spec["Field"]] = f"{found / len(group) * 100:.0f}%" if len(group) else "—"
            rows.append(row)

    overall = {"Company": "OVERALL", "Records": len(population)}
    for spec in CLOSEOUT_FIELD_SPECS:
        found = int(_closeout_field_available_mask(population, spec).sum()) if not population.empty else 0
        overall[spec["Field"]] = f"{found / len(population) * 100:.0f}%" if len(population) else "—"
    rows.append(overall)
    return pd.DataFrame(rows)


def discovery_submission_summary(qa_frame: pd.DataFrame) -> pd.DataFrame:
    """Show research discoveries and recorded external-submission status."""
    if not isinstance(qa_frame, pd.DataFrame) or qa_frame.empty or "Directory Discovery Status" not in qa_frame.columns:
        return pd.DataFrame(columns=[
            "Building Name", "Management/Owner", "Street Address", "City",
            "External Submission Status", "Verification Status",
        ])
    discoveries = qa_frame.loc[
        qa_frame["Directory Discovery Status"].astype(str).eq("Newly Discovered")
    ].copy()
    if discoveries.empty:
        return pd.DataFrame(columns=[
            "Building Name", "Management/Owner", "Street Address", "City",
            "External Submission Status", "Verification Status",
        ])
    status_map = {
        "Entered": "Submitted Externally",
        "Needs Correction": "Needs Correction",
        "Not Entered": "Not Recorded",
    }
    discoveries["External Submission Status"] = discoveries.get(
        "Directory Entry Status", pd.Series("Not Entered", index=discoveries.index)
    ).astype(str).map(status_map).fillna("Not Recorded")
    columns = [
        "Building Name", "Management/Owner", "Street Address", "City",
        "External Submission Status", "Verification Status",
    ]
    return discoveries[[column for column in columns if column in discoveries.columns]].reset_index(drop=True)


def _project_material_change_counts(qa_frame: pd.DataFrame, registry=None, baseline=None) -> dict:
    """Count research records with material Starting Data differences by company."""
    registry = normalize_company_registry(registry)
    if not isinstance(baseline, pd.DataFrame):
        baseline = current_starting_source_records()
    counts = {}
    if not isinstance(qa_frame, pd.DataFrame) or qa_frame.empty:
        return counts

    material_labels = {
        "Changed — verify", "Possible typo / minor text change",
        "Added in research", "Missing from research",
    }
    for _, company in registry.iterrows():
        company_id = safe_text(company.get("Company ID", ""))
        company_name = safe_text(company.get("Management/Owner", ""))
        research = qa_frame.loc[
            qa_frame["Company ID"].astype(str).eq(company_id)
        ].copy() if company_id and "Company ID" in qa_frame.columns else qa_frame.iloc[0:0].copy()
        source = company_source_records_for_research(
            baseline,
            company_id=company_id,
            company_name=company_name,
        )
        if source.empty or research.empty:
            counts[company_id] = 0
            continue
        reconciliation = company_source_presence_reconciliation(source, research)
        differences = company_source_field_comparison(source, research, reconciliation=reconciliation)
        if differences.empty:
            counts[company_id] = 0
            continue
        material = differences.loc[differences["Comparison"].isin(material_labels)].copy()
        if material.empty:
            counts[company_id] = 0
        elif "Matched Research Record ID" in material.columns:
            ids = material["Matched Research Record ID"].astype(str).replace("", pd.NA).dropna()
            counts[company_id] = int(ids.nunique()) if not ids.empty else int(material["Matched Research Record"].nunique())
        else:
            counts[company_id] = int(material["Matched Research Record"].nunique())
    return counts


def closeout_research_coverage_matrix(qa_frame: pd.DataFrame, registry=None, baseline=None) -> pd.DataFrame:
    """Create the project deliverables company matrix."""
    registry = normalize_company_registry(registry)
    if not isinstance(baseline, pd.DataFrame):
        baseline = current_starting_source_records()
    change_counts = _project_material_change_counts(qa_frame, registry, baseline=baseline)
    rows = []
    represented = set()

    for _, company in registry.iterrows():
        company_id = safe_text(company.get("Company ID", ""))
        company_name = safe_text(company.get("Management/Owner", "")) or company_id or "Unassigned"
        represented.add(company_id)
        group = qa_frame.loc[qa_frame["Company ID"].astype(str).eq(company_id)].copy() if company_id and "Company ID" in qa_frame.columns else qa_frame.iloc[0:0].copy()
        source = company_source_records_for_research(
            baseline,
            company_id=company_id,
            company_name=company_name,
        )
        existing = int(group["Directory Discovery Status"].eq("Existing Source Record").sum()) if "Directory Discovery Status" in group.columns else 0
        discoveries = int(group["Directory Discovery Status"].eq("Newly Discovered").sum()) if "Directory Discovery Status" in group.columns else 0
        submitted = int((
            group.get("Directory Discovery Status", pd.Series("", index=group.index)).astype(str).eq("Newly Discovered")
            & group.get("Directory Entry Status", pd.Series("", index=group.index)).astype(str).eq("Entered")
        ).sum()) if not group.empty else 0
        needs_review = int(deliverable_needs_review_mask(group).sum()) if not group.empty else 0
        changed = int(change_counts.get(company_id, 0))
        rows.append({
            "Company": company_name,
            "Source Records": len(source),
            "Research Records": len(group),
            "Existing Matches": existing,
            "New Discoveries": discoveries,
            "Changed Existing Records": changed,
            "Discoveries Submitted": submitted,
            "Needs Review": needs_review,
        })

    return pd.DataFrame(rows)


def coverage_matrix_totals(coverage_matrix: pd.DataFrame) -> dict:
    """Return safe project/company totals without changing the analytical matrix itself."""
    numeric_columns = [
        "Source Records",
        "Research Records",
        "Existing Matches",
        "New Discoveries",
        "Changed Existing Records",
        "Discoveries Submitted",
        "Needs Review",
    ]
    totals = {"Company": "TOTAL"}
    for column in numeric_columns:
        if isinstance(coverage_matrix, pd.DataFrame) and column in coverage_matrix.columns:
            totals[column] = int(pd.to_numeric(coverage_matrix[column], errors="coerce").fillna(0).sum())
        else:
            totals[column] = 0
    return totals


def coverage_matrix_with_total_row(coverage_matrix: pd.DataFrame) -> pd.DataFrame:
    """Append one presentation-only TOTAL row to the research coverage matrix.

    The underlying analytical matrix is intentionally left unchanged so downstream
    calculations that sum its columns do not double-count totals.
    """
    if not isinstance(coverage_matrix, pd.DataFrame):
        return pd.DataFrame()
    if coverage_matrix.empty:
        return coverage_matrix.copy()
    total_row = pd.DataFrame([coverage_matrix_totals(coverage_matrix)])
    for column in coverage_matrix.columns:
        if column not in total_row.columns:
            total_row[column] = ""
    return pd.concat([coverage_matrix.copy(), total_row[coverage_matrix.columns]], ignore_index=True)


def verification_followup_summary(qa_frame: pd.DataFrame) -> pd.DataFrame:
    """Aggregate the most useful project deliverables verification gaps."""
    population = closeout_research_population(qa_frame)
    total = len(population)
    rows = []

    def add(issue, mask, why):
        count = int(mask.fillna(False).sum()) if isinstance(mask, pd.Series) else 0
        rows.append({
            "Verification / Follow-Up Item": issue,
            "Affected Records": count,
            "Share of Eligible Records": f"{(count / total * 100):.1f}%" if total else "0.0%",
        })

    if total:
        availability = population.get("Rental Availability Status", pd.Series("", index=population.index)).astype(str)
        add(
            "Rental availability unclear or not checked",
            availability.isin({"", "Not Checked", "Status Unclear — Needs Review"}),
            "Current website inventory and current rental availability are separate facts. Unclear statuses should be verified or presented with a freshness warning rather than assumed available.",
        )
        add(
            "Storeys / height classification unresolved",
            unresolved_mask(population.get("Number of Storeys", pd.Series(pd.NA, index=population.index))),
            "Datablix should not infer Low-, Mid-, or High-rise from appearance, unit count, or marketing language when a reliable storey count is unavailable.",
        )
        add(
            "Apartment count unresolved",
            unresolved_mask(population.get("Number of Apartments", pd.Series(pd.NA, index=population.index))),
            "A blank apartment count is preferable to using advertised vacancies, floor-plan counts, or estimates as the property's total residential inventory.",
        )
        geo = population.get("Geographic Scope Status", pd.Series("", index=population.index)).astype(str)
        add(
            "Geographic / address review",
            geo.isin({"Needs Geographic Review", "Not Checked", ""}),
            "The project is limited to properties physically inside the City of Ottawa, so uncertain municipal evidence requires follow-up rather than assumption.",
        )
        discovery = population.get("Directory Discovery Status", pd.Series("", index=population.index)).astype(str)
        add(
            "Possible duplicates",
            discovery.eq("Possible Duplicate"),
            "Possible duplicates should be reconciled before reporting discovery counts or preparing final directory entries.",
        )
    return pd.DataFrame(rows)


def project_closeout_report_html(qa_frame: pd.DataFrame, registry=None, scope_label="All companies", baseline=None) -> bytes:
    """Create one portable HTML evidence report for final project drafting."""
    registry = normalize_company_registry(registry)
    project_name = safe_text(st.session_state.get(S_PROJECT_NAME, "Datablix project")) or "Datablix project"
    coverage_matrix = closeout_research_coverage_matrix(qa_frame, registry, baseline=baseline)
    field_matrix = field_availability_matrix(qa_frame, registry)
    findings = field_availability_summary(qa_frame)
    discoveries = discovery_submission_summary(qa_frame)
    followup = verification_followup_summary(qa_frame)
    methodology_table = methodology_and_limitations_report(qa_frame, scope_label)
    summary_table = report_summary(qa_frame, registry, scope_label=scope_label, baseline=baseline)
    final_matrix = final_project_data_matrix(qa_frame, registry, scope_label=scope_label, baseline=baseline)
    recommendations = directory_recommendations_with_coverage(qa_frame)

    def table(frame):
        if not isinstance(frame, pd.DataFrame) or frame.empty:
            return "<p><em>No rows in this section.</em></p>"
        return frame.to_html(index=False, border=0, classes="db-table", escape=True)

    generated = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M")
    html = f"""<!doctype html>
<html><head><meta charset=\"utf-8\"><title>{escape(project_name)} — Datablix Deliverables</title>
<style>
body{{font-family:Arial,Helvetica,sans-serif;max-width:1200px;margin:36px auto;padding:0 24px;color:#111827;line-height:1.45}}
h1{{margin-bottom:.2rem}} h2{{margin-top:2rem;border-bottom:1px solid #d1d5db;padding-bottom:.35rem}}
.note{{background:#f3f4f6;padding:12px 14px;border-left:4px solid #6b7280;margin:18px 0}}
.db-table{{border-collapse:collapse;width:100%;font-size:12px;margin:12px 0 24px}}
.db-table th,.db-table td{{border:1px solid #d1d5db;padding:7px;vertical-align:top;text-align:left}}
.db-table th{{background:#111827;color:white}}
.small{{color:#4b5563;font-size:12px}}
</style></head><body>
<h1>{escape(project_name)}</h1>
<p><strong>Datablix Final Project Data Matrix</strong><br><span class=\"small\">Scope: {escape(scope_label)} · Generated {escape(generated)}</span></p>
<div class=\"note\"><strong>Reporting boundary:</strong> This report summarizes the research dataset, source comparison, review status, and recorded external-submission tracking. It does not claim totals for separately maintained datasets that are not directly accessible in Datablix.</div>
<h2>1. Final Project Data Matrix</h2>{table(final_matrix)}
<h2>2. Project Overview</h2>{table(summary_table)}
<h2>3. Research Coverage Matrix</h2>{table(coverage_matrix)}
<h2>4. New Discoveries and External Submission Tracking</h2>{table(discoveries)}
<h2>5. Field Availability Matrix</h2>{table(field_matrix)}
<h2>6. Findings and Recommendations</h2>{table(recommendations)}
<h2>7. Verification and Follow-Up</h2>{table(followup)}
<h2>8. Research Methodology and Limitations</h2>{table(methodology_table)}
<p class=\"small\">Generated by Datablix from the current reviewed project state. Review generated interpretations before using them in final project outputs.</p>
</body></html>"""
    return html.encode("utf-8")


def freeze_reporting_snapshot() -> dict:
    """Freeze the current project state used by all deliverable exports."""
    working = st.session_state.get(S_WORKING)
    if not isinstance(working, pd.DataFrame):
        working = pd.DataFrame(columns=INTERNAL_COLUMNS)
    registry = normalize_company_registry(st.session_state.get(S_COMPANIES))
    baseline = current_starting_source_records()
    snapshot = {
        "created_at": datetime.now().astimezone().isoformat(timespec="seconds"),
        "working": working.copy(deep=True),
        "registry": registry.copy(deep=True),
        "baseline": baseline.copy(deep=True) if isinstance(baseline, pd.DataFrame) else pd.DataFrame(),
        "baseline_fingerprint": dataframe_content_fingerprint(baseline),
        "source_hash": safe_text(
            (st.session_state.get(S_SOURCE_BASELINE_META, {}) or {}).get("source_hash", "")
        ),
    }
    st.session_state[S_REPORTING_SNAPSHOT] = snapshot
    autosave_current_project()
    return snapshot


def clear_reporting_snapshot() -> None:
    """Return Deliverables to the live project state."""
    st.session_state.pop(S_REPORTING_SNAPSHOT, None)
    autosave_current_project()


def reporting_snapshot_state() -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, str, bool]:
    """Return the reporting state, never pairing a snapshot with stale Starting Data.

    A frozen snapshot remains valid while research changes, which is its purpose.
    If the active Starting Data baseline changes, however, the old snapshot is
    automatically invalidated so source counts always come from the current source.
    """
    snapshot = st.session_state.get(S_REPORTING_SNAPSHOT)
    live_baseline = current_starting_source_records()

    if isinstance(snapshot, dict) and isinstance(snapshot.get("working"), pd.DataFrame):
        frozen_baseline = snapshot.get("baseline")
        if not isinstance(frozen_baseline, pd.DataFrame):
            frozen_baseline = pd.DataFrame()

        frozen_fingerprint = (
            safe_text(snapshot.get("baseline_fingerprint", ""))
            or dataframe_content_fingerprint(frozen_baseline)
        )
        live_fingerprint = dataframe_content_fingerprint(live_baseline)

        if frozen_fingerprint != live_fingerprint:
            # Legacy or current snapshot refers to a different Starting Data source.
            # Remove it rather than silently reporting old source-record counts.
            st.session_state.pop(S_REPORTING_SNAPSHOT, None)
            autosave_current_project()
        else:
            frozen_working = normalize_workflow(prepare_data(snapshot["working"].copy()))
            frozen_qa = qa_checks(frozen_working) if not frozen_working.empty else frozen_working.copy()
            frozen_registry = normalize_company_registry(snapshot.get("registry"))
            created = safe_text(snapshot.get("created_at", ""))
            label = created.replace("T", " ")[:16] if created else "Frozen snapshot"
            return frozen_qa, frozen_registry, frozen_baseline.copy(), label, True

    live_working = st.session_state.get(S_WORKING)
    if not isinstance(live_working, pd.DataFrame):
        live_working = pd.DataFrame(columns=INTERNAL_COLUMNS)
    live_qa = qa_checks(live_working.copy()) if not live_working.empty else live_working.copy()
    live_registry = normalize_company_registry(st.session_state.get(S_COMPANIES))
    return live_qa, live_registry, live_baseline, "Live project state", False


def _deliverable_html(title: str, sections: list[tuple[str, object]], note: str = "") -> bytes:
    """Create a simple portable HTML deliverable from DataFrames or text blocks."""
    project_name = safe_text(st.session_state.get(S_PROJECT_NAME, "Datablix project")) or "Datablix project"

    def render(value):
        if isinstance(value, pd.DataFrame):
            if value.empty:
                return "<p><em>No rows in this section.</em></p>"
            return value.to_html(index=False, border=0, classes="db-table", escape=True)
        return f"<div class='text-block'>{escape(str(value or '')).replace(chr(10), '<br>')}</div>"

    section_html = "\n".join(
        f"<h2>{escape(str(heading))}</h2>{render(value)}"
        for heading, value in sections
    )
    note_html = f"<div class='note'>{escape(note)}</div>" if note else ""
    generated = datetime.now().astimezone().strftime("%Y-%m-%d %H:%M")
    html = f"""<!doctype html>
<html><head><meta charset="utf-8"><title>{escape(project_name)} — {escape(title)}</title>
<style>
body{{font-family:Arial,Helvetica,sans-serif;max-width:1200px;margin:36px auto;padding:0 24px;color:#111827;line-height:1.45}}
h1{{margin-bottom:.15rem}} h2{{margin-top:1.8rem;border-bottom:1px solid #d1d5db;padding-bottom:.3rem}}
.small{{color:#4b5563;font-size:12px}} .note{{background:#f3f4f6;padding:12px 14px;border-left:4px solid #6b7280;margin:18px 0}}
.db-table{{border-collapse:collapse;width:100%;font-size:12px;margin:12px 0 24px}}
.db-table th,.db-table td{{border:1px solid #d1d5db;padding:7px;vertical-align:top;text-align:left}}
.db-table th{{background:#111827;color:white}} .text-block{{white-space:normal}}
</style></head><body>
<h1>{escape(project_name)}</h1>
<p><strong>{escape(title)}</strong><br><span class="small">Generated {escape(generated)}</span></p>
{note_html}
{section_html}
</body></html>"""
    return html.encode("utf-8")


def draft_profiles_table(qa_frame: pd.DataFrame) -> pd.DataFrame:
    """Return one draft profile for every researched property row."""
    columns = [
        "Record ID", "Building Name", "Management/Owner", "Street Address",
        "Directory Discovery Status", "Current Inventory Status",
        "Rental Availability Status", "Verification Status", "Record Decision",
        "Profile Draft",
    ]
    if not isinstance(qa_frame, pd.DataFrame) or qa_frame.empty:
        return pd.DataFrame(columns=columns)

    candidates = qa_frame.copy()
    sort_fields = [
        field for field in ["Management/Owner", "Building Name", "Street Address"]
        if field in candidates.columns
    ]
    if sort_fields:
        candidates = candidates.sort_values(sort_fields, kind="stable")

    rows = []
    for _, row in candidates.iterrows():
        rows.append({
            "Record ID": row.get("Record ID", ""),
            "Building Name": row.get("Building Name", ""),
            "Management/Owner": row.get("Management/Owner", ""),
            "Street Address": row.get("Street Address", ""),
            "Directory Discovery Status": row.get("Directory Discovery Status", ""),
            "Current Inventory Status": row.get("Current Inventory Status", ""),
            "Rental Availability Status": row.get("Rental Availability Status", ""),
            "Verification Status": row.get("Verification Status", ""),
            "Record Decision": row.get("Record Decision", ""),
            "Profile Draft": community_profile_text(row),
        })
    return pd.DataFrame(rows, columns=columns)


def final_profile_candidates_table(qa_frame: pd.DataFrame) -> pd.DataFrame:
    """Return the separate filtered profile-candidate view."""
    population = deliverable_research_population(qa_frame)
    return draft_profiles_table(population)


def final_project_data_matrix(
    qa_frame: pd.DataFrame,
    registry=None,
    scope_label: str = "All companies",
    baseline=None,
) -> pd.DataFrame:
    """Create a matrix-first factual summary for final project reporting."""
    registry = normalize_company_registry(registry)
    coverage_matrix = closeout_research_coverage_matrix(
        qa_frame, registry, baseline=baseline
    )
    field_summary = field_availability_summary(qa_frame)
    company_count = len(registry) if not registry.empty else (
        int(qa_frame["Company ID"].astype(str).replace("", pd.NA).dropna().nunique())
        if isinstance(qa_frame, pd.DataFrame) and not qa_frame.empty and "Company ID" in qa_frame.columns
        else 0
    )
    total = len(qa_frame) if isinstance(qa_frame, pd.DataFrame) else 0
    eligible_population = deliverable_research_population(qa_frame)
    eligible_total = len(eligible_population)
    existing = int(qa_frame["Directory Discovery Status"].eq("Existing Source Record").sum()) if total and "Directory Discovery Status" in qa_frame.columns else 0
    discoveries = int(qa_frame["Directory Discovery Status"].eq("Newly Discovered").sum()) if total and "Directory Discovery Status" in qa_frame.columns else 0
    changed = int(coverage_matrix["Changed Existing Records"].sum()) if not coverage_matrix.empty else 0
    approved = int(approved_for_export_mask(qa_frame).sum()) if total else 0
    verified = int(qa_frame["Verification Status"].eq("Verified").sum()) if total and "Verification Status" in qa_frame.columns else 0
    needs_review = int(deliverable_needs_review_mask(qa_frame).sum()) if total else 0
    submitted = int((
        qa_frame.get("Directory Discovery Status", pd.Series("", index=qa_frame.index)).astype(str).eq("Newly Discovered")
        & qa_frame.get("Directory Entry Status", pd.Series("", index=qa_frame.index)).astype(str).eq("Entered")
    ).sum()) if total else 0

    discovery_status = qa_frame.get(
        "Directory Discovery Status", pd.Series("", index=qa_frame.index)
    ).astype(str) if total else pd.Series(dtype=str)
    inventory_status = qa_frame.get(
        "Current Inventory Status", pd.Series("", index=qa_frame.index)
    ).astype(str) if total else pd.Series(dtype=str)
    rental_status = qa_frame.get(
        "Rental Availability Status", pd.Series("", index=qa_frame.index)
    ).astype(str) if total else pd.Series(dtype=str)
    research_status = qa_frame.get(
        "Research Status", pd.Series("", index=qa_frame.index)
    ).astype(str) if total else pd.Series(dtype=str)

    needs_classification = int(discovery_status.eq("Needs Classification").sum()) if total else 0
    possible_duplicates = int(discovery_status.eq("Possible Duplicate").sum()) if total else 0
    excluded_not_current = int(discovery_status.eq("Excluded / Not Current").sum()) if total else 0
    rented = int(rental_status.eq("Rented").sum()) if total else 0
    current_inventory = int(inventory_status.eq("Current").sum()) if total else 0
    inventory_review = int(inventory_status.eq("Review").sum()) if total else 0
    completed = int(research_status.eq("Completed").sum()) if total else 0

    rows = [
        {"Matrix Area": "Project Scope", "Data Point": "Companies researched", "Result": company_count, "Denominator / Scope": "", "Rate %": "", "Assessment": "Coverage", "What the Data Shows": f"{company_count:,} organization workspace(s) represented in {scope_label}.", "Deliverable Use": "Research summary"},
        {"Matrix Area": "Project Scope", "Data Point": "Research properties", "Result": total, "Denominator / Scope": "", "Rate %": "", "Assessment": "Coverage", "What the Data Shows": f"{total:,} researched property record(s) are represented.", "Deliverable Use": "Research scope"},
        {"Matrix Area": "Full Research Status", "Data Point": "Completed research records", "Result": completed, "Denominator / Scope": total, "Rate %": round(completed / total * 100, 1) if total else 0.0, "Assessment": "Research status", "What the Data Shows": f"{completed:,} record(s) are marked Completed.", "Deliverable Use": "Full research"},
        {"Matrix Area": "Full Research Status", "Data Point": "Current inventory", "Result": current_inventory, "Denominator / Scope": total, "Rate %": round(current_inventory / total * 100, 1) if total else 0.0, "Assessment": "Inventory status", "What the Data Shows": f"{current_inventory:,} record(s) are marked Current inventory.", "Deliverable Use": "Full research"},
        {"Matrix Area": "Full Research Status", "Data Point": "Inventory review", "Result": inventory_review, "Denominator / Scope": total, "Rate %": round(inventory_review / total * 100, 1) if total else 0.0, "Assessment": "Inventory status", "What the Data Shows": f"{inventory_review:,} record(s) remain in inventory review.", "Deliverable Use": "Full research"},
        {"Matrix Area": "Full Research Status", "Data Point": "Rented properties", "Result": rented, "Denominator / Scope": total, "Rate %": round(rented / total * 100, 1) if total else 0.0, "Assessment": "Rental availability", "What the Data Shows": f"{rented:,} researched record(s) are marked Rented; this does not erase them from the research history.", "Deliverable Use": "Full research"},
        {"Matrix Area": "Full Research Status", "Data Point": "Needs discovery classification", "Result": needs_classification, "Denominator / Scope": total, "Rate %": round(needs_classification / total * 100, 1) if total else 0.0, "Assessment": "Discovery status", "What the Data Shows": f"{needs_classification:,} record(s) still need discovery classification.", "Deliverable Use": "Full research"},
        {"Matrix Area": "Full Research Status", "Data Point": "Possible duplicates", "Result": possible_duplicates, "Denominator / Scope": total, "Rate %": round(possible_duplicates / total * 100, 1) if total else 0.0, "Assessment": "Discovery status", "What the Data Shows": f"{possible_duplicates:,} record(s) are retained as possible duplicates for audit visibility.", "Deliverable Use": "Full research"},
        {"Matrix Area": "Full Research Status", "Data Point": "Excluded / not current", "Result": excluded_not_current, "Denominator / Scope": total, "Rate %": round(excluded_not_current / total * 100, 1) if total else 0.0, "Assessment": "Discovery status", "What the Data Shows": f"{excluded_not_current:,} record(s) are excluded/not current but remain visible in the full research record.", "Deliverable Use": "Full research"},
        {"Matrix Area": "Project Scope", "Data Point": "Valid/current directory records", "Result": eligible_total, "Denominator / Scope": total, "Rate %": round(eligible_total / total * 100, 1) if total else 0.0, "Assessment": "Included", "What the Data Shows": f"{eligible_total:,} researched record(s) are valid/current in the filtered directory view.", "Deliverable Use": "Final directory view"},
        {"Matrix Area": "Source Reconciliation", "Data Point": "Existing source matches", "Result": existing, "Denominator / Scope": total, "Rate %": round(existing / total * 100, 1) if total else 0.0, "Assessment": "Matched", "What the Data Shows": f"{existing:,} record(s) matched the Starting Data.", "Deliverable Use": "Research summary"},
        {"Matrix Area": "Source Reconciliation", "Data Point": "New discoveries", "Result": discoveries, "Denominator / Scope": total, "Rate %": round(discoveries / total * 100, 1) if total else 0.0, "Assessment": "New", "What the Data Shows": f"{discoveries:,} record(s) are classified as newly discovered relative to Starting Data.", "Deliverable Use": "Research summary"},
        {"Matrix Area": "Source Reconciliation", "Data Point": "Changed existing records", "Result": changed, "Denominator / Scope": existing, "Rate %": round(changed / existing * 100, 1) if existing else 0.0, "Assessment": "Changed", "What the Data Shows": f"{changed:,} matched record(s) contain material researched differences.", "Deliverable Use": "Research summary"},
        {"Matrix Area": "External Submission", "Data Point": "Discoveries submitted", "Result": submitted, "Denominator / Scope": discoveries, "Rate %": round(submitted / discoveries * 100, 1) if discoveries else 0.0, "Assessment": "Submitted", "What the Data Shows": f"{submitted:,} discovery/discoveries are recorded as externally submitted.", "Deliverable Use": "Submission tracking"},
        {"Matrix Area": "Review & Verification", "Data Point": "Human verified", "Result": verified, "Denominator / Scope": total, "Rate %": round(verified / total * 100, 1) if total else 0.0, "Assessment": "Verified", "What the Data Shows": f"{verified:,} record(s) are marked Verified.", "Deliverable Use": "Verification tracker"},
        {"Matrix Area": "Review & Verification", "Data Point": "Approved for Export", "Result": approved, "Denominator / Scope": total, "Rate %": round(approved / total * 100, 1) if total else 0.0, "Assessment": "Workflow", "What the Data Shows": f"{approved:,} record(s) meet the stricter optional export-approval rule.", "Deliverable Use": "Export workflow"},
        {"Matrix Area": "Review & Verification", "Data Point": "Needs review", "Result": needs_review, "Denominator / Scope": total, "Rate %": round(needs_review / total * 100, 1) if total else 0.0, "Assessment": "Open", "What the Data Shows": f"{needs_review:,} active record(s) have genuine unresolved review or verification items.", "Deliverable Use": "Verification tracker"},
    ]

    for _, row in field_summary.iterrows():
        rows.append({
            "Matrix Area": "Field Availability",
            "Data Point": row["Field"],
            "Result": int(row["Found"]),
            "Denominator / Scope": int(row["Found"] + row["Missing"]),
            "Rate %": float(row["Coverage %"]),
            "Assessment": row["Availability Assessment"],
            "What the Data Shows": row["Research Finding"],
            "Deliverable Use": "Structure & search recommendations",
        })
    return pd.DataFrame(rows)


def deliverable_research_dataset(qa_frame: pd.DataFrame) -> pd.DataFrame:
    """Return every researched property row for Deliverable 1.

    Deliverable 1 is the complete research record, not a final-directory filter.
    It therefore preserves existing Starting Data matches, new discoveries,
    rented records, exclusions, possible duplicates, and review items together
    with the statuses that explain each record's outcome.
    """
    if not isinstance(qa_frame, pd.DataFrame) or qa_frame.empty:
        return listing_export(qa_frame.iloc[0:0].copy()) if isinstance(qa_frame, pd.DataFrame) else pd.DataFrame()

    population = qa_frame.copy()
    sort_fields = [
        field for field in ["Management/Owner", "Building Name", "Street Address"]
        if field in population.columns
    ]
    if sort_fields:
        population = population.sort_values(sort_fields, kind="stable")

    dataset = listing_export(population)
    # Preserve workflow/audit fields that explain why a researched record may or
    # may not appear in the final directory view.
    for field in [
        "Research Status", "Source Status", "Record Decision",
        "Discovery Status Source", "QA Status", "QA Flags",
    ]:
        if field in population.columns and field not in dataset.columns:
            dataset[field] = population[field]
    return dataset


def deliverable_final_directory_dataset(qa_frame: pd.DataFrame) -> pd.DataFrame:
    """Return the existing valid/current directory-oriented view for Deliverable 1."""
    if not isinstance(qa_frame, pd.DataFrame) or qa_frame.empty:
        return listing_export(qa_frame.iloc[0:0].copy()) if isinstance(qa_frame, pd.DataFrame) else pd.DataFrame()
    population = deliverable_research_population(qa_frame)
    return listing_export(population)


def research_dataset_workbook(qa_frame: pd.DataFrame) -> bytes:
    """Create a plain editable workbook with full research and a filtered directory view."""
    full_research = deliverable_research_dataset(qa_frame)
    final_directory = deliverable_final_directory_dataset(qa_frame)
    return excel_bytes({
        "Full Research": full_research,
        "Final Directory View": final_directory,
    })


def organization_research_summary_workbook(
    qa_frame: pd.DataFrame,
    registry=None,
    baseline=None,
) -> bytes:
    """Export company coverage plus discovery tracking as one workbook."""
    return excel_bytes({
        "All Research by Company": full_company_research_summary(qa_frame, registry),
        "Research Coverage": closeout_research_coverage_matrix(qa_frame, registry, baseline=baseline),
        "Discoveries": discovery_submission_summary(qa_frame),
        "Research Status Summary": full_research_status_summary(qa_frame),
    })


def recommendations_workbook(qa_frame: pd.DataFrame, registry=None) -> bytes:
    """Export field availability evidence and recommendations."""
    return excel_bytes({
        "Full Research Availability": field_availability_matrix(qa_frame, registry, include_all=True),
        "Full Research Findings": field_availability_summary(qa_frame, include_all=True),
        "Directory Candidate Findings": field_availability_summary(qa_frame),
        "Recommendations": directory_recommendations_with_coverage(qa_frame),
    })


def verification_tracker_workbook(qa_frame: pd.DataFrame) -> bytes:
    """Export source verification records and project-level follow-up counts."""
    return excel_bytes({
        "Full Verification Tracker": source_verification_tracker(qa_frame),
        "Research Status Summary": full_research_status_summary(qa_frame),
        "Active Follow-Up Summary": verification_followup_summary(qa_frame),
    })


def profiles_html(qa_frame: pd.DataFrame) -> bytes:
    """Export all draft record profiles into one portable HTML file."""
    profiles = draft_profiles_table(qa_frame)
    final_candidates = final_profile_candidates_table(qa_frame)
    return _deliverable_html(
        "Draft Record Profiles",
        [
            ("Full Research Profiles", profiles),
            ("Filtered Final Profile Candidates", final_candidates),
        ],
        "Full Research Profiles preserves every researched property. The filtered candidate section is secondary and does not remove records from the research history.",
    )


def methodology_html(qa_frame: pd.DataFrame, scope_label: str) -> bytes:
    """Export methodology and limitations as a portable HTML file."""
    method = methodology_and_limitations_report(qa_frame, scope_label)
    return _deliverable_html(
        "Methodology & Limitations",
        [
            ("Full Research Status Summary", full_research_status_summary(qa_frame)),
            ("Methodology & Limitations", method),
        ],
        "The methodology report is grounded in the complete research state. Any final-directory filtering is treated as a separate downstream use.",
    )


def deliverables_package_bytes(
    qa_frame: pd.DataFrame,
    registry=None,
    scope_label: str = "All companies",
    baseline=None,
) -> bytes:
    """Create one ZIP containing independently reusable project deliverables."""
    registry = normalize_company_registry(registry)
    project_name = safe_filename(
        safe_text(st.session_state.get(S_PROJECT_NAME, "Datablix project")) or "Datablix project"
    )
    scope_name = safe_filename(scope_label)
    dataset = deliverable_research_dataset(qa_frame)
    matrix = final_project_data_matrix(qa_frame, registry, scope_label, baseline=baseline)
    presentation = presentation_summary_text(
        qa_frame, registry, scope_label, baseline=baseline
    )
    package = io.BytesIO()
    with zipfile.ZipFile(package, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(
            "01_Research_Dataset.xlsx",
            research_dataset_workbook(qa_frame),
        )
        archive.writestr("01_Research_Dataset.csv", csv_bytes(dataset))
        archive.writestr(
            "02_Organization_Research_Summary.xlsx",
            organization_research_summary_workbook(qa_frame, registry, baseline=baseline),
        )
        archive.writestr(
            "03_Source_Verification_Tracker.xlsx",
            verification_tracker_workbook(qa_frame),
        )
        archive.writestr("04_Draft_Record_Profiles.html", profiles_html(qa_frame))
        archive.writestr(
            "05_Structure_Search_Recommendations.xlsx",
            recommendations_workbook(qa_frame, registry),
        )
        archive.writestr(
            "06_Methodology_Limitations.html",
            methodology_html(qa_frame, scope_label),
        )
        archive.writestr(
            "07_Final_Project_Data_Matrix.html",
            project_closeout_report_html(
                qa_frame, registry, scope_label=scope_label, baseline=baseline
            ),
        )
        archive.writestr("07_Final_Project_Data_Matrix.csv", csv_bytes(matrix))
        archive.writestr("Presentation_Summary.txt", presentation.encode("utf-8"))
        archive.writestr(
            "README.txt",
            (
                f"Datablix Deliverables Package\n"
                f"Project: {project_name}\n"
                f"Scope: {scope_label}\n"
                f"Generated: {datetime.now().astimezone().isoformat(timespec='seconds')}\n\n"
                f"Research records in scope: {len(qa_frame):,}\n"
                f"Full Research rows in Deliverable 1: {len(qa_frame):,}\n"
                f"Filtered final-directory rows: {len(deliverable_research_population(qa_frame)):,}\n\n"
                "Each file is generated from the same Datablix reporting state so the figures remain aligned.\n"
                "Every deliverable is grounded in the full research state, including Starting Data matches, new discoveries, rented, excluded/not-current, duplicate, and review records where applicable.\n"
                "Filtered final-use views remain secondary and are labeled separately; they do not replace or erase the full research evidence.\n"
                "Approved for Export remains a separate optional workflow status.\n"
            ).encode("utf-8"),
        )
    return package.getvalue()


def issue_summary(df):
    counts = {}
    for text in df["QA Flags"].fillna(""):
        for item in str(text).split("; "):
            if item and item != "No rental property data issues found": counts[item] = counts.get(item, 0) + 1
    rows = []
    for item, count in counts.items():
        severity, _, issue = item.partition(": ")
        rows.append({"Severity": severity, "Issue": issue or item, "Affected Records": count})
    return pd.DataFrame(rows).sort_values(["Severity", "Affected Records"], ascending=[True, False]) if rows else pd.DataFrame(columns=["Severity", "Issue", "Affected Records"])


def project_summary(df):
    return pd.DataFrame([
        {"Metric": "Records", "Value": len(df), "Interpretation": "Rows in the working rental property dataset."},
        {"Metric": "Management/owner organizations", "Value": resolved(df["Management/Owner"]).dropna().astype(str).str.strip().nunique(), "Interpretation": "Distinct recorded organizations."},
        {"Metric": "Records with usable core identity", "Value": int(df["Core Gap Count"].eq(0).sum()), "Interpretation": "Records with management/owner, street address, and city."},
        {"Metric": "Verified records", "Value": int(df["Verification Status"].eq("Verified").sum()), "Interpretation": "Records marked as human-verified."},
        {"Metric": "Approved for Export", "Value": int(approved_for_export_mask(df).sum()), "Interpretation": "Records explicitly completed, human-verified, kept, and free of critical data blockers."},
        {"Metric": "Newly discovered records", "Value": int(df["Directory Discovery Status"].eq("Newly Discovered").sum()) if "Directory Discovery Status" in df.columns else 0, "Interpretation": "Building records not matched to the starting source dataset."},
        {"Metric": "Open research gaps", "Value": int(df["Research Gap Count"].sum()), "Interpretation": "Unconfirmed listing fields."},
    ])


def structure_recommendations():
    rows = [
        ("Identity", "Apartment Building Name", "Where available", "Text", "Recognizable building or property name", "Search"),
        ("Location", "Street Address", "Required", "Text", "Primary building address", "Search"),
        ("Location", "City and Postal Code", "Required", "Formatted location", "City, province code, and postal code", "Search/Filter"),
        ("Property", "Property Type", "Where available", "Controlled text", "Residential property form", "Category/Filter"),
        ("Property", "Building Classification", "Where available", "Controlled text", "Height band derived from storeys", "Filter"),
        ("Property", "Storeys", "Where available", "Whole number", "Number of building storeys", "Sort/Filter"),
        ("Property", "Number of Apartments", "Where available", "Whole number", "Apartment count", "Sort/Filter"),
        ("Ownership", "Apartment Building Management/Owner", "Required", "Controlled text", "Responsible organization", "Filter"),
        ("Contact", "Phone Number", "Where available", "Phone", "Available contact number", "Search"),
        ("Contact", "Email Contact", "Where available", "Email", "Available email contact", "Search"),
        ("Contact", "WebSite", "Recommended", "URL", "Property or company webpage", "Link"),
        ("Research", "Source URL", "Required for verification", "URL", "Exact supporting page", "Link"),
        ("Research", "Date Researched", "Required for verified records", "Date", "Freshness trail", "Filter"),
        ("Research", "Researcher", "Required for verified records", "Controlled text", "Accountability", "Filter"),
        ("Research", "Verification Status", "Required", "Controlled status", "Human review outcome", "Filter"),
        ("Research", "Directory Discovery Status", "Required for project audit", "Controlled status", "Distinguishes existing source records, newly discovered buildings, duplicates, and excluded records", "Filter"),
        ("Research", "Missing Information", "Automatically generated", "System text", "Lists current research fields that remain unconfirmed", "No"),
        ("Workflow", "Record Decision", "Required before final use", "Controlled status", "Keep, update, duplicate, or remove", "Filter"),
        ("Workflow", "Directory Entry Status", "Required during final entry", "Controlled status", "Tracks Not Entered, Entered, or Needs Correction", "Filter"),
    ]
    return pd.DataFrame(rows, columns=["Field Group", "Field", "Requirement", "Recommended Type", "Purpose", "Platform Use"])


def methodology(df, name, sheet):
    return pd.DataFrame([
        {"Section": "Purpose", "Report Text": "Organize listing information into a consistent, searchable structure using the records provided and publicly available sources."},
        {"Section": "Input reviewed", "Report Text": f"Workspace: {name}. Worksheet: {sheet or 'not specified'}. Records reviewed: {len(df):,}."},
        {"Section": "Core record view", "Report Text": "The Building Listings sheet keeps the main rental property, location, ownership, and contact fields together in a concise view."},
        {"Section": "Method", "Report Text": "Match imported headings, preserve original columns, check identity and formats, track sources and verification, and keep review decisions explicit."},
        {"Section": "Limitations", "Report Text": "Public information may be incomplete, outdated, duplicated, or inconsistent. Automated checks support review but do not replace human judgment."},
        {"Section": "Suggested next checks", "Report Text": "Work through high-priority records, confirm sources, document unavailable information, and read through generated text before use."},
    ])













def company_progress_summary(qa_frame, registry=None):
    registry = normalize_company_registry(registry)
    rows = []
    represented_ids = set()

    def discovery_count(group, status):
        if group.empty or "Directory Discovery Status" not in group.columns:
            return 0
        return int(group["Directory Discovery Status"].astype(str).eq(status).sum())

    for _, company in registry.iterrows():
        company_id = str(company["Company ID"]).strip()
        represented_ids.add(company_id)
        group = qa_frame.loc[qa_frame["Company ID"].astype(str).eq(company_id)]
        rows.append({
            "Company ID": company_id,
            "Management/Owner": company["Management/Owner"],
            "Main Website": company["Main Website"],
            "Scope Type": company["Scope Type"],
            "Company Status": company["Company Status"],
            "Building Records": len(group),
            "Existing Source Records": discovery_count(group, "Existing Source Record"),
            "Newly Discovered": discovery_count(group, "Newly Discovered"),
            "Needs Classification": discovery_count(group, "Needs Classification"),
            "Possible Duplicates": discovery_count(group, "Possible Duplicate"),
            "Excluded / Not Current": discovery_count(group, "Excluded / Not Current"),
            "Completed Records": int(group["Research Status"].eq("Completed").sum()) if not group.empty else 0,
            "Verified Records": int(group["Verification Status"].eq("Verified").sum()) if not group.empty else 0,
            "Records Passing QA": int(group["QA Status"].eq("Pass").sum()) if not group.empty else 0,
            "Approved for Export": int(approved_for_export_mask(group).sum()) if not group.empty else 0,
            "Entered in Directory": int(group["Directory Entry Status"].eq("Entered").sum()) if not group.empty else 0,
            "Open QA Issues": int(group["QA Flag Count"].sum()) if not group.empty else 0,
            "Open Field Gaps": int(group["Research Gap Count"].sum()) if not group.empty else 0,
            "Still in Review": int((~approved_for_export_mask(group) & ~group["Record Decision"].eq("Remove")).sum()) if not group.empty else 0,
        })

    unregistered = qa_frame.loc[
        ~qa_frame["Company ID"].astype(str).isin(represented_ids)
    ]
    for owner, group in unregistered.assign(
        _owner=display_values(unregistered["Management/Owner"], "Unassigned")
    ).groupby("_owner"):
        rows.append({
            "Company ID": "",
            "Management/Owner": owner,
            "Main Website": "",
            "Scope Type": "Imported",
            "Company Status": "Researching",
            "Building Records": len(group),
            "Existing Source Records": discovery_count(group, "Existing Source Record"),
            "Newly Discovered": discovery_count(group, "Newly Discovered"),
            "Needs Classification": discovery_count(group, "Needs Classification"),
            "Possible Duplicates": discovery_count(group, "Possible Duplicate"),
            "Excluded / Not Current": discovery_count(group, "Excluded / Not Current"),
            "Completed Records": int(group["Research Status"].eq("Completed").sum()),
            "Verified Records": int(group["Verification Status"].eq("Verified").sum()),
            "Records Passing QA": int(group["QA Status"].eq("Pass").sum()),
            "Approved for Export": int(approved_for_export_mask(group).sum()),
            "Entered in Directory": int(group["Directory Entry Status"].eq("Entered").sum()),
            "Open QA Issues": int(group["QA Flag Count"].sum()),
            "Open Field Gaps": int(group["Research Gap Count"].sum()),
            "Still in Review": int((~approved_for_export_mask(group) & ~group["Record Decision"].eq("Remove")).sum()),
        })
    return pd.DataFrame(rows)



def source_verification_tracker(qa_frame):
    """Create the project source and verification tracker."""
    columns = [
        "Record ID", "Building Name", "Management/Owner",
        "Directory Discovery Status", "Street Address", "City", "Postal Code",
        "Source URL", "Date Researched", "Researcher", "Source Status",
        "Verification Status", "Missing Information", "Reviewer Notes",
        "Follow-up Priority", "Record Decision", "Directory Entry Status", "Export Status",
    ]
    available = [column for column in columns if column in qa_frame.columns]
    tracker = qa_frame[available].copy()
    if not tracker.empty:
        tracker = tracker.sort_values(
            by=["Export Status", "Follow-up Priority", "Management/Owner", "Building Name"],
            kind="stable",
        )
    return tracker.reset_index(drop=True)


def _profile_value(row, field, blank="Not publicly confirmed"):
    value = row.get(field, "")
    return blank if is_unresolved(value) else str(value).strip()


def community_profile_text(row) -> str:
    """Create a copy-ready draft apartment community profile from one reviewed record."""
    location = formatted_location(row)
    lines = [
        f"# {_profile_value(row, 'Building Name', 'Apartment Building')}",
        "",
        f"**Address:** {_profile_value(row, 'Street Address')} "
        + (f"— {location}" if not is_unresolved(location) else ""),
        f"**Management / Owner:** {_profile_value(row, 'Management/Owner')}",
        f"**Property Type:** {_profile_value(row, 'Property Type')}",
        f"**Building Classification:** {_profile_value(row, 'Building Classification')}",
        f"**Storeys:** {_profile_value(row, 'Number of Storeys')}",
        f"**Number of Apartments:** {_profile_value(row, 'Number of Apartments')}",
        f"**Phone:** {_profile_value(row, 'Phone')}",
        f"**Email:** {_profile_value(row, 'Primary Email')}",
        f"**Website:** {_profile_value(row, 'Website')}",
        "",
        "## Additional property information",
        f"**Suite Types:** {_profile_value(row, 'Suite Types')}",
        f"**Amenities:** {_profile_value(row, 'Amenities')}",
        f"**Parking:** {_profile_value(row, 'Parking')}",
        f"**Laundry:** {_profile_value(row, 'Laundry')}",
        f"**Utilities:** {_profile_value(row, 'Utilities')}",
        f"**Accessibility:** {_profile_value(row, 'Accessibility')}",
        f"**Pet Policy:** {_profile_value(row, 'Pet Policy')}",
        "",
        "## Research note",
        f"**Missing / unclear information:** {_profile_value(row, 'Missing Information', 'None documented')}",
        f"**Verification:** {_profile_value(row, 'Verification Status')}",
        f"**Source:** {_profile_value(row, 'Source URL')}",
    ]
    return "\n".join(lines)


def directory_recommendations_with_coverage(qa_frame):
    """Return evidence-first structure and search recommendations.

    Every recommendation is tied to observed field availability in the current
    research scope and includes a concise rationale.
    """
    findings = field_availability_summary(qa_frame)
    columns = [
        "Field", "Found", "Missing", "Coverage %", "Availability Assessment",
        "Research Finding", "Decision Value", "Suggested Directory Use",
        "Recommendation", "Rationale",
    ]
    return findings[[column for column in columns if column in findings.columns]].copy()


def methodology_and_limitations_report(qa_frame, scope_label):
    """Generate a project-specific research methodology and limitations summary."""
    unavailable_sources = int(qa_frame["Source Status"].eq("Unavailable").sum()) if not qa_frame.empty else 0
    missing_source = int(unresolved_mask(qa_frame["Source URL"]).sum()) if not qa_frame.empty else 0
    excluded = int(qa_frame["Directory Discovery Status"].eq("Excluded / Not Current").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    new_records = int(qa_frame["Directory Discovery Status"].eq("Newly Discovered").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    submitted = int((
        qa_frame.get("Directory Discovery Status", pd.Series("", index=qa_frame.index)).astype(str).eq("Newly Discovered")
        & qa_frame.get("Directory Entry Status", pd.Series("", index=qa_frame.index)).astype(str).eq("Entered")
    ).sum()) if not qa_frame.empty else 0
    return pd.DataFrame([
        {
            "Section": "Research scope",
            "Report Text": f"Scope analysed: {scope_label}. Datablix treats each company as a separate research workspace and aggregates saved research evidence for project-level deliverables.",
        },
        {
            "Section": "Research method",
            "Report Text": "Research follows an inventory-first process: establish the current official portfolio, research confirmed/current properties deeply, use secondary public sources only for permitted gaps, then import structured CSV results for human review.",
        },
        {
            "Section": "Inclusion criteria",
            "Report Text": "A property is included only when there is meaningful property evidence and sufficient current-inventory support within the City of Ottawa scope. A loading URL alone is not treated as proof of current inventory or current rental availability.",
        },
        {
            "Section": "Duplicate and discovery method",
            "Report Text": f"Records are compared against the Starting Data using normalized address, postal-code, city, and building-name evidence. {new_records:,} current record(s) in this scope are classified as newly discovered relative to the Starting Data available to Datablix.",
        },
        {
            "Section": "External submission tracking",
            "Report Text": f"Datablix records whether discoveries were marked Entered through an external submission process. {submitted:,} newly discovered record(s) in this scope are currently recorded as externally submitted. This is workflow tracking and does not independently confirm separately maintained external datasets.",
        },
        {
            "Section": "Verification method",
            "Report Text": "Imported AI/scanner findings remain candidates until a person completes research, verifies the record, records supporting evidence, and resolves any required review decisions.",
        },
        {
            "Section": "Missing information",
            "Report Text": "Unconfirmed values remain blank and are documented as research gaps rather than guessed. Missing information is tracked separately from formatting or QA errors.",
        },
        {
            "Section": "Limitations",
            "Report Text": f"Public websites may be incomplete, stale, blocked, JavaScript-dependent, or inconsistent. This scope currently contains {unavailable_sources:,} record(s) with unavailable sources, {missing_source:,} record(s) without a recorded source URL, and {excluded:,} excluded/not-current record(s).",
        },
        {
            "Section": "External data-access boundary",
            "Report Text": "Datablix reports only datasets available within the workspace. Research records, source matches, discoveries, changes, and recorded submissions are summarized without claiming totals for separately maintained external datasets.",
        },
        {
            "Section": "Assumptions",
            "Report Text": "Official company/property sources are treated as primary evidence; XML sitemaps are discovery evidence rather than proof of current inventory; orphan pages without meaningful property evidence are ignored as non-record pages; explicit rental-status evidence is kept separate from website inventory presence.",
        },
        {
            "Section": "Recommended next steps",
            "Report Text": "Resolve high-priority follow-ups, confirm remaining source evidence, review possible duplicates, document information that cannot be publicly confirmed, preserve Datablix as the research evidence trail, and follow the applicable governance process for any separately maintained external dataset.",
        },
    ])


def presentation_summary_text(qa_frame, registry, scope_label, baseline=None) -> str:
    """Create a copy-ready factual summary for the final project presentation."""
    approved = int(approved_for_export_mask(qa_frame).sum())
    submitted = int((
        qa_frame.get("Directory Discovery Status", pd.Series("", index=qa_frame.index)).astype(str).eq("Newly Discovered")
        & qa_frame.get("Directory Entry Status", pd.Series("", index=qa_frame.index)).astype(str).eq("Entered")
    ).sum()) if not qa_frame.empty else 0
    needs_correction = int(qa_frame["Directory Entry Status"].eq("Needs Correction").sum()) if "Directory Entry Status" in qa_frame.columns else 0
    still_review = int(deliverable_needs_review_mask(qa_frame).sum())
    existing = int(qa_frame["Directory Discovery Status"].eq("Existing Source Record").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    discovered = int(qa_frame["Directory Discovery Status"].eq("Newly Discovered").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    needs_classification = int(qa_frame["Directory Discovery Status"].eq("Needs Classification").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    duplicates = int(qa_frame["Directory Discovery Status"].eq("Possible Duplicate").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    excluded = int(qa_frame["Directory Discovery Status"].eq("Excluded / Not Current").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    company_count = int(qa_frame["Company ID"].astype(str).replace("", pd.NA).dropna().nunique())
    if scope_label == "All companies" and not normalize_company_registry(registry).empty:
        company_count = len(normalize_company_registry(registry))

    critical = int(qa_frame["QA Status"].eq("Critical").sum()) if "QA Status" in qa_frame.columns else 0
    warnings = int(qa_frame["Warning Count"].sum()) if "Warning Count" in qa_frame.columns else 0
    research_gaps = int(qa_frame["Research Gap Count"].sum()) if "Research Gap Count" in qa_frame.columns else 0
    verified = int(qa_frame["Verification Status"].eq("Verified").sum()) if "Verification Status" in qa_frame.columns else 0
    company_matrix = closeout_research_coverage_matrix(qa_frame, registry, baseline=baseline)
    changed = int(company_matrix["Changed Existing Records"].sum()) if not company_matrix.empty else 0

    findings = field_availability_summary(qa_frame)
    strongest = findings.sort_values("Coverage %", ascending=False).head(3) if not findings.empty else pd.DataFrame()
    weakest = findings.sort_values("Coverage %", ascending=True).head(3) if not findings.empty else pd.DataFrame()
    strongest_text = [
        f"{row['Field']} ({row['Coverage %']:.1f}% coverage)"
        for _, row in strongest.iterrows()
    ]
    weakest_text = [
        f"{row['Field']} ({row['Coverage %']:.1f}% coverage)"
        for _, row in weakest.iterrows()
    ]

    return f"""# Deliverables Summary — {scope_label}

## Scope and research coverage
- Companies represented: {company_count}
- Research property records analysed: {len(qa_frame)}
- Existing Starting Data matches: {existing}
- Newly discovered records: {discovered}
- Existing records with material researched differences: {changed}
- Records still needing discovery classification: {needs_classification}
- Possible duplicates flagged: {duplicates}
- Excluded / not-current records: {excluded}

## Review and external-submission tracking
- Approved for Export: {approved}
- Newly discovered records marked Submitted Externally: {submitted}
- Directory-entry records needing correction: {needs_correction}
- Records with genuine unresolved review items: {still_review}

## Current quality position
- Records with critical issues: {critical}
- Warnings: {warnings}
- Open research gaps: {research_gaps}
- Human-verified records: {verified}

## Strongest publicly available directory fields
{chr(10).join(f'- {item}' for item in strongest_text) if strongest_text else '- No field-availability results available.'}

## Largest public-data gaps
{chr(10).join(f'- {item}' for item in weakest_text) if weakest_text else '- No field-availability results available.'}

## Key methodology
- Inventory-first public-source research.
- Official company/property sources used as primary evidence.
- Starting Data comparison performed inside Datablix after research import.
- Missing information documented rather than inferred.
- Website inventory presence kept separate from current rental availability.
- Human verification retained before export or external-submission tracking.

## Reporting boundary
- Datablix reports the research dataset and recorded submission tracking.
- Datablix does not claim totals for any external master dataset that is not directly accessible.

## Recommended next steps
- Resolve remaining high-priority follow-ups and possible duplicates.
- Use field-availability evidence to prioritize reliable search/filter options and identify data-collection gaps.
- Preserve Datablix as the research evidence trail and treat any separately maintained external dataset according to its own governance process.
"""


def project_deliverables_table():
    """Map each reusable deliverable to its Datablix generator."""
    return pd.DataFrame([
        {"Deliverable": "1. Research Dataset", "Datablix Location": "Deliverables → Research dataset", "Generated Output": "Full research + filtered final directory view", "Export": "Excel + CSV"},
        {"Deliverable": "2. Organization Research Summary", "Datablix Location": "Deliverables → Research coverage", "Generated Output": "All research by company + coverage + discoveries + status summary", "Export": "Excel"},
        {"Deliverable": "3. Source & Verification Tracker", "Datablix Location": "Deliverables → Verification", "Generated Output": "Full verification tracker + status summary + active follow-up view", "Export": "Excel"},
        {"Deliverable": "4. Draft Record Profiles", "Datablix Location": "Deliverables → Profiles", "Generated Output": "Full research profiles + filtered final profile candidates", "Export": "HTML"},
        {"Deliverable": "5. Structure & Search Recommendations", "Datablix Location": "Deliverables → Recommendations", "Generated Output": "Full-research availability + candidate findings + recommendations", "Export": "Excel"},
        {"Deliverable": "6. Methodology & Limitations", "Datablix Location": "Deliverables → Methodology", "Generated Output": "Full-research status + methodology, assumptions, limitations, next steps", "Export": "HTML"},
        {"Deliverable": "7. Final Project Data Matrix", "Datablix Location": "Deliverables → Final matrix", "Generated Output": "Full-research project matrix + clearly separated filtered-use metrics", "Export": "HTML + CSV"},
    ])


def report_summary(qa_frame, registry=None, scope_label="All companies", baseline=None):
    registry = normalize_company_registry(registry)
    company_count = int(qa_frame["Company ID"].astype(str).replace("", pd.NA).dropna().nunique())
    if scope_label == "All companies" and not registry.empty:
        company_count = len(registry)
    approved_count = int(approved_for_export_mask(qa_frame).sum())
    submitted_count = int((
        qa_frame.get("Directory Discovery Status", pd.Series("", index=qa_frame.index)).astype(str).eq("Newly Discovered")
        & qa_frame.get("Directory Entry Status", pd.Series("", index=qa_frame.index)).astype(str).eq("Entered")
    ).sum()) if not qa_frame.empty else 0
    issue_count = int(qa_frame["QA Flag Count"].sum())
    unresolved_count = int(deliverable_needs_review_mask(qa_frame).sum())
    cities = sorted(set(resolved(qa_frame["City"]).dropna().astype(str).str.strip()))
    existing_count = int(qa_frame["Directory Discovery Status"].eq("Existing Source Record").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    discovered_count = int(qa_frame["Directory Discovery Status"].eq("Newly Discovered").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    needs_classification_count = int(qa_frame["Directory Discovery Status"].eq("Needs Classification").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    duplicate_count = int(qa_frame["Directory Discovery Status"].eq("Possible Duplicate").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    excluded_count = int(qa_frame["Directory Discovery Status"].eq("Excluded / Not Current").sum()) if "Directory Discovery Status" in qa_frame.columns else 0
    verified_count = int(qa_frame["Verification Status"].eq("Verified").sum()) if "Verification Status" in qa_frame.columns else 0
    company_matrix = closeout_research_coverage_matrix(qa_frame, registry, baseline=baseline)
    changed_count = int(company_matrix["Changed Existing Records"].sum()) if not company_matrix.empty else 0

    rows = [
        {"Section": "Scope", "Report Text": f"Analysis scope: {scope_label}. Companies represented or assigned: {company_count:,}. Research property records analysed: {len(qa_frame):,}."},
        {"Section": "Research contribution", "Report Text": f"The current scope contains {existing_count:,} existing Starting Data match(es), {discovered_count:,} newly discovered record(s), and {changed_count:,} existing record(s) with material researched differences. {needs_classification_count:,} record(s) still need discovery classification, {duplicate_count:,} are possible duplicate(s), and {excluded_count:,} are excluded/not-current."},
        {"Section": "External submission tracking", "Report Text": f"{submitted_count:,} newly discovered record(s) are currently marked as externally submitted. Datablix records the submission workflow but does not independently verify separately maintained external datasets."},
        {"Section": "Data quality", "Report Text": f"The current audit contains {issue_count:,} rule-based quality findings. {verified_count:,} record(s) are human verified, {approved_count:,} are Approved for Export, and {unresolved_count:,} have genuine unresolved review or verification items."},
        {"Section": "Method", "Report Text": "Companies were researched separately using an inventory-first public-source method. Structured research results were imported into Datablix, compared with Starting Data, and reviewed by a person before final use."},
        {"Section": "Assumptions", "Report Text": "A loading property URL is not proof of current inventory; website inventory presence is not the same as current rental availability; unavailable information is documented rather than invented; and official company/property sources are primary evidence."},
        {"Section": "Limitations", "Report Text": "Public information may be incomplete, outdated, blocked, duplicated, JavaScript-dependent, or inconsistent. Datablix reports only datasets available within the workspace and does not claim totals for separately maintained external datasets."},
        {"Section": "Recommended next actions", "Report Text": "Resolve high-priority follow-ups, use the field-availability matrix to prioritize reliable search and filter options, flag high-value fields with weak public coverage for further data collection, and use the downloadable project data matrix to support final project outputs."},
    ]
    return pd.DataFrame(rows)


def project_info_dataframe(qa_frame, registry):
    registry = normalize_company_registry(registry)
    return pd.DataFrame([
        {"Setting": "Project Name", "Value": st.session_state.get(S_PROJECT_NAME, "Datablix master project")},
        {"Setting": "Saved At", "Value": datetime.now().isoformat(timespec="seconds")},
        {"Setting": "Companies in Scope", "Value": len(registry)},
        {"Setting": "Building Records", "Value": len(qa_frame)},
        {"Setting": "Approved for Export", "Value": int(approved_for_export_mask(qa_frame).sum())},
        {"Setting": "Source", "Value": st.session_state.get(S_SOURCE_TYPE, "Workspace")},
        {"Setting": "Source Baseline Records", "Value": _safe_int((st.session_state.get(S_SOURCE_BASELINE_META, {}) or {}).get("source_records", 0))},
        {"Setting": "Current Source Version", "Value": safe_text((st.session_state.get(S_SOURCE_BASELINE_META, {}) or {}).get("version_label", "v1")) or "v1"},
        {"Setting": "Source Versions Preserved", "Value": len(_source_versions_state())},
        {"Setting": "Source Assignment Sheet", "Value": safe_text((st.session_state.get(S_SOURCE_BASELINE_META, {}) or {}).get("assignment_sheet", ""))},
        {"Setting": "Datablix Project Format", "Value": "2"},
    ])


def project_workbook_bytes():
    working = st.session_state.get(S_WORKING)
    if not isinstance(working, pd.DataFrame):
        working = normalize_workflow(pd.DataFrame(columns=INTERNAL_COLUMNS))
    working, registry = synchronize_company_registry(
        working,
        st.session_state.get(S_COMPANIES),
    )
    st.session_state[S_WORKING] = working
    st.session_state[S_COMPANIES] = registry
    if not working.empty:
        qa_frame = qa_checks(working)
    else:
        qa_frame = working.copy()
        for column, default in {
            "QA Flags": "No rental property data issues found",
            "QA Flag Count": 0,
            "Research Gap Count": 0,
            "Record Readiness": "Needs Research",
            "QA Status": "Pass",
        }.items():
            qa_frame[column] = default
    source_baseline = st.session_state.get(S_ORIGINAL)
    source_meta = st.session_state.get(S_SOURCE_BASELINE_META, {})
    if not isinstance(source_meta, dict) or not source_meta:
        source_baseline = pd.DataFrame()
    if not isinstance(source_baseline, pd.DataFrame):
        source_baseline = pd.DataFrame()

    classification_rules = st.session_state.get(
        S_CLASSIFICATION_RULES
    )
    if not isinstance(classification_rules, pd.DataFrame):
        classification_rules = pd.DataFrame()

    source_meta_frame = (
        pd.DataFrame([source_meta])
        if isinstance(source_meta, dict) and source_meta
        else pd.DataFrame()
    )

    source_versions = _source_versions_state()
    source_versions_meta = _source_versions_meta_frame(
        source_versions
    )
    source_version_records = _source_version_records_frame(
        source_versions
    )
    source_version_rules = _source_version_rules_frame(
        source_versions
    )
    sheets = {
        "Project Info": project_info_dataframe(qa_frame, registry),
        "Company Registry": registry,
        "Working Data": working,
        "Source Baseline": source_baseline,
        "Source Baseline Meta": source_meta_frame,
        "Source Versions Meta": source_versions_meta,
        "Source Version Records": source_version_records,
        "Source Version Rules": source_version_rules,
        "Classification Rules": classification_rules,
        "Current QA": qa_frame,
        "Company Analysis": company_progress_summary(qa_frame, registry) if not qa_frame.empty else pd.DataFrame(),
        "Report Summary": report_summary(qa_frame, registry),
        "Scan History": st.session_state.get(S_SCAN_HISTORY, pd.DataFrame()),
        "Scan Candidates": st.session_state.get(S_SCAN_CANDIDATES, pd.DataFrame()),
        "Scan Pages": st.session_state.get(S_SCAN_PAGES, pd.DataFrame()),
    }
    return excel_bytes(sheets)


def load_project_workbook(uploaded):
    data = uploaded.getvalue()
    with pd.ExcelFile(io.BytesIO(data), engine="openpyxl") as workbook:
        if "Working Data" not in workbook.sheet_names:
            raise ValueError(
                "This workbook is not a resumable Datablix project. Open a file containing a 'Working Data' sheet, or use Open file for an ordinary directory workbook."
            )
        working = prepare_data(pd.read_excel(workbook, sheet_name="Working Data"))
        for column in INTERNAL_COLUMNS:
            if column not in working.columns:
                working[column] = pd.NA
        working = ensure_ids(normalize_workflow(working))
        source_baseline = (
            prepare_data(pd.read_excel(workbook, sheet_name="Source Baseline"))
            if "Source Baseline" in workbook.sheet_names
            else pd.DataFrame()
        )
        if not source_baseline.empty:
            for column in INTERNAL_COLUMNS:
                if column not in source_baseline.columns:
                    source_baseline[column] = pd.NA
            source_baseline = ensure_ids(normalize_workflow(source_baseline))
        source_meta_sheet = (
            pd.read_excel(workbook, sheet_name="Source Baseline Meta")
            if "Source Baseline Meta" in workbook.sheet_names
            else pd.DataFrame()
        )
        source_meta = (
            {
                str(key): ("" if pd.isna(value) else value)
                for key, value in source_meta_sheet.iloc[0].to_dict().items()
            }
            if not source_meta_sheet.empty
            else {}
        )
        classification_rules = (
            pd.read_excel(workbook, sheet_name="Classification Rules")
            if "Classification Rules" in workbook.sheet_names
            else pd.DataFrame()
        )
        source_versions_meta_sheet = (
            pd.read_excel(
                workbook,
                sheet_name="Source Versions Meta",
            )
            if "Source Versions Meta" in workbook.sheet_names
            else pd.DataFrame()
        )
        source_version_records_sheet = (
            pd.read_excel(
                workbook,
                sheet_name="Source Version Records",
            )
            if "Source Version Records" in workbook.sheet_names
            else pd.DataFrame()
        )
        source_version_rules_sheet = (
            pd.read_excel(
                workbook,
                sheet_name="Source Version Rules",
            )
            if "Source Version Rules" in workbook.sheet_names
            else pd.DataFrame()
        )

        source_versions = _restore_source_versions_from_workbook(
            source_versions_meta_sheet,
            source_version_records_sheet,
            source_version_rules_sheet,
        )

        # Older saved projects automatically become source v1.
        if (
            not source_versions
            and isinstance(source_baseline, pd.DataFrame)
            and not source_baseline.empty
        ):
            migrated_meta = dict(source_meta)
            migrated_meta.setdefault("version_number", 1)
            migrated_meta.setdefault("version_label", "v1")
            migrated_meta.setdefault("is_original", True)
            migrated_meta.setdefault("is_active", True)

            source_versions = [{
                "version_number": 1,
                "version_label": "v1",
                "is_original": True,
                "is_active": True,
                "meta": migrated_meta,
                "records": source_baseline.copy(),
                "rules": classification_rules.copy(),
            }]
        registry = (
            pd.read_excel(workbook, sheet_name="Company Registry")
            if "Company Registry" in workbook.sheet_names
            else empty_company_registry()
        )
        scan_history = (
            pd.read_excel(workbook, sheet_name="Scan History")
            if "Scan History" in workbook.sheet_names
            else pd.DataFrame()
        )
        scan_candidates = (
            pd.read_excel(workbook, sheet_name="Scan Candidates")
            if "Scan Candidates" in workbook.sheet_names
            else pd.DataFrame()
        )
        scan_pages = (
            pd.read_excel(workbook, sheet_name="Scan Pages")
            if "Scan Pages" in workbook.sheet_names
            else pd.DataFrame()
        )
        project_name = safe_filename(uploaded.name).replace("_", " ").title()
        if "Project Info" in workbook.sheet_names:
            project_info = pd.read_excel(workbook, sheet_name="Project Info")
            if {"Setting", "Value"}.issubset(project_info.columns):
                name_rows = project_info.loc[project_info["Setting"].eq("Project Name"), "Value"]
                if not name_rows.empty and str(name_rows.iloc[0]).strip():
                    project_name = str(name_rows.iloc[0]).strip()

    working, registry = synchronize_company_registry(working, registry)
    mapping = pd.DataFrame({
        "Datablix Field": INTERNAL_COLUMNS,
        "Imported Column(s)": INTERNAL_COLUMNS,
        "Mapping Status": "Saved project field",
    })
    signature = f"project:{uploaded.name}:{hashlib.sha256(data).hexdigest()}"
    st.session_state.pop(S_FILE, None)
    open_workspace(
        working,
        mapping,
        signature,
        uploaded.name,
        "Working Data",
        "Saved Datablix project",
        uploaded.name,
        message=f"Resumed {project_name} with {len(working):,} building record(s).",
        registry=registry,
    )
    st.session_state[S_PROJECT_NAME] = project_name
    st.session_state[S_COMPANIES] = registry
    st.session_state[S_SOURCE_VERSIONS] = source_versions

    active_source = next(
        (
            version
            for version in reversed(source_versions)
            if version.get("is_active")
        ),
        source_versions[-1] if source_versions else None,
    )

    if active_source is not None:
        st.session_state[S_ORIGINAL] = active_source[
            "records"
        ].copy()
        st.session_state[S_SOURCE_BASELINE_META] = dict(
            active_source.get("meta", {})
        )
        st.session_state[S_CLASSIFICATION_RULES] = (
            active_source["rules"].copy()
        )
    else:
        if (
            isinstance(source_baseline, pd.DataFrame)
            and not source_baseline.empty
        ):
            st.session_state[S_ORIGINAL] = source_baseline
        st.session_state[S_SOURCE_BASELINE_META] = source_meta
        st.session_state[S_CLASSIFICATION_RULES] = (
            classification_rules
        )
    st.session_state[S_SCAN_HISTORY] = scan_history
    st.session_state[S_SCAN_CANDIDATES] = scan_candidates
    st.session_state[S_SCAN_PAGES] = scan_pages
    st.session_state[S_PROJECT_LOADED] = True
    st.session_state[S_CLOUD_PROJECT_ID] = str(uuid.uuid4())
    st.session_state.pop(S_CLOUD_STATE_HASH, None)
    if not registry.empty:
        active_id = str(st.session_state.get(S_ACTIVE_COMPANY, "")).strip()
        if active_id not in set(registry["Company ID"].astype(str)):
            st.session_state[S_ACTIVE_COMPANY] = registry.iloc[0]["Company ID"]


# =========================================================
# Session operations
# =========================================================

def open_workspace(
    mapped,
    mapping,
    signature,
    name,
    sheet,
    source_type,
    source_ref="",
    selector="",
    message="Rental property workspace opened.",
    registry=None,
):
    if st.session_state.get(S_FILE) != signature:
        starting_registry = (
            normalize_company_registry(registry)
            if isinstance(registry, pd.DataFrame)
            else empty_company_registry()
        )
        mapped, registry = synchronize_company_registry(mapped, starting_registry)

        # A newly opened source is a new project context. Do not silently carry
        # companies or scan results from the previous project.
        st.session_state[S_FILE] = signature
        st.session_state[S_ORIGINAL] = mapped.copy()
        st.session_state[S_WORKING] = mapped.copy()
        st.session_state[S_MAPPING] = mapping
        st.session_state[S_NAME] = name
        st.session_state[S_SHEET] = sheet or ""
        st.session_state[S_SOURCE_TYPE] = source_type
        st.session_state[S_SOURCE_REF] = source_ref
        st.session_state[S_SELECTOR] = selector
        st.session_state[S_EDIT_COUNT] = 0
        st.session_state[S_COMPANIES] = registry
        st.session_state[S_PROJECT_NAME] = (
            safe_filename(name).replace("_", " ").title()
            or "Datablix master project"
        )
        st.session_state[S_SCAN_HISTORY] = pd.DataFrame()
        st.session_state[S_SCAN_CANDIDATES] = pd.DataFrame()
        st.session_state[S_SCAN_PAGES] = pd.DataFrame()
        st.session_state[S_SOURCE_BASELINE_META] = {}
        st.session_state[S_SOURCE_VERSIONS] = []
        st.session_state[S_CLASSIFICATION_RULES] = pd.DataFrame()
        st.session_state.pop(S_REPORTING_SNAPSHOT, None)
        st.session_state[S_PROJECT_LOADED] = True
        st.session_state[S_CLOUD_PROJECT_ID] = str(uuid.uuid4())
        st.session_state.pop(S_CLOUD_STATE_HASH, None)
        st.session_state[S_ACTIVE_COMPANY] = (
            registry.iloc[0]["Company ID"] if not registry.empty else ""
        )

        # Clear only the current scanner UI/session cache. Historical scan logs
        # remain in the project tables above.
        for key in list(st.session_state):
            if (
                key.startswith("website_scan_")
                or key.startswith("full_scan_")
                or key.startswith("_db_company_scan_")
                or key in {"confirm_clear_full_scan", "_db_active_scan_company"}
            ):
                st.session_state.pop(key, None)

        st.session_state[S_FLASH] = message


def looks_like_company_assignment(df):
    """Return True when rows describe assigned companies rather than buildings."""
    imported = prepare_data(df)
    has_company = bool(source_columns(imported, ALIASES["Management/Owner"]))
    if not has_company:
        return False

    # City, province, or postal columns may describe a company's office and do
    # not prove that each row is a residential rental property. Treat strong property
    # identity fields as the deciding signal.
    has_building_name = bool(source_columns(imported, ALIASES["Building Name"]))
    has_street_address = bool(source_columns(imported, ALIASES["Street Address"]))
    has_apartment_count = bool(
        source_columns(imported, ALIASES["Number of Apartments"])
    )
    has_building_records = (
        has_building_name or has_street_address or has_apartment_count
    )
    return not has_building_records


def company_registry_from_assignment(df):
    """Build the company registry from an assignment/company-list worksheet."""
    imported = prepare_data(df)
    owner_columns = source_columns(imported, ALIASES["Management/Owner"])
    if not owner_columns:
        raise ValueError(
            "Datablix could not find a company or management-owner column. "
            "Use a heading such as Assigned Company, Management Company, Owner, or Company."
        )

    def values_for(aliases):
        columns = source_columns(imported, aliases)
        return (
            combine_columns(imported, columns)
            if columns
            else pd.Series(pd.NA, index=imported.index, dtype="object")
        )

    owners = combine_columns(imported, owner_columns)
    company_ids = values_for(ALIASES["Company ID"])
    websites = values_for([
        "Main Website", "Company Website", "Portfolio Website",
        "Website", "WebSite", "Website / Source URL",
    ])
    scope_types = values_for(["Scope Type", "Assignment Type"])
    assigned_dates = values_for(["Date Assigned", "Assignment Date"])
    statuses = values_for(["Company Status", "Status"])
    notes = values_for(["Company Notes", "Assignment Notes", "Notes", "Reviewer Notes"])

    rows_by_name = {}
    for index in imported.index:
        owner = "" if is_unresolved(owners.loc[index]) else re.sub(
            r"\s+", " ", str(owners.loc[index])
        ).strip()
        if not owner:
            continue
        key = company_name_key(owner)
        row = rows_by_name.setdefault(key, {
            "Company ID": "",
            "Management/Owner": owner,
            "Main Website": "",
            "Scope Type": "Initial assignment",
            "Date Assigned": date.today().isoformat(),
            "Company Status": "Not started",
            "Notes": "",
        })

        candidates = {
            "Company ID": company_ids.loc[index],
            "Main Website": websites.loc[index],
            "Scope Type": scope_types.loc[index],
            "Date Assigned": assigned_dates.loc[index],
            "Company Status": statuses.loc[index],
            "Notes": notes.loc[index],
        }
        for field, value in candidates.items():
            if is_unresolved(value):
                continue
            clean = str(value).strip()
            if field == "Notes" and row[field] and clean not in row[field]:
                row[field] = f"{row[field]} | {clean}"
            elif not row[field] or field in {"Scope Type", "Company Status"}:
                row[field] = clean

    if not rows_by_name:
        raise ValueError("No company names were found in the selected worksheet.")

    registry = pd.DataFrame(list(rows_by_name.values()))
    registry["Scope Type"] = normalize_choice(
        registry["Scope Type"], COMPANY_SCOPE_TYPES, "Initial assignment"
    )
    registry["Company Status"] = normalize_choice(
        registry["Company Status"], COMPANY_STATUSES, "Not started"
    )
    return normalize_company_registry(registry)


def open_assignment_project(
    df,
    data,
    name,
    sheet,
    source_type,
    source_ref="",
    selector="",
):
    registry = company_registry_from_assignment(df)
    working = normalize_workflow(pd.DataFrame(columns=INTERNAL_COLUMNS))
    mapping = pd.DataFrame({
        "Datablix Field": COMPANY_COLUMNS,
        "Imported Column(s)": COMPANY_COLUMNS,
        "Mapping Status": "Company assignment field",
    })
    signature = f"assignment:{name}:{sheet}:{hashlib.sha256(data).hexdigest()}"
    open_workspace(
        working,
        mapping,
        signature,
        name,
        sheet,
        source_type,
        source_ref,
        selector,
        message=(
            f"Project registered with {len(registry):,} assigned company or owner "
            "record(s). Select a company and prepare its external AI research prompt."
        ),
        registry=registry,
    )


def load_upload(uploaded, sheet=None):
    df, data = read_upload(uploaded, sheet)
    if looks_like_company_assignment(df):
        open_assignment_project(
            df, data, uploaded.name, sheet, "Uploaded assignment file", uploaded.name
        )
        return "company_assignment"

    validate_input(df)
    mapped, mapping = map_schema(df)
    signature = f"{uploaded.name}:{sheet}:{hashlib.sha256(data).hexdigest()}"
    open_workspace(
        mapped,
        mapping,
        signature,
        uploaded.name,
        sheet,
        "Uploaded building file",
        uploaded.name,
        message=(
            f"Opened {uploaded.name} with {len(mapped):,} building record(s). "
            "Companies were registered from the management-owner fields."
        ),
    )
    return "building_records"



SOURCE_WORKBOOK_RESERVED_SHEETS = {
    "apartmentbuildings",
    "listofcompanies",
    "buildingclassifications",
    "projectinfo",
    "companyregistry",
    "workingdata",
    "currentqa",
}


def source_assignment_sheet_candidates(uploaded) -> list[str]:
    """Return likely researcher/assignment tabs from a multi-sheet source workbook."""
    names = excel_sheet_names(uploaded)
    candidates = [
        name for name in names
        if norm_header(name) not in SOURCE_WORKBOOK_RESERVED_SHEETS
    ]
    return candidates or names


def _clean_assignment_company_name(value: str) -> str:
    """Reduce explanatory assignment labels to the actual company name."""
    clean = re.sub(r"\s+", " ", safe_text(value)).strip()
    clean = re.sub(r"\s+is\s+also\s+.*$", "", clean, flags=re.IGNORECASE).strip()
    return clean


def _company_core_key(value: str) -> str:
    """Create a conservative matching key for company-name variants."""
    clean = _clean_assignment_company_name(value).lower()
    tokens = re.findall(r"[a-z0-9]+", clean)
    stopwords = {
        "the", "property", "properties", "management", "manager", "managers",
        "apartment", "apartments", "reit", "land", "holdings", "holding",
        "group", "inc", "incorporated", "ltd", "limited", "corp", "corporation",
        "company", "companies", "service", "services", "realty", "real", "estate",
    }
    core = [token for token in tokens if token not in stopwords]
    return "".join(core) if core else "".join(tokens)


def _company_core_matches(left: str, right: str) -> bool:
    a = _company_core_key(left)
    b = _company_core_key(right)
    if not a or not b:
        return False
    if a == b:
        return True
    return min(len(a), len(b)) >= 4 and (a in b or b in a)


def _assignment_registry_from_block_sheet(data: bytes, sheet_name: str) -> pd.DataFrame:
    """Read block-style assignment tabs where each company occupies one group of rows."""
    raw = pd.read_excel(
        io.BytesIO(data),
        sheet_name=sheet_name,
        header=None,
        engine="openpyxl",
    )
    if raw.empty:
        raise ValueError("The selected assignment sheet is empty.")

    first_column = raw.iloc[:, 0].tolist()
    blocks = []
    current = []
    for value in first_column:
        if pd.isna(value) or not str(value).strip():
            if current:
                blocks.append(current)
                current = []
            continue
        current.append(str(value).strip())
    if current:
        blocks.append(current)

    rows = []
    for block in blocks:
        if not block:
            continue
        raw_name = block[0]
        company_name = _clean_assignment_company_name(raw_name)
        if not company_name:
            continue

        # The first company website in the block is retained as a starting reference.
        urls = [
            value for value in block
            if str(value).strip().lower().startswith(("http://", "https://"))
        ]
        website = next(
            (
                url for url in urls
                if "fifty-five-plus.com/apartment" not in url.lower()
            ),
            urls[0] if urls else "",
        )
        note = ""
        if company_name != raw_name:
            note = f"Assignment label: {raw_name}"

        rows.append({
            "Company ID": "",
            "Management/Owner": company_name,
            "Main Website": website,
            "Scope Type": "Initial assignment",
            "Date Assigned": date.today().isoformat(),
            "Company Status": "Not started",
            "Notes": note,
        })

    if not rows:
        raise ValueError(
            "Datablix could not identify company blocks in the selected assignment sheet."
        )
    return pd.DataFrame(rows)


def _merge_assignment_registry(existing, incoming) -> pd.DataFrame:
    """Merge an assignment list into the current registry while preserving existing IDs and prompts."""
    registry = normalize_company_registry(existing)
    incoming = incoming.copy()

    for _, row in incoming.iterrows():
        name = safe_text(row.get("Management/Owner", ""))
        if not name:
            continue

        match_index = None
        for idx, existing_row in registry.iterrows():
            if _company_core_matches(name, existing_row.get("Management/Owner", "")):
                match_index = idx
                break

        if match_index is not None:
            if not safe_text(registry.at[match_index, "Main Website"]):
                registry.at[match_index, "Main Website"] = safe_text(row.get("Main Website", ""))
            registry.at[match_index, "Scope Type"] = "Initial assignment"
            if not safe_text(registry.at[match_index, "Date Assigned"]):
                registry.at[match_index, "Date Assigned"] = safe_text(row.get("Date Assigned", ""))
            incoming_note = safe_text(row.get("Notes", ""))
            if incoming_note and incoming_note not in safe_text(registry.at[match_index, "Notes"]):
                existing_note = safe_text(registry.at[match_index, "Notes"])
                registry.at[match_index, "Notes"] = (
                    f"{existing_note} | {incoming_note}" if existing_note else incoming_note
                )
            continue

        new_row = {column: "" for column in COMPANY_COLUMNS}
        for column in COMPANY_COLUMNS:
            if column in row.index and not is_unresolved(row.get(column)):
                new_row[column] = row.get(column)
        new_row["Company ID"] = next_company_id(registry)
        new_row["Management/Owner"] = name
        new_row["Scope Type"] = "Initial assignment"
        new_row["Company Status"] = safe_text(row.get("Company Status", "Not started"), "Not started") or "Not started"
        registry = pd.concat([registry, pd.DataFrame([new_row])], ignore_index=True)

    return normalize_company_registry(registry)


def _registry_match_index(owner_name: str, registry: pd.DataFrame):
    for idx, row in registry.iterrows():
        if _company_core_matches(owner_name, row.get("Management/Owner", "")):
            return idx
    return None


def _read_source_table_with_detected_header(
    data: bytes,
    sheet_name: str,
) -> pd.DataFrame:
    """Read a project source table even when headings do not start on row 1."""
    raw = pd.read_excel(
        io.BytesIO(data),
        sheet_name=sheet_name,
        header=None,
        engine="openpyxl",
    )
    if raw.empty:
        return pd.DataFrame()

    alias_groups = [
        ALIASES["Building Name"],
        ALIASES["Management/Owner"],
        ALIASES["Street Address"],
        ALIASES["City"],
        COMBINED_LOCATION_ALIASES,
        ALIASES["Website"],
        ALIASES["Phone"],
        ALIASES["Number of Apartments"],
        ALIASES["Property Type"],
        ALIASES["Building Classification"],
    ]
    normalized_groups = [
        {norm_header(alias) for alias in aliases if norm_header(alias)}
        for aliases in alias_groups
    ]

    best_row = None
    best_score = 0
    for idx in range(min(len(raw), 25)):
        row_values = [
            norm_header(value)
            for value in raw.iloc[idx].tolist()
            if safe_text(value)
        ]
        row_values = [value for value in row_values if value]
        if not row_values:
            continue

        score = 0
        for group in normalized_groups:
            if any(
                cell in group
                or any(len(alias) >= 4 and (alias in cell or cell in alias) for alias in group)
                for cell in row_values
            ):
                score += 1
        if score > best_score:
            best_score = score
            best_row = idx

    if best_row is None or best_score < 2:
        return pd.DataFrame()

    return prepare_data(
        pd.read_excel(
            io.BytesIO(data),
            sheet_name=sheet_name,
            header=int(best_row),
            engine="openpyxl",
        )
    )


def _source_sheet_structure_score(
    data: bytes,
    sheet_name: str,
) -> tuple[int, bool]:
    """Score a worksheet using the property table Datablix can actually parse."""
    try:
        sample = _read_source_table_with_detected_header(data, sheet_name)
    except Exception:
        return 0, False

    if sample.empty or len(sample.columns) == 0:
        return 0, False

    groups = {
        "building": ALIASES["Building Name"],
        "owner": ALIASES["Management/Owner"],
        "address": ALIASES["Street Address"],
        "city": ALIASES["City"],
        "location": COMBINED_LOCATION_ALIASES,
        "website": ALIASES["Website"],
        "phone": ALIASES["Phone"],
        "apartments": ALIASES["Number of Apartments"],
        "property_type": ALIASES["Property Type"],
        "classification": ALIASES["Building Classification"],
    }
    found = {key: bool(source_columns(sample, aliases)) for key, aliases in groups.items()}
    score = sum(int(value) for value in found.values())
    identity_signals = sum(int(found[key]) for key in ["building", "owner", "address", "city", "location"])
    strong_identity = identity_signals >= 2 or (
        found["owner"] and (found["website"] or found["phone"] or found["apartments"])
    )
    return score, strong_identity


def _find_source_building_sheet(
    data: bytes,
    sheet_names: list[str],
    assignment_sheet: str = "",
) -> str | None:
    """Find the project building table by both worksheet name and schema."""
    if not sheet_names:
        return None

    preferred_tokens = [
        "apartmentbuildings",
        "apartmentbuilding",
        "buildingdirectory",
        "apartmentdirectory",
        "buildingdata",
        "propertydirectory",
        "properties",
        "buildings",
        "listings",
    ]

    # First try likely worksheet names, but verify their columns.
    named_candidates = []
    for name in sheet_names:
        if len(sheet_names) > 1 and safe_text(name) == safe_text(assignment_sheet):
            continue

        normalized = norm_header(name)
        if any(token in normalized for token in preferred_tokens):
            score, strong = _source_sheet_structure_score(
                data,
                name,
            )
            if strong:
                named_candidates.append((score, name))

    if named_candidates:
        named_candidates.sort(
            key=lambda item: item[0],
            reverse=True,
        )
        return named_candidates[0][1]

    # Fallback: ignore the title entirely and inspect worksheet structure.
    reserved = {
        "listofcompanies",
        "buildingclassifications",
        "classificationrules",
    }
    if len(sheet_names) > 1 and safe_text(assignment_sheet):
        reserved.add(norm_header(assignment_sheet))

    structural_candidates = []
    for name in sheet_names:
        if norm_header(name) in reserved:
            continue

        score, strong = _source_sheet_structure_score(
            data,
            name,
        )
        if strong:
            structural_candidates.append((score, name))

    if not structural_candidates:
        return None

    structural_candidates.sort(
        key=lambda item: item[0],
        reverse=True,
    )
    return structural_candidates[0][1]


def _find_classification_sheet(
    data: bytes,
    sheet_names: list[str],
) -> str | None:
    """Find building-classification rules by title or Type/Height structure."""
    named = next(
        (
            name
            for name in sheet_names
            if (
                "buildingclassifications" in norm_header(name)
                or "classificationrules" in norm_header(name)
            )
        ),
        None,
    )
    if named:
        return named

    for name in sheet_names:
        try:
            raw = pd.read_excel(
                io.BytesIO(data),
                sheet_name=name,
                header=None,
                nrows=25,
                engine="openpyxl",
            )
        except Exception:
            continue

        if raw.empty or raw.shape[1] < 2:
            continue

        for idx in raw.index:
            first = norm_header(raw.iloc[idx, 0])
            second = norm_header(raw.iloc[idx, 1])
            if first == "type" and "height" in second:
                return name

    return None


def _parse_classification_rules(data: bytes, sheet_name: str | None) -> pd.DataFrame:
    if not sheet_name:
        return pd.DataFrame(columns=["Type", "Typical Height"])
    raw = pd.read_excel(
        io.BytesIO(data),
        sheet_name=sheet_name,
        header=None,
        engine="openpyxl",
    )
    if raw.empty or raw.shape[1] < 2:
        return pd.DataFrame(columns=["Type", "Typical Height"])

    header_row = None
    for idx in raw.index:
        first = norm_header(raw.iloc[idx, 0])
        second = norm_header(raw.iloc[idx, 1])
        if first == "type" and "height" in second:
            header_row = idx
            break
    if header_row is None:
        return pd.DataFrame(columns=["Type", "Typical Height"])

    rows = []
    for idx in range(header_row + 1, len(raw)):
        type_value = raw.iloc[idx, 0]
        height_value = raw.iloc[idx, 1]
        if pd.isna(type_value):
            continue
        type_text = str(type_value).strip()
        if not type_text:
            continue
        rows.append({
            "Type": type_text,
            "Typical Height": "" if pd.isna(height_value) else str(height_value).strip(),
        })
    return pd.DataFrame(rows)


def _source_baseline_from_workbook(
    data: bytes,
    assignment_sheet: str = "",
    existing_registry=None,
):
    """Read project-wide Starting Data without requiring a special workbook layout."""
    with pd.ExcelFile(io.BytesIO(data), engine="openpyxl") as workbook:
        sheet_names = workbook.sheet_names

    building_sheet = _find_source_building_sheet(data, sheet_names, assignment_sheet=assignment_sheet)
    classification_sheet = _find_classification_sheet(data, sheet_names)

    current_registry = normalize_company_registry(
        existing_registry if isinstance(existing_registry, pd.DataFrame) else empty_company_registry()
    )
    incoming_registry = empty_company_registry()

    if (
        safe_text(assignment_sheet)
        and len(sheet_names) > 1
        and safe_text(assignment_sheet) != safe_text(building_sheet)
    ):
        try:
            incoming_registry = _assignment_registry_from_block_sheet(data, assignment_sheet)
        except Exception:
            incoming_registry = empty_company_registry()

    registry = _merge_assignment_registry(current_registry, incoming_registry)
    rules = _parse_classification_rules(data, classification_sheet)

    if not building_sheet:
        return pd.DataFrame(columns=INTERNAL_COLUMNS), registry, rules, pd.DataFrame(), ""

    source_df = _read_source_table_with_detected_header(data, building_sheet)
    if source_df.empty:
        return pd.DataFrame(columns=INTERNAL_COLUMNS), registry, rules, pd.DataFrame(), building_sheet

    try:
        validate_input(source_df)
        mapped, mapping = map_schema(source_df)
    except Exception:
        return pd.DataFrame(columns=INTERNAL_COLUMNS), registry, rules, pd.DataFrame(), building_sheet

    if registry.empty:
        try:
            registry = company_registry_from_assignment(source_df)
        except Exception:
            registry = empty_company_registry()

    baseline = mapped.copy()
    if not registry.empty:
        for idx, row in baseline.iterrows():
            match_index = _registry_match_index(safe_text(row.get("Management/Owner", "")), registry)
            if match_index is None:
                continue
            company = registry.loc[match_index]
            baseline.at[idx, "Company ID"] = company["Company ID"]
            baseline.at[idx, "Management/Owner"] = company["Management/Owner"]

    for column in INTERNAL_COLUMNS:
        if column not in baseline.columns:
            baseline[column] = pd.NA

    baseline = normalize_workflow(baseline)
    baseline["Directory Discovery Status"] = "Existing Source Record"
    baseline["Directory Entry Status"] = "Not Entered"
    baseline["Research Status"] = "Imported - Needs Review"
    baseline["Verification Status"] = "Needs Review"
    baseline["Record Decision"] = "Undecided"
    baseline = ensure_ids(baseline)
    return baseline, registry, rules, mapping, building_sheet


def _safe_int(value, default=0) -> int:
    number = pd.to_numeric(value, errors="coerce")
    return default if pd.isna(number) else int(number)


def _repair_active_source_baseline_from_raw_if_needed() -> bool:
    """Rebuild an old saved structured baseline from its preserved raw workbook.

    Starting Data is immutable evidence: the raw workbook is the source of truth for
    how many rows were originally supplied. Older Datablix builds may have saved a
    structured baseline whose company linkage/counts were incomplete. When raw bytes
    are available, re-run the *current* parser once and replace only the derived
    structured baseline. Research rows, notes and review decisions are preserved.
    """
    versions = st.session_state.get(S_SOURCE_VERSIONS, [])
    if not isinstance(versions, list) or not versions:
        return False

    active_index = None
    for index in range(len(versions) - 1, -1, -1):
        item = versions[index]
        if isinstance(item, dict) and bool(item.get("is_active")):
            active_index = index
            break
    if active_index is None:
        for index in range(len(versions) - 1, -1, -1):
            if isinstance(versions[index], dict):
                active_index = index
                break
    if active_index is None:
        return False

    active = dict(versions[active_index])
    meta = dict(active.get("meta", {}) or {})
    parse_version = _safe_int(meta.get("structured_parse_version", 0))
    if parse_version >= SOURCE_BASELINE_PARSE_VERSION:
        return False

    raw_bytes = active.get("raw_bytes", b"")
    if not isinstance(raw_bytes, (bytes, bytearray)) or not raw_bytes:
        # Legacy projects without preserved raw bytes cannot be reconstructed safely.
        # Leave their current baseline untouched rather than inventing missing rows.
        return False

    assignment_sheet = safe_text(meta.get("assignment_sheet", ""))
    current_registry = normalize_company_registry(
        st.session_state.get(S_COMPANIES, empty_company_registry())
    )

    repaired_source, repaired_registry, rules, mapping, building_sheet = (
        _source_baseline_from_workbook(
            bytes(raw_bytes),
            assignment_sheet=assignment_sheet,
            existing_registry=current_registry,
        )
    )
    if not isinstance(repaired_source, pd.DataFrame) or repaired_source.empty:
        return False

    relevant = _source_records_for_project_companies(
        repaired_source,
        repaired_registry,
    )

    meta.update({
        "building_sheet": safe_text(building_sheet),
        "assigned_companies": len(repaired_registry),
        "source_records": len(repaired_source),
        "project_company_source_records": len(relevant),
        "classification_rules": len(rules) if isinstance(rules, pd.DataFrame) else 0,
        "structured_parse_version": SOURCE_BASELINE_PARSE_VERSION,
        "structured_reparse_status": "rebuilt from preserved original workbook",
        "structured_reparsed_at": datetime.now().isoformat(timespec="seconds"),
    })

    active["records"] = repaired_source.copy()
    active["rules"] = rules.copy() if isinstance(rules, pd.DataFrame) else pd.DataFrame()
    active["meta"] = meta
    versions[active_index] = active

    st.session_state[S_SOURCE_VERSIONS] = versions
    st.session_state[S_ORIGINAL] = repaired_source.copy()
    st.session_state[S_COMPANIES] = normalize_company_registry(repaired_registry)
    st.session_state[S_MAPPING] = mapping if isinstance(mapping, pd.DataFrame) else pd.DataFrame()
    st.session_state[S_CLASSIFICATION_RULES] = (
        rules.copy() if isinstance(rules, pd.DataFrame) else pd.DataFrame()
    )
    st.session_state[S_SOURCE_BASELINE_META] = dict(meta)

    # Recompare current research against the repaired source without injecting source
    # rows into the research dataset or discarding any researcher-entered values.
    working = st.session_state.get(S_WORKING)
    if isinstance(working, pd.DataFrame):
        repaired_working = ensure_ids(normalize_workflow(working.copy()))
        st.session_state[S_WORKING] = classify_discovery_status(
            repaired_working,
            repaired_source,
        )

    # Any frozen report may contain the old structured baseline and is no longer valid.
    st.session_state.pop(S_REPORTING_SNAPSHOT, None)
    autosave_current_project()
    return True


def _source_versions_state() -> list[dict]:
    """Return normalized Starting Data history, migrating older projects when needed."""
    raw = st.session_state.get(S_SOURCE_VERSIONS, [])
    versions = raw if isinstance(raw, list) else []

    normalized = []
    for item in versions:
        if not isinstance(item, dict):
            continue

        records = item.get("records")
        rules = item.get("rules")
        meta = item.get("meta", {})
        raw_bytes = item.get("raw_bytes", b"")
        raw_filename = safe_text(item.get("raw_filename", meta.get("workbook_name", "")))

        if not isinstance(raw_bytes, (bytes, bytearray)):
            raw_bytes = b""

        if not isinstance(records, pd.DataFrame):
            records = pd.DataFrame()
        if not isinstance(rules, pd.DataFrame):
            rules = pd.DataFrame()
        if not isinstance(meta, dict):
            meta = {}

        number = _safe_int(
            item.get("version_number", meta.get("version_number", 0))
        )
        if number <= 0:
            number = len(normalized) + 1

        normalized.append({
            "version_number": number,
            "version_label": safe_text(
                item.get("version_label", meta.get("version_label", f"v{number}"))
            ) or f"v{number}",
            "is_original": bool(item.get("is_original", number == 1)),
            "is_active": bool(item.get("is_active", False)),
            "meta": dict(meta),
            "records": records.copy(),
            "rules": rules.copy(),
            "raw_bytes": bytes(raw_bytes),
            "raw_filename": raw_filename,
        })

    # Migrate pre-v30 projects to a single preserved v1.
    if not normalized:
        legacy_meta = st.session_state.get(S_SOURCE_BASELINE_META, {})
        legacy_records = st.session_state.get(S_ORIGINAL)
        legacy_rules = st.session_state.get(S_CLASSIFICATION_RULES)

        if (
            isinstance(legacy_meta, dict)
            and legacy_meta
            and isinstance(legacy_records, pd.DataFrame)
            and not legacy_records.empty
        ):
            meta = dict(legacy_meta)
            meta.setdefault("version_number", 1)
            meta.setdefault("version_label", "v1")
            meta.setdefault("is_original", True)
            meta.setdefault("is_active", True)

            normalized = [{
                "version_number": 1,
                "version_label": "v1",
                "is_original": True,
                "is_active": True,
                "meta": meta,
                "records": legacy_records.copy(),
                "rules": (
                    legacy_rules.copy()
                    if isinstance(legacy_rules, pd.DataFrame)
                    else pd.DataFrame()
                ),
                "raw_bytes": b"",
                "raw_filename": safe_text(meta.get("workbook_name", "")),
            }]

    if normalized and not any(bool(v.get("is_active")) for v in normalized):
        newest = max(
            range(len(normalized)),
            key=lambda i: normalized[i]["version_number"],
        )
        normalized[newest]["is_active"] = True
        normalized[newest]["meta"]["is_active"] = True

    normalized.sort(key=lambda v: int(v.get("version_number", 0)))
    st.session_state[S_SOURCE_VERSIONS] = normalized
    return normalized


def _delete_source_version(version_number: int) -> dict:
    """Delete one NON-CURRENT Starting Data version.

    Current research records are intentionally not altered.
    """
    versions = _source_versions_state()
    target = next(
        (
            version
            for version in versions
            if int(version.get("version_number", 0) or 0)
            == int(version_number)
        ),
        None,
    )

    if target is None:
        return {
            "deleted": False,
            "reason": "Source version was not found.",
        }

    if bool(target.get("is_active")):
        return {
            "deleted": False,
            "reason": (
                "The current source cannot be deleted. "
                "Add or activate another source version first."
            ),
        }

    remaining = [
        version
        for version in versions
        if int(version.get("version_number", 0) or 0)
        != int(version_number)
    ]

    st.session_state[S_SOURCE_VERSIONS] = remaining

    # The active source and current research stay untouched.
    active = next(
        (
            version
            for version in reversed(remaining)
            if bool(version.get("is_active"))
        ),
        None,
    )

    if active is not None:
        st.session_state[S_ORIGINAL] = active["records"].copy()
        st.session_state[S_SOURCE_BASELINE_META] = dict(
            active.get("meta", {})
        )
        st.session_state[S_CLASSIFICATION_RULES] = (
            active["rules"].copy()
            if isinstance(active.get("rules"), pd.DataFrame)
            else pd.DataFrame()
        )

    autosave_current_project()

    return {
        "deleted": True,
        "version_label": safe_text(
            target.get("version_label", f"v{version_number}")
        ) or f"v{version_number}",
        "remaining_versions": len(remaining),
    }


def _active_source_version() -> dict | None:
    versions = _source_versions_state()
    for version in reversed(versions):
        if bool(version.get("is_active")):
            return version
    return versions[-1] if versions else None


def clear_current_starting_source() -> tuple[bool, str]:
    """Remove all Starting Data from the project while preserving research records."""
    had_source = bool(_source_versions_state() or st.session_state.get(S_SOURCE_BASELINE_META))
    st.session_state[S_SOURCE_VERSIONS] = []
    st.session_state[S_ORIGINAL] = pd.DataFrame(columns=INTERNAL_COLUMNS)
    st.session_state[S_SOURCE_BASELINE_META] = {}
    st.session_state[S_CLASSIFICATION_RULES] = pd.DataFrame()
    st.session_state[S_SOURCE_TYPE] = ""
    st.session_state[S_SOURCE_REF] = ""
    st.session_state[S_SHEET] = ""
    st.session_state.pop(S_REPORTING_SNAPSHOT, None)

    current = st.session_state.get(S_WORKING, pd.DataFrame(columns=INTERNAL_COLUMNS))
    if isinstance(current, pd.DataFrame):
        st.session_state[S_WORKING] = classify_discovery_status(
            ensure_ids(normalize_workflow(current.copy())),
            None,
        )

    autosave_current_project()
    if had_source:
        return True, "Starting Data removed. Your research records were preserved."
    return False, "No Starting Data was loaded."


def _source_versions_meta_frame(versions: list[dict]) -> pd.DataFrame:
    rows = []
    for version in versions:
        meta = dict(version.get("meta", {}))
        rows.append({
            "Version Number": int(version.get("version_number", 0) or 0),
            "Version": safe_text(version.get("version_label", "")),
            "Is Original": bool(version.get("is_original", False)),
            "Is Active": bool(version.get("is_active", False)),
            "Workbook": safe_text(meta.get("workbook_name", "")),
            "Assignment Sheet": safe_text(meta.get("assignment_sheet", "")),
            "Building Sheet": safe_text(meta.get("building_sheet", "")),
            "Imported At": safe_text(meta.get("imported_at", "")),
            "Assigned Companies": _safe_int(meta.get("assigned_companies", 0)),
            "Source Records": _safe_int(meta.get("source_records", 0)),
            "Project Company Source Records": _safe_int(meta.get("project_company_source_records", meta.get("source_records", 0))),
            "Source Mode": safe_text(meta.get("source_mode", "Structured records")),
            "Classification Rules": _safe_int(meta.get("classification_rules", 0)),
            "Source Hash": safe_text(meta.get("source_hash", "")),
        })
    return pd.DataFrame(rows)


def _source_version_records_frame(versions: list[dict]) -> pd.DataFrame:
    frames = []
    for version in versions:
        records = version.get("records")
        if not isinstance(records, pd.DataFrame) or records.empty:
            continue

        frame = records.copy()
        frame.insert(0, "__Source Version", safe_text(version.get("version_label", "")))
        frame.insert(
            1,
            "__Source Version Number",
            int(version.get("version_number", 0) or 0),
        )
        frame.insert(2, "__Is Original", bool(version.get("is_original", False)))
        frame.insert(3, "__Is Active", bool(version.get("is_active", False)))
        frames.append(frame)

    return (
        pd.concat(frames, ignore_index=True, sort=False)
        if frames
        else pd.DataFrame()
    )


def _source_version_rules_frame(versions: list[dict]) -> pd.DataFrame:
    frames = []
    for version in versions:
        rules = version.get("rules")
        if not isinstance(rules, pd.DataFrame) or rules.empty:
            continue

        frame = rules.copy()
        frame.insert(0, "__Source Version", safe_text(version.get("version_label", "")))
        frame.insert(
            1,
            "__Source Version Number",
            int(version.get("version_number", 0) or 0),
        )
        frames.append(frame)

    return (
        pd.concat(frames, ignore_index=True, sort=False)
        if frames
        else pd.DataFrame()
    )


def _restore_source_versions_from_workbook(
    meta_frame: pd.DataFrame,
    records_frame: pd.DataFrame,
    rules_frame: pd.DataFrame,
) -> list[dict]:
    """Reconstruct saved Starting Data history from a Datablix project workbook."""
    if not isinstance(meta_frame, pd.DataFrame) or meta_frame.empty:
        return []

    versions = []
    for _, row in meta_frame.iterrows():
        number = _safe_int(row.get("Version Number", 0))
        if number <= 0:
            continue

        label = safe_text(row.get("Version", "")) or f"v{number}"

        if isinstance(records_frame, pd.DataFrame) and not records_frame.empty:
            version_numbers = pd.to_numeric(
                records_frame["__Source Version Number"],
                errors="coerce",
            ).fillna(0).astype(int)

            subset = records_frame.loc[version_numbers.eq(number)].copy()
            subset = subset.drop(
                columns=[
                    "__Source Version",
                    "__Source Version Number",
                    "__Is Original",
                    "__Is Active",
                ],
                errors="ignore",
            )

            if not subset.empty:
                for column in INTERNAL_COLUMNS:
                    if column not in subset.columns:
                        subset[column] = pd.NA
                subset = ensure_ids(normalize_workflow(subset))
        else:
            subset = pd.DataFrame()

        if isinstance(rules_frame, pd.DataFrame) and not rules_frame.empty:
            rule_numbers = pd.to_numeric(
                rules_frame["__Source Version Number"],
                errors="coerce",
            ).fillna(0).astype(int)

            rule_subset = rules_frame.loc[rule_numbers.eq(number)].copy()
            rule_subset = rule_subset.drop(
                columns=["__Source Version", "__Source Version Number"],
                errors="ignore",
            )
        else:
            rule_subset = pd.DataFrame()

        is_original_raw = row.get("Is Original", number == 1)
        is_active_raw = row.get("Is Active", False)

        is_original = (
            number == 1
            if pd.isna(is_original_raw)
            else bool(is_original_raw)
        )
        is_active = (
            False
            if pd.isna(is_active_raw)
            else bool(is_active_raw)
        )

        meta = {
            "version_number": number,
            "version_label": label,
            "is_original": is_original,
            "is_active": is_active,
            "workbook_name": safe_text(row.get("Workbook", "")),
            "assignment_sheet": safe_text(row.get("Assignment Sheet", "")),
            "building_sheet": safe_text(row.get("Building Sheet", "")),
            "imported_at": safe_text(row.get("Imported At", "")),
            "assigned_companies": _safe_int(row.get("Assigned Companies", 0)),
            "source_records": _safe_int(row.get("Source Records", 0)),
            "classification_rules": _safe_int(row.get("Classification Rules", 0)),
            "source_hash": safe_text(row.get("Source Hash", "")),
        }

        versions.append({
            "version_number": number,
            "version_label": label,
            "is_original": is_original,
            "is_active": is_active,
            "meta": meta,
            "records": subset,
            "rules": rule_subset,
        })

    versions.sort(key=lambda v: int(v.get("version_number", 0)))

    if versions and not any(bool(v.get("is_active")) for v in versions):
        versions[-1]["is_active"] = True
        versions[-1]["meta"]["is_active"] = True

    return versions


def _source_records_for_project_companies(
    source_records: pd.DataFrame,
    registry: pd.DataFrame,
) -> pd.DataFrame:
    """Return source rows relevant to registered project companies."""
    if not isinstance(source_records, pd.DataFrame) or source_records.empty:
        return pd.DataFrame(columns=source_records.columns if isinstance(source_records, pd.DataFrame) else [])
    if not isinstance(registry, pd.DataFrame) or registry.empty:
        return source_records.copy()

    matched_rows = []
    for _, row in source_records.iterrows():
        match_index = _registry_match_index(safe_text(row.get("Management/Owner", "")), registry)
        if match_index is None:
            continue
        company = registry.loc[match_index]
        copied = row.copy()
        copied["Company ID"] = company["Company ID"]
        copied["Management/Owner"] = company["Management/Owner"]
        matched_rows.append(copied)

    if not matched_rows:
        return pd.DataFrame(columns=source_records.columns)
    return ensure_ids(normalize_workflow(pd.DataFrame(matched_rows)))


def _merge_source_baseline_with_working(current: pd.DataFrame, baseline: pd.DataFrame) -> pd.DataFrame:
    """Preserve current research while adding unmatched source records and marking source matches."""
    current = normalize_workflow(current.copy()) if isinstance(current, pd.DataFrame) else normalize_workflow(pd.DataFrame(columns=INTERNAL_COLUMNS))
    baseline = normalize_workflow(baseline.copy())

    if current.empty:
        return classify_discovery_status(ensure_ids(baseline), baseline)

    out = current.copy()
    fill_fields = [
        "Management/Owner", "Street Address", "Address Line 2", "City", "Province",
        "Postal Code", "Country", "Phone", "Primary Email", "Secondary Email",
        "Website", "Number of Apartments", "Number of Storeys",
        "Property Type", "Building Classification", "Source URL",
    ]

    current_key_map = {}
    for idx, row in out.iterrows():
        for key in _discovery_keys_for_row(row):
            current_key_map.setdefault(key, []).append(idx)

    rows_to_add = []
    for _, source_row in baseline.iterrows():
        source_keys = _discovery_keys_for_row(source_row)
        matches = []
        for key in source_keys:
            matches.extend(current_key_map.get(key, []))
        matches = list(dict.fromkeys(matches))

        if matches:
            target = matches[0]
            out.at[target, "Directory Discovery Status"] = "Existing Source Record"
            if is_unresolved(out.at[target, "Directory Entry Status"]):
                out.at[target, "Directory Entry Status"] = "Not Entered"
            for field in fill_fields:
                if field in out.columns and field in source_row.index:
                    if is_unresolved(out.at[target, field]) and not is_unresolved(source_row.get(field)):
                        out.at[target, field] = source_row.get(field)
            if is_unresolved(out.at[target, "Company ID"]):
                out.at[target, "Company ID"] = source_row.get("Company ID")
        else:
            rows_to_add.append(source_row)

    if rows_to_add:
        out = pd.concat([out, pd.DataFrame(rows_to_add)], ignore_index=True, sort=False)

    # Force any stale origin label to be reconsidered against this source baseline.
    out["Directory Discovery Status"] = out["Directory Discovery Status"].replace(
        {"Existing Client Record": "Existing Source Record"}
    )
    out = ensure_ids(normalize_workflow(out))
    return classify_discovery_status(out, baseline)


def import_source_baseline_workbook(uploaded, assignment_sheet: str = "") -> dict:
    """Replace the project's Starting Data with one current baseline.

    Previous source files/baselines are removed from project state. Existing research
    records are preserved and reclassified against the new baseline.
    """
    data = uploaded.getvalue()
    source_hash = hashlib.sha256(data).hexdigest()
    current_registry = st.session_state.get(S_COMPANIES, empty_company_registry())

    project_source, registry, rules, mapping, building_sheet = _source_baseline_from_workbook(
        data, assignment_sheet, current_registry
    )

    previous_versions = _source_versions_state()
    replaced_existing = bool(previous_versions or st.session_state.get(S_SOURCE_BASELINE_META))

    relevant = (
        _source_records_for_project_companies(project_source, registry)
        if not project_source.empty
        else pd.DataFrame()
    )
    source_mode = (
        "Structured records + original file"
        if not project_source.empty
        else "Original project source file"
    )

    meta = {
        "version_number": 1,
        "version_label": "Current",
        "is_original": True,
        "is_active": True,
        "workbook_name": uploaded.name,
        "assignment_sheet": safe_text(assignment_sheet),
        "building_sheet": safe_text(building_sheet),
        "imported_at": datetime.now().isoformat(timespec="seconds"),
        "assigned_companies": len(registry),
        "source_records": len(project_source),
        "project_company_source_records": len(relevant),
        "classification_rules": len(rules),
        "source_hash": source_hash,
        "source_mode": source_mode,
        "raw_source_available": True,
        "structured_parse_version": SOURCE_BASELINE_PARSE_VERSION,
        "structured_reparse_status": "fresh import",
    }

    current_source = {
        "version_number": 1,
        "version_label": "Current",
        "is_original": True,
        "is_active": True,
        "meta": dict(meta),
        "records": project_source.copy(),
        "rules": rules.copy(),
        "raw_bytes": bytes(data),
        "raw_filename": uploaded.name,
    }

    # One project = one current Starting Data source. Replacing it removes all older source files.
    st.session_state[S_SOURCE_VERSIONS] = [current_source]

    # Starting Data remains a comparison baseline; do not inject source rows into research results.
    current = st.session_state.get(S_WORKING, pd.DataFrame(columns=INTERNAL_COLUMNS))
    research_records = ensure_ids(normalize_workflow(current.copy()))
    research_records = classify_discovery_status(
        research_records,
        project_source if not project_source.empty else None,
    )

    st.session_state[S_ORIGINAL] = project_source.copy()
    st.session_state[S_WORKING] = research_records
    st.session_state[S_COMPANIES] = normalize_company_registry(registry)
    st.session_state[S_MAPPING] = mapping if isinstance(mapping, pd.DataFrame) else pd.DataFrame()
    st.session_state[S_SOURCE_TYPE] = "Current project source file"
    st.session_state[S_SOURCE_REF] = uploaded.name
    st.session_state[S_SHEET] = safe_text(assignment_sheet)
    st.session_state[S_CLASSIFICATION_RULES] = rules
    st.session_state[S_SOURCE_BASELINE_META] = dict(meta)
    # A frozen report contains a copy of the previous source baseline.  Once the
    # Starting Data changes, that report is no longer valid for source counts.
    st.session_state.pop(S_REPORTING_SNAPSHOT, None)
    st.session_state[S_PROJECT_LOADED] = True

    if st.session_state.get(S_FILE) in {None, "", "blank-workspace"}:
        st.session_state[S_FILE] = "project-source:" + source_hash
        st.session_state[S_NAME] = uploaded.name

    if not registry.empty:
        active_id = safe_text(st.session_state.get(S_ACTIVE_COMPANY, ""))
        if active_id not in set(registry["Company ID"].astype(str)):
            st.session_state[S_ACTIVE_COMPANY] = registry.iloc[0]["Company ID"]

    autosave_current_project()
    return {
        "assigned_companies": len(registry),
        "source_records": len(project_source),
        "project_company_source_records": len(relevant),
        "working_records": len(research_records),
        "classification_rules": len(rules),
        "version_number": 1,
        "version_label": "Current",
        "source_mode": source_mode,
        "duplicate_version": False,
        "replaced_existing": replaced_existing,
    }

def load_google(url, selector="", force=False):
    df, data, name, sheet = read_google_sheet(url, selector)
    signature = f"{name}:{sheet}:{hashlib.sha256(data).hexdigest()}"
    if not force and st.session_state.get(S_FILE) == signature:
        st.session_state[S_FLASH] = "This Google Sheet is already open. Your session edits were kept."
        return False
    if force:
        st.session_state.pop(S_FILE, None)

    if looks_like_company_assignment(df):
        open_assignment_project(
            df,
            data,
            name,
            sheet,
            "Google Sheet assignment",
            str(url).strip(),
            str(selector).strip(),
        )
    else:
        validate_input(df)
        mapped, mapping = map_schema(df)
        open_workspace(
            mapped,
            mapping,
            signature,
            name,
            sheet,
            "Google Sheet building file",
            str(url).strip(),
            str(selector).strip(),
            (
                "Opened the Google Sheet as a building-record working copy. "
                "The original Sheet is never edited."
            ),
        )
    return True


def blank_workspace():
    df = normalize_workflow(pd.DataFrame(columns=INTERNAL_COLUMNS))
    mapping = pd.DataFrame({"Datablix Field": INTERNAL_COLUMNS, "Imported Column(s)": INTERNAL_COLUMNS, "Mapping Status": "Template field"})
    st.session_state.pop(S_FILE, None)
    st.session_state[S_COMPANIES] = empty_company_registry()
    st.session_state[S_ACTIVE_COMPANY] = ""
    st.session_state[S_PROJECT_NAME] = "Datablix master project"
    open_workspace(
        df, mapping, "blank-workspace", "datablix_rental_property_research.csv",
        "", "Blank project", message="Created a blank Datablix project. Add a company before starting research.",
        registry=empty_company_registry(),
    )


def create_manual_project(
    project_name: str,
    first_company: str = "",
    company_website: str = "",
    company_notes: str = "",
):
    """Register a project manually and optionally add its first company."""
    clean_project_name = re.sub(r"\s+", " ", str(project_name or "")).strip()
    if not clean_project_name:
        raise ValueError("Enter a project name.")

    blank_workspace()
    st.session_state[S_PROJECT_NAME] = clean_project_name
    st.session_state[S_NAME] = clean_project_name
    st.session_state[S_SOURCE_TYPE] = "Manually registered project"
    st.session_state[S_FILE] = (
        "manual-project:"
        + hashlib.sha256(
            f"{clean_project_name}|{datetime.now().isoformat()}".encode("utf-8")
        ).hexdigest()
    )

    company_id = ""
    created = False
    clean_company = re.sub(r"\s+", " ", str(first_company or "")).strip()
    if clean_company:
        company_id, created = add_company_to_project(
            clean_company,
            str(company_website or "").strip(),
            "Initial assignment",
            str(company_notes or "").strip(),
        )

    st.session_state[S_FLASH] = (
        f"Registered {clean_project_name} and added {clean_company} as {company_id}."
        if clean_company and created
        else f"Registered {clean_project_name}. Add or select a company to begin research."
    )
    return company_id


def start_demo_workspace() -> None:
    """Load an editable, session-only rental property demonstration."""
    create_manual_project("Ottawa Rental Property Research Demo")
    st.session_state[S_DEMO_MODE] = True
    st.session_state[S_PROJECT_ROLE] = "owner"
    st.session_state.pop(S_CLOUD_PROJECT_ID, None)
    st.session_state.pop(S_CLOUD_STATE_HASH, None)

    companies = [
        ("North River Property Management", "https://example.com/north-river", "Initial assignment"),
        ("Capital Key Apartments", "https://example.com/capital-key", "Initial assignment"),
        ("Maple Court Residential", "https://example.com/maple-court", "Added later"),
    ]
    company_ids = {}
    for name, website, scope in companies:
        cid, _ = add_company_to_project(name, website, scope, "Fictional company used for the Datablix demonstration.")
        company_ids[name] = cid

    today = date.today().isoformat()
    rows = [
        {"Building Name":"Riverside Place", "Street Address":"120 Demo Street", "City":"Ottawa", "Province":"Ontario", "Postal Code":"K1A 0A1", "Management/Owner":"North River Property Management", "Company ID":company_ids["North River Property Management"], "Phone":"613-555-0101", "Primary Email":"leasing@example.com", "Website":"https://example.com/north-river/riverside", "Number of Apartments":84, "Source URL":"https://example.com/north-river/riverside", "Date Researched":today, "Researcher":"Demo Researcher", "Verification Status":"Verified", "Research Status":"Completed", "Record Decision":"Keep"},
        {"Building Name":"Capital View Apartments", "Street Address":"245 Sample Avenue", "City":"Ottawa", "Province":"Ontario", "Postal Code":"K1B 2B2", "Management/Owner":"Capital Key Apartments", "Company ID":company_ids["Capital Key Apartments"], "Phone":"613-555-0112", "Website":"https://example.com/capital-key/capital-view", "Number of Apartments":126, "Source URL":"https://example.com/capital-key/capital-view", "Date Researched":today, "Researcher":"Demo Researcher", "Verification Status":"Needs Review", "Research Status":"In Progress", "Record Decision":"Needs Review"},
        {"Building Name":"Maple Court", "Street Address":"88 Research Road", "City":"Ottawa", "Province":"Ontario", "Postal Code":"K1C 3C3", "Management/Owner":"Maple Court Residential", "Company ID":company_ids["Maple Court Residential"], "Phone":"613-555-0124", "Primary Email":"contact@example.com", "Website":"https://example.com/maple-court", "Number of Apartments":48, "Source URL":"https://example.com/maple-court", "Date Researched":today, "Researcher":"Demo Researcher", "Verification Status":"Verified", "Research Status":"Completed", "Record Decision":"Keep"},
        {"Building Name":"Capital View Apartments", "Street Address":"245 Sample Ave", "City":"Ottawa", "Province":"ON", "Postal Code":"K1B2B2", "Management/Owner":"Capital Key Apartments", "Company ID":company_ids["Capital Key Apartments"], "Website":"https://example.com/capital-key", "Source URL":"", "Date Researched":today, "Researcher":"Demo Researcher", "Verification Status":"Not Verified", "Research Status":"Needs Follow-up", "Record Decision":"Possible Duplicate"},
    ]
    frame = prepare_data(pd.DataFrame(rows))
    for column in INTERNAL_COLUMNS:
        if column not in frame.columns:
            frame[column] = pd.NA
    frame = ensure_ids(normalize_workflow(frame))
    st.session_state[S_WORKING] = frame
    st.session_state[S_ORIGINAL] = frame.copy()
    st.session_state[S_ACTIVE_COMPANY] = company_ids["North River Property Management"]
    st.session_state[S_SOURCE_TYPE] = "Demo workspace"
    st.session_state[S_SOURCE_REF] = "Fictional sample rental property information"
    st.session_state[S_FLASH] = "Demo workspace opened. Changes are temporary and are not saved."
    st.session_state["db_section"] = "Research projects & companies"


def return_to_project_start() -> None:
    """Leave the active project without deleting its permanent cloud copy.

    Authentication is preserved so the user can immediately choose another
    saved project or create a new one.
    """
    preserved = {
        key: st.session_state.get(key)
        for key in [
            S_AUTH_USER_ID,
            S_AUTH_EMAIL,
            S_AUTH_ACCESS_TOKEN,
            S_AUTH_REFRESH_TOKEN,
        ]
        if st.session_state.get(key) not in (None, "")
    }

    clear_autosaved_project()

    prefixes = ("db_", "website_scan", "full_scan")
    for key in list(st.session_state.keys()):
        if str(key).startswith(prefixes):
            st.session_state.pop(key, None)

    for key, value in preserved.items():
        st.session_state[key] = value

    if preserved:
        # Prevent the project we just left from being restored automatically.
        # The project-start screen can still list all accessible cloud projects.
        st.session_state[S_SKIP_CLOUD_RESTORE] = True


def generate_id(df):
    existing = set(resolved(df["Record ID"]).dropna().astype(str).str.strip())
    n = 1
    while f"DB-NEW-{n:03d}" in existing: n += 1
    return f"DB-NEW-{n:03d}"


def save_edits(edited, columns):
    """Save review edits against stable Record IDs, not only dataframe positions.

    Review tables are filtered views of the working data. Using Record ID as the
    primary key prevents an edit from being written to the wrong row when a view
    has been filtered, re-ordered, or rebuilt during a Streamlit rerun.
    """
    working = st.session_state[S_WORKING].copy()
    editable_columns = [c for c in columns if c in edited.columns and c in working.columns]

    working_ids = (
        working["Record ID"].astype("string").fillna("").str.strip()
        if "Record ID" in working.columns
        else pd.Series("", index=working.index, dtype="string")
    )
    id_counts = working_ids[working_ids.ne("")].value_counts()
    id_to_index = {
        record_id: index
        for index, record_id in working_ids.items()
        if record_id and int(id_counts.get(record_id, 0)) == 1
    }

    for edited_index, edited_row in edited.iterrows():
        target_index = None
        record_id = str(edited_row.get("Record ID", "") or "").strip()
        if record_id:
            target_index = id_to_index.get(record_id)

        # Safe fallback for legacy rows or duplicate/missing IDs.
        if target_index is None and edited_index in working.index:
            target_index = edited_index
        if target_index is None:
            continue

        previous_discovery_status = safe_text(
            working.at[target_index, "Directory Discovery Status"]
            if "Directory Discovery Status" in working.columns
            else ""
        )

        for column in editable_columns:
            working.at[target_index, column] = edited_row[column]

        # If a reviewer explicitly changes discovery status, remember that choice.
        # Choosing Needs Classification hands the row back to automatic logic; any
        # resolved classification becomes a manual override and survives reruns.
        if "Directory Discovery Status" in editable_columns:
            new_discovery_status = safe_text(working.at[target_index, "Directory Discovery Status"])
            if new_discovery_status != previous_discovery_status:
                working.at[target_index, "Discovery Status Source"] = (
                    "Automatic"
                    if new_discovery_status == "Needs Classification"
                    else "Manual"
                )

    working["Province"] = working["Province"].apply(canonical_province)
    working["Postal Code"] = working["Postal Code"].apply(postal_code)
    st.session_state[S_WORKING] = normalize_workflow(prepare_data(working))
    st.session_state[S_EDIT_COUNT] = st.session_state.get(S_EDIT_COUNT, 0) + 1
    refreshed_qa = qa_checks(st.session_state[S_WORKING].copy())
    approved_count = int(approved_for_export_mask(refreshed_qa).sum())
    st.session_state[S_FLASH] = (
        f"Changes saved. Quality checks refreshed; {approved_count:,} record(s) "
        "are currently Approved for Export."
    )


def update_directory_entry_status(record_id: str, status: str) -> bool:
    """Update one approved record's final directory-entry tracking status."""
    if status not in DIRECTORY_ENTRY_STATUSES:
        return False
    working = st.session_state.get(S_WORKING)
    if not isinstance(working, pd.DataFrame) or working.empty:
        return False
    mask = working["Record ID"].astype(str).eq(str(record_id))
    if int(mask.sum()) != 1:
        return False
    working.loc[mask, "Directory Entry Status"] = status
    st.session_state[S_WORKING] = normalize_workflow(working)
    st.session_state[S_EDIT_COUNT] = st.session_state.get(S_EDIT_COUNT, 0) + 1
    autosave_current_project()
    return True


def directory_entry_record_label(row) -> str:
    """Create a compact navigation label for the directory-entry assistant."""
    parts = []
    for field in ["Building Name", "Street Address", "City"]:
        value = row.get(field, "")
        if not is_unresolved(value):
            parts.append(str(value).strip())
    return " · ".join(parts) if parts else str(row.get("Record ID", "Record"))


# =========================================================
# Interface
# =========================================================


def go_to(section_name: str) -> None:
    """Move to another primary area on the next Streamlit rerun."""
    st.session_state["db_section"] = section_name


def render_page_heading(label: str, title: str, description: str) -> None:
    """Render a consistent, accessible page introduction."""
    st.markdown(
        f"""
        <section class="db-page-head" aria-label="{escape(title)}">
            <div class="db-eyebrow">{escape(label)}</div>
            <h2>{escape(title)}</h2>
            <p>{escape(description)}</p>
        </section>
        """,
        unsafe_allow_html=True,
    )


def render_guidance(title: str, message: str) -> None:
    """Place short decision-support copy beside the task it explains."""
    st.markdown(
        f'<div class="db-guidance"><strong>{escape(title)}</strong>'
        f'<span>{escape(message)}</span></div>',
        unsafe_allow_html=True,
    )


def smart_expander(
    title: str,
    *,
    count: int | None = None,
    status: str = "",
    expanded: bool = False,
):
    """Create a consistently labelled progressive-disclosure section.

    Keep context, KPIs, warnings and next actions visible. Use this helper for
    supporting evidence, long tables, optional tools, settings and history.
    Streamlit still executes collapsed content, so this improves information
    hierarchy rather than computation time.
    """
    details = []
    if count is not None:
        try:
            details.append(f"{int(count):,}")
        except (TypeError, ValueError):
            details.append(str(count))
    if str(status or "").strip():
        details.append(str(status).strip())
    label = title + (" · " + " · ".join(details) if details else "")
    return st.expander(label, expanded=bool(expanded))


def _review_company(qa_frame: pd.DataFrame) -> tuple[str | None, pd.DataFrame]:
    """Choose the company whose records and live quality results are shown in Review."""
    registry = normalize_company_registry(st.session_state.get(S_COMPANIES))
    if qa_frame is None or qa_frame.empty or registry.empty:
        return None, qa_frame.iloc[0:0].copy() if isinstance(qa_frame, pd.DataFrame) else pd.DataFrame()

    available = registry.loc[
        registry["Company ID"].astype(str).isin(set(qa_frame["Company ID"].astype(str)))
    ].copy()
    if available.empty:
        return None, qa_frame.iloc[0:0].copy()

    company_ids = available["Company ID"].astype(str).tolist()
    active_id = str(st.session_state.get(S_ACTIVE_COMPANY, "")).strip()
    default_index = company_ids.index(active_id) if active_id in company_ids else 0
    selected_id = st.selectbox(
        "Company for review",
        company_ids,
        index=default_index,
        format_func=lambda company_id: company_label(
            available.loc[available["Company ID"].eq(company_id)].iloc[0]
        ),
        key="db_review_company",
        help="Select the company whose records you want to review and verify.",
    )
    selected_qa = qa_frame.loc[
        qa_frame["Company ID"].astype(str).eq(str(selected_id))
    ].copy()
    return str(selected_id), selected_qa




def render_review_progress(qa_frame: pd.DataFrame, company_id: str | None) -> None:
    """Show live review progress and a direct path to exporting approved records."""
    if not company_id:
        return
    company_qa = qa_frame.loc[
        qa_frame["Company ID"].astype(str).eq(str(company_id))
    ].copy()
    if company_qa.empty:
        return

    st.divider()
    st.markdown("### Review progress")
    st.caption(
        "Use one clear finish line: Approved for Export. A record reaches it after Completed + Verified + Keep with no critical data blocker."
    )

    total_records = len(company_qa)
    approved_count = int(approved_for_export_mask(company_qa).sum())
    excluded_count = int(company_qa["Record Decision"].eq("Remove").sum())
    still_reviewing = int(
        (~approved_for_export_mask(company_qa) & ~company_qa["Record Decision"].eq("Remove")).sum()
    )

    live_metrics = st.columns(4)
    live_metrics[0].metric("Total records", f"{total_records:,}")
    live_metrics[1].metric("Approved for Export", f"{approved_count:,}")
    live_metrics[2].metric("Still in review", f"{still_reviewing:,}")
    live_metrics[3].metric("Excluded", f"{excluded_count:,}")

    if approved_count:
        st.success(
            f"{approved_count:,} of {total_records:,} record(s) are approved for export for this company."
        )
        if st.button(
            f"Export {approved_count:,} approved record(s)",
            type="primary",
            width="stretch",
            key=f"db_review_export_approved_{company_id}",
        ):
            st.session_state["db_export_scope_mode"] = "One company"
            st.session_state["db_export_company"] = str(company_id)
            st.session_state["db_custom_export_scope"] = "Approved for Export"
            go_to("Downloads")
            st.rerun()
    else:
        st.info(
            "No records are approved for export yet. Finish the review fields for the records you want to deliver."
        )

    st.markdown("#### Current quality")
    quality_metrics = st.columns(4)
    quality_metrics[0].metric("Critical issues", f"{int(company_qa['QA Status'].eq('Critical').sum()):,}")
    quality_metrics[1].metric("Warnings", f"{int(company_qa['Warning Count'].sum()):,}")
    quality_metrics[2].metric("Open research gaps", f"{int(company_qa['Research Gap Count'].sum()):,}")
    quality_metrics[3].metric("Human verified", f"{int(company_qa['Verification Status'].eq('Verified').sum()):,}")


    quality_exception_count = (
        int(company_qa["QA Status"].eq("Critical").sum())
        + int(company_qa["Warning Count"].sum())
        + int(company_qa["Research Gap Count"].sum())
    )
    with smart_expander(
        "Quality details",
        count=quality_exception_count,
        status="open signals",
        expanded=False,
    ):
        detail_tabs = st.tabs(["Current issues", "Research progress", "Field coverage"])
        with detail_tabs[0]:
            issues = issue_summary(company_qa)
            if issues.empty:
                st.success("No data-quality issues are currently flagged for this company.")
            else:
                st.dataframe(issues, width="stretch", hide_index=True)
            attention_columns = [
                "Record ID", "Working Record Label", "QA Status", "QA Flags",
                "Research Gaps", "Follow-up Priority", "Record Readiness", "Export Status",
            ]
            attention_columns = [c for c in attention_columns if c in company_qa.columns]
            attention = company_qa[
                ~approved_for_export_mask(company_qa)
                & ~company_qa["Record Decision"].eq("Remove")
            ][attention_columns]
            st.markdown("#### Records still in review")
            st.dataframe(attention, width="stretch", hide_index=True, height=360)
        with detail_tabs[1]:
            st.dataframe(research_log(company_qa).head(250), width="stretch", hide_index=True, height=460)
        with detail_tabs[2]:
            st.dataframe(field_coverage(company_qa), width="stretch", hide_index=True)


def render_report_navigation(active_section: str) -> None:
    """Keep analysis and saving inside one understandable report stage."""
    columns = st.columns(2)
    options = [
        ("Analysis & report", "Analysis & report"),
        ("Downloads", "Export"),
    ]
    for column, (section_name, label) in zip(columns, options):
        with column:
            if st.button(
                label,
                type="primary" if active_section == section_name else "secondary",
                width="stretch",
                key=f"db_report_subnav_{norm_header(section_name)}",
            ):
                go_to(section_name)
                st.rerun()


def recommended_next_action(qa_frame: pd.DataFrame | None) -> tuple[str, str, str, str]:
    """Return a practical next action based on the current workspace."""
    if qa_frame is None or qa_frame.empty:
        return (
            "Begin company research",
            "Select a registered company, scan its permitted public website, or add a building manually.",
            "Research projects & companies",
            "Choose company and method",
        )

    critical_count = int(qa_frame["QA Status"].eq("Critical").sum())
    review_count = int(qa_frame["Verification Status"].eq("Needs Review").sum())
    follow_up_count = int(qa_frame["Record Readiness"].isin([
        "Duplicate Review", "Needs Follow-up", "Fix Critical Data"
    ]).sum())
    approved_count = int(approved_for_export_mask(qa_frame).sum())

    if critical_count:
        return (
            "Fix critical records first",
            f"{critical_count:,} record(s) are missing core identity details or carry a critical conflict.",
            "Review records",
            "Fix critical records",
        )
    if follow_up_count:
        return (
            "Clear the high-priority follow-ups",
            f"{follow_up_count:,} record(s) need a duplicate decision, a source follow-up, or a key correction.",
            "Review records",
            "Open review queue",
        )
    if review_count:
        return (
            "Verify the reviewed candidates",
            f"{review_count:,} record(s) are waiting for a human verification decision.",
            "Review records",
            "Verify candidates",
        )
    active_count = int((~qa_frame["Record Decision"].eq("Remove")).sum())
    if approved_count < active_count:
        return (
            "Check progress and remaining gaps",
            "See which research is incomplete, which details are missing, and how fresh each source is.",
            "Review records",
            "Review quality progress",
        )
    return (
        "Export your selected columns",
        "Every active record is approved for export. Choose the fields you need and download a CSV.",
        "Downloads",
        "Open custom export",
    )



def _company_records_for_progress(
    records: pd.DataFrame,
    *,
    company_id: str,
    company_name: str,
    company_website: str = "",
) -> pd.DataFrame:
    """Return records linked to a company, repairing legacy linkage gaps defensively.

    Company ID remains the primary key. Older saved projects can contain valid
    Hazelview research rows under a blank, obsolete, or URL-derived Company ID.
    When the exact ID produces no rows, Datablix safely falls back to the canonical
    owner name and official company domain so completed work is not displayed as
    Not started.
    """
    if not isinstance(records, pd.DataFrame) or records.empty:
        return pd.DataFrame(columns=INTERNAL_COLUMNS)

    frame = records.copy()
    for column in ["Company ID", "Management/Owner", "Company Website", "Website", "Source URL"]:
        if column not in frame.columns:
            frame[column] = pd.NA

    clean_id = safe_text(company_id)
    if clean_id:
        exact = frame.loc[
            frame["Company ID"].fillna("").astype(str).str.strip().eq(clean_id)
        ].copy()
        if not exact.empty:
            return exact

    target_name = company_name_key(company_name)
    if target_name:
        by_name = frame.loc[
            frame["Management/Owner"].apply(company_name_key).eq(target_name)
        ].copy()
        if not by_name.empty:
            return by_name

    target_domain = _company_domain_key(company_website)
    if target_domain:
        domain_mask = pd.Series(False, index=frame.index, dtype="bool")
        for column in ["Company Website", "Website", "Source URL"]:
            domain_mask |= frame[column].apply(_company_domain_key).eq(target_domain)
        by_domain = frame.loc[domain_mask].copy()
        if not by_domain.empty:
            return by_domain

    return frame.iloc[0:0].copy()


def company_progress_snapshot(company_row: pd.Series, records: pd.DataFrame) -> dict:
    """Return a user-facing progress model without confusing exclusion with no work.

    All linked records count as research activity. Only directory-eligible records
    are used for verification and export-readiness calculations. This distinction
    prevents a fully researched company whose records were excluded from appearing
    as ``Not started``.
    """
    company_id = str(company_row.get("Company ID", "")).strip()
    company_name = str(company_row.get("Management/Owner", "")).strip() or "Unnamed company"
    website = str(company_row.get("Main Website", "")).strip()
    stored_status = str(company_row.get("Company Status", "Not started")).strip()

    group = _company_records_for_progress(
        records,
        company_id=company_id,
        company_name=company_name,
        company_website=website,
    )

    if group.empty:
        collected = active_count = excluded = reviewed = verified = ready = 0
        attention = critical = follow_up = 0
        progress = 0.0
        display_status = stored_status if stored_status in COMPANY_STATUSES else "Not started"
    else:
        qa_columns = {"Record Readiness", "QA Status", "Warning Count"}
        company_qa = group.copy() if qa_columns.issubset(group.columns) else qa_checks(group)

        # Research count includes every linked record, including excluded rows.
        collected = len(company_qa)
        excluded_mask = company_qa["Record Readiness"].eq("Excluded from Listings")
        excluded = int(excluded_mask.sum())
        active = company_qa.loc[~excluded_mask].copy()
        active_count = len(active)

        # Review completion is measured across all researched rows. An explicitly
        # excluded row has already received a human/workflow decision.
        reviewed_mask_all = (
            company_qa["Research Status"].eq("Completed")
            | company_qa["Verification Status"].eq("Verified")
            | company_qa["Record Decision"].isin(
                ["Keep", "Update", "Possible Duplicate", "Remove"]
            )
            | excluded_mask
        )
        reviewed = int(reviewed_mask_all.sum())

        # Verification/readiness apply only to eligible directory records.
        verified = int(active["Verification Status"].eq("Verified").sum()) if active_count else 0
        ready = int(ready_mask(active).sum()) if active_count else 0
        critical = int(active["QA Status"].eq("Critical").sum()) if active_count else 0
        follow_up_mask = active["Record Readiness"].isin(
            ["Duplicate Review", "Needs Follow-up", "Fix Critical Data", "Needs Data Review", "Needs Update"]
        ) if active_count else pd.Series(False, index=active.index, dtype="bool")
        follow_up = int(follow_up_mask.sum())
        attention_mask = (~ready_mask(active)) & (
            active["QA Status"].isin(["Critical", "Review"]) | follow_up_mask
        ) if active_count else pd.Series(False, index=active.index, dtype="bool")
        attention = int(attention_mask.sum())
        progress = verified / active_count if active_count else 0.0

        explicit_complete = stored_status in {"Complete", "Complete with limitations"}
        all_research_reviewed = reviewed == collected
        calculated_complete = active_count > 0 and verified == active_count and attention == 0

        if active_count == 0:
            # Research exists, but no row is eligible for the directory.
            display_status = (
                "Complete with limitations"
                if explicit_complete or all_research_reviewed
                else "Researching"
            )
        elif explicit_complete:
            display_status = stored_status
        elif calculated_complete:
            display_status = "Complete"
        elif critical or follow_up:
            display_status = "Needs attention"
        elif active["Research Status"].isin(
            ["Imported - Needs Review", "Not Started", "In Progress"]
        ).any():
            display_status = "Researching"
        elif verified < active_count:
            display_status = "Ready for QA"
        else:
            display_status = "Researching"

    unverified = max(active_count - verified, 0)
    if not website and collected == 0:
        next_title = "Add the company website"
        next_copy = "Register the official website, or add a known building manually."
        next_section = "Research projects & companies"
        next_button = "Add website"
    elif collected == 0:
        next_title = "Start company research"
        next_copy = "Scan the public website or register the first building manually."
        next_section = "Website scanner"
        next_button = "Start research"
    elif active_count == 0 and excluded:
        next_title = "Research completed with no eligible listings"
        next_copy = f"All {excluded:,} researched record(s) are excluded from the directory. Review the exclusion evidence if needed."
        next_section = "Review records"
        next_button = "Review exclusions"
    elif attention:
        next_title = "Resolve records needing attention"
        next_copy = f"Review {attention:,} active record(s) with missing details, evidence, or decisions."
        next_section = "Review records"
        next_button = "Review & quality"
    elif unverified:
        next_title = "Complete human verification"
        next_copy = f"Verify the remaining {unverified:,} eligible record(s)."
        next_section = "Review records"
        next_button = "Verify records"
    elif display_status in {"Complete", "Complete with limitations"}:
        next_title = "Company research is complete"
        next_copy = "Review the project summary or continue with another company."
        next_section = "Analysis & report"
        next_button = "View project report"
    else:
        next_title = "Continue company research"
        next_copy = "Review the collected records and document any remaining gaps."
        next_section = "Review records"
        next_button = "Continue research"

    return {
        "company_id": company_id,
        "company_name": company_name,
        "website": website,
        "stored_status": stored_status,
        "status": display_status,
        "collected": collected,
        "active_count": active_count,
        "excluded": excluded,
        "reviewed": reviewed,
        "verified": verified,
        "ready": ready,
        "attention": attention,
        "critical": critical,
        "follow_up": follow_up,
        "progress": progress,
        "progress_percent": int(round(progress * 100)),
        "complete": display_status in {"Complete", "Complete with limitations"},
        "next_title": next_title,
        "next_copy": next_copy,
        "next_section": next_section,
        "next_button": next_button,
    }


def project_progress_snapshot(registry: pd.DataFrame, records: pd.DataFrame) -> dict:
    """Summarize company completion and record health for the active project."""
    registry = normalize_company_registry(registry)
    if isinstance(records, pd.DataFrame) and not records.empty:
        qa_columns = {"Record Readiness", "QA Status", "Warning Count"}
        qa_records = (
            records.copy()
            if qa_columns.issubset(records.columns)
            else qa_checks(records)
        )
    else:
        qa_records = pd.DataFrame(columns=INTERNAL_COLUMNS)

    rows = [
        company_progress_snapshot(company, qa_records)
        for _, company in registry.iterrows()
    ]
    total_companies = len(rows)
    completed = sum(row["complete"] for row in rows)
    not_started = sum(row["status"] == "Not started" for row in rows)
    needs_attention = sum(row["status"] == "Needs attention" for row in rows)
    in_progress = max(total_companies - completed - not_started, 0)

    if not qa_records.empty:
        active_qa = qa_records.loc[
            ~qa_records["Record Readiness"].eq("Excluded from Listings")
        ]
        buildings = len(active_qa)
        verified_records = int(active_qa["Verification Status"].eq("Verified").sum())
        project_follow_up = active_qa["Record Readiness"].isin(
            [
                "Duplicate Review",
                "Needs Follow-up",
                "Fix Critical Data",
                "Needs Data Review",
                "Needs Update",
            ]
        )
        attention_records = int(
            (
                (~ready_mask(active_qa))
                & (
                    active_qa["QA Status"].isin(["Critical", "Review"])
                    | project_follow_up
                )
            ).sum()
        )
    else:
        buildings = verified_records = attention_records = 0

    return {
        "companies": total_companies,
        "completed": completed,
        "not_started": not_started,
        "in_progress": in_progress,
        "needs_attention_companies": needs_attention,
        "buildings": buildings,
        "verified_records": verified_records,
        "attention_records": attention_records,
        "progress": completed / total_companies if total_companies else 0.0,
        "progress_percent": int(round(completed / total_companies * 100)) if total_companies else 0,
        "company_rows": rows,
    }


def company_progress_table(
    registry: pd.DataFrame,
    records: pd.DataFrame,
    snapshot: dict | None = None,
) -> pd.DataFrame:
    """Return a project-home table with one understandable row per company."""
    snapshot = snapshot or project_progress_snapshot(registry, records)
    rows = []
    for item in snapshot["company_rows"]:
        rows.append({
            "Company": item["company_name"],
            "Website": item["website"] or "Not registered",
            "Buildings": item["collected"],
            "Reviewed": item["reviewed"],
            "Verified": item["verified"],
            "Needs attention": item["attention"],
            "Progress": f"{item['progress_percent']}%" if item["collected"] else "Not started",
            "Status": item["status"],
            "Research prompt": (
                "Saved"
                if not registry.loc[registry["Company ID"].astype(str).eq(item["company_id"]), "Research Prompt"].fillna("").astype(str).str.strip().eq("").all()
                else "Not saved (optional)"
            ),
            "Next action": item["next_title"],
            "Company ID": item["company_id"],
        })
    return pd.DataFrame(rows)


def _sidebar_company_rows(company_rows: list[dict], active_company_id: str) -> str:
    """Render all companies as a compact, non-interactive progress list."""
    status_order = {
        "Needs attention": 0,
        "Researching": 1,
        "Ready for review": 2,
        "Not started": 3,
        "Complete": 4,
    }
    ordered = sorted(
        company_rows,
        key=lambda row: (
            0 if row["company_id"] == active_company_id else 1,
            status_order.get(row["status"], 9),
            row["company_name"].lower(),
        ),
    )
    blocks = []
    for row in ordered:
        selected = " selected" if row["company_id"] == active_company_id else ""
        status_class = re.sub(r"[^a-z]+", "-", row["status"].lower()).strip("-")
        progress_label = (
            f"{row['progress_percent']}% verified"
            if row["collected"]
            else "Research not started"
        )
        blocks.append(
            f'<div class="db-company-progress-row{selected}">'
            f'<div class="db-company-progress-head">'
            f'<span class="db-company-progress-name">{escape(row["company_name"])}</span>'
            f'<span class="db-company-status {status_class}">{escape(row["status"])}</span>'
            f'</div>'
            f'<div class="db-company-progress-meta">'
            f'{row["collected"]:,} buildings · {escape(progress_label)}'
            f'</div>'
            f'<div class="db-mini-progress" aria-label="{escape(progress_label)}">'
            f'<span style="width:{row["progress_percent"]}%"></span>'
            f'</div>'
            f'</div>'
        )
    return "".join(blocks)


def _normalize_analytics_records(records) -> pd.DataFrame:
    """Return analytics-ready records regardless of how project data was restored."""
    if records is None:
        frame = pd.DataFrame(columns=INTERNAL_COLUMNS)
    elif isinstance(records, pd.DataFrame):
        frame = records.copy()
    elif isinstance(records, list):
        frame = pd.DataFrame(records)
    elif isinstance(records, dict):
        # A column-oriented dictionary and a single-record dictionary both occur
        # in saved project payloads, so support both forms safely.
        try:
            frame = pd.DataFrame(records)
        except ValueError:
            frame = pd.DataFrame([records])
    else:
        frame = pd.DataFrame(columns=INTERNAL_COLUMNS)

    if frame.empty and len(frame.columns) == 0:
        frame = pd.DataFrame(columns=INTERNAL_COLUMNS)

    return normalize_workflow(frame)


def _current_source_display_details() -> dict:
    """Return the name, version and row count for the active Starting Data baseline."""
    records = current_starting_source_records()
    versions = st.session_state.get(S_SOURCE_VERSIONS, [])
    active = None
    if isinstance(versions, list):
        active = next(
            (
                item for item in reversed(versions)
                if isinstance(item, dict) and bool(item.get("is_active"))
            ),
            None,
        )
        if active is None:
            active = next(
                (item for item in reversed(versions) if isinstance(item, dict)),
                None,
            )

    meta = dict(active.get("meta", {})) if isinstance(active, dict) else {}
    filename = safe_text(
        active.get("raw_filename", "") if isinstance(active, dict) else ""
    ) or safe_text(meta.get("workbook_name", ""))
    version_label = safe_text(
        active.get("version_label", "") if isinstance(active, dict) else ""
    ) or safe_text(meta.get("version_label", "")) or "Current"

    if not filename:
        legacy_meta = st.session_state.get(S_SOURCE_BASELINE_META, {})
        if isinstance(legacy_meta, dict):
            filename = safe_text(legacy_meta.get("workbook_name", ""))
            version_label = (
                safe_text(legacy_meta.get("version_label", ""))
                or version_label
            )

    return {
        "records": records,
        "filename": Path(filename).name if filename else "Current Starting Data",
        "version_label": version_label,
        "label": (
            f"{Path(filename).name if filename else 'Current Starting Data'} · "
            f"{version_label}"
        ),
        "row_count": len(records),
    }


def _analytics_company_reference(row, registry: pd.DataFrame) -> tuple[str, str]:
    """Resolve one source/research row to a stable company key and display name."""
    if not isinstance(registry, pd.DataFrame):
        registry = empty_company_registry()
    company_id = safe_text(row.get("Company ID", ""))
    owner = safe_text(row.get("Management/Owner", ""))

    if company_id and not registry.empty:
        exact = registry.loc[registry["Company ID"].astype(str).eq(company_id)]
        if not exact.empty:
            canonical_name = safe_text(exact.iloc[0].get("Management/Owner", ""))
            # Trust the ID only when the row has no owner label or the owner label
            # agrees with the canonical registry company.  Otherwise treat the ID
            # as stale and resolve the row by name below.
            if not owner or not canonical_name or _company_core_matches(owner, canonical_name):
                return company_id, canonical_name or owner or "Unnamed company"

    if owner and not registry.empty:
        match_index = _registry_match_index(owner, registry)
        if match_index is not None:
            company = registry.loc[match_index]
            resolved_id = safe_text(company.get("Company ID", ""))
            resolved_name = safe_text(company.get("Management/Owner", "")) or owner
            return resolved_id or f"company:{norm_header(resolved_name)}", resolved_name

    if owner:
        return f"unregistered:{norm_header(owner)}", owner
    return "unassigned", "Unassigned source records"


def _active_research_for_comparison(records: pd.DataFrame) -> pd.DataFrame:
    """Exclude records already removed from the current researched portfolio."""
    if not isinstance(records, pd.DataFrame) or records.empty:
        return pd.DataFrame(columns=INTERNAL_COLUMNS)
    frame = normalize_workflow(records)
    excluded = (
        frame["Record Decision"].eq("Remove")
        | frame["Directory Discovery Status"].eq("Excluded / Not Current")
    )
    return frame.loc[~excluded].copy()


def source_reconciliation_snapshot(
    registry: pd.DataFrame,
    records: pd.DataFrame,
    source_records: pd.DataFrame,
) -> dict:
    """Compare current website research with the active Starting Data in both directions.

    Research-to-source classification already identifies existing and new research rows.
    This helper also performs the reverse comparison needed to identify source rows that
    have not yet been matched by any current research record.
    """
    registry = normalize_company_registry(registry)
    research = _normalize_analytics_records(records)
    source = (
        coalesce_duplicate_columns(source_records.copy())
        if isinstance(source_records, pd.DataFrame)
        else pd.DataFrame()
    )

    if source.empty:
        empty_table = pd.DataFrame(columns=[
            "Company ID", "Company", "Source records", "Research records",
            "Matched", "New", "Source-only", "Excluded",
            "Needs classification", "Possible duplicates", "Verified",
            "Net verified change",
        ])
        return {
            "available": False,
            "records": research,
            "active_research": _active_research_for_comparison(research),
            "source_records": 0,
            "research_records": len(_active_research_for_comparison(research)),
            "matched_source": 0,
            "source_only": 0,
            "newly_discovered": 0,
            "needs_classification": 0,
            "possible_duplicates": 0,
            "excluded": 0,
            "verified": 0,
            "company_table": empty_table,
            "status_table": pd.DataFrame(columns=["Reconciliation status", "Records"]),
        }

    # Re-evaluate automatic discovery classifications against the currently active
    # source baseline so the dashboard never reports results against an old file.
    comparison_records = classify_discovery_status(research, source)
    active_research = _active_research_for_comparison(comparison_records)

    source_identities, source_indexes = _build_discovery_source_index(source)
    matched_source_positions: set[int] = set()
    for position, row in enumerate(active_research.to_dict(orient="records")):
        identity = _row_identity(row)
        candidate_positions = _candidate_source_positions(identity, source_indexes)
        best_position = -1
        best_score = 0
        for source_position in candidate_positions:
            score, _reason = _source_match_score_from_identities(
                identity, source_identities[source_position]
            )
            if score > best_score:
                best_position = source_position
                best_score = score
                if best_score >= 100:
                    break

        # A manual Existing Source Record decision can accept a plausible match,
        # but an automatic match still requires the normal strong threshold.
        status = safe_text(row.get("Directory Discovery Status", ""))
        manual_existing = (
            safe_text(row.get("Discovery Status Source", "")) == "Manual"
            and status == "Existing Source Record"
        )
        threshold = 72 if manual_existing else 88
        if best_position >= 0 and best_score >= threshold:
            matched_source_positions.add(best_position)

    company_rows: dict[str, dict] = {}

    def ensure_company(key: str, name: str) -> dict:
        if key not in company_rows:
            company_rows[key] = {
                "Company ID": key,
                "Company": name,
                "Source records": 0,
                "Research records": 0,
                "Matched": 0,
                "New": 0,
                "Source-only": 0,
                "Excluded": 0,
                "Needs classification": 0,
                "Possible duplicates": 0,
                "Verified": 0,
                "Net verified change": 0,
            }
        return company_rows[key]

    # Keep all registered companies visible, including companies with no records yet.
    for _, company in registry.iterrows():
        company_id = safe_text(company.get("Company ID", ""))
        company_name = safe_text(company.get("Management/Owner", "")) or "Unnamed company"
        ensure_company(company_id or f"company:{norm_header(company_name)}", company_name)

    source_records_list = source.to_dict(orient="records")
    for position, row in enumerate(source_records_list):
        key, name = _analytics_company_reference(row, registry)
        company = ensure_company(key, name)
        company["Source records"] += 1
        if position in matched_source_positions:
            company["Matched"] += 1
        else:
            company["Source-only"] += 1

    for row in active_research.to_dict(orient="records"):
        key, name = _analytics_company_reference(row, registry)
        company = ensure_company(key, name)
        company["Research records"] += 1
        status = safe_text(row.get("Directory Discovery Status", ""))
        if status == "Newly Discovered":
            company["New"] += 1
        elif status == "Needs Classification":
            company["Needs classification"] += 1
        elif status == "Possible Duplicate":
            company["Possible duplicates"] += 1
        if safe_text(row.get("Verification Status", "")) == "Verified":
            company["Verified"] += 1

    excluded_research = comparison_records.loc[
        comparison_records["Record Decision"].eq("Remove")
        | comparison_records["Directory Discovery Status"].eq("Excluded / Not Current")
    ]
    for row in excluded_research.to_dict(orient="records"):
        key, name = _analytics_company_reference(row, registry)
        ensure_company(key, name)["Excluded"] += 1

    for company in company_rows.values():
        company["Net verified change"] = (
            company["Verified"] - company["Source records"]
        )

    company_table = pd.DataFrame(company_rows.values())
    if not company_table.empty:
        company_table = company_table.sort_values(
            ["Source-only", "Needs classification", "New", "Company"],
            ascending=[False, False, False, True],
        ).reset_index(drop=True)

    newly_discovered = int(
        active_research["Directory Discovery Status"].eq("Newly Discovered").sum()
    )
    needs_classification = int(
        active_research["Directory Discovery Status"].eq("Needs Classification").sum()
    )
    possible_duplicates = int(
        active_research["Directory Discovery Status"].eq("Possible Duplicate").sum()
    )
    verified = int(active_research["Verification Status"].eq("Verified").sum())
    excluded = len(excluded_research)
    source_only = max(len(source) - len(matched_source_positions), 0)

    status_table = pd.DataFrame({
        "Reconciliation status": [
            "Matched source records",
            "Newly discovered",
            "Source-only / unmatched",
            "Needs classification",
            "Possible duplicates",
            "Excluded / not current",
        ],
        "Records": [
            len(matched_source_positions),
            newly_discovered,
            source_only,
            needs_classification,
            possible_duplicates,
            excluded,
        ],
    })

    return {
        "available": True,
        "records": comparison_records,
        "active_research": active_research,
        "source_records": len(source),
        "research_records": len(active_research),
        "matched_source": len(matched_source_positions),
        "source_only": source_only,
        "newly_discovered": newly_discovered,
        "needs_classification": needs_classification,
        "possible_duplicates": possible_duplicates,
        "excluded": excluded,
        "verified": verified,
        "company_table": company_table,
        "status_table": status_table,
    }


def _analytics_percent(numerator: int | float, denominator: int | float) -> int:
    """Return a whole-number percentage without risking divide-by-zero errors."""
    return round((float(numerator) / float(denominator)) * 100) if denominator else 0


def _render_analytics_header(source_details: dict, reconciliation: dict) -> None:
    """Introduce the dashboard with one clear purpose and the active comparison baseline."""
    baseline = (
        source_details.get("label", "Current Starting Data")
        if reconciliation.get("available")
        else "No active source baseline"
    )
    baseline_note = (
        f"{int(source_details.get('row_count', 0)):,} structured source records"
        if reconciliation.get("available")
        else "Add Starting Data to activate reconciliation analytics"
    )
    st.markdown(
        f"""
        <section class="db-analytics-hero" aria-label="Analytics dashboard introduction">
            <div class="db-analytics-hero-copy">
                <div class="db-analytics-eyebrow">PROJECT INTELLIGENCE</div>
                <h2>Analytics dashboard</h2>
                <p>Understand project progress, compare current research with the active source file, and focus the next review on the records that matter most.</p>
            </div>
            <div class="db-analytics-baseline" title="Active comparison baseline">
                <span>ACTIVE BASELINE</span>
                <strong>{escape(str(baseline))}</strong>
                <small>{escape(str(baseline_note))}</small>
            </div>
        </section>
        """,
        unsafe_allow_html=True,
    )


def _render_analytics_kpis(items: list[dict]) -> None:
    """Render every analytics KPI in one compact horizontal line.

    UX decision:
    - KPI cards never wrap into a second row inside Analytics.
    - On narrow screens the row scrolls horizontally instead of changing the
      information hierarchy or making one metric appear less important.
    - The number of equal-width columns is passed to CSS so every KPI line is
      balanced whether it contains four or five cards.
    """
    if not items:
        return

    cards = []
    for item in items:
        label = escape(str(item.get("label", "Metric")))
        value = escape(str(item.get("value", "0")))
        helper = escape(str(item.get("helper", "")))
        tone = re.sub(r"[^a-z]+", "-", str(item.get("tone", "neutral")).lower()).strip("-")
        icon_map = {
            "Companies": "▦",
            "Source records": "≡",
            "Research records": "⌕",
            "Verified": "✓",
            "Needs attention": "!",
            "Project readiness": "↗",
            "Researched": "⌕",
            "Reviewed": "✓",
            "Source coverage": "↔",
            "Passing QA": "✓",
        }
        icon = escape(icon_map.get(str(item.get("label", "")), "•"))
        cards.append(
            f'<article class="db-analytics-kpi {tone}">'
            f'<div class="db-analytics-kpi-top"><span class="db-analytics-kpi-icon">{icon}</span>'
            f'<div class="db-analytics-kpi-label">{label}</div></div>'
            f'<div class="db-analytics-kpi-value">{value}</div>'
            f'<div class="db-analytics-kpi-helper">{helper}</div>'
            f'</article>'
        )

    count = max(len(cards), 1)
    st.markdown(
        f'<div class="db-analytics-kpi-scroll" role="region" aria-label="Analytics key performance indicators">'
        f'<div class="db-analytics-kpi-grid" style="--db-kpi-count:{count}">'
        + "".join(cards)
        + "</div></div>",
        unsafe_allow_html=True,
    )


def _render_analytics_section(title: str, description: str, eyebrow: str = "") -> None:
    """Create a consistent visual hierarchy before each analytics group."""
    eyebrow_html = (
        f'<div class="db-analytics-section-eyebrow">{escape(eyebrow)}</div>'
        if eyebrow
        else ""
    )
    st.markdown(
        f"""
        <div class="db-analytics-section-head">
            <div>
                {eyebrow_html}
                <h3>{escape(title)}</h3>
                <p>{escape(description)}</p>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _render_analytics_callout(title: str, copy: str, chips: list[str] | None = None, tone: str = "accent") -> None:
    """Surface one concise interpretation instead of leaving users to decode charts alone."""
    chip_html = "".join(
        f'<span class="db-analytics-chip">{escape(str(chip))}</span>'
        for chip in (chips or [])
        if safe_text(chip)
    )
    tone_class = re.sub(r"[^a-z]+", "-", safe_text(tone).lower()).strip("-") or "accent"
    st.markdown(
        f"""
        <aside class="db-analytics-callout {tone_class}">
            <div>
                <strong>{escape(title)}</strong>
                <p>{escape(copy)}</p>
            </div>
            <div class="db-analytics-chip-row">{chip_html}</div>
        </aside>
        """,
        unsafe_allow_html=True,
    )


def _render_chart_empty_state(title: str, message: str) -> None:
    with st.container(border=True):
        st.markdown(f"#### {title}")
        st.markdown(
            f'<div class="db-analytics-empty"><strong>No chart to display</strong><span>{escape(message)}</span></div>',
            unsafe_allow_html=True,
        )


def _render_horizontal_bar_chart(
    data: pd.DataFrame,
    category: str,
    value: str,
    title: str,
    value_title: str = "Records",
    color: str = "#1287CE",
    sort_order: str = "descending",
    description: str = "",
    include_zero: bool = False,
) -> None:
    """Render a readable horizontal ranking chart inside a consistent dashboard panel."""
    if not isinstance(data, pd.DataFrame) or data.empty:
        _render_chart_empty_state(title, "No records are available for this view yet.")
        return

    frame = data[[category, value]].copy()
    frame[category] = frame[category].fillna("Unspecified").astype(str)
    frame[value] = pd.to_numeric(frame[value], errors="coerce").fillna(0)
    if not include_zero:
        frame = frame.loc[frame[value].gt(0)]
    if frame.empty:
        _render_chart_empty_state(title, "All values are currently zero.")
        return

    height = max(180, min(520, 32 * len(frame) + 62))
    with st.container(border=True):
        st.markdown(f"#### {title}")
        if description:
            st.caption(description)
        st.vega_lite_chart(
            frame,
            {
                "height": height,
                "mark": {"type": "bar", "cornerRadiusEnd": 5},
                "encoding": {
                    "y": {
                        "field": category,
                        "type": "nominal",
                        "sort": {"field": value, "order": sort_order},
                        "axis": {"title": None, "labelLimit": 270, "labelPadding": 8},
                    },
                    "x": {
                        "field": value,
                        "type": "quantitative",
                        "axis": {"title": value_title, "tickMinStep": 1, "gridOpacity": 0.18},
                    },
                    "color": {"value": color},
                    "tooltip": [
                        {"field": category, "type": "nominal", "title": category},
                        {"field": value, "type": "quantitative", "format": ",.0f", "title": value_title},
                    ],
                },
                "config": {
                    "view": {"stroke": None},
                    "axis": {"labelFontSize": 12, "titleFontSize": 12},
                },
            },
            use_container_width=True,
        )


def _render_grouped_company_chart(data: pd.DataFrame, title: str, description: str = "") -> None:
    """Compare source and research counts with a compact dumbbell chart.

    A dumbbell chart makes the direction and size of each company-level gap
    easier to read than two adjacent bars while using less visual ink.
    Grey represents the current source baseline and blue represents current
    research throughout the dashboard.
    """
    if not isinstance(data, pd.DataFrame) or data.empty:
        _render_chart_empty_state(title, "No source or research records are available yet.")
        return

    frame = data[["Company", "Source records", "Research records"]].copy()
    frame["Source records"] = pd.to_numeric(frame["Source records"], errors="coerce").fillna(0)
    frame["Research records"] = pd.to_numeric(frame["Research records"], errors="coerce").fillna(0)
    frame = frame.loc[frame[["Source records", "Research records"]].sum(axis=1).gt(0)]
    if frame.empty:
        _render_chart_empty_state(title, "No source or research records are available yet.")
        return

    frame["Gap"] = frame["Research records"] - frame["Source records"]
    frame["Minimum"] = frame[["Source records", "Research records"]].min(axis=1)
    frame["Maximum"] = frame[["Source records", "Research records"]].max(axis=1)
    height = max(220, min(590, 42 * frame["Company"].nunique() + 74))

    with st.container(border=True):
        st.markdown(f"#### {title}")
        if description:
            st.caption(description)
        st.vega_lite_chart(
            frame,
            {
                "height": height,
                "layer": [
                    {
                        "mark": {"type": "rule", "strokeWidth": 3, "color": "#CBD5E1"},
                        "encoding": {
                            "y": {
                                "field": "Company",
                                "type": "nominal",
                                "sort": {"field": "Gap", "order": "ascending"},
                                "axis": {"title": None, "labelLimit": 270, "labelPadding": 8},
                            },
                            "x": {
                                "field": "Minimum",
                                "type": "quantitative",
                                "axis": {"title": "Records", "tickMinStep": 1, "gridOpacity": 0.16},
                            },
                            "x2": {"field": "Maximum"},
                        },
                    },
                    {
                        "mark": {"type": "point", "filled": True, "size": 115, "color": "#94A3B8"},
                        "encoding": {
                            "y": {
                                "field": "Company",
                                "type": "nominal",
                                "sort": {"field": "Gap", "order": "ascending"},
                            },
                            "x": {"field": "Source records", "type": "quantitative"},
                            "tooltip": [
                                {"field": "Company", "type": "nominal"},
                                {"field": "Source records", "type": "quantitative", "format": ",.0f"},
                                {"field": "Research records", "type": "quantitative", "format": ",.0f"},
                                {"field": "Gap", "type": "quantitative", "format": "+,.0f"},
                            ],
                        },
                    },
                    {
                        "mark": {"type": "point", "filled": True, "size": 125, "color": "#1287CE"},
                        "encoding": {
                            "y": {
                                "field": "Company",
                                "type": "nominal",
                                "sort": {"field": "Gap", "order": "ascending"},
                            },
                            "x": {"field": "Research records", "type": "quantitative"},
                            "tooltip": [
                                {"field": "Company", "type": "nominal"},
                                {"field": "Source records", "type": "quantitative", "format": ",.0f"},
                                {"field": "Research records", "type": "quantitative", "format": ",.0f"},
                                {"field": "Gap", "type": "quantitative", "format": "+,.0f"},
                            ],
                        },
                    },
                ],
                "config": {
                    "view": {"stroke": None},
                    "axis": {"labelFontSize": 12, "titleFontSize": 12},
                },
            },
            use_container_width=True,
        )
        st.markdown(
            '<div class="db-chart-legend">'
            '<span><i class="source"></i>Current source</span>'
            '<span><i class="research"></i>Current research</span>'
            '</div>',
            unsafe_allow_html=True,
        )



def _render_reconciliation_status_chart(
    data: pd.DataFrame,
    title: str,
    description: str = "",
) -> None:
    """Show reconciliation outcomes with stable semantic colours and direct labels."""
    if not isinstance(data, pd.DataFrame) or data.empty:
        _render_chart_empty_state(title, "No reconciliation results are available yet.")
        return

    category = "Reconciliation status"
    value = "Records"
    frame = data[[category, value]].copy()
    frame[category] = frame[category].fillna("Unspecified").astype(str)
    frame[value] = pd.to_numeric(frame[value], errors="coerce").fillna(0)
    frame = frame.loc[frame[value].gt(0)]
    if frame.empty:
        _render_chart_empty_state(title, "All reconciliation values are currently zero.")
        return

    order = [
        "Matched source",
        "Newly discovered",
        "Source-only / unmatched",
        "Needs classification",
        "Possible duplicates",
        "Excluded / not current",
    ]
    colours = ["#1287CE", "#16835F", "#C27C0E", "#C27C0E", "#94A3B8", "#64748B"]
    height = max(220, min(430, 38 * len(frame) + 66))

    with st.container(border=True):
        st.markdown(f"#### {title}")
        if description:
            st.caption(description)
        st.vega_lite_chart(
            frame,
            {
                "height": height,
                "layer": [
                    {
                        "mark": {"type": "bar", "cornerRadiusEnd": 5},
                        "encoding": {
                            "y": {
                                "field": category,
                                "type": "nominal",
                                "sort": order,
                                "axis": {"title": None, "labelLimit": 235, "labelPadding": 8},
                            },
                            "x": {
                                "field": value,
                                "type": "quantitative",
                                "axis": {"title": "Records", "tickMinStep": 1, "gridOpacity": 0.16},
                            },
                            "color": {
                                "field": category,
                                "type": "nominal",
                                "scale": {"domain": order, "range": colours},
                                "legend": None,
                            },
                            "tooltip": [
                                {"field": category, "type": "nominal", "title": "Status"},
                                {"field": value, "type": "quantitative", "format": ",.0f"},
                            ],
                        },
                    },
                    {
                        "mark": {"type": "text", "align": "left", "baseline": "middle", "dx": 6, "fontSize": 11},
                        "encoding": {
                            "y": {"field": category, "type": "nominal", "sort": order},
                            "x": {"field": value, "type": "quantitative"},
                            "text": {"field": value, "type": "quantitative", "format": ",.0f"},
                        },
                    },
                ],
                "config": {
                    "view": {"stroke": None},
                    "axis": {"labelFontSize": 12, "titleFontSize": 12},
                },
            },
            use_container_width=True,
        )

def _render_verification_progress_chart(project_chart: pd.DataFrame) -> None:
    """Show mutually exclusive verification stages as a 100% stacked company view."""
    title = "Verification progress by company"
    if not isinstance(project_chart, pd.DataFrame) or project_chart.empty:
        _render_chart_empty_state(title, "Add company research records to activate this chart.")
        return

    rows = []
    for _, row in project_chart.iterrows():
        collected = int(row.get("Buildings", 0) or 0)
        reviewed = min(int(row.get("Reviewed", 0) or 0), collected)
        verified = min(int(row.get("Verified", 0) or 0), collected)
        reviewed_not_verified = max(reviewed - verified, 0)
        not_reviewed = max(collected - reviewed, 0)
        rate = (verified / collected) if collected else 0.0
        for stage, count in [
            ("Verified", verified),
            ("Reviewed, not verified", reviewed_not_verified),
            ("Not reviewed", not_reviewed),
        ]:
            rows.append({
                "Company": row.get("Company", ""),
                "Stage": stage,
                "Records": count,
                "Verification rate": rate,
            })

    frame = pd.DataFrame(rows)
    height = max(210, min(580, 40 * project_chart["Company"].nunique() + 62))
    with st.container(border=True):
        st.markdown(f"#### {title}")
        st.caption("Companies are ordered from the lowest verification rate to the highest.")
        st.vega_lite_chart(
            frame,
            {
                "height": height,
                "mark": {"type": "bar", "cornerRadiusEnd": 4},
                "encoding": {
                    "y": {
                        "field": "Company",
                        "type": "nominal",
                        "sort": {"field": "Verification rate", "order": "ascending"},
                        "axis": {"title": None, "labelLimit": 270, "labelPadding": 8},
                    },
                    "x": {
                        "field": "Records",
                        "type": "quantitative",
                        "stack": "normalize",
                        "axis": {"title": "Share of researched records", "format": ".0%", "gridOpacity": 0.15},
                    },
                    "color": {
                        "field": "Stage",
                        "type": "nominal",
                        "scale": {
                            "domain": ["Verified", "Reviewed, not verified", "Not reviewed"],
                            "range": ["#16835F", "#C27C0E", "#CBD5E1"],
                        },
                        "legend": {"title": None, "orient": "top", "direction": "horizontal"},
                    },
                    "tooltip": [
                        {"field": "Company", "type": "nominal"},
                        {"field": "Stage", "type": "nominal"},
                        {"field": "Records", "type": "quantitative", "format": ",.0f"},
                        {"field": "Verification rate", "type": "quantitative", "format": ".0%"},
                    ],
                },
                "config": {
                    "view": {"stroke": None},
                    "axis": {"labelFontSize": 12, "titleFontSize": 12},
                    "legend": {"labelFontSize": 11},
                },
            },
            use_container_width=True,
        )



def _render_stage_composition_chart(
    data: pd.DataFrame,
    title: str = "Research progress",
    description: str = "",
) -> None:
    """Render one mutually exclusive 100% stacked bar instead of three separate bars."""
    if not isinstance(data, pd.DataFrame) or data.empty:
        _render_chart_empty_state(title, "No research stages are available yet.")
        return

    frame = data[["Stage", "Records"]].copy()
    frame["Records"] = pd.to_numeric(frame["Records"], errors="coerce").fillna(0)
    if frame["Records"].sum() <= 0:
        _render_chart_empty_state(title, "No active records are available yet.")
        return
    frame["Portfolio"] = "Current company"

    order = ["Verified", "Reviewed, not verified", "Not reviewed"]
    colours = ["#16835F", "#C27C0E", "#CBD5E1"]
    frame["Stage order"] = frame["Stage"].map({stage: position for position, stage in enumerate(order)})
    total = int(frame["Records"].sum())

    with st.container(border=True):
        st.markdown(f"#### {title}")
        if description:
            st.caption(description)
        st.vega_lite_chart(
            frame,
            {
                "height": 92,
                "mark": {"type": "bar", "cornerRadius": 5, "height": 30},
                "encoding": {
                    "y": {"field": "Portfolio", "type": "nominal", "axis": None},
                    "x": {
                        "field": "Records",
                        "type": "quantitative",
                        "stack": "normalize",
                        "axis": {"title": None, "format": ".0%", "gridOpacity": 0.12},
                    },
                    "color": {
                        "field": "Stage",
                        "type": "nominal",
                        "scale": {"domain": order, "range": colours},
                        "legend": {"title": None, "orient": "top", "direction": "horizontal"},
                    },
                    "order": {"field": "Stage order", "type": "quantitative", "sort": "ascending"},
                    "tooltip": [
                        {"field": "Stage", "type": "nominal"},
                        {"field": "Records", "type": "quantitative", "format": ",.0f"},
                    ],
                },
                "config": {
                    "view": {"stroke": None},
                    "axis": {"labelFontSize": 11},
                    "legend": {"labelFontSize": 11},
                },
            },
            use_container_width=True,
        )
        st.caption(f"{total:,} active research record(s) represented; stages do not overlap.")

def _missing_information_summary(records: pd.DataFrame, limit: int = 8) -> pd.DataFrame:
    """Count the most frequent unresolved fields in the current research records."""
    if not isinstance(records, pd.DataFrame) or records.empty:
        return pd.DataFrame(columns=["Missing field", "Records"])
    counts: dict[str, int] = {}
    for value in records.get("Missing Information", pd.Series(dtype="object")):
        for field in [part.strip() for part in safe_text(value).split(",") if part.strip()]:
            counts[field] = counts.get(field, 0) + 1
    rows = sorted(counts.items(), key=lambda item: (-item[1], item[0]))[:limit]
    return pd.DataFrame(rows, columns=["Missing field", "Records"])


def _project_priority_message(project: dict, reconciliation: dict) -> tuple[str, str, list[str], str]:
    """Translate dashboard totals into one plain-language priority statement."""
    attention = int(project.get("attention_records", 0))
    source_only = int(reconciliation.get("source_only", 0))
    needs_classification = int(reconciliation.get("needs_classification", 0))
    newly_discovered = int(reconciliation.get("newly_discovered", 0))

    if source_only or needs_classification:
        return (
            "Reconciliation needs attention",
            "Resolve unmatched source entries and ambiguous discovery classifications before treating the current portfolio as complete.",
            [
                f"{source_only:,} source-only",
                f"{needs_classification:,} need classification",
                f"{attention:,} quality attention",
            ],
            "warning",
        )
    if attention:
        return (
            "Research is reconciled; quality review remains",
            "The source comparison is in good shape. Focus next on records that still have verification or data-quality issues.",
            [f"{attention:,} need attention", f"{newly_discovered:,} newly discovered"],
            "accent",
        )
    if project.get("buildings", 0):
        return (
            "Project is review-ready",
            "No current reconciliation or quality exceptions are showing in the dashboard.",
            [f"{project.get('verified_records', 0):,} verified", f"{newly_discovered:,} newly discovered"],
            "positive",
        )
    return (
        "Start with company research",
        "Add or import building records to activate verification, source-comparison and quality analytics.",
        [],
        "neutral",
    )


def _health_status(percent: int, exceptions: int = 0, *, empty: bool = False) -> str:
    """Translate a percentage and exception count into a calm health label."""
    if empty:
        return "Not started"
    if exceptions > 0:
        return "Needs attention"
    if percent >= 90:
        return "Healthy"
    if percent >= 70:
        return "In progress"
    return "At risk"


def _render_health_checklist(title: str, items: list[dict], description: str = "") -> None:
    """Render a plain-language checklist instead of another chart."""
    with st.container(border=True):
        st.markdown(f"#### {title}")
        if description:
            st.caption(description)
        for item in items:
            state = str(item.get("state", "pending")).lower()
            symbol = {"pass": "✓", "warning": "!", "pending": "○"}.get(state, "○")
            label = escape(str(item.get("label", "")))
            detail = escape(str(item.get("detail", "")))
            css_state = state if state in {"pass", "warning", "pending"} else "pending"
            st.markdown(
                f"""
                <div class="db-health-check {css_state}">
                    <span class="db-health-check-icon">{symbol}</span>
                    <div><strong>{label}</strong>{f'<small>{detail}</small>' if detail else ''}</div>
                </div>
                """,
                unsafe_allow_html=True,
            )


def _project_health_matrix(
    registry: pd.DataFrame,
    records: pd.DataFrame,
    project: dict,
    reconciliation: dict,
) -> pd.DataFrame:
    """Combine research, verification, QA and source health into one company matrix."""
    progress = company_progress_table(registry, records, project)
    source_table = reconciliation.get("company_table", pd.DataFrame()).copy()
    source_lookup = {}
    if not source_table.empty:
        source_lookup = {
            str(row.get("Company ID", "")): row
            for _, row in source_table.iterrows()
        }

    rows = []
    for snapshot in project.get("company_rows", []):
        company_id = str(snapshot.get("company_id", ""))
        source = source_lookup.get(company_id)
        collected = int(snapshot.get("collected", 0))
        reviewed = int(snapshot.get("reviewed", 0))
        verified = int(snapshot.get("verified", 0))
        attention = int(snapshot.get("attention", 0))
        research_pct = _analytics_percent(reviewed, collected)
        verification_pct = _analytics_percent(verified, collected)

        source_records = int(source["Source records"]) if source is not None else 0
        matched = int(source["Matched"]) if source is not None else 0
        source_only = int(source["Source-only"]) if source is not None else 0
        needs_classification = int(source["Needs classification"]) if source is not None else 0
        duplicates = int(source["Possible duplicates"]) if source is not None else 0
        source_exceptions = source_only + needs_classification + duplicates
        source_pct = _analytics_percent(matched, source_records) if reconciliation.get("available") else 0

        quality = "Good" if attention == 0 and collected else (
            f"{attention} issue{'s' if attention != 1 else ''}" if attention else "Not started"
        )
        source_result = (
            "No baseline"
            if not reconciliation.get("available")
            else "Reconciled" if source_exceptions == 0
            else f"{source_exceptions} exception{'s' if source_exceptions != 1 else ''}"
        )
        rows.append({
            "Company": snapshot.get("company_name", "Unnamed company"),
            "Research": f"{research_pct}%" if collected else "Not started",
            "Source match": f"{source_pct}%" if reconciliation.get("available") and source_records else "—",
            "Verification": f"{verification_pct}%" if collected else "Not started",
            "Quality": quality,
            "Source result": source_result,
            "Status": snapshot.get("status", "Not started"),
            "Next action": snapshot.get("next_title", "Continue research"),
        })
    return pd.DataFrame(rows)


def _source_comparison_checklist(reconciliation: dict) -> list[dict]:
    """Translate source reconciliation results into plain-English completion checks."""
    available = bool(reconciliation.get("available", True))
    source_count = int(reconciliation.get("source_records", 0))
    research_count = int(reconciliation.get("research_records", 0))
    matched = int(reconciliation.get("matched_source", 0))
    new = int(reconciliation.get("newly_discovered", 0))
    source_only = int(reconciliation.get("source_only", 0))
    review = int(reconciliation.get("needs_classification", 0))
    duplicates = int(reconciliation.get("possible_duplicates", 0))
    matched_rate = _analytics_percent(matched, source_count) if source_count else 0

    return [
        {
            "state": "pass" if available and source_count else "pending",
            "label": f"Current source baseline contains {source_count:,} record(s)" if available and source_count else "Current source baseline is not available",
            "detail": "These records form the comparison baseline." if available and source_count else "Add Starting Data to activate source reconciliation.",
        },
        {
            "state": "pass" if research_count >= source_count and research_count else ("warning" if research_count else "pending"),
            "label": f"Current research contains {research_count:,} active record(s)",
            "detail": f"Research is {'at or above' if research_count >= source_count else 'below'} the {source_count:,}-record source baseline.",
        },
        {
            "state": "pass" if source_count and matched == source_count else ("warning" if source_count else "pending"),
            "label": "Every source record has a confirmed research match" if source_count and matched == source_count else f"{matched:,} of {source_count:,} source record(s) are matched",
            "detail": f"Current source-match coverage is {matched_rate}%.",
        },
        {
            "state": "pass" if new else "pending",
            "label": f"{new:,} newly discovered record(s) identified" if new else "No newly discovered records are currently confirmed",
            "detail": "New records are findings that did not receive a credible match in the active source file.",
        },
        {
            "state": "pass" if source_only == 0 and available else ("warning" if available else "pending"),
            "label": "No source-only records remain" if available and source_only == 0 else f"{source_only:,} source-only record(s) remain",
            "detail": "Source-only records have not yet been reconciled with current website research.",
        },
        {
            "state": "pass" if review == 0 and duplicates == 0 else "warning",
            "label": "No ambiguous or duplicate comparison results remain" if review == 0 and duplicates == 0 else f"{review:,} need classification · {duplicates:,} possible duplicate(s)",
            "detail": "Resolve ambiguous matches and possible duplicates before marking reconciliation complete.",
        },
    ]


def _render_source_comparison_summary(reconciliation: dict, *, company_name: str = "") -> None:
    """Show source comparison as an actionable checklist instead of KPI cards."""
    scope = f" for {company_name}" if company_name else ""
    _render_health_checklist(
        f"Source comparison{scope}",
        _source_comparison_checklist(reconciliation),
        "Each check explains whether the active source file and current research are fully reconciled.",
    )


def _company_checklist(
    selected_row: pd.Series,
    snapshot: dict,
    company_qa: pd.DataFrame,
    company_source,
    reconciliation_available: bool,
) -> list[dict]:
    """Return the most important company checks in plain English."""
    website = safe_text(selected_row.get("Main Website", ""))
    total = len(company_qa)
    critical = int(company_qa["QA Status"].eq("Critical").sum()) if total else 0
    source_links = int(company_qa["Source URL"].fillna("").astype(str).str.strip().ne("").sum()) if total else 0
    postal_missing = int(unresolved_mask(company_qa["Postal Code"]).sum()) if total and "Postal Code" in company_qa else 0
    classification_missing = int(unresolved_mask(company_qa["Building Classification"]).sum()) if total and "Building Classification" in company_qa else 0
    source_only = int(company_source["Source-only"]) if company_source is not None else 0
    needs_classification = int(company_source["Needs classification"]) if company_source is not None else 0

    return [
        {
            "state": "pass" if website else "warning",
            "label": "Official company website is registered" if website else "Official company website is missing",
            "detail": website or "Add the official website before continuing research.",
        },
        {
            "state": "pass" if snapshot["collected"] else "pending",
            "label": f"{snapshot['collected']:,} active building record(s) collected",
            "detail": "Website research has produced active records." if snapshot["collected"] else "Start website research or add a building manually.",
        },
        {
            "state": "pass" if total and source_links == total else ("warning" if total else "pending"),
            "label": "Every record has source evidence" if total and source_links == total else f"{max(total - source_links, 0):,} record(s) lack a source URL",
            "detail": f"{source_links:,} of {total:,} records are source-linked.",
        },
        {
            "state": "pass" if postal_missing == 0 and total else ("warning" if postal_missing else "pending"),
            "label": "All active records have postal codes" if postal_missing == 0 and total else f"{postal_missing:,} postal code(s) are missing",
            "detail": "Postal codes support matching and directory entry.",
        },
        {
            "state": "pass" if classification_missing == 0 and total else ("warning" if classification_missing else "pending"),
            "label": "All building classifications are complete" if classification_missing == 0 and total else f"{classification_missing:,} classification(s) are unresolved",
            "detail": "Use reliable storey-count evidence before classifying.",
        },
        {
            "state": "pass" if critical == 0 and total else ("warning" if critical else "pending"),
            "label": "No critical QA issues" if critical == 0 and total else f"{critical:,} critical QA record(s)",
            "detail": "Critical records should be resolved before export.",
        },
        {
            "state": "pass" if reconciliation_available and source_only == 0 and needs_classification == 0 else ("warning" if reconciliation_available else "pending"),
            "label": "Source reconciliation is complete" if reconciliation_available and source_only == 0 and needs_classification == 0 else (
                f"{source_only + needs_classification:,} source comparison exception(s)" if reconciliation_available else "Source comparison is not active"
            ),
            "detail": "Resolve source-only and ambiguous records before marking the company complete.",
        },
    ]


def render_project_company_analytics(registry: pd.DataFrame, records: pd.DataFrame) -> None:
    """Render chart-free project and company health analytics."""
    registry = normalize_company_registry(registry)
    records = _normalize_analytics_records(records)
    source_details = _current_source_display_details()
    reconciliation = source_reconciliation_snapshot(registry, records, source_details["records"])
    comparison_records = reconciliation["records"]
    project = project_progress_snapshot(registry, comparison_records)

    _render_analytics_header(source_details, reconciliation)
    project_tab, company_tab = st.tabs(["Project dashboard", "Company dashboard"])

    # ================================================================
    # PROJECT HEALTH
    # One matrix, one checklist and one exact source comparison replace
    # every project-level chart.
    # ================================================================
    with project_tab:
        verification_rate = _analytics_percent(project["verified_records"], project["buildings"])
        source_coverage = _analytics_percent(reconciliation["matched_source"], reconciliation["source_records"])
        _render_analytics_kpis([
            {"label": "Companies", "value": f"{project['companies']:,}", "helper": f"{project['completed']:,} complete", "tone": "accent"},
            {"label": "Source records", "value": f"{reconciliation['source_records']:,}" if reconciliation["available"] else "—", "helper": "Active baseline", "tone": "neutral"},
            {"label": "Research records", "value": f"{reconciliation['research_records']:,}", "helper": "Active building records", "tone": "accent"},
            {"label": "Verified", "value": f"{project['verified_records']:,}", "helper": f"{verification_rate}% verification", "tone": "positive" if verification_rate >= 80 else "accent"},
            {"label": "Needs attention", "value": f"{project['attention_records']:,}", "helper": "QA or review follow-up", "tone": "warning" if project["attention_records"] else "positive"},
            {"label": "Project readiness", "value": f"{project['progress_percent']:,}%", "helper": f"{project['completed']:,} of {project['companies']:,} companies complete", "tone": "positive" if project["progress_percent"] >= 80 else "accent"},
        ])

        priority_title, priority_copy, priority_chips, priority_tone = _project_priority_message(project, reconciliation)
        _render_analytics_callout(priority_title, priority_copy, priority_chips, priority_tone)

        _render_analytics_section(
            "Project health",
            "Compare company readiness with source reconciliation in one dashboard view.",
            "PROJECT READINESS",
        )
        health_matrix = _project_health_matrix(registry, comparison_records, project, reconciliation)
        health_left, health_right = st.columns([1.7, 1], gap="large")
        with health_left:
            with st.container(border=True):
                st.markdown("#### Company health matrix")
                st.caption("Research, source matching, verification, quality and the next action for every company.")
                if health_matrix.empty:
                    st.info("Add companies to activate the project health matrix.")
                else:
                    st.dataframe(health_matrix, width="stretch", hide_index=True, height=315)
        with health_right:
            _render_source_comparison_summary(reconciliation)

        companies_missing_websites = int(registry["Main Website"].fillna("").astype(str).str.strip().eq("").sum()) if not registry.empty else 0
        project_checks = [
            {
                "state": "pass" if project["companies"] else "pending",
                "label": f"{project['companies']:,} company record(s) are registered" if project["companies"] else "No companies are registered",
                "detail": "Company records define the project research scope.",
            },
            {
                "state": "pass" if companies_missing_websites == 0 and project["companies"] else ("warning" if companies_missing_websites else "pending"),
                "label": "Every company has an official website" if companies_missing_websites == 0 and project["companies"] else f"{companies_missing_websites:,} company website(s) missing",
                "detail": "Official websites anchor the research workflow.",
            },
            {
                "state": "pass" if reconciliation["available"] else "pending",
                "label": "Current source baseline is active" if reconciliation["available"] else "Current source baseline is not active",
                "detail": source_details["label"] if reconciliation["available"] else "Add Starting Data to activate source comparison.",
            },
            {
                "state": "pass" if reconciliation["available"] and reconciliation["matched_source"] == reconciliation["source_records"] else ("warning" if reconciliation["available"] else "pending"),
                "label": "Every source record has a confirmed research match" if reconciliation["available"] and reconciliation["matched_source"] == reconciliation["source_records"] else f"{reconciliation['matched_source']:,} of {reconciliation['source_records']:,} source record(s) matched",
                "detail": "Matched records confirm that current research aligns with the active source file.",
            },
            {
                "state": "pass" if reconciliation["source_only"] == 0 and reconciliation["available"] else ("warning" if reconciliation["available"] else "pending"),
                "label": "No source-only records remain" if reconciliation["available"] and reconciliation["source_only"] == 0 else f"{reconciliation['source_only']:,} source-only record(s) remain",
                "detail": "Unmatched source records require review before completion.",
            },
            {
                "state": "pass" if reconciliation["needs_classification"] == 0 and reconciliation["possible_duplicates"] == 0 else "warning",
                "label": "No ambiguous or duplicate source results remain" if reconciliation["needs_classification"] == 0 and reconciliation["possible_duplicates"] == 0 else f"{reconciliation['needs_classification']:,} need classification · {reconciliation['possible_duplicates']:,} possible duplicate(s)",
                "detail": "Resolve ambiguous matches and duplicate candidates before final reconciliation.",
            },
            {
                "state": "pass" if reconciliation["newly_discovered"] else "pending",
                "label": f"{reconciliation['newly_discovered']:,} newly discovered record(s) identified" if reconciliation["newly_discovered"] else "No newly discovered records are currently confirmed",
                "detail": "New discoveries are current properties not credibly matched to the active source file.",
            },
            {
                "state": "pass" if project["attention_records"] == 0 and project["buildings"] else ("warning" if project["attention_records"] else "pending"),
                "label": "No active QA or verification exceptions" if project["attention_records"] == 0 and project["buildings"] else f"{project['attention_records']:,} record(s) need attention",
                "detail": "Resolve these records before final export.",
            },
            {
                "state": "pass" if project["completed"] == project["companies"] and project["companies"] else "pending",
                "label": "All companies are complete" if project["completed"] == project["companies"] and project["companies"] else f"{project['completed']:,} of {project['companies']:,} companies complete",
                "detail": "Completion requires verified records and no unresolved attention.",
            },
        ]
        _render_health_checklist(
            "Data quality & workflow checklist",
            project_checks,
            "Research readiness, source reconciliation, quality and completion checks for final close-out.",
        )

        if reconciliation["available"]:
            with smart_expander(
                "Detailed source comparison by company",
                count=len(reconciliation["company_table"]),
                status="companies",
                expanded=False,
            ):
                st.dataframe(
                    reconciliation["company_table"].drop(columns=["Company ID"]),
                    width="stretch",
                    hide_index=True,
                )

    # ================================================================
    # COMPANY HEALTH
    # The selected company gets important metrics, a health matrix,
    # a checklist and an exact source comparison. No charts are used.
    # ================================================================
    with company_tab:
        if registry.empty:
            st.info("Add a company to see company health analytics.")
            return

        company_ids = registry["Company ID"].astype(str).tolist()
        active_id = str(st.session_state.get(S_ACTIVE_COMPANY, "")).strip()
        company_index = company_ids.index(active_id) if active_id in company_ids else 0

        with st.container(border=True):
            selector_col, context_col = st.columns([1.8, 1], gap="large")
            with selector_col:
                selected_id = st.selectbox(
                    "Company to analyze",
                    company_ids,
                    index=company_index,
                    format_func=lambda company_id: company_label(
                        registry.loc[registry["Company ID"].eq(company_id)].iloc[0]
                    ),
                    key="db_analytics_company_selector",
                )
            selected_row = registry.loc[registry["Company ID"].eq(selected_id)].iloc[0]
            snapshot = company_progress_snapshot(selected_row, comparison_records)
            with context_col:
                st.markdown(
                    f"""
                    <div class="db-analytics-company-context">
                        <span>CURRENT STATUS</span>
                        <strong>{escape(snapshot['status'])}</strong>
                        <small>{snapshot['progress_percent']:,}% verification progress</small>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

        company_records = _company_records_for_progress(
            comparison_records,
            company_id=selected_id,
            company_name=safe_text(selected_row.get("Management/Owner")),
            company_website=safe_text(selected_row.get("Main Website")),
        )
        company_qa = qa_checks(company_records) if not company_records.empty else pd.DataFrame()
        active_company_qa = (
            company_qa.loc[~company_qa["Record Readiness"].eq("Excluded from Listings")].copy()
            if not company_qa.empty else pd.DataFrame()
        )
        source_links = int(company_qa["Source URL"].fillna("").astype(str).str.strip().ne("").sum()) if not company_qa.empty else 0
        source_rate = _analytics_percent(source_links, len(company_qa)) if len(company_qa) else 0
        passing = int(company_qa["QA Status"].eq("Pass").sum()) if not company_qa.empty else 0
        critical = int(company_qa["QA Status"].eq("Critical").sum()) if not company_qa.empty else 0

        company_reconciliation = (
            reconciliation["company_table"].loc[
                reconciliation["company_table"]["Company ID"].astype(str).eq(selected_id)
            ] if not reconciliation["company_table"].empty else pd.DataFrame()
        )
        company_source = company_reconciliation.iloc[0] if not company_reconciliation.empty else None
        company_source_records = int(company_source["Source records"]) if company_source is not None else 0
        company_matched = int(company_source["Matched"]) if company_source is not None else 0
        company_source_pct = _analytics_percent(company_matched, company_source_records)

        _render_analytics_kpis([
            {"label": "Researched", "value": f"{snapshot['collected']:,}", "helper": f"{snapshot['active_count']:,} eligible · {snapshot['excluded']:,} excluded", "tone": "accent"},
            {"label": "Reviewed", "value": f"{snapshot['reviewed']:,}", "helper": "Human review completed", "tone": "neutral"},
            {"label": "Verified", "value": f"{snapshot['verified']:,}", "helper": f"{snapshot['progress_percent']:,}% verified", "tone": "positive" if snapshot["progress_percent"] >= 80 else "accent"},
            {"label": "Source coverage", "value": f"{source_rate}%" if len(company_qa) else "—", "helper": f"{source_links:,} source-linked", "tone": "positive" if source_rate >= 90 else "accent"},
            {"label": "Passing QA", "value": f"{passing:,}", "helper": f"{critical:,} critical", "tone": "warning" if critical else "positive"},
            {"label": "Needs attention", "value": f"{snapshot['attention']:,}", "helper": snapshot["next_title"], "tone": "warning" if snapshot["attention"] else "positive"},
        ])

        _render_analytics_section(
            "Company health matrix",
            "The selected company is assessed across research, source reconciliation, verification, quality and directory readiness.",
            "COMPANY READINESS",
        )
        directory_entered = int(active_company_qa["Directory Entry Status"].eq("Entered").sum()) if not active_company_qa.empty else 0
        active_total = len(active_company_qa)
        company_source_exceptions = (
            int(company_source["Source-only"]) + int(company_source["Needs classification"]) + int(company_source["Possible duplicates"])
            if company_source is not None else 0
        )
        health_rows = pd.DataFrame([
            {
                "Health area": "Research review",
                "Result": f"{snapshot['reviewed']:,} of {snapshot['collected']:,}",
                "Completion": f"{_analytics_percent(snapshot['reviewed'], snapshot['collected'])}%" if snapshot['collected'] else "Not started",
                "Status": _health_status(_analytics_percent(snapshot['reviewed'], snapshot['collected']), empty=snapshot['collected'] == 0),
                "Action": "Complete remaining record review" if snapshot['reviewed'] < snapshot['collected'] else "No action required",
            },
            {
                "Health area": "Source reconciliation",
                "Result": f"{company_matched:,} of {company_source_records:,}",
                "Completion": f"{company_source_pct}%" if reconciliation['available'] and company_source_records else "—",
                "Status": _health_status(company_source_pct, company_source_exceptions, empty=not reconciliation['available']),
                "Action": "Resolve source comparison exceptions" if company_source_exceptions else "No action required",
            },
            {
                "Health area": "Verification",
                "Result": f"{snapshot['verified']:,} of {snapshot['collected']:,}",
                "Completion": f"{snapshot['progress_percent']}%" if snapshot['collected'] else "Not started",
                "Status": _health_status(snapshot['progress_percent'], snapshot['attention'], empty=snapshot['collected'] == 0),
                "Action": "Verify remaining records" if snapshot['verified'] < snapshot['collected'] else "No action required",
            },
            {
                "Health area": "Data quality",
                "Result": f"{passing:,} pass · {critical:,} critical",
                "Completion": f"{_analytics_percent(passing, len(company_qa))}% pass" if len(company_qa) else "Not started",
                "Status": "Healthy" if len(company_qa) and critical == 0 and snapshot['attention'] == 0 else ("Needs attention" if len(company_qa) else "Not started"),
                "Action": "Resolve QA exceptions" if critical or snapshot['attention'] else "No action required",
            },
            {
                "Health area": "Directory entry",
                "Result": f"{directory_entered:,} of {active_total:,}",
                "Completion": f"{_analytics_percent(directory_entered, active_total)}%" if active_total else "Not started",
                "Status": _health_status(_analytics_percent(directory_entered, active_total), empty=active_total == 0),
                "Action": "Enter approved records" if directory_entered < active_total else "Complete",
            },
        ])
        st.dataframe(health_rows, width="stretch", hide_index=True)

        company_checks = _company_checklist(
            selected_row,
            snapshot,
            company_qa,
            company_source,
            reconciliation["available"],
        )
        if reconciliation["available"]:
            company_new = int(company_source["New"]) if company_source is not None else 0
            company_source_only = int(company_source["Source-only"]) if company_source is not None else 0
            company_needs_classification = int(company_source["Needs classification"]) if company_source is not None else 0
            company_possible_duplicates = int(company_source["Possible duplicates"]) if company_source is not None else 0
            company_checks.extend([
                {
                    "state": "pass" if company_source_records and company_matched == company_source_records else ("warning" if company_source_records else "pending"),
                    "label": "Every company source record has a confirmed research match" if company_source_records and company_matched == company_source_records else f"{company_matched:,} of {company_source_records:,} company source record(s) matched",
                    "detail": f"Current source-match coverage is {company_source_pct}%.",
                },
                {
                    "state": "pass" if company_source_only == 0 else "warning",
                    "label": "No source-only company records remain" if company_source_only == 0 else f"{company_source_only:,} source-only company record(s) remain",
                    "detail": "Review source entries that are not yet linked to current company research.",
                },
                {
                    "state": "pass" if company_needs_classification == 0 and company_possible_duplicates == 0 else "warning",
                    "label": "No ambiguous or duplicate company matches remain" if company_needs_classification == 0 and company_possible_duplicates == 0 else f"{company_needs_classification:,} need classification · {company_possible_duplicates:,} possible duplicate(s)",
                    "detail": "Resolve these exceptions before marking the company fully reconciled.",
                },
                {
                    "state": "pass" if company_new else "pending",
                    "label": f"{company_new:,} newly discovered company record(s) identified" if company_new else "No newly discovered company records are currently confirmed",
                    "detail": "New discoveries are company properties not credibly matched to the active source file.",
                },
            ])
        else:
            company_checks.append({
                "state": "pending",
                "label": "Company source comparison is not active",
                "detail": "Add the active Starting Data baseline to compare this company with the source file.",
            })

        _render_health_checklist(
            "Overall company checklist",
            company_checks,
            "Company setup, research evidence, source reconciliation, data quality and completion are combined in one checklist.",
        )

        if not company_qa.empty:
            with smart_expander(
                "Selected company records",
                count=len(company_qa),
                status="records",
                expanded=False,
            ):
                detail_columns = [
                    column for column in [
                        "Building Name", "Street Address", "Postal Code",
                        "Directory Discovery Status", "Verification Status",
                        "QA Status", "Missing Information", "Source URL",
                    ] if column in company_qa.columns
                ]
                st.dataframe(company_qa[detail_columns], width="stretch", hide_index=True)

        _render_analytics_callout(
            snapshot["next_title"],
            snapshot["next_copy"],
            [f"Status: {snapshot['status']}", f"{snapshot['attention']:,} need attention"],
            "warning" if snapshot["attention"] else "accent",
        )
        if st.button(
            snapshot["next_button"],
            type="primary",
            width="stretch",
            key=f"db_analytics_next_{selected_id}",
        ):
            st.session_state[S_ACTIVE_COMPANY] = selected_id
            go_to(snapshot["next_section"])
            st.rerun()


def render_project_progress_sidebar() -> None:
    """Render a navigation-first SaaS sidebar while preserving project controls."""
    st.markdown(
        """
        <div class="db-side-brand">
            <div class="db-side-brand-mark">D</div>
            <div>
                <div class="db-side-brand-name">DATABLIX</div>
                <div class="db-side-brand-sub">RESEARCH INTELLIGENCE</div>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if S_WORKING not in st.session_state:
        st.markdown('<div class="db-side-empty">No project is open yet.</div>', unsafe_allow_html=True)
        return

    records, registry = synchronize_company_registry(
        st.session_state[S_WORKING],
        st.session_state.get(S_COMPANIES),
    )
    st.session_state[S_WORKING] = records
    st.session_state[S_COMPANIES] = registry

    project_name = str(
        st.session_state.get(S_PROJECT_NAME, "Datablix project")
    ).strip() or "Datablix project"
    project = project_progress_snapshot(registry, records)
    active = active_company_row()
    active_name = (
        safe_text(active.get("Management/Owner", ""))
        if active is not None
        else "No company selected"
    )

    st.markdown(
        f"""
        <div class="db-side-context-card">
            <span class="db-side-kicker">CURRENT PROJECT</span>
            <strong>{escape(project_name)}</strong>
            <div class="db-side-context-meta">
                <span>{project['companies']:,} companies</span>
                <span>{project['buildings']:,} records</span>
            </div>
        </div>
        <div class="db-side-context-card company">
            <span class="db-side-kicker">SELECTED COMPANY</span>
            <strong>{escape(active_name)}</strong>
        </div>
        """,
        unsafe_allow_html=True,
    )

    st.markdown('<div class="db-side-section-label">WORKSPACE</div>', unsafe_allow_html=True)
    sidebar_navigation = [
        ("Dashboard", "▦  Dashboard"),
        ("Research projects & companies", "◇  Project"),
        ("Website scanner", "⌕  Research"),
        ("Review records", "✓  Review"),
        ("Analysis & report", "▤  Deliverables"),
        ("Downloads", "⇩  Export"),
    ]
    current = str(st.session_state.get("db_section", "Dashboard"))
    for section_key, label in sidebar_navigation:
        if st.button(
            label,
            type="primary" if current == section_key else "secondary",
            width="stretch",
            key=f"db_side_nav_{norm_header(section_key)}",
        ):
            go_to(section_key)
            st.rerun()

    st.markdown('<div class="db-side-section-label progress">PROGRESS</div>', unsafe_allow_html=True)
    progress_pct = int(round(float(project.get("progress", 0.0)) * 100))
    st.markdown(
        f"""
        <div class="db-side-progress-card">
            <div class="db-side-progress-head">
                <span>Project readiness</span><strong>{progress_pct}%</strong>
            </div>
            <div class="db-side-progress-track"><span style="width:{max(0, min(progress_pct, 100))}%"></span></div>
            <div class="db-side-progress-stats">
                <span><strong>{project['completed']:,}</strong><small>Complete</small></span>
                <span><strong>{project['in_progress']:,}</strong><small>In progress</small></span>
                <span><strong>{project['attention_records']:,}</strong><small>Attention</small></span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if active is not None:
        company = company_progress_snapshot(active, records)
        st.markdown(
            f"""<div class="db-side-next">
                <span>NEXT RECOMMENDED ACTION</span>
                <strong>{escape(company['next_title'])}</strong>
                <small>{escape(company['next_copy'])}</small>
            </div>""",
            unsafe_allow_html=True,
        )
        if st.button(
            company["next_button"],
            width="stretch",
            key=f"db_sidebar_continue_{company['company_id']}",
        ):
            go_to(company["next_section"])
            st.rerun()

    st.markdown('<div class="db-side-section-label account">ACCOUNT</div>', unsafe_allow_html=True)
    if st.session_state.get(S_DEMO_MODE):
        with st.expander("Demo workspace", expanded=False):
            st.caption("Sample information only. Changes are temporary.")
            if st.button("Leave Demo", width="stretch", key="db_sidebar_leave_demo"):
                return_to_project_start()
                st.rerun()
        project_id = ""
    else:
        project_id = str(st.session_state.get(S_CLOUD_PROJECT_ID, "")).strip()
        role_label = st.session_state.get(S_PROJECT_ROLE, "owner").title()
        with st.expander(f"{role_label} · Account", expanded=False):
            st.caption(current_user_email())
            if st.button("Sign out", width="stretch", key="db_sidebar_sign_out"):
                sign_out_datablix()
                st.rerun()

    if project_id and st.session_state.get(S_PROJECT_ROLE) == "owner":
        with st.expander("Share project", expanded=False):
            member_email = st.text_input("Team member email", key="db_share_member_email")
            member_role = st.selectbox("Access", ["editor", "viewer"], format_func=str.title, key="db_share_member_role")
            if st.button("Save access", width="stretch", key="db_save_member_access"):
                ok, message = add_project_member(project_id, member_email, member_role)
                (st.success if ok else st.error)(message)
            members = list_project_members(project_id)
            if members:
                for member in members:
                    email = str(member.get("member_email", ""))
                    role = str(member.get("role", "viewer")).title()
                    cols = st.columns([3, 1])
                    cols[0].caption(f"{email} · {role}")
                    if cols[1].button("×", key=f"db_remove_member_{hashlib.md5(email.encode()).hexdigest()[:8]}"):
                        if remove_project_member(project_id, email):
                            st.rerun()


st.html("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@500;600;700;800&display=swap');

:root{
    --db-accent:#1287CE;            /* sky blue, matching the Datablix logo; deep enough for white button text */
    --db-accent-strong:#0E6BA4;
    --db-accent-soft:rgba(18,135,206,.09);
    --db-accent-edge:rgba(18,135,206,.45);
    --db-ink:#1C272E;
    --db-border:rgba(28,39,46,.14);
    --db-soft:rgba(28,39,46,.035);
    --db-soft-strong:rgba(28,39,46,.07);
    --db-display:'Sora','Source Sans Pro',sans-serif;
}
@media(prefers-color-scheme:dark){
    :root{
        --db-accent:#4FB6F0;
        --db-accent-strong:#79C8F5;
        --db-accent-soft:rgba(79,182,240,.12);
        --db-accent-edge:rgba(79,182,240,.5);
        --db-border:rgba(255,255,255,.14);
        --db-soft:rgba(255,255,255,.04);
        --db-soft-strong:rgba(255,255,255,.075);
    }
}

.block-container{
    max-width:1380px;
    padding-top:1rem;
    padding-bottom:4rem;
}

/* Type: Sora carries the identity in headings and the brand mark only. */
h1,h2,h3{
    font-family:var(--db-display);
    letter-spacing:-.02em;
}
h2{margin-bottom:.1rem}

.db-brand{
    text-align:center;
    margin:.15rem auto 1.3rem;
}
.db-brand-name{
    font-family:var(--db-display);
    font-size:2.15rem;
    font-weight:800;
    letter-spacing:-.04em;
    line-height:1.05;
}
.db-brand-name::after{
    content:"";
    display:block;
    width:2.6rem;
    height:3px;
    margin:.45rem auto 0;
    border-radius:2px;
    background:var(--db-accent);
}
.db-tag{
    margin-top:.4rem;
    font-size:1.03rem;
    font-weight:600;
    opacity:.85;
}
.db-subtag{
    margin-top:.2rem;
    font-size:.9rem;
    opacity:.62;
}

.db-eyebrow{
    margin-top:.25rem;
    margin-bottom:-.35rem;
    font-size:.68rem;
    font-weight:750;
    letter-spacing:.1em;
    text-transform:uppercase;
    color:var(--db-accent);
    opacity:.95;
}

/* Workspace ledger strip: the one place session state is always visible. */
.db-workspace-strip{
    display:flex;
    flex-wrap:wrap;
    gap:.55rem 1.35rem;
    align-items:center;
    padding:.72rem 1rem;
    margin:.25rem 0 1rem;
    border:1px solid var(--db-border);
    border-left:4px solid var(--db-accent-edge);
    border-radius:10px;
    background:var(--db-soft);
    font-size:.88rem;
}
.db-workspace-strip strong{font-weight:700}
.db-workspace-strip .db-num{
    font-variant-numeric:tabular-nums;
    font-weight:700;
    color:var(--db-accent);
}

.db-step-line{
    margin:.2rem 0 1rem;
    font-size:.82rem;
    letter-spacing:.02em;
    opacity:.66;
}
.db-card-copy{min-height:3.6rem}

[data-testid="stSidebar"]{
    border-right:1px solid var(--db-border);
}

div[data-testid="stMetric"]{
    background:var(--db-soft);
    border:1px solid var(--db-border);
    border-top:3px solid var(--db-accent-edge);
    border-radius:12px;
    padding:.8rem .9rem;
    min-height:100px;
}
div[data-testid="stMetric"] label{font-weight:650}
div[data-testid="stMetricValue"]{
    font-variant-numeric:tabular-nums;
    font-family:var(--db-display);
    letter-spacing:-.02em;
}

div[data-testid="stFileUploader"]{
    border:1px dashed var(--db-accent-edge);
    border-radius:11px;
    padding:.3rem .6rem .7rem;
    background:var(--db-accent-soft);
}

div[data-testid="stExpander"],
div[data-testid="stDataFrame"],
div[data-testid="stDataEditor"]{
    border:1px solid var(--db-border);
    border-radius:10px;
    overflow:hidden;
}

.stButton>button,.stDownloadButton>button{
    border-radius:9px;
    font-weight:650;
    min-height:2.55rem;
}
button[data-testid="stBaseButton-primary"],
button[kind="primary"]{
    background:var(--db-accent) !important;
    border-color:var(--db-accent) !important;
    color:#fff !important;
}
button[data-testid="stBaseButton-primary"]:hover,
button[kind="primary"]:hover{
    background:var(--db-accent-strong) !important;
    border-color:var(--db-accent-strong) !important;
}
button[data-testid="stBaseButton-primaryFormSubmit"]{
    background:var(--db-accent) !important;
    border-color:var(--db-accent) !important;
    color:#fff !important;
}
@media(prefers-color-scheme:dark){
    button[data-testid="stBaseButton-primary"],
    button[kind="primary"],
    button[data-testid="stBaseButton-primaryFormSubmit"]{
        color:#0B1D2A !important;
    }
}
.stButton>button:focus-visible,
.stDownloadButton>button:focus-visible{
    outline:2px solid var(--db-accent);
    outline-offset:2px;
}

.stProgress > div > div > div > div{
    background-color:var(--db-accent);
}

button[data-testid="stSidebarCollapseButton"]{
    width:auto !important;
    min-width:7.4rem !important;
    justify-content:flex-start !important;
    gap:.35rem !important;
}
button[data-testid="stSidebarCollapseButton"]::after{
    content:"Menu";
    font-size:.86rem;
    font-weight:700;
    white-space:nowrap;
    opacity:.84;
}

@media (max-width:900px){
    .db-card-copy{min-height:auto}
    button[data-testid="stSidebarCollapseButton"]{min-width:6.7rem !important}
}
@media (prefers-reduced-motion:reduce){
    *{transition:none !important;animation:none !important}
}

/* Page-level hierarchy: the heading and its explanation read as one unit. */
.db-page-head{
    max-width:920px;
    margin:.3rem 0 1.25rem;
    padding:0 0 1rem;
    border-bottom:1px solid var(--db-border);
}
.db-page-head h2{
    margin:.3rem 0 .35rem;
    font-family:var(--db-display);
    font-size:clamp(1.7rem,3vw,2.25rem);
    line-height:1.12;
}
.db-page-head p{
    max-width:820px;
    margin:0;
    font-size:.98rem;
    line-height:1.55;
    opacity:.72;
}

/* Five-button workflow navigation.
   Arrows are drawn in the gaps with CSS so they do not consume Streamlit columns. */
.st-key-db_nav_row div[data-testid="stHorizontalBlock"]{
    flex-wrap:nowrap !important;
    gap:1.8rem !important;
    align-items:center !important;
}
.st-key-db_nav_row div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"]{
    position:relative !important;
    flex:1 1 0 !important;
    min-width:0 !important;
    width:auto !important;
}
.st-key-db_nav_row div[data-testid="stHorizontalBlock"] > div[data-testid="stColumn"]:not(:last-child)::after{
    content:"→";
    position:absolute;
    right:-1.15rem;
    top:50%;
    transform:translateY(-50%);
    color:var(--db-muted);
    font-size:.72rem;
    opacity:.5;
    pointer-events:none;
    z-index:4;
}
.st-key-db_nav_row button{
    width:100% !important;
    white-space:nowrap !important;
    min-width:0 !important;
}
.db-nav-context{
    margin:.35rem 0 1.15rem;
    padding:.2rem .15rem;
    font-size:.86rem;
    line-height:1.45;
    opacity:.72;
}
.db-nav-context strong{
    color:var(--db-accent-strong);
    font-weight:750;
}

/* Contextual help is visually quieter than an alert and stronger than a caption. */
.db-guidance{
    display:flex;
    flex-wrap:wrap;
    gap:.25rem .55rem;
    align-items:baseline;
    margin:.35rem 0 .9rem;
    padding:.7rem .85rem;
    border-left:3px solid var(--db-accent-edge);
    border-radius:7px;
    background:var(--db-accent-soft);
    font-size:.88rem;
    line-height:1.45;
}
.db-guidance strong{font-weight:750}
.db-guidance span{opacity:.76}


/* Analytics workspace: responsive hierarchy, calm cards and decision-focused status. */
.db-analytics-hero{
    display:flex;
    align-items:flex-start;
    justify-content:space-between;
    gap:1.5rem;
    margin:.15rem 0 1.1rem;
    padding:1.15rem 1.2rem;
    border:1px solid var(--db-border);
    border-left:4px solid var(--db-accent);
    border-radius:14px;
    background:var(--db-soft);
}
.db-analytics-hero-copy{max-width:780px}
.db-analytics-eyebrow,
.db-analytics-section-eyebrow{
    margin-bottom:.22rem;
    font-size:.67rem;
    font-weight:800;
    letter-spacing:.1em;
    color:var(--db-accent-strong);
}
.db-analytics-hero h2{
    margin:0 0 .35rem;
    font-family:var(--db-display);
    font-size:clamp(1.55rem,2.8vw,2.15rem);
    line-height:1.12;
    letter-spacing:-.025em;
}
.db-analytics-hero p{
    max-width:760px;
    margin:0;
    font-size:.95rem;
    line-height:1.55;
    opacity:.72;
}
.db-analytics-baseline{
    flex:0 1 320px;
    display:flex;
    flex-direction:column;
    gap:.2rem;
    min-width:230px;
    padding:.78rem .85rem;
    border:1px solid var(--db-border);
    border-radius:10px;
    background:var(--db-accent-soft);
}
.db-analytics-baseline span,
.db-analytics-source-strip span,
.db-analytics-company-context span{
    font-size:.64rem;
    font-weight:800;
    letter-spacing:.08em;
    color:var(--db-accent-strong);
}
.db-analytics-baseline strong{
    overflow-wrap:anywhere;
    font-size:.84rem;
    line-height:1.35;
}
.db-analytics-baseline small,
.db-analytics-company-context small{
    font-size:.74rem;
    opacity:.66;
}

.db-analytics-kpi-scroll{
    width:100%;
    overflow-x:auto;
    overflow-y:hidden;
    margin:.82rem 0 1rem;
    padding-bottom:.12rem;
    scrollbar-width:thin;
}
.db-analytics-kpi-grid{
    display:grid;
    grid-template-columns:repeat(var(--db-kpi-count),minmax(138px,1fr));
    gap:.58rem;
    min-width:max-content;
}
.db-analytics-kpi{
    position:relative;
    min-height:96px;
    padding:.68rem .72rem;
    border:1px solid var(--db-border);
    border-top:3px solid rgba(100,110,125,.35);
    border-radius:10px;
    background:var(--db-soft);
}
.db-analytics-kpi.accent{border-top-color:var(--db-accent)}
.db-analytics-kpi.positive{border-top-color:#16835F}
.db-analytics-kpi.warning{border-top-color:#C27C0E}
.db-analytics-kpi.neutral{border-top-color:rgba(100,110,125,.42)}
.db-analytics-kpi-label{
    min-height:1.75em;
    font-size:.68rem;
    font-weight:720;
    line-height:1.3;
    opacity:.72;
}
.db-analytics-kpi-value{
    margin:.2rem 0 .18rem;
    font-family:var(--db-display);
    font-size:clamp(1.25rem,1.8vw,1.65rem);
    font-weight:750;
    letter-spacing:-.035em;
    line-height:1.1;
    font-variant-numeric:tabular-nums;
    overflow-wrap:anywhere;
}
.db-analytics-kpi-helper{
    font-size:.67rem;
    line-height:1.3;
    opacity:.62;
}


.db-chart-legend{
    display:flex;
    flex-wrap:wrap;
    gap:1rem;
    margin:.15rem 0 .25rem;
    font-size:.72rem;
    opacity:.72;
}
.db-chart-legend span{display:inline-flex;align-items:center;gap:.38rem}
.db-chart-legend i{
    width:.58rem;
    height:.58rem;
    border-radius:50%;
    display:inline-block;
}
.db-chart-legend i.source{background:#94A3B8}
.db-chart-legend i.research{background:#1287CE}

.db-analytics-section-head{
    display:flex;
    align-items:flex-end;
    justify-content:space-between;
    gap:1rem;
    margin:1.55rem 0 .72rem;
    padding-bottom:.55rem;
    border-bottom:1px solid var(--db-border);
}
.db-analytics-section-head h3{
    margin:0 0 .2rem;
    font-family:var(--db-display);
    font-size:1.15rem;
    letter-spacing:-.015em;
}
.db-analytics-section-head p{
    max-width:760px;
    margin:0;
    font-size:.84rem;
    line-height:1.45;
    opacity:.66;
}

.db-analytics-callout{
    display:flex;
    align-items:center;
    justify-content:space-between;
    gap:1rem;
    margin:.65rem 0 1rem;
    padding:.82rem .9rem;
    border:1px solid var(--db-accent-edge);
    border-left:4px solid var(--db-accent);
    border-radius:10px;
    background:var(--db-accent-soft);
}
.db-analytics-callout.warning{
    border-color:rgba(194,124,14,.42);
    border-left-color:#C27C0E;
    background:rgba(194,124,14,.07);
}
.db-analytics-callout.positive{
    border-color:rgba(22,131,95,.38);
    border-left-color:#16835F;
    background:rgba(22,131,95,.07);
}
.db-analytics-callout.neutral{
    border-color:var(--db-border);
    border-left-color:rgba(100,110,125,.5);
    background:var(--db-soft);
}
.db-analytics-callout strong{
    display:block;
    margin-bottom:.18rem;
    font-size:.88rem;
}
.db-analytics-callout p{
    max-width:760px;
    margin:0;
    font-size:.8rem;
    line-height:1.45;
    opacity:.72;
}
.db-analytics-chip-row{
    display:flex;
    flex-wrap:wrap;
    justify-content:flex-end;
    gap:.35rem;
}
.db-analytics-chip{
    display:inline-flex;
    align-items:center;
    min-height:1.65rem;
    padding:.22rem .5rem;
    border:1px solid var(--db-border);
    border-radius:999px;
    background:rgba(255,255,255,.45);
    font-size:.68rem;
    font-weight:720;
    white-space:nowrap;
}
@media(prefers-color-scheme:dark){
    .db-analytics-chip{background:rgba(255,255,255,.04)}
}

.db-analytics-source-strip{
    display:grid;
    grid-template-columns:minmax(0,2fr) repeat(2,minmax(120px,.65fr));
    gap:.7rem;
    margin:.15rem 0 .9rem;
    padding:.72rem .82rem;
    border:1px solid var(--db-border);
    border-radius:10px;
    background:var(--db-soft);
}
.db-analytics-source-strip>div{
    display:flex;
    flex-direction:column;
    gap:.2rem;
    min-width:0;
}
.db-analytics-source-strip strong{
    overflow-wrap:anywhere;
    font-size:.82rem;
    font-variant-numeric:tabular-nums;
}
.db-analytics-company-context{
    display:flex;
    flex-direction:column;
    gap:.18rem;
    margin-top:.2rem;
    padding:.65rem .72rem;
    border-left:3px solid var(--db-accent);
    background:var(--db-accent-soft);
    border-radius:7px;
}
.db-analytics-company-context strong{font-size:.9rem}
.db-analytics-empty{
    display:flex;
    flex-direction:column;
    gap:.18rem;
    min-height:150px;
    align-items:center;
    justify-content:center;
    padding:1rem;
    text-align:center;
    border:1px dashed var(--db-border);
    border-radius:8px;
    background:var(--db-soft);
}
.db-analytics-empty strong{font-size:.84rem}
.db-analytics-empty span{font-size:.76rem;opacity:.62}

/* Make dashboard tabs feel like primary views rather than incidental controls. */
div[data-testid="stTabs"] button[data-baseweb="tab"]{
    min-height:2.65rem;
    padding-left:1rem;
    padding-right:1rem;
    font-weight:700;
}
div[data-testid="stTabs"] div[data-baseweb="tab-list"]{
    gap:.25rem;
    border-bottom:1px solid var(--db-border);
}

@media(max-width:780px){
    .db-analytics-hero,
    .db-analytics-callout{
        flex-direction:column;
        align-items:stretch;
    }
    .db-analytics-baseline{min-width:0;flex-basis:auto}
    .db-analytics-chip-row{justify-content:flex-start}
    .db-analytics-source-strip{grid-template-columns:1fr}
    .db-analytics-kpi-grid{
        grid-template-columns:repeat(var(--db-kpi-count),minmax(132px,1fr));
    }
}
@media(max-width:480px){
    .db-analytics-kpi-grid{
        grid-template-columns:repeat(var(--db-kpi-count),minmax(126px,1fr));
    }
    .db-analytics-kpi{min-height:92px}
}


/* Compact progress summaries keep the sidebar informative rather than form-heavy. */
.db-next-action{
    display:flex;
    flex-direction:column;
    gap:.28rem;
    margin:.65rem 0 .65rem;
    padding:.78rem .82rem;
    border:1px solid var(--db-accent-edge);
    border-radius:10px;
    background:var(--db-accent-soft);
    line-height:1.4;
}
.db-next-action-label{
    font-size:.68rem;
    font-weight:800;
    letter-spacing:.08em;
    color:var(--db-accent-strong);
}
.db-next-action span{font-size:.82rem;opacity:.75}
.db-company-progress-row{
    margin:0 0 .58rem;
    padding:.64rem .68rem;
    border:1px solid var(--db-border);
    border-radius:9px;
    background:var(--db-soft);
}
.db-company-progress-row.selected{
    border-color:var(--db-accent-edge);
    background:var(--db-accent-soft);
    box-shadow:inset 3px 0 0 var(--db-accent);
}
.db-company-progress-head{
    display:flex;
    align-items:flex-start;
    justify-content:space-between;
    gap:.5rem;
}
.db-company-progress-name{
    min-width:0;
    font-size:.82rem;
    font-weight:750;
    line-height:1.25;
}
.db-company-progress-meta{
    margin:.2rem 0 .38rem;
    font-size:.72rem;
    opacity:.68;
}
.db-company-status{
    flex:0 0 auto;
    padding:.15rem .38rem;
    border-radius:999px;
    font-size:.62rem;
    font-weight:750;
    white-space:nowrap;
    background:rgba(90,100,115,.12);
}
.db-company-status.complete{background:rgba(38,145,85,.14)}
.db-company-status.needs-attention{background:rgba(205,91,65,.14)}
.db-company-status.ready-for-review{background:rgba(194,139,28,.14)}
.db-company-status.researching{background:var(--db-accent-soft)}
.db-mini-progress{
    height:.32rem;
    overflow:hidden;
    border-radius:999px;
    background:rgba(100,110,125,.15);
}
.db-mini-progress span{
    display:block;
    height:100%;
    border-radius:inherit;
    background:var(--db-accent);
}

/* Keep navigation scannable and equal in height. */
div[data-testid="stHorizontalBlock"] .stButton>button{
    line-height:1.2;
}
[data-testid="stTabs"] button{
    font-weight:650;
}
[data-testid="stCaptionContainer"]{
    line-height:1.45;
}

@media (max-width:760px){
    .db-workspace-strip{gap:.4rem .8rem}
}


/* =========================================================
   v96 SaaS shell — visual-only upgrade
   ========================================================= */
html, body, [data-testid="stAppViewContainer"]{
    background:#F5F7FB !important;
    color:#172033;
}
[data-testid="stHeader"]{
    background:rgba(245,247,251,.92) !important;
    border-bottom:1px solid rgba(15,23,42,.06);
}
.block-container{
    max-width:1500px;
    padding:1.45rem 2rem 4rem;
}

/* Sidebar shell */
[data-testid="stSidebar"]{
    width:288px !important;
    min-width:288px !important;
    background:#0C1824 !important;
    border-right:0 !important;
    box-shadow:8px 0 30px rgba(15,23,42,.14) !important;
}
[data-testid="stSidebar"] > div:first-child,
[data-testid="stSidebar"] [data-testid="stSidebarContent"]{
    background:#0C1824 !important;
}
[data-testid="stSidebar"] > div{
    background:#0C1824 !important;
}
[data-testid="stSidebar"] [data-testid="stSidebarContent"]{
    padding-top:.5rem;
}
[data-testid="stSidebar"] p,
[data-testid="stSidebar"] label,
[data-testid="stSidebar"] small,
[data-testid="stSidebar"] [data-testid="stCaptionContainer"]{
    color:#AFC0D0 !important;
}
.db-side-brand{
    display:flex;
    align-items:center;
    gap:.68rem;
    padding:.35rem .15rem 1rem;
    margin-bottom:.4rem;
    border-bottom:1px solid rgba(255,255,255,.08);
}
.db-side-brand-mark{
    display:flex;
    align-items:center;
    justify-content:center;
    width:2.05rem;
    height:2.05rem;
    border-radius:.62rem;
    background:linear-gradient(145deg,#35A7EA,#0E6BA4);
    color:white;
    font-family:var(--db-display);
    font-weight:800;
    box-shadow:0 8px 22px rgba(18,135,206,.28);
}
.db-side-brand-name{
    color:#F8FBFF;
    font-family:var(--db-display);
    font-weight:800;
    letter-spacing:.08em;
    font-size:.92rem;
}
.db-side-brand-sub{
    margin-top:.08rem;
    color:#6F879A;
    font-size:.52rem;
    font-weight:700;
    letter-spacing:.12em;
}
.db-side-empty{
    padding:.8rem;
    border:1px solid rgba(255,255,255,.08);
    border-radius:.7rem;
    color:#9FB2C2;
    background:rgba(255,255,255,.03);
}
.db-side-context-card{
    display:flex;
    flex-direction:column;
    gap:.22rem;
    padding:.72rem .78rem;
    margin:.45rem 0;
    border:1px solid rgba(255,255,255,.075);
    border-radius:.75rem;
    background:rgba(255,255,255,.035);
}
.db-side-context-card.company{margin-top:.35rem}
.db-side-context-card strong{
    color:#F4F8FC;
    font-size:.82rem;
    line-height:1.35;
    overflow-wrap:anywhere;
}
.db-side-kicker,
.db-side-section-label{
    color:#71889A;
    font-size:.56rem;
    font-weight:800;
    letter-spacing:.12em;
}
.db-side-context-meta{
    display:flex;
    gap:.65rem;
    margin-top:.2rem;
    color:#8EA4B6;
    font-size:.63rem;
}
.db-side-section-label{
    margin:1.05rem .12rem .4rem;
}
.db-side-section-label.progress{margin-top:1.15rem}
.db-side-section-label.account{margin-top:1.1rem}

[data-testid="stSidebar"] .stButton > button{
    min-height:2.35rem;
    justify-content:flex-start !important;
    padding:.44rem .7rem;
    border:1px solid transparent !important;
    border-radius:.58rem !important;
    box-shadow:none !important;
    font-size:.78rem;
    font-weight:650;
}
[data-testid="stSidebar"] button[data-testid="stBaseButton-secondary"]{
    color:#AFC0D0 !important;
    background:transparent !important;
}
[data-testid="stSidebar"] button[data-testid="stBaseButton-secondary"]:hover{
    color:#F7FBFF !important;
    background:rgba(255,255,255,.06) !important;
}
[data-testid="stSidebar"] button[data-testid="stBaseButton-primary"]{
    color:#FFFFFF !important;
    background:linear-gradient(135deg,#1287CE,#0E6BA4) !important;
    border-color:rgba(79,182,240,.24) !important;
    box-shadow:0 7px 20px rgba(3,105,161,.22) !important;
}
.db-side-progress-card{
    padding:.72rem .78rem;
    border:1px solid rgba(255,255,255,.075);
    border-radius:.75rem;
    background:rgba(255,255,255,.035);
}
.db-side-progress-head{
    display:flex;
    justify-content:space-between;
    gap:.6rem;
    color:#AFC0D0;
    font-size:.67rem;
}
.db-side-progress-head strong{color:#F5F9FD;font-variant-numeric:tabular-nums}
.db-side-progress-track{
    height:.32rem;
    margin:.52rem 0 .68rem;
    border-radius:999px;
    background:rgba(255,255,255,.09);
    overflow:hidden;
}
.db-side-progress-track span{
    display:block;
    height:100%;
    border-radius:999px;
    background:linear-gradient(90deg,#35A7EA,#5BC0EB);
}
.db-side-progress-stats{
    display:grid;
    grid-template-columns:repeat(3,1fr);
    gap:.3rem;
}
.db-side-progress-stats span{display:flex;flex-direction:column;gap:.08rem}
.db-side-progress-stats strong{color:#F5F9FD;font-size:.82rem}
.db-side-progress-stats small{color:#70889B !important;font-size:.55rem}
.db-side-next{
    display:flex;
    flex-direction:column;
    gap:.2rem;
    margin:.5rem 0 .45rem;
    padding:.68rem .72rem;
    border-left:3px solid #35A7EA;
    border-radius:.5rem;
    background:rgba(18,135,206,.11);
}
.db-side-next span{color:#68B9EB;font-size:.54rem;font-weight:800;letter-spacing:.1em}
.db-side-next strong{color:#F3F8FC;font-size:.72rem;line-height:1.35}
.db-side-next small{color:#8FA6B8 !important;font-size:.62rem;line-height:1.4}
[data-testid="stSidebar"] details{
    border-color:rgba(255,255,255,.08) !important;
    background:rgba(255,255,255,.025) !important;
}

/* Main workspace */
.db-workspace-strip{
    margin:0 0 1.05rem;
    padding:.62rem .85rem;
    border:1px solid #E3E9F0;
    border-left:0;
    border-radius:9px;
    background:#FFFFFF;
    box-shadow:0 1px 3px rgba(15,23,42,.03);
    color:#526173;
    font-size:.76rem;
}
.db-workspace-strip .db-num{color:#0E6BA4}
.db-nav-context{display:none}
.st-key-db_nav_row{display:none !important}

.db-page-head{
    max-width:none;
    margin:.15rem 0 1rem;
    padding:0 0 .78rem;
    border-bottom:0;
}
.db-page-head h2{font-size:clamp(1.55rem,2.5vw,2rem)}
.db-page-head p{max-width:920px;color:#667588;opacity:1}

/* Cards and bordered containers get the soft SaaS treatment. */
div[data-testid="stVerticalBlockBorderWrapper"]{
    border-color:#E2E8F0 !important;
    border-radius:12px !important;
    background:#FFFFFF;
    box-shadow:0 1px 3px rgba(15,23,42,.025);
}
div[data-testid="stMetric"]{
    background:#FFFFFF;
    border:1px solid #E2E8F0;
    border-top:0;
    border-radius:12px;
    box-shadow:0 1px 3px rgba(15,23,42,.03);
}
div[data-testid="stDataFrame"],div[data-testid="stDataEditor"]{
    border-color:#E2E8F0;
    background:#FFFFFF;
    border-radius:10px;
}

/* Dashboard hero */
.db-analytics-hero{
    align-items:center;
    margin:.1rem 0 .95rem;
    padding:.2rem 0 .65rem;
    border:0;
    border-radius:0;
    background:transparent;
}
.db-analytics-eyebrow{color:#4B86A8}
.db-analytics-hero h2{
    font-size:clamp(1.65rem,2.7vw,2.18rem);
    color:#172033;
}
.db-analytics-hero p{color:#687789;opacity:1}
.db-analytics-baseline{
    flex:0 1 300px;
    border:1px solid #DCE6EF;
    background:#FFFFFF;
    box-shadow:0 1px 3px rgba(15,23,42,.035);
}
.db-analytics-baseline span{color:#4B86A8}

/* KPI cards mimic the reference dashboard. */
.db-analytics-kpi-scroll{margin:.4rem 0 1rem}
.db-analytics-kpi-grid{gap:.72rem}
.db-analytics-kpi{
    min-height:118px;
    padding:.82rem .88rem;
    border:1px solid #E0E7EF;
    border-top:0;
    border-radius:12px;
    background:#FFFFFF;
    box-shadow:0 2px 7px rgba(15,23,42,.035);
}
.db-analytics-kpi::after{
    content:"";
    position:absolute;
    left:.88rem;
    right:.88rem;
    bottom:.72rem;
    height:3px;
    border-radius:99px;
    background:#D9E5EE;
}
.db-analytics-kpi.accent::after{background:#1287CE}
.db-analytics-kpi.positive::after{background:#16835F}
.db-analytics-kpi.warning::after{background:#C27C0E}
.db-analytics-kpi-top{
    display:flex;
    align-items:center;
    gap:.46rem;
}
.db-analytics-kpi-icon{
    display:inline-flex;
    align-items:center;
    justify-content:center;
    width:1.55rem;
    height:1.55rem;
    border-radius:.48rem;
    background:#EDF6FC;
    color:#0E6BA4;
    font-weight:800;
    font-size:.72rem;
}
.db-analytics-kpi.positive .db-analytics-kpi-icon{background:#E9F6F1;color:#16835F}
.db-analytics-kpi.warning .db-analytics-kpi-icon{background:#FFF5DD;color:#A86406}
.db-analytics-kpi.neutral .db-analytics-kpi-icon{background:#F0F3F6;color:#607080}
.db-analytics-kpi-label{min-height:auto;color:#526173;opacity:1}
.db-analytics-kpi-value{margin:.42rem 0 .12rem;font-size:1.72rem;color:#172033}
.db-analytics-kpi-helper{color:#7B8997;opacity:1;padding-bottom:.62rem}

.db-analytics-callout{
    border:1px solid #B9D8EA;
    border-left:4px solid #1287CE;
    border-radius:11px;
    background:#F0F8FD;
    box-shadow:0 1px 3px rgba(15,23,42,.025);
}
.db-analytics-section-head{border-bottom:0;margin:1.25rem 0 .55rem;padding-bottom:0}
.db-analytics-section-head h3{color:#243044}
.db-analytics-section-head p{color:#718093;opacity:1}

/* Tabs look like a clean segmented control. */
div[data-testid="stTabs"] div[data-baseweb="tab-list"]{
    display:inline-flex;
    gap:.18rem;
    padding:.2rem;
    border:1px solid #DFE6EE;
    border-radius:10px;
    background:#EEF2F6;
}
div[data-testid="stTabs"] button[data-baseweb="tab"]{
    min-height:2.25rem;
    padding:.4rem .8rem;
    border-radius:7px;
    font-size:.78rem;
}
div[data-testid="stTabs"] button[aria-selected="true"]{
    background:#FFFFFF;
    box-shadow:0 1px 3px rgba(15,23,42,.07);
}

.db-health-check{padding:.53rem 0}
.db-health-check strong{font-size:.79rem;color:#334155}
.db-health-check small{font-size:.68rem;color:#7B8997;line-height:1.35}

@media(max-width:1050px){
    [data-testid="stSidebar"]{width:260px !important;min-width:260px !important}
    .block-container{padding-left:1.2rem;padding-right:1.2rem}
}


/* v57 chart-free health dashboard */
.db-health-check { display:flex; gap:.7rem; align-items:flex-start; padding:.65rem 0; border-bottom:1px solid rgba(148,163,184,.22); }
.db-health-check:last-child { border-bottom:0; }
.db-health-check-icon { display:inline-flex; align-items:center; justify-content:center; width:1.35rem; height:1.35rem; border-radius:999px; font-weight:800; flex:0 0 auto; }
.db-health-check.pass .db-health-check-icon { color:#166534; background:#dcfce7; }
.db-health-check.warning .db-health-check-icon { color:#92400e; background:#fef3c7; }
.db-health-check.pending .db-health-check-icon { color:#475569; background:#e2e8f0; }
.db-health-check strong { display:block; font-size:.92rem; }
.db-health-check small { display:block; margin-top:.12rem; color:#64748b; line-height:1.35; }
</style>
""")
render_public_entry_gate()
render_auth_gate()
if user_is_authenticated():
    restore_autosaved_project()
if S_WORKING not in st.session_state:
    render_brand_header()
if st.session_state.get(S_DEMO_MODE):
    st.info("Demo workspace: sample information only. Changes are temporary and will not be saved.")


# -----------------------------
# Sidebar: project and company progress
# -----------------------------
with st.sidebar:
    render_project_progress_sidebar()
    st.caption(f"Datablix build: {DATABLIX_BUILD.rsplit('-', 1)[-1]}")


# -----------------------------
# Landing screen
# -----------------------------
if S_WORKING not in st.session_state:
    render_page_heading(
        "DATABLIX",
        "Your Rental Property Research & Data Audit Platform",
        "Transform public rental property research into clear, reliable, and review-ready information. Organize projects, research residential rental properties and property companies, check every finding against its public source, and generate clear analytics and reports for decision-making.",
    )
    render_guidance(
        "From public-source research to trusted property records",
        "Create a project, save each owner or management company under it, collect building observations, verify source evidence, resolve data-quality issues, and track progress through project and company analytics.",
    )

    journey = st.radio(
        "What would you like to do?",
        ["Start a new project", "Continue an existing project"],
        horizontal=True,
        key="db_landing_journey",
    )

    if journey == "Continue an existing project":
        with st.container(border=True):
            st.subheader("Continue a saved Datablix project")
            cloud_projects = list_cloud_projects()
            if cloud_projects:
                project_labels = {
                    f"{row.get('project_name', 'Datablix project')} — {str(row.get('updated_at', ''))[:16].replace('T', ' ')}": str(row.get('project_id', ''))
                    for row in cloud_projects
                }
                selected_cloud_label = st.selectbox(
                    "Projects saved permanently",
                    list(project_labels.keys()),
                    key="db_cloud_project_selector",
                )
                if st.button(
                    "Open selected project",
                    type="primary",
                    width="stretch",
                    key="db_open_cloud_project",
                ):
                    st.session_state.pop(S_SKIP_CLOUD_RESTORE, None)
                    if restore_cloud_project(project_labels[selected_cloud_label]):
                        st.rerun()
                    else:
                        st.error("The cloud project could not be opened.")
                st.divider()
            elif not cloud_persistence_available():
                st.info("Permanent cloud saving will activate after Supabase secrets are added.")
            st.write(
                "You can also open a master project workbook to restore its companies, building records, scan history, and progress."
            )
            landing_project = st.file_uploader(
                "Saved Datablix project",
                type=["xlsx"],
                key="db_landing_project_upload",
            )
            if landing_project is not None and st.button(
                "Continue project",
                type="primary",
                width="stretch",
                key="db_landing_resume_project",
            ):
                try:
                    load_project_workbook(landing_project)
                    st.rerun()
                except Exception as error:
                    st.error(str(error))
    else:
        start_method = st.radio(
            "How would you like to create the project?",
            [
                "Import assignment file",
                "Create manually",
                "Connect Google Sheet",
            ],
            horizontal=True,
            key="db_landing_start_method",
        )

        if start_method == "Import assignment file":
            with st.container(border=True):
                st.subheader("Create a project from an assignment file")
                st.write(
                    "Upload the spreadsheet supplied for the project. Datablix will first identify whether it contains assigned companies, existing building records, or both."
                )
                landing_upload = st.file_uploader(
                    "Assignment or building-data file",
                    type=["csv", "xlsx"],
                    key="db_landing_upload",
                )
                landing_sheet = None
                if landing_upload is not None:
                    if landing_upload.name.lower().endswith(".xlsx"):
                        landing_names = excel_sheet_names(landing_upload)
                        landing_sheet = st.selectbox(
                            "Worksheet containing the assignment",
                            landing_names,
                            index=preferred_sheet(landing_names),
                            key="db_landing_sheet",
                        )
                    if st.button(
                        "Create project from file",
                        type="primary",
                        width="stretch",
                        key="db_landing_import_project",
                    ):
                        try:
                            load_upload(landing_upload, landing_sheet)
                            go_to("Research projects & companies")
                            st.rerun()
                        except Exception as error:
                            st.error(str(error))

        elif start_method == "Create manually":
            with st.container(border=True):
                st.subheader("Create an empty project")
                st.write(
                    "Name the project first. Adding the first company now is optional; additional companies are registered from the Project page."
                )
                with st.form("db_landing_manual_project_form"):
                    landing_manual_project = st.text_input(
                        "Project name",
                        placeholder="Example: Rental Property Research Project",
                    )
                    landing_manual_company = st.text_input(
                        "First company or owner (optional)",
                        placeholder="Example: ABC Property Management",
                    )
                    landing_manual_website = st.text_input(
                        "Company website (optional)",
                        placeholder="https://example.ca",
                    )
                    landing_manual_notes = st.text_area(
                        "Notes (optional)",
                        height=90,
                    )
                    landing_manual_submit = st.form_submit_button(
                        "Create project",
                        type="primary",
                        width="stretch",
                    )
                if landing_manual_submit:
                    try:
                        create_manual_project(
                            landing_manual_project,
                            landing_manual_company,
                            landing_manual_website,
                            landing_manual_notes,
                        )
                        go_to("Research projects & companies")
                        st.rerun()
                    except Exception as error:
                        st.error(str(error))

        else:
            with st.container(border=True):
                st.subheader("Create a project from a Google Sheet")
                st.write(
                    "Use a viewable Sheet containing assigned companies or building records. Datablix opens a separate working copy and never edits the original Sheet."
                )
                with st.form("landing_google_form"):
                    landing_url = st.text_input(
                        "Google Sheets link",
                        placeholder="https://docs.google.com/spreadsheets/d/...",
                    )
                    landing_selector = st.text_input(
                        "Worksheet name or tab ID (optional)",
                        placeholder="Example: Apartment Buildings or 0",
                    )
                    landing_submit = st.form_submit_button(
                        "Create project from Sheet",
                        type="primary",
                        width="stretch",
                    )
                if landing_submit:
                    try:
                        if load_google(landing_url, landing_selector):
                            go_to("Research projects & companies")
                            st.rerun()
                    except Exception as error:
                        st.error(str(error))

    with smart_expander("How Datablix works", expanded=False):
        flow_columns = st.columns(4)
        flow_items = [
            ("Project", "Create or open the container for the assignment."),
            ("Company", "Register and select one company inside the project."),
            ("Research", "Generate the company research prompt, import the completed spreadsheet, or add a building manually."),
            ("Finish", "Review & quality, report, and export the project."),
        ]
        for column, (heading, copy) in zip(flow_columns, flow_items):
            with column:
                with st.container(border=True):
                    st.markdown(f"**{heading}**")
                    st.caption(copy)
    st.stop()


if S_FLASH in st.session_state:
    st.toast(st.session_state.pop(S_FLASH), icon="✅")

working, project_registry = synchronize_company_registry(
    st.session_state[S_WORKING].copy(),
    st.session_state.get(S_COMPANIES),
)
_comparison_baseline = current_starting_source_records()
working = classify_discovery_status(
    working,
    _comparison_baseline if not _comparison_baseline.empty else None,
)
st.session_state[S_WORKING] = working
st.session_state[S_COMPANIES] = project_registry
has_records = not working.empty
qa = qa_checks(working) if has_records else None

# -----------------------------
# Primary navigation
# -----------------------------
all_sections = [
    "Dashboard",
    "Research projects & companies",
    "Website scanner",
    "Review records",
    "Analysis & report",
    "Downloads",
]
primary_sections = all_sections.copy()
NAV_LABELS = {
    "Dashboard": "Dashboard",
    "Research projects & companies": "Project",
    "Website scanner": "Research",
    "Review records": "Review",
    "Analysis & report": "Deliverables",
    "Downloads": "Export",
}
PRIMARY_ACTIVE_SECTION = {section_name: section_name for section_name in all_sections}
legacy_sections = {
    "Review & edit": "Review records",
    "Research": "Website scanner",
    "Data quality": "Review records",
    "Export": "Downloads",
    "Review and edit records": "Review records",
    "Progress & quality": "Review records",
    "Progress and data quality": "Review records",
    "Download your work": "Downloads",
    "Analysis": "Analysis & report",
    "Report": "Analysis & report",
    "Overview": "Dashboard",
}
current_section = st.session_state.get("db_section", "Dashboard")
current_section = legacy_sections.get(current_section, current_section)
if current_section not in all_sections:
    current_section = "Research projects & companies"
st.session_state["db_section"] = current_section

project_name_display = str(
    st.session_state.get(S_PROJECT_NAME, "Datablix project")
).strip() or "Datablix project"
active_header_company = active_company_row()
active_header_name = (
    str(active_header_company.get("Management/Owner", "")).strip()
    if active_header_company is not None
    else "No company selected"
)
workspace_source = st.session_state.get(S_SOURCE_TYPE, "Project")
workspace_name = st.session_state.get(S_NAME, "project")
workspace_sheet = st.session_state.get(S_SHEET, "")
workspace_display = workspace_name + (
    f" · {workspace_sheet}" if workspace_sheet else ""
)

st.markdown(
    (
        '<div class="db-workspace-strip">'
        f'<span><strong>Project:</strong> {escape(project_name_display)}</span>'
        f'<span><strong>Selected company:</strong> {escape(active_header_name)}</span>'
        f'<span><strong>Companies:</strong> <span class="db-num">{len(project_registry):,}</span></span>'
        f'<span><strong>Buildings:</strong> <span class="db-num">{len(working):,}</span></span>'
        '</div>'
    ),
    unsafe_allow_html=True,
)

visible_active_section = PRIMARY_ACTIVE_SECTION[st.session_state["db_section"]]

# Progress is shown directly in the main navigation so there is only one workflow row.
active_company_id = (
    str(active_header_company.get("Company ID", "")).strip()
    if active_header_company is not None
    else ""
)
company_records = (
    working.loc[working["Company ID"].astype(str).eq(active_company_id)].copy()
    if active_company_id and "Company ID" in working.columns
    else working.iloc[0:0].copy()
)
company_qa = (
    qa.loc[qa["Company ID"].astype(str).eq(active_company_id)].copy()
    if active_company_id and isinstance(qa, pd.DataFrame) and "Company ID" in qa.columns
    else pd.DataFrame()
)
review_population = (
    company_qa.loc[~company_qa["Record Decision"].eq("Remove")].copy()
    if not company_qa.empty and "Record Decision" in company_qa.columns
    else company_qa
)

NAV_DESCRIPTIONS = {
    "Dashboard": "Monitor project health, source reconciliation, review progress, and the next actions that matter most.",
    "Research projects & companies": "Set up your project and company workspaces.",
    "Website scanner": "Research the selected company and add or import building records.",
    "Review records": "Review & Quality — verify records, resolve quality issues, and approve clean records for export.",
    "Analysis & report": "Close out the project with research coverage, field availability, justified directory recommendations, verification gaps, and a final evidence report.",
    "Downloads": "Choose the company, records, and columns, preview them, then download CSV.",
}

# Primary navigation is rendered in the left SaaS sidebar.
section = st.session_state["db_section"]
if st.session_state.get(S_PROJECT_ROLE) == "viewer":
    st.info("You have view-only access to this project. Ask the owner for Editor access to make changes.")


if not has_records and section in ["Analysis & report", "Downloads"]:
    st.info(
        "This project has no building records yet. Select a company, generate its research prompt, import the completed spreadsheet, or add the first building manually."
    )
    action_a, action_b = st.columns(2)
    if action_a.button("Open company research", type="primary", width="stretch"):
        go_to("Website scanner")
        st.rerun()
    if action_b.button("Add building manually", width="stretch"):
        st.session_state[S_MANUAL_ENTRY_OPEN] = True
        go_to("Review records")
        st.rerun()
    st.stop()


# -----------------------------
# Dashboard
# -----------------------------
if section == "Dashboard":
    render_project_company_analytics(project_registry, working)

# -----------------------------
# Project and company setup
# -----------------------------
elif section == "Research projects & companies":
    project_context_token = hashlib.sha256(
        str(st.session_state.get(S_FILE, "project")).encode("utf-8")
    ).hexdigest()[:10]
    render_page_heading(
        "PROJECT",
        "Research project",
        "Set up the project, import its starting data once, manage companies, and continue from the next recommended action.",
    )
    render_guidance(
        "One research project contains many saved company workspaces.",
        "Each company keeps its website, editable research prompt, imported building records, review progress, and optional scanner history under the same project.",
    )

    project_snapshot = project_progress_snapshot(project_registry, working)
    with st.container(border=True):
        project_header, project_edit = st.columns([3, 1], vertical_alignment="center")
        with project_header:
            st.caption("CURRENT PROJECT")
            st.subheader(
                str(st.session_state.get(S_PROJECT_NAME, "Datablix project")).strip()
                or "Datablix project"
            )
            if project_snapshot["companies"]:
                st.progress(
                    project_snapshot["progress"],
                    text=(
                        f"{project_snapshot['completed']:,} of "
                        f"{project_snapshot['companies']:,} companies complete"
                    ),
                )
            else:
                st.progress(0.0, text="Add the first company to begin")
        with project_edit:
            st.caption("Project structure")
            st.markdown("**Project → Company → Buildings**")

        project_metrics = st.columns(4)
        project_metrics[0].metric("Companies", f"{project_snapshot['companies']:,}")
        project_metrics[1].metric("Complete", f"{project_snapshot['completed']:,}")
        project_metrics[2].metric("Buildings", f"{project_snapshot['buildings']:,}")
        project_metrics[3].metric(
            "Need attention", f"{project_snapshot['attention_records']:,}"
        )
        st.caption(
            f"{project_snapshot['in_progress']:,} companies in progress · "
            f"{project_snapshot['not_started']:,} not started · "
            f"{project_snapshot['verified_records']:,} building records verified"
        )

    with st.expander("Edit project name", expanded=False):
        with st.form(f"db_project_name_form_{project_context_token}"):
            project_name_main = st.text_input(
                "Project name",
                value=st.session_state.get(S_PROJECT_NAME, "Datablix project"),
            )
            save_project_name = st.form_submit_button(
                "Save project name",
                type="primary",
                width="stretch",
            )
        if save_project_name:
            clean_project_name = project_name_main.strip()
            if clean_project_name:
                st.session_state[S_PROJECT_NAME] = clean_project_name
                st.session_state[S_FLASH] = "Project name saved."
                st.rerun()
            else:
                st.error("Enter a project name.")

    # One-time Starting Data setup. Internally this creates the source baseline
    # used to distinguish records that already existed from later discoveries.
    source_meta = st.session_state.get(
        S_SOURCE_BASELINE_META,
        {},
    )
    if not isinstance(source_meta, dict):
        source_meta = {}

    source_versions = _source_versions_state()

    with st.container(border=True):
        st.markdown("### Starting Data")

        if source_meta:
            assigned_count = _safe_int(
                source_meta.get("assigned_companies", 0)
            )
            source_count = _safe_int(
                source_meta.get("source_records", 0)
            )
            assignment_name = safe_text(
                source_meta.get("assignment_sheet", "")
            )
            workbook_name = (
                safe_text(
                    source_meta.get(
                        "workbook_name",
                        "Source workbook",
                    ),
                    "Source workbook",
                )
                or "Source workbook"
            )
            current_version = (
                safe_text(
                    source_meta.get("version_label", "")
                )
                or f"v{_safe_int(source_meta.get('version_number', 1), 1)}"
            )

            st.success(
                f"Starting data ready · {workbook_name} · {source_count:,} structured record(s)"
            )
            st.caption(
                "This is the only current Starting Data baseline for the project. "
                "AI website research does not receive this file; Datablix compares imported research against it afterward."
            )

            with st.expander("View starting data details", expanded=False):
                source_metrics = st.columns(2)
                source_metrics[0].metric("Current structured rows", f"{source_count:,}")
                source_metrics[1].metric("Assigned companies", f"{assigned_count:,}")
                st.caption(
                    f"Assignment: {assignment_name or 'Not recorded'} · Current workbook: {workbook_name}"
                )
                rules = st.session_state.get(S_CLASSIFICATION_RULES)
                if isinstance(rules, pd.DataFrame) and not rules.empty:
                    st.markdown("**Current building classification rules**")
                    st.dataframe(rules, width="stretch", hide_index=True)

            with st.expander("Manage current source", expanded=False):
                st.code(
                    f"{workbook_name} · {source_count:,} records",
                    language=None,
                )
                st.caption(
                    "Replacing this source automatically removes all older Starting Data files/baselines. "
                    "Your saved research records are preserved and re-compared against the replacement source."
                )
                remove_source_confirm = st.checkbox(
                    "Confirm removal of the current Starting Data",
                    key="db_remove_current_source_confirm",
                )
                if st.button(
                    "Remove current source",
                    type="secondary",
                    disabled=not remove_source_confirm,
                    key="db_remove_current_source_button",
                    width="stretch",
                ):
                    removed, message = clear_current_starting_source()
                    if removed:
                        st.session_state[S_FLASH] = message
                        st.rerun()
                    st.info(message)

            starting_data_expander_label = "Replace starting data"
            starting_data_expanded = False
        else:
            st.info(
                "Import the current project-wide Starting Data once. Datablix keeps it as a comparison baseline; "
                "the AI website-research prompt stays independent of it."
            )
            starting_data_expander_label = "Import starting data"
            starting_data_expanded = True

        with st.expander(
            starting_data_expander_label,
            expanded=starting_data_expanded,
        ):
            st.write(
                "Upload the project-wide Starting Data workbook. Datablix will keep only this current source "
                "for comparison. Replacing an existing source removes the old source baseline but preserves research records."
            )
            if source_meta:
                st.warning(
                    "Replacing Starting Data will remove all previously saved source files/baselines for this project. "
                    "Your company research, evidence, notes, and review work will remain."
                )
            source_workbook_upload = st.file_uploader(
                "Project source file",
                type=["xlsx"],
                key=f"db_source_baseline_upload_{project_context_token}",
            )
            if source_workbook_upload is not None:
                source_assignment_sheet = ""
                try:
                    workbook_sheets = excel_sheet_names(source_workbook_upload)
                    assignment_options = source_assignment_sheet_candidates(source_workbook_upload)
                except Exception as error:
                    workbook_sheets = []
                    assignment_options = []
                    st.error(str(error))

                if len(workbook_sheets) > 1:
                    company_scope_options = ["Use current project companies"] + [
                        f"Use assignment sheet: {name}" for name in assignment_options
                    ]
                    company_scope_choice = st.selectbox(
                        "Company scope for matching",
                        company_scope_options,
                        key=f"db_source_company_scope_{project_context_token}",
                        help=(
                            "Starting Data belongs to the whole project. Use your current Datablix "
                            "company list by default. Choose an assignment sheet only when you need "
                            "Datablix to add or update companies from that sheet."
                        ),
                    )
                    if company_scope_choice.startswith("Use assignment sheet: "):
                        source_assignment_sheet = company_scope_choice.split(": ", 1)[1]
                elif len(workbook_sheets) == 1:
                    st.caption(
                        f"Single project-source worksheet detected: {workbook_sheets[0]}. "
                        "No assignment worksheet is required."
                    )

                if has_records:
                    st.caption(
                        "Your current research records will not be discarded. Datablix keeps the "
                        "project source separately, reconciles matching source records, and preserves "
                        "your reviewed values."
                    )

                import_label = "Replace starting data" if source_meta else "Import starting data"
                if st.button(import_label, type="primary", width="stretch", key=f"db_import_source_baseline_{project_context_token}"):
                    try:
                        result = import_source_baseline_workbook(source_workbook_upload, source_assignment_sheet)
                        structured_count = int(result.get("source_records", 0) or 0)
                        relevant_count = int(result.get("project_company_source_records", 0) or 0)
                        action_word = "replaced" if result.get("replaced_existing") else "imported"
                        if structured_count:
                            st.session_state[S_FLASH] = (
                                f"Starting Data {action_word}. Datablix parsed {structured_count:,} project source row(s), "
                                f"with {relevant_count:,} matching the current project companies. "
                                "This is now the only source baseline; saved research was preserved and re-compared against it."
                            )
                        else:
                            st.session_state[S_FLASH] = (
                                f"Starting Data {action_word}. The uploaded workbook is now the only source baseline. "
                                "Datablix could not reliably convert it into structured building rows, so automatic record comparison may be limited."
                            )
                        st.rerun()
                    except Exception as error:
                        st.error("Starting data import could not be completed. " + str(error))

    st.subheader("Companies in this project")
    registry_main = normalize_company_registry(st.session_state.get(S_COMPANIES))
    company_table = company_progress_table(registry_main, working, project_snapshot)
    if company_table.empty:
        st.info(
            "No companies are registered. Add the first company below, or start a different project and import its assignment file."
        )
    else:
        with smart_expander(
            "All companies in this project",
            count=len(company_table),
            status="progress overview",
            expanded=False,
        ):
            st.dataframe(
                company_table.drop(columns=["Company ID"]),
                width="stretch",
                hide_index=True,
                column_config={
                    "Company": st.column_config.TextColumn("Company", width="large"),
                    "Website": st.column_config.TextColumn("Website", width="medium"),
                    "Buildings": st.column_config.NumberColumn("Buildings", format="%d"),
                    "Reviewed": st.column_config.NumberColumn("Reviewed", format="%d"),
                    "Verified": st.column_config.NumberColumn("Verified", format="%d"),
                    "Needs attention": st.column_config.NumberColumn(
                        "Needs attention", format="%d"
                    ),
                    "Progress": st.column_config.TextColumn("Progress"),
                    "Status": st.column_config.TextColumn("Status"),
                    "Research prompt": st.column_config.TextColumn(
                        "Research prompt", width="small"
                    ),
                    "Next action": st.column_config.TextColumn(
                        "Next action", width="large"
                    ),
                },
            )

        st.subheader("Choose the company to work on")
        main_ids = registry_main["Company ID"].astype(str).tolist()
        main_selector_key = f"db_main_active_company_{project_context_token}"
        pending_main_id = str(
            st.session_state.pop(S_PENDING_ACTIVE_COMPANY, "")
        ).strip()
        if pending_main_id in main_ids:
            st.session_state[S_ACTIVE_COMPANY] = pending_main_id
            st.session_state.pop(main_selector_key, None)

        current_main_id = str(st.session_state.get(S_ACTIVE_COMPANY, "")).strip()
        main_index = main_ids.index(current_main_id) if current_main_id in main_ids else 0
        selected_main_id = st.selectbox(
            "Company to research",
            main_ids,
            index=main_index,
            format_func=lambda company_id: company_label(
                registry_main.loc[
                    registry_main["Company ID"].eq(company_id)
                ].iloc[0]
            ),
            key=main_selector_key,
        )
        st.session_state[S_ACTIVE_COMPANY] = selected_main_id
        selected_company_row = registry_main.loc[
            registry_main["Company ID"].eq(selected_main_id)
        ].iloc[0]
        selected_snapshot = company_progress_snapshot(selected_company_row, working)

        with st.container(border=True):
            selected_head, selected_progress = st.columns(
                [2.2, 1], vertical_alignment="center"
            )
            with selected_head:
                st.caption("SELECTED COMPANY")
                st.subheader(selected_snapshot["company_name"])
                st.caption(
                    f"{selected_snapshot['company_id']} · "
                    f"{selected_snapshot['status']} · "
                    f"Website: {selected_snapshot['website'] or 'Not registered'}"
                )
            with selected_progress:
                if selected_snapshot["collected"]:
                    st.progress(
                        selected_snapshot["progress"],
                        text=(
                            f"{selected_snapshot['verified']:,} of "
                            f"{selected_snapshot['collected']:,} verified"
                        ),
                    )
                else:
                    st.progress(0.0, text="Research not started")

            selected_metrics = st.columns(4)
            selected_metrics[0].metric(
                "Collected", f"{selected_snapshot['collected']:,}"
            )
            selected_metrics[1].metric(
                "Reviewed", f"{selected_snapshot['reviewed']:,}"
            )
            selected_metrics[2].metric(
                "Verified", f"{selected_snapshot['verified']:,}"
            )
            selected_metrics[3].metric(
                "Need attention", f"{selected_snapshot['attention']:,}"
            )

            st.markdown(
                f'<div class="db-next-action">'
                f'<div class="db-next-action-label">NEXT RECOMMENDED ACTION</div>'
                f'<strong>{escape(selected_snapshot["next_title"])}</strong>'
                f'<span>{escape(selected_snapshot["next_copy"])}</span>'
                f'</div>',
                unsafe_allow_html=True,
            )

            next_action_col, alternate_action_col = st.columns(2)
            if next_action_col.button(
                selected_snapshot["next_button"],
                type="primary",
                width="stretch",
                key=f"db_project_next_{selected_main_id}",
            ):
                go_to(selected_snapshot["next_section"])
                st.rerun()
            alternate_label = (
                "Add building manually"
                if selected_snapshot["next_section"] != "Review records"
                else "Open company research"
            )
            if alternate_action_col.button(
                alternate_label,
                width="stretch",
                key=f"db_project_alternate_{selected_main_id}",
            ):
                if alternate_label == "Add building manually":
                    st.session_state[S_MANUAL_ENTRY_OPEN] = True
                    go_to("Review records")
                else:
                    go_to("Website scanner")
                st.rerun()

        company_setup_incomplete = (
            not bool(selected_snapshot["website"])
            or selected_snapshot["status"] == "Not started"
        )
        with smart_expander(
            "Edit selected company details",
            status="setup incomplete" if company_setup_incomplete else "optional",
            expanded=company_setup_incomplete,
        ):
            with st.form(
                f"db_main_company_details_{project_context_token}_{selected_main_id}"
            ):
                detail_left, detail_right = st.columns(2)
                selected_website = detail_left.text_input(
                    "Official company website",
                    value=str(selected_company_row.get("Main Website", "")).strip(),
                    placeholder="https://example.ca",
                )
                selected_status_value = str(
                    selected_company_row.get("Company Status", "Not started")
                )
                selected_status_index = (
                    COMPANY_STATUSES.index(selected_status_value)
                    if selected_status_value in COMPANY_STATUSES
                    else 0
                )
                selected_company_status = detail_right.selectbox(
                    "Internal company status",
                    COMPANY_STATUSES,
                    index=selected_status_index,
                    help="Datablix presents a simplified status in progress views while preserving this detailed status in the project data.",
                )
                selected_related_links = st.text_area(
                    "Related official links (optional)",
                    value=str(
                        selected_company_row.get("Related Official Links", "")
                    ),
                    height=105,
                    placeholder=(
                        "One official URL per line, for example:\n"
                        "https://property.example.ca/\n"
                        "https://example.ca/property-name/"
                    ),
                    help=(
                        "Use for official property microsites, subdomains, or property paths "
                        "that belong to this company. Leave blank for ordinary one-site companies."
                    ),
                )
                selected_special_website_notes = st.text_area(
                    "Special website notes (optional)",
                    value=str(
                        selected_company_row.get("Special Website Notes", "")
                    ),
                    height=90,
                    placeholder=(
                        "Example: Property communities use separate official subdomains "
                        "but belong to the same management company."
                    ),
                )
                selected_company_notes = st.text_area(
                    "General company notes",
                    value=str(selected_company_row.get("Notes", "")),
                    height=90,
                )
                save_company_details = st.form_submit_button(
                    "Save company details",
                    type="primary",
                    width="stretch",
                )
            if save_company_details:
                normalized_related_links, invalid_related_links = (
                    _parse_related_official_links(
                        selected_related_links,
                        selected_website,
                    )
                )
                if invalid_related_links:
                    st.error(
                        "Fix these related official links before saving: "
                        + ", ".join(invalid_related_links)
                    )
                    st.stop()
                registry_main.loc[
                    registry_main["Company ID"].eq(selected_main_id),
                    [
                        "Main Website", "Related Official Links",
                        "Special Website Notes", "Company Status", "Notes",
                    ],
                ] = [
                    selected_website.strip(),
                    "\n".join(normalized_related_links),
                    selected_special_website_notes.strip(),
                    selected_company_status,
                    selected_company_notes.strip(),
                ]
                st.session_state[S_COMPANIES] = normalize_company_registry(
                    registry_main
                )
                st.session_state[S_FLASH] = "Company details saved."
                autosave_current_project()
                st.rerun()

        project_id_for_company_delete = str(
            st.session_state.get(S_CLOUD_PROJECT_ID, "")
        ).strip()
        company_delete_role = str(
            st.session_state.get(S_PROJECT_ROLE, "")
            or (
                project_access_role(project_id_for_company_delete)
                if project_id_for_company_delete
                else "owner"
            )
        ).strip().lower()
        company_delete_allowed = (
            st.session_state.get(S_DEMO_MODE)
            or not project_id_for_company_delete
            or company_delete_role == "owner"
        )

        with st.expander("Delete selected company", expanded=False):
            if not company_delete_allowed:
                st.caption(
                    "Only the project owner can permanently delete a company "
                    "from this shared project."
                )
            else:
                company_rows_to_delete = _company_row_mask(
                    working,
                    company_id=selected_main_id,
                    company_name=selected_snapshot["company_name"],
                )
                company_record_count = int(company_rows_to_delete.sum())

                st.warning(
                    f'This removes "{selected_snapshot["company_name"]}" from this project '
                    f'and permanently removes its {company_record_count:,} associated '
                    "building record(s), scan history, and saved "
                    "company scanner state. The project and other companies are not affected."
                )

                delete_company_ack = st.checkbox(
                    "I understand that this company and its associated project data will be deleted",
                    key=f"db_delete_company_ack_{selected_main_id}",
                )
                delete_company_name = st.text_input(
                    f'Type the company name to confirm: "{selected_snapshot["company_name"]}"',
                    key=f"db_delete_company_name_{selected_main_id}",
                    autocomplete="off",
                )
                company_name_matches = (
                    delete_company_name.strip()
                    == str(selected_snapshot["company_name"]).strip()
                )
                if delete_company_name.strip() and not company_name_matches:
                    st.caption(
                        "The confirmation name does not match the selected company."
                    )

                if st.button(
                    "Delete selected company permanently",
                    type="secondary",
                    width="stretch",
                    disabled=not (delete_company_ack and company_name_matches),
                    key=f"db_delete_company_button_{selected_main_id}",
                ):
                    deleted, delete_message, delete_stats = delete_company_from_project(
                        selected_main_id
                    )
                    if deleted:
                        st.session_state[S_FLASH] = (
                            f'{delete_message} '
                            f'{delete_stats.get("records_removed", 0):,} building record(s) '
                            "were removed. The rest of the project is unchanged."
                        )
                        st.rerun()
                    else:
                        st.error(delete_message)

    with st.expander(
        "Add another company to this project",
        expanded=registry_main.empty,
    ):
        st.write(
            "Register the company here so all future scans and building records can inherit the correct project and company context."
        )
        with st.form("db_main_add_company_form", clear_on_submit=True):
            company_form_left, company_form_right = st.columns(2)
            main_new_company = company_form_left.text_input(
                "Company or owner name",
                placeholder="Example: ABC Property Management",
            )
            main_new_website = company_form_right.text_input(
                "Official website (optional)",
                placeholder="https://example.ca",
            )
            main_new_scope = company_form_left.selectbox(
                "How was it added?",
                ["Initial assignment", "Added later"],
            )
            main_new_related_links = company_form_right.text_area(
                "Related official links (optional)",
                height=90,
                placeholder="One official URL per line",
                help=(
                    "Add official property microsites, subdomains, or property paths only "
                    "when they are already known. Leave blank for one-site companies."
                ),
            )
            main_new_special_notes = company_form_left.text_area(
                "Special website notes (optional)",
                height=75,
            )
            main_new_notes = company_form_right.text_area(
                "General notes (optional)",
                height=75,
            )
            main_add_company = st.form_submit_button(
                "Add company to project",
                type="primary",
                width="stretch",
            )
        if main_add_company:
            try:
                normalized_new_links, invalid_new_links = (
                    _parse_related_official_links(
                        main_new_related_links,
                        main_new_website,
                    )
                )
                if invalid_new_links:
                    raise ValueError(
                        "Fix these related official links: "
                        + ", ".join(invalid_new_links)
                    )
                new_company_id, company_created = add_company_to_project(
                    main_new_company,
                    main_new_website,
                    main_new_scope,
                    main_new_notes,
                    related_official_links="\n".join(normalized_new_links),
                    special_website_notes=main_new_special_notes,
                )
                st.session_state[S_FLASH] = (
                    f"Added {main_new_company.strip()} as {new_company_id}."
                    if company_created
                    else f"{main_new_company.strip()} was already registered and is now selected."
                )
                st.rerun()
            except Exception as error:
                st.error(str(error))

    st.divider()
    with st.expander("Project administration", expanded=False):
        st.caption(
            "Save the current master project before replacing it in this browser session."
        )
        administration_columns = st.columns(2)
        if administration_columns[0].button(
            "Save project",
            width="stretch",
            key="db_project_admin_save",
        ):
            go_to("Downloads")
            st.rerun()

        confirm_new_project = st.checkbox(
            "I saved my work and want to open or create a different project",
            key="db_confirm_return_to_project_start",
        )
        if administration_columns[1].button(
            "Start a different project",
            disabled=not confirm_new_project,
            width="stretch",
            key="db_return_to_project_start",
        ):
            return_to_project_start()
            st.rerun()

        project_id_for_admin = str(
            st.session_state.get(S_CLOUD_PROJECT_ID, "")
        ).strip()
        project_name_for_admin = str(
            st.session_state.get(S_PROJECT_NAME, "Datablix project")
        ).strip() or "Datablix project"
        project_role_for_admin = str(
            st.session_state.get(S_PROJECT_ROLE, "")
            or project_access_role(project_id_for_admin)
        ).strip().lower()

        if (
            project_id_for_admin
            and project_role_for_admin == "owner"
            and not st.session_state.get(S_DEMO_MODE)
        ):
            st.divider()
            st.markdown("#### Danger zone")
            st.warning(
                "Deleting a project is permanent. Its saved cloud workspace and "
                "project-member access records will be removed. Other projects are not affected."
            )

            delete_acknowledged = st.checkbox(
                "I understand that this project will be permanently deleted",
                key="db_confirm_project_delete_ack",
            )
            delete_name = st.text_input(
                f'Type the project name to confirm: "{project_name_for_admin}"',
                key="db_confirm_project_delete_name",
                autocomplete="off",
            )
            name_matches = (
                delete_name.strip() == project_name_for_admin
            )

            if delete_name.strip() and not name_matches:
                st.caption("The confirmation name does not match the current project.")

            if st.button(
                "Delete this project permanently",
                type="secondary",
                width="stretch",
                disabled=not (delete_acknowledged and name_matches),
                key="db_delete_project_permanently",
            ):
                deleted_project_name = project_name_for_admin
                deleted, delete_message = delete_cloud_project(project_id_for_admin)
                if deleted:
                    return_to_project_start()
                    st.session_state[S_FLASH] = (
                        f'Project "{deleted_project_name}" was permanently deleted.'
                    )
                    st.rerun()
                else:
                    st.error(delete_message)
        elif project_id_for_admin and project_role_for_admin != "owner":
            st.divider()
            st.caption(
                "Only the project owner can permanently delete this shared project."
            )


# -----------------------------
# Overview
# -----------------------------
elif section == "Overview":
    render_page_heading(
        "WORKSPACE",
        "Workspace overview",
        "See what has been collected, what still needs review, and what is approved for export.",
    )

    next_title, next_copy, next_section, next_button = recommended_next_action(qa)
    with st.container(border=True):
        next_left, next_right = st.columns([2.2, 1], vertical_alignment="center")
        with next_left:
            st.subheader(next_title)
            st.write(next_copy)
        with next_right:
            if st.button(
                next_button,
                type="primary",
                width="stretch",
                key="db_overview_next",
            ):
                go_to(next_section)
                st.rerun()

    if not has_records:
        st.info(
            "This workspace is empty. Generate a company research prompt, import a completed spreadsheet, or add a listing manually to begin."
        )
        quick_project, quick_scan, quick_manual = st.columns(3)
        if quick_project.button(
            "Manage project & companies",
            width="stretch",
            key="overview_project_empty",
        ):
            go_to("Research projects & companies")
            st.rerun()
        if quick_scan.button(
            "Open company research",
            type="primary",
            width="stretch",
            key="overview_scan_empty",
        ):
            go_to("Website scanner")
            st.rerun()
        if quick_manual.button(
            "Add building manually",
            width="stretch",
            key="overview_manual_empty",
        ):
            st.session_state[S_MANUAL_ENTRY_OPEN] = True
            go_to("Review records")
            st.rerun()
    else:
        metric_columns = st.columns(4)
        metric_columns[0].metric("Records", f"{len(qa):,}")
        metric_columns[1].metric("Approved for Export", f"{int(approved_for_export_mask(qa).sum()):,}")
        metric_columns[2].metric(
            "Need attention",
            f"{int((~ready_mask(qa) & ~qa['Record Readiness'].eq('Excluded from Listings')).sum()):,}",
        )
        metric_columns[3].metric(
            "Human verified",
            f"{int(qa['Verification Status'].eq('Verified').sum()):,}",
        )

        completed = int(qa["Research Status"].eq("Completed").sum())
        st.progress(
            completed / len(qa),
            text=f"Research complete: {completed:,} of {len(qa):,} records",
        )

        quick_1, quick_2, quick_3, quick_4 = st.columns(4)
        if quick_1.button("Manage companies", width="stretch"):
            go_to("Research projects & companies")
            st.rerun()
        if quick_2.button("Company research", width="stretch"):
            go_to("Website scanner")
            st.rerun()
        if quick_3.button("Add building manually", width="stretch"):
            st.session_state[S_MANUAL_ENTRY_OPEN] = True
            go_to("Review records")
            st.rerun()
        if quick_4.button("Review & quality", width="stretch"):
            go_to("Review records")
            st.rerun()

        preview_count = min(len(qa), 5)
        with smart_expander(
            "Listing preview",
            count=preview_count,
            status=f"of {len(qa):,} records",
            expanded=False,
        ):
            st.caption(
                "Each record follows the required field-and-value layout. Required listing fields appear first, followed by confirmed additional findings."
            )
            render_listing_preview(qa, limit=5)

        mapped_count = int(
            st.session_state[S_MAPPING]["Mapping Status"].ne("Not found").sum()
        )
        with smart_expander(
            "Workspace details and column matching",
            count=mapped_count,
            status="fields mapped",
            expanded=False,
        ):
            detail_columns = st.columns(3)
            detail_columns[0].metric("Source type", workspace_source)
            detail_columns[1].metric("Original columns", f"{len(working.columns):,}")
            detail_columns[2].metric(
                "Mapped fields",
                f"{mapped_count:,}",
            )
            st.caption(
                "Your original columns remain in the working data. This table shows how imported headings were matched to consistent rental property fields."
            )
            st.dataframe(
                st.session_state[S_MAPPING],
                width="stretch",
                hide_index=True,
                height=360,
            )


# -----------------------------
# Company research
# -----------------------------
elif section == "Website scanner":
    active_company = active_company_row()
    if active_company is None:
        render_page_heading(
            "RESEARCH",
            "Select a company before researching",
            "Each research prompt, imported deliverable, and optional website scan must belong to one registered company.",
        )
        st.error(
            "No company is selected. Register or select a company so every imported finding remains attached to the correct organization."
        )
        missing_company_setup, missing_company_manual = st.columns(2)
        if missing_company_setup.button(
            "Register or select company",
            type="primary",
            width="stretch",
            key="db_research_missing_company_setup",
        ):
            go_to("Research projects & companies")
            st.rerun()
        if missing_company_manual.button(
            "Add building manually instead",
            width="stretch",
            key="db_research_missing_company_manual",
        ):
            st.session_state[S_MANUAL_ENTRY_OPEN] = True
            go_to("Review records")
            st.rerun()
        st.stop()

    company_id = str(active_company["Company ID"]).strip()
    company_name = str(active_company["Management/Owner"]).strip()
    company_website = str(active_company.get("Main Website", "")).strip()
    company_related_links = str(
        active_company.get("Related Official Links", "") or ""
    ).strip()
    company_special_website_notes = str(
        active_company.get("Special Website Notes", "") or ""
    ).strip()

    render_page_heading(
        "RESEARCH",
        "Company website research",
        "Generate one strong editable prompt, use it with the AI tool of your choice, and import the completed CSV into Datablix for validation and human review.",
    )
    st.caption(f"Workspace build: {DATABLIX_BUILD}")

    st.subheader("Company workspace")
    with st.container(border=True):
        context_left, context_right = st.columns([2.3, 1], vertical_alignment="center")
        with context_left:
            st.caption("ACTIVE COMPANY")
            st.markdown(f"**{company_name}** · {company_id}")
            st.caption(f"Official website: {company_website or 'Not recorded yet'}")
            related_link_count = len(
                _parse_related_official_links(
                    company_related_links,
                    company_website,
                )[0]
            )
            if related_link_count:
                st.caption(
                    f"Additional official entry points: {related_link_count} saved"
                )
        with context_right:
            if st.button(
                "Edit company details",
                width="stretch",
                key=f"db_research_edit_company_{company_id}",
            ):
                go_to("Research projects & companies")
                st.rerun()

    st.subheader("1. Prepare the website research prompt")
    st.caption(
        "Datablix personalizes one comprehensive prompt for this company. The prompt requires CSV output and remains editable before you copy or download it."
    )

    company_rows = working.loc[
        working["Company ID"].astype("string").fillna("").str.strip().eq(company_id)
    ].copy()

    active_source_version = _active_source_version()
    source_meta = (
        dict(active_source_version.get("meta", {}))
        if isinstance(active_source_version, dict)
        else {}
    )
    project_source_records = (
        active_source_version.get("records")
        if isinstance(active_source_version, dict)
        else pd.DataFrame()
    )
    if not isinstance(project_source_records, pd.DataFrame):
        project_source_records = pd.DataFrame()
    has_project_source = bool(source_meta) and not project_source_records.empty

    if has_project_source:
        st.info(
            f"Datablix has a current Starting Data baseline with {len(project_source_records):,} structured record(s). "
            "It is NOT included in this AI prompt. Datablix will compare the imported research CSV against it afterward."
        )
    else:
        st.warning(
            "No structured Starting Data baseline is currently available. Website research can still proceed, "
            "but Datablix cannot classify imported rows as existing versus newly discovered until a source is loaded."
        )

    default_scope = PROJECT_GEOGRAPHIC_SCOPE
    default_source_policy = (
        "PROPERTY DISCOVERY AND ORDINARY FIELD RESEARCH: use the selected company's official website first, including confirmed official property pages, subdomains, and microsites under the same registrable root domain. Treat them as one company and do not create a company per hostname. "
        "The project includes current official residential listings inside the City of Ottawa municipal boundary, including apartment buildings or units, condominium rentals, townhomes, duplexes, and garden homes. Website inventory presence and rental availability are separate facts. Do not exclude these recognized property forms merely because they are not conventional apartment buildings. Retain current detached single-family homes for human scope review and identify them clearly in Reviewer Notes. Company labels such as Ottawa Region or National Capital Region are not geographic proof. "
        "RENTAL AVAILABILITY: inspect every current property card/listing for an explicit status. Normalize to Available Now, Available Future, Rented, or Status Unclear — Needs Review. A listing can remain on a For Rent page after it is rented; page placement, page existence, a displayed rent, or a working URL is not availability proof. Keep explicit Rented rows when they remain on a current official inventory page so Datablix can reconcile them, but never count them as active/new rental inventory. "
        "PROPERTY TYPE: use official website or source evidence for Apartment, Condominium, Townhome, Duplex, Garden Home, Semi-detached, Detached, or another clearly stated residential form. Store these labels only in Property Type; use | only when the same property/community genuinely contains multiple supported forms. Never place property-form labels in Building Classification. Do not infer property form from appearance. "
        "PO BOX / MAILING ADDRESS: search official Contact, Corporate, Legal, Privacy, Accessibility, footer, tenant-document, payment, PDF, and form pages. If still missing, Google/search engines may locate reliable underlying evidence. Keep mailing and PO Box information separate from the physical property address and record the source, evidence, and confidence. "
        "GEOGRAPHIC POSITION: after an official-site candidate is identified, Google Maps/geocoding and a City of Ottawa boundary check may verify latitude, longitude, municipality, and scope. Never geocode a PO Box to establish property location. "
        "POSTAL CODE: exact-address external recovery is allowed when the official property page omits it. "
        "NUMBER OF APARTMENTS: research total residential inventory for EVERY Current/Review property. Never use available/vacant/advertised units, floor-plan counts, or visual estimates. If the official property page is silent, complete the exact-property recovery hierarchy: other official pages/documents/PDFs first; then exact-address municipal/planning/government/public records; then only a reputable exact-address property/industry source. Verify that the candidate total matches the exact building/complex scope represented by the row; do not transfer community/portfolio totals to a single building. Open underlying sources—snippets/AI summaries are not evidence. Resolve conflicts by source authority, exact-property match, date, context, and scope; otherwise leave blank and mark Conflict — Needs Review. Do not finish a blank count until the mandatory recovery sequence is documented. "
        "NUMBER OF STOREYS: research the total storey count for EVERY Current/Review property and populate the dedicated Storey Count Search Status, Source URL, Evidence, and Confidence fields. Use exact-property official pages/documents first, then exact-address planning/building/public records, then only reputable exact-address secondary property/industry sources. Accept storey/story/floor/level wording only when it clearly states TOTAL building height; never treat an apartment's floor location, unit number, Tower marketing, elevator, photo, Street View, windows, or balconies as storey evidence. Do not independently count basements, underground parking, mezzanines, podium/mechanical levels, or rooftops unless the source includes them. For mixed-height multi-building communities, leave a single storey value blank unless one count truly applies to the row and use Conflict — Needs Review when no single count safely represents the row. Datablix, not the AI, derives Building Classification after import from verified Number of Storeys. "
        "Search snippets alone are not evidence; open the underlying source. External evidence must not discover extra properties, override official current-inventory evidence, or override explicit official rental-status evidence."
    )
    default_priority_notes = (
        "Scan the selected company's official website for CURRENT residential listing inventory physically located within the City of Ottawa only, including apartment buildings or units, condominium rentals, townhomes, duplexes, and garden homes. Do not exclude a current recognized property form merely because it is not a conventional apartment building. Retain current detached single-family homes for human scope review and identify them in Reviewer Notes. Follow confirmed official property subdomains and microsites that share the company's registrable root domain; keep them under the selected company and store them as Property Website sources. Exclude Carleton Place and every other independent municipality even when grouped under an Ottawa-area page. "
        "For every current listing, perform a mandatory rental-status pass. Explicitly capture FOR RENT/AVAILABLE as Available Now, future dated/month availability as Available Future, RENTED/LEASED/NO LONGER AVAILABLE as Rented, and unresolved/conflicting signals as Status Unclear — Needs Review. Preserve the site's wording in Rental Availability Detail and the exact official context/URL in Rental Availability Evidence. A For Rent page can contain Rented listings; never use page placement, a displayed rent, or a live URL as availability proof. "
        "For every candidate, verify the exact physical address and geographic position. Record Latitude, Longitude, Geocoded Municipality, Geographic Scope Status, evidence, and confidence. "
        "Conduct an exhaustive PO Box/mailing-address search across official contact, corporate, legal, privacy, accessibility, footer, PDF, form, rent-payment, and tenant-document pages. If official sources remain incomplete, use Google to locate reliable underlying evidence. Never place a PO Box or corporate mailing address in Street Address. "
        "Recover missing Postal Code only from an exact civic-address match. For EVERY Current/Review property, perform a mandatory count pass for both Number of Apartments and Number of Storeys. Apartment count must be TOTAL residential inventory for the exact row/property scope—not available/vacant units, floor plans, a community/portfolio total incorrectly applied to one building, or a visual estimate. Storeys must describe TOTAL building height—not a unit's floor, unit number, Tower marketing, elevator, photo, Street View, windows, or balconies. For each unresolved count, search official exact-property pages/documents first, then exact-address planning/municipal/government/public records, then only reputable exact-address secondary property/industry sources; open the underlying evidence and do not rely on snippets or AI summaries. Reconcile conflicts using source authority, exact-property identity, date, context, and scope; otherwise leave the value unresolved and document the conflict. For mixed-height multi-building communities, do not force one storey count. Populate Storey Count Search Status, Source URL, Evidence, and Confidence for every completed storey search. Keep Property Type separate from Building Classification. Do not return Building Classification in the AI CSV; Datablix derives it after import from verified Number of Storeys: Low-rise = 1–4, Mid-rise = 5–11, High-rise = 12+. Preserve secondary-source details in the dedicated count evidence fields and Reviewer Notes."
    )
    default_output_notes = (
        "Return exactly one downloadable CSV file only. Use one row per unique company-leased property record—not one row per URL—and keep the exact requested headings in the exact requested order. Keep the root/corporate URL in Company Website, the exact property page or official subdomain in Property Website, and the strongest evidence page in Source URL. When multiple civic addresses share one property/complex name and the same leasing page/contact/process, keep them together in one combined-address row rather than splitting them. "
        "Preserve blanks for genuinely unknown values. Keep Current and Review website inventory records in the same CSV, including explicit Rented listings that still appear on a current official inventory/For Rent page so Datablix can reconcile them. Use Rental Availability Status to distinguish active/upcoming listings from Rented listings. Do not count Rented rows as active/new rental discoveries. "
        "Do not create any CSV row for orphan/empty/generic or stale sitemap-only pages that lack meaningful current property-specific evidence. "
        "Do not return Excel, Google Sheets, JSON, PDF, Markdown tables, or a narrative instead of the CSV. The CSV must be ready for direct Datablix import."
    )

    # Geographic scope is project-wide and fixed. Do not restore older regional values.
    saved_scope = PROJECT_GEOGRAPHIC_SCOPE
    saved_source_policy = default_source_policy
    stored_priority_notes = str(active_company.get("Prompt Priority Notes", "") or "").strip()
    saved_priority_notes = (
        default_priority_notes
        if _contains_legacy_regional_scope(stored_priority_notes)
        else (stored_priority_notes or default_priority_notes)
    )
    saved_output_notes = str(active_company.get("Prompt Output Notes", "") or "").strip() or default_output_notes

    prompt_updated = str(active_company.get("Prompt Updated", "") or "").strip()
    prompt_saved = bool(prompt_updated)
    with smart_expander(
        "Customize company research rules",
        status=(f"saved {prompt_updated}" if prompt_saved else "not saved"),
        expanded=not prompt_saved,
    ):
        prompt_left, prompt_right = st.columns(2)
        geographic_scope = PROJECT_GEOGRAPHIC_SCOPE
        prompt_left.text_input(
            "Geographic scope",
            value=PROJECT_GEOGRAPHIC_SCOPE,
            key=f"db_prompt_scope_{company_id}",
            disabled=True,
            help="Fixed project scope: current residential rentals within the configured geographic boundary and property-type rules.",
        )
        prompt_left.caption(
            "Project-wide rule: include only physical properties inside the City of Ottawa municipal boundary; exclude all independent nearby municipalities."
        )

        prompt_right.caption("Starting Data comparison is handled inside Datablix after you import the completed CSV.")

        source_policy = prompt_left.text_area(
            "Source policy",
            value=saved_source_policy,
            height=180,
            key=f"db_prompt_sources_{company_id}",
            disabled=True,
            help="Project-wide guardrails apply the configured geographic scope, source requirements, address validation, postal-code recovery, unit-count research, and storey/classification evidence rules.",
        )
        priority_notes = prompt_right.text_area(
            "Company-specific priorities or exclusions",
            value=saved_priority_notes,
            height=150,
            key=f"db_prompt_priority_{company_id}",
            help="Use this for company-specific exclusions, special website structure, or research priorities.",
        )
        output_notes = st.text_area(
            "Deliverable instructions",
            value=saved_output_notes,
            height=105,
            key=f"db_prompt_output_{company_id}",
            help="Persistent company-specific content rules. The exactly-one consolidated CSV output format is enforced by the master prompt and cannot be overridden here.",
        )
        ai_tool_used = st.text_input(
            "AI tool used (optional)",
            value=str(active_company.get("AI Tool Used", "") or "").strip(),
            placeholder="Example: ChatGPT, Claude, Gemini or Copilot",
            key=f"db_prompt_ai_tool_{company_id}",
        )

    generated_prompt = build_company_website_research_prompt(
        company_name=company_name,
        company_website=company_website,
        related_official_links=company_related_links,
        special_website_notes=company_special_website_notes,
        geographic_scope=geographic_scope,
        priority_notes=priority_notes,
        source_policy=source_policy,
        output_notes=output_notes,
    )

    # The master prompt is regenerated whenever any dynamic company context changes.
    # A fingerprint in the widget key forces Streamlit to refresh the text area instead
    # of retaining a stale prompt from the same company.
    prompt_fingerprint = hashlib.sha256(generated_prompt.encode("utf-8")).hexdigest()[:12]
    master_key_prefix = f"db_master_prompt_{company_id}_"
    master_prompt_key = f"{master_key_prefix}{prompt_fingerprint}"
    for session_key in list(st.session_state.keys()):
        if str(session_key).startswith(master_key_prefix) and session_key != master_prompt_key:
            st.session_state.pop(session_key, None)

    with smart_expander(
        "Review or edit the full master research prompt",
        status="advanced",
        expanded=False,
    ):
        editable_prompt = st.text_area(
            "Editable master research prompt",
            value=generated_prompt,
            height=650,
            key=master_prompt_key,
            help=(
                "Automatically rebuilt when the selected company, website, or saved prompt settings change. "
                "For persistent custom rules, edit the fields above and save them to the company workspace."
            ),
        )
        st.caption(
            "The company name and website refresh automatically. Starting Data is intentionally excluded from the AI prompt. "
            "Persistent research rules are stored separately per company; the full prompt saved below is an audit snapshot."
        )

    previous_prompt_snapshot = str(active_company.get("Research Prompt", "") or "").strip()
    if previous_prompt_snapshot and previous_prompt_snapshot != generated_prompt:
        with smart_expander("Previous saved prompt snapshot", status="audit history", expanded=False):
            st.caption(
                "Kept for audit/history. Datablix no longer uses an old full prompt as the source of truth, so company changes cannot leave the active prompt stale."
            )
            st.code(previous_prompt_snapshot, language="markdown")

    st.caption(
        f"Prompt settings saved to this company: {prompt_updated}"
        if prompt_updated
        else "This company's prompt settings have not yet been saved."
    )
    if st.button(
        "Save prompt settings to company workspace",
        type="primary",
        width="stretch",
        key=f"db_save_company_prompt_{company_id}",
    ):
        registry_prompt = normalize_company_registry(st.session_state.get(S_COMPANIES))
        company_mask = registry_prompt["Company ID"].astype(str).eq(company_id)
        registry_prompt.loc[company_mask, "Prompt Scope"] = PROJECT_GEOGRAPHIC_SCOPE
        registry_prompt.loc[company_mask, "Prompt Source Policy"] = source_policy.strip()
        registry_prompt.loc[company_mask, "Prompt Priority Notes"] = priority_notes.strip()
        registry_prompt.loc[company_mask, "Prompt Output Notes"] = output_notes.strip()
        registry_prompt.loc[company_mask, "Research Prompt"] = editable_prompt
        registry_prompt.loc[company_mask, "Prompt Updated"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        registry_prompt.loc[company_mask, "AI Tool Used"] = ai_tool_used.strip()
        registry_prompt.loc[company_mask, "Company Status"] = registry_prompt.loc[company_mask, "Company Status"].replace("Not started", "Researching")
        st.session_state[S_COMPANIES] = normalize_company_registry(registry_prompt)
        st.session_state[S_FLASH] = f"Research prompt settings saved under {company_name}."
        autosave_current_project()
        st.rerun()

    with smart_expander("Copy-ready prompt", status="copy or inspect", expanded=False):
        st.caption("Use the copy icon in the code block after finishing your edits above.")
        st.code(editable_prompt, language="markdown")
    prompt_download_name = f"{safe_filename(company_name)}_website_research_prompt.txt"
    research_template_df = ai_research_template(
        company_name,
        company_website,
    )

    prompt_actions = st.columns(2)
    prompt_actions[0].download_button(
        "Download website research prompt",
        data=editable_prompt.encode("utf-8"),
        file_name=prompt_download_name,
        mime="text/plain",
        type="primary",
        width="stretch",
        key=f"db_download_prompt_{company_id}_{prompt_fingerprint}",
    )
    prompt_actions[1].download_button(
        "Download CSV template",
        data=csv_bytes(research_template_df),
        file_name=f"{safe_filename(company_name)}_research_template.csv",
        mime="text/csv",
        width="stretch",
        key=f"db_download_template_{company_id}_{prompt_fingerprint}",
    )

    st.caption(
        "Do not upload the project Starting Data to the AI research tool. Research the website independently, "
        "then import the completed CSV here. Datablix performs the project-source comparison after import."
    )

    st.divider()
    st.subheader("2. Import the single completed CSV research deliverable")
    st.caption(
        "Exactly one consolidated CSV is the required research-deliverable format. The AI file contains website research only. When imported, Datablix first consolidates clearly shared leasing-property rows, then compares normalized civic addresses, compound addresses, municipality and Ottawa-locality labels, postal codes, names and property URLs against the current Starting Data baseline before any row can be called newly discovered."
    )
    import_tabs = st.tabs(["Upload CSV or Excel", "Connect Google Sheet"])
    with import_tabs[0]:
        research_upload = st.file_uploader(
            "Completed research spreadsheet",
            type=["csv", "xlsx"],
            key=f"db_external_research_upload_{company_id}",
        )
        selected_sheet = None
        if research_upload is not None and research_upload.name.lower().endswith(".xlsx"):
            try:
                sheet_names = excel_sheet_names(research_upload)
                selected_sheet = st.selectbox(
                    "Worksheet",
                    sheet_names,
                    index=preferred_sheet(sheet_names),
                    key=f"db_external_research_sheet_{company_id}",
                )
            except Exception as error:
                st.error(f"Datablix could not inspect this workbook: {error}")
        if st.button(
            "Import spreadsheet into review",
            type="primary",
            width="stretch",
            disabled=research_upload is None,
            key=f"db_external_research_import_{company_id}",
        ):
            try:
                imported_df, _ = read_upload(research_upload, selected_sheet)
                added_count = append_external_research_results(
                    imported_df,
                    company_id=company_id,
                    company_name=company_name,
                    company_website=company_website,
                )
                st.session_state[S_EDIT_COUNT] = st.session_state.get(S_EDIT_COUNT, 0) + added_count
                st.session_state[S_FLASH] = (
                    f"Imported {added_count:,} website-research row(s) for {company_name}. Datablix compared them with the current Starting Data when available; review the comparison, evidence, duplicates, and missing information next."
                )
                go_to("Review records")
                st.rerun()
            except Exception as error:
                st.error(str(error))

    with import_tabs[1]:
        with st.form(f"db_external_google_form_{company_id}"):
            google_url = st.text_input(
                "Shareable Google Sheets link",
                placeholder="https://docs.google.com/spreadsheets/d/...",
            )
            google_selector = st.text_input(
                "Worksheet name or gid (optional)",
                placeholder="Research Results",
            )
            import_google = st.form_submit_button(
                "Import Google Sheet into review",
                type="primary",
                width="stretch",
            )
        if import_google:
            try:
                imported_df, _, _, _ = read_google_sheet(google_url, google_selector)
                added_count = append_external_research_results(
                    imported_df,
                    company_id=company_id,
                    company_name=company_name,
                    company_website=company_website,
                )
                st.session_state[S_EDIT_COUNT] = st.session_state.get(S_EDIT_COUNT, 0) + added_count
                st.session_state[S_FLASH] = (
                    f"Imported {added_count:,} Google Sheets research row(s) for {company_name}. Datablix compared them with the current Starting Data when available; review the findings before verification."
                )
                go_to("Review records")
                st.rerun()
            except Exception as error:
                st.error(str(error))

    st.divider()
    with smart_expander(
        "Optional Datablix website scanner",
        status="coverage and cross-checking",
        expanded=False,
    ):
        st.caption(
            "The scanner is no longer the primary research method. Use it when you need a second source of page coverage, want to compare AI findings against the live site, or need to investigate possible omissions. Scanner findings still require human review."
        )
        saved_scanner_entry_points = [company_website] if company_website else []
        saved_scanner_entry_points.extend(
            _parse_related_official_links(
                company_related_links,
                company_website,
            )[0]
        )
        if saved_scanner_entry_points:
            st.caption("Saved official scan entry points:")
            st.code("\n".join(saved_scanner_entry_points), language="text")
            if len(saved_scanner_entry_points) > 1:
                st.info(
                    "Start with the main website. If a known microsite is not reached, "
                    "run a follow-up scan with that saved official link while keeping "
                    "this same company selected."
                )
        scan_result = render_website_scanner_panel(
            working_data_key=S_WORKING,
            active_company_id=company_id,
            active_company_name=company_name,
            active_company_website=company_website,
            scan_history_key=S_SCAN_HISTORY,
            scan_candidates_key=S_SCAN_CANDIDATES,
            scan_pages_key=S_SCAN_PAGES,
        )
        if scan_result:
            merged = st.session_state.get(S_WORKING, pd.DataFrame()).copy()
            for column in INTERNAL_COLUMNS:
                if column not in merged.columns:
                    merged[column] = pd.NA
            merged = ensure_ids(normalize_workflow(prepare_data(merged)))
            merged, registry = synchronize_company_registry(
                merged,
                st.session_state.get(S_COMPANIES),
            )
            registry.loc[
                registry["Company ID"].eq(company_id), "Company Status"
            ] = "Researching"
            st.session_state[S_WORKING] = merged
            st.session_state[S_COMPANIES] = normalize_company_registry(registry)
            st.session_state[S_EDIT_COUNT] = (
                st.session_state.get(S_EDIT_COUNT, 0)
                + int(scan_result.get("added", 0))
            )
            st.session_state[S_FLASH] = (
                f"Added {int(scan_result.get('added', 0))} scanner cross-check record(s) for {company_name}. Compare them with the imported research before verification."
            )
            go_to("Review records")
            st.rerun()

# -----------------------------
# Review records
# -----------------------------
elif section == "Review records":
    render_page_heading(
        "REVIEW & QUALITY",
        "Review records and track quality",
        "Correct, verify, and approve records while Datablix refreshes the current quality checks after every save.",
    )
    st.caption(f"Workspace build: {DATABLIX_BUILD}")
    render_guidance(
        "Review, verify, approve.",
        "Use the live quality checks to resolve critical issues, document research gaps, and move reviewed records toward Approved for Export.",
    )

    review_company_id, _review_company_qa = _review_company(qa) if has_records else (None, pd.DataFrame())

    st.divider()
    st.markdown("### Review and verify records")
    render_guidance(
        "Blank values stay neutral.",
        "A blank means the information has not been confirmed; it does not automatically mean the feature or detail is unavailable.",
    )

    review_scope_qa = (
        qa.loc[qa["Company ID"].astype(str).eq(str(review_company_id))].copy()
        if has_records and review_company_id
        else qa.copy() if has_records else pd.DataFrame()
    )
    filtered = review_scope_qa.copy()

    if has_records and not review_scope_qa.empty and _secret_value("GOOGLE_MAPS_API_KEY"):
        geo_col, geo_help_col = st.columns([1, 2])
        with geo_col:
            run_geo_validation = st.button(
                "Validate Ottawa geography",
                key=f"db_validate_geo_{review_company_id or 'project'}",
                width="stretch",
                help="Geocode unresolved physical addresses and check the configured geographic scope. PO Box fields are never used.",
            )
        with geo_help_col:
            boundary_mode = (
                "exact configured boundary polygon"
                if _secret_value("OTTAWA_BOUNDARY_GEOJSON_URL")
                else "Google municipality/address components"
            )
            st.caption(f"Validation method: Google Maps geocoding plus {boundary_mode}.")
        if run_geo_validation:
            working = st.session_state[S_WORKING].copy()
            target_mask = (
                working["Company ID"].astype(str).eq(str(review_company_id))
                if review_company_id
                else pd.Series(True, index=working.index)
            )
            enriched = enrich_geographic_scope(working.loc[target_mask].copy(), max_requests=250)
            for column in enriched.columns:
                if column not in working.columns:
                    working[column] = pd.NA
            working.loc[target_mask, enriched.columns] = enriched.to_numpy()
            st.session_state[S_WORKING] = normalize_workflow(working)
            st.session_state[S_EDIT_COUNT] = st.session_state.get(S_EDIT_COUNT, 0) + 1
            st.session_state[S_FLASH] = "Geographic validation completed for unresolved physical addresses. Review all outside-boundary and low-confidence results."
            autosave_current_project()
            st.rerun()

    if has_records and not review_scope_qa.empty:
        approved_now = int(approved_for_export_mask(review_scope_qa).sum())
        still_now = int((~approved_for_export_mask(review_scope_qa) & ~review_scope_qa["Record Decision"].eq("Remove")).sum())
        existing_now = int(review_scope_qa["Directory Discovery Status"].eq("Existing Source Record").sum())
        discovered_now = int(review_scope_qa["Directory Discovery Status"].eq("Newly Discovered").sum())
        needs_origin_now = int(review_scope_qa["Directory Discovery Status"].eq("Needs Classification").sum())
        summary_cols = st.columns(5)
        summary_cols[0].metric("Company records", f"{len(review_scope_qa):,}")
        summary_cols[1].metric("Existing source", f"{existing_now:,}")
        summary_cols[2].metric("Newly discovered", f"{discovered_now:,}")
        summary_cols[3].metric("Approved for Export", f"{approved_now:,}")
        summary_cols[4].metric("Still in review", f"{still_now:,}")
        st.caption(
            "Discovery status compares each working record with the project's starting building dataset. "
            "Correct the classification during review when needed."
        )
        if needs_origin_now:
            st.info(
                f"{needs_origin_now:,} record(s) still need discovery classification because Datablix could not safely determine whether they were in the starting source building list."
            )

        # Starting Data vs saved research: visible property-level change matrix.
        render_company_source_presence_reconciliation(
            str(review_company_id),
            review_scope_qa,
        )

        search_col, focus_col = st.columns([2, 1])
        search_text = search_col.text_input(
            "Search records",
            placeholder="Rental property, owner, address, city, or record ID",
            key="db_review_search",
        )
        focus = focus_col.selectbox(
            "Focus",
            ["Still in review", "All records", "Ready for review", "Verified", "Approved for Export"],
            key="db_review_focus",
        )

        mask = pd.Series(True, index=review_scope_qa.index)
        if search_text.strip():
            search_blob = (
                review_scope_qa[[
                    "Record ID", "Building Name", "Management/Owner", "Street Address",
                    "City", "Postal Code", "Primary Email", "Phone"
                ]]
                .astype("string")
                .fillna("")
                .agg(" ".join, axis=1)
                .str.lower()
            )
            mask &= search_blob.str.contains(search_text.strip().lower(), regex=False)

        if focus == "Still in review":
            mask &= ~approved_for_export_mask(review_scope_qa) & ~review_scope_qa["Record Decision"].eq("Remove")
        elif focus == "Ready for review":
            mask &= review_scope_qa["Research Status"].eq("Ready for Review") | review_scope_qa["Verification Status"].eq("Needs Review")
        elif focus == "Verified":
            mask &= review_scope_qa["Verification Status"].eq("Verified")
        elif focus == "Approved for Export":
            mask &= approved_for_export_mask(review_scope_qa)

        with st.expander("More filters"):
            filter_row1 = st.columns(3)
            quality_filter = filter_row1[0].multiselect(
                "Listing quality",
                sorted(display_values(review_scope_qa["QA Status"]).unique()),
                help="Leave blank to include every quality status.",
            )
            owner_filter = filter_row1[1].multiselect(
                "Management or owner",
                sorted(display_values(review_scope_qa["Management/Owner"]).unique()),
                help="Leave blank to include every organization.",
            )
            research_filter = filter_row1[2].multiselect(
                "Research status",
                sorted(display_values(review_scope_qa["Research Status"]).unique()),
                help="Leave blank to include every research status.",
            )
            filter_row2 = st.columns(3)
            verification_filter = filter_row2[0].multiselect(
                "Verification status",
                sorted(display_values(review_scope_qa["Verification Status"]).unique()),
                help="Leave blank to include every verification status.",
            )
            readiness_filter = filter_row2[1].multiselect(
                "Record readiness",
                sorted(display_values(review_scope_qa["Record Readiness"]).unique()),
                help="Leave blank to include every readiness status.",
            )
            discovery_filter = filter_row2[2].multiselect(
                "Discovery status",
                sorted(display_values(review_scope_qa["Directory Discovery Status"]).unique()),
                help="Compare existing source records with newly discovered, duplicate, or excluded records.",
            )

        if quality_filter:
            mask &= display_values(review_scope_qa["QA Status"]).isin(quality_filter)
        if owner_filter:
            mask &= display_values(review_scope_qa["Management/Owner"]).isin(owner_filter)
        if research_filter:
            mask &= display_values(review_scope_qa["Research Status"]).isin(research_filter)
        if verification_filter:
            mask &= display_values(review_scope_qa["Verification Status"]).isin(verification_filter)
        if readiness_filter:
            mask &= display_values(review_scope_qa["Record Readiness"]).isin(readiness_filter)
        if discovery_filter:
            mask &= display_values(review_scope_qa["Directory Discovery Status"]).isin(discovery_filter)

        filtered = review_scope_qa.loc[mask].copy()
        st.caption(f"Showing {len(filtered):,} of {len(review_scope_qa):,} records for the selected company.")

        review_tabs = st.tabs(["Review queue", "Edit fields"])

        with review_tabs[0]:
            if filtered.empty:
                st.info(
                    "No records match this search and focus. Clear the search box or switch the focus to All records."
                )
            else:
                st.caption(
                    "Review the quality issue, research gap, source status, and readiness columns before changing a record."
                )
                inspect_columns = [
                    "Record ID", "Working Record Label", "Management/Owner",
                    "Street Address", "City", "Postal Code", "Directory Discovery Status",
                    "Research Status", "Verification Status", "QA Status", "QA Flags", "Research Gaps",
                    "Follow-up Priority", "Record Readiness", "Export Status",
                ]
                inspect = filtered[inspect_columns].rename(
                    columns={
                        "Management/Owner": "Management / Owner",
                        "Working Record Label": "Record",
                    }
                )
                st.dataframe(
                    inspect,
                    width="stretch",
                    hide_index=True,
                    height=520,
                )

        with review_tabs[1]:
            if filtered.empty:
                st.info(
                    "No records match this search and focus. Widen the filters to continue editing."
                )
            else:
                edit_presets = {
                    "Required listing information": [
                        "Building Name", "Management/Owner", "Street Address", "Address Line 2",
                        "City", "Province", "Postal Code", "Property Type",
                        "Number of Storeys", "Building Classification",
                        "Number of Apartments", "Rental Rate Range",
                    ],
                    "Contact and source information": [
                        "Phone", "Primary Email", "Secondary Email", "Website", "Source URL",
                        "Apartment Count Search Status", "Apartment Count Source URL",
                        "Apartment Count Evidence", "Apartment Count Confidence",
                        "Storey Count Search Status", "Storey Count Source URL",
                        "Storey Count Evidence", "Storey Count Confidence",
                        "Date Researched", "Researcher", "Source Status",
                    ],
                    "Research and verification": [
                        "Directory Discovery Status", "Research Status",
                        "Verification Status", "Record Decision",
                        "Directory Entry Status", "Reviewer Notes",
                    ],
                }
                preset = st.selectbox(
                    "Fields to review",
                    [*edit_presets.keys(), "Custom fields"],
                    help="Choose a focused field group or select Custom fields to review a different combination of listing information.",
                    key="db_edit_preset",
                )
                if preset == "Custom fields":
                    edit_fields = st.multiselect(
                        "Fields to edit",
                        [c for c in INTERNAL_COLUMNS if c not in {"Record ID", "Missing Information", "Discovery Status Source"}],
                        default=[
                            "Building Name", "Management/Owner", "Phone", "Primary Email",
                            "Website", "Directory Discovery Status", "Research Status",
                            "Verification Status", "Record Decision",
                            "Directory Entry Status",
                        ],
                        key="db_custom_edit_fields",
                    )
                else:
                    edit_fields = edit_presets[preset]

                # Keep the exact supporting source visible for every observation row,
                # regardless of the selected review preset. The editable Source URL
                # stores the evidence link; Check Source provides a consistent,
                # one-click link for verification.
                filtered = filtered.copy()
                filtered["Check Source"] = filtered["Source URL"].where(
                    ~unresolved_mask(filtered["Source URL"]),
                    pd.NA,
                )
                context = ["Record ID", "Working Record Label"] + edit_fields + [
                    "Source URL", "Check Source", "Missing Information", "Research Gaps",
                    "QA Status", "Record Readiness", "Export Status"
                ]
                context = list(dict.fromkeys(c for c in context if c in filtered.columns))
                locked = [
                    c for c in context
                    if c in [
                        "Record ID", "Working Record Label", "Check Source",
                        "Missing Information", "Research Gaps", "QA Status", "Record Readiness", "Export Status"
                    ]
                ]
                edited = st.data_editor(
                    filtered[context],
                    width="stretch",
                    hide_index=True,
                    height=520,
                    num_rows="fixed",
                    disabled=locked,
                    column_config={
                        "Record ID": st.column_config.TextColumn(
                            "Record ID",
                            width="small",
                            pinned=True,
                            help="Pinned so the record identifier stays visible while you scroll horizontally.",
                        ),
                        "Working Record Label": st.column_config.TextColumn(
                            "Working Record Label",
                            width="medium",
                            pinned=True,
                            help="Pinned so the working property label stays visible while you scroll horizontally.",
                        ),
                        "Building Name": st.column_config.TextColumn("Apartment Building Name"),
                        "Management/Owner": st.column_config.TextColumn("Management / Owner", width="large"),
                        "Phone": st.column_config.TextColumn("Phone Number"),
                        "Primary Email": st.column_config.TextColumn("Email Contact", width="large"),
                        "Website": st.column_config.TextColumn("Website", width="large"),
                        "Source URL": st.column_config.LinkColumn(
                            "Source URL",
                            width="large",
                            help="Exact official page supporting this observation. You can edit the URL and open it directly.",
                        ),
                        "Check Source": st.column_config.LinkColumn(
                            "Check Source",
                            width="medium",
                            display_text="Open source",
                            help="Open the supporting page for this observation in a new tab.",
                        ),
                        "Missing Information": st.column_config.TextColumn(
                            "Missing Information",
                            width="large",
                            help="Automatically generated from the currently blank research fields. Add explanations in Reviewer Notes.",
                        ),
                        "Date Researched": st.column_config.DateColumn("Date Researched", format="YYYY-MM-DD"),
                        "Directory Discovery Status": st.column_config.SelectboxColumn(
                            "Directory Discovery Status",
                            options=DISCOVERY_STATUSES,
                            required=True,
                            help="Existing Source Record means it matches the starting source dataset; Newly Discovered means Datablix did not find a starting-data match. A reviewer change is saved as a manual override and will not be replaced on rerun.",
                        ),
                        "Research Status": st.column_config.SelectboxColumn(
                            "Research Status", options=RESEARCH_STATUSES, required=True
                        ),
                        "Source Status": st.column_config.SelectboxColumn(
                            "Source Status", options=SOURCE_STATUSES, required=True
                        ),
                        "Verification Status": st.column_config.SelectboxColumn(
                            "Verification Status", options=VERIFICATION_STATUSES, required=True
                        ),
                        "Record Decision": st.column_config.SelectboxColumn(
                            "Record Decision", options=RECORD_DECISIONS, required=True
                        ),
                        "Directory Entry Status": st.column_config.SelectboxColumn(
                            "Directory Entry Status",
                            options=DIRECTORY_ENTRY_STATUSES,
                            required=True,
                            help="Track whether this approved listing has been submitted through the external entry process.",
                        ),
                    },
                    key=(
                        f"editor_{st.session_state.get(S_EDIT_COUNT, 0)}_"
                        f"{hashlib.sha1('|'.join(edit_fields).encode()).hexdigest()[:8]}"
                    ),
                )
                save_col, save_note = st.columns([1, 2])
                with save_col:
                    save_changes = st.button(
                        "Save changes",
                        type="primary",
                        width="stretch",
                        key="db_save_edits",
                    )
                with save_note:
                    st.caption(
                        "Saving updates the working copy, refreshes quality checks, and recalculates Approved for Export automatically."
                    )
                if save_changes:
                    save_edits(edited, [c for c in edit_fields if c in edited.columns])
                    st.rerun()

    if has_records:
        render_review_progress(qa_checks(st.session_state[S_WORKING].copy()), review_company_id)

# -----------------------------
# Analysis and report
# -----------------------------
elif section == "Analysis & report":
    render_page_heading(
        "DELIVERABLES",
        "Generate and refine project outputs",
        "Use one reporting state to produce each deliverable separately, refine them independently, or download the complete package.",
    )
    st.caption(f"Workspace build: {DATABLIX_BUILD}")

    deliverable_qa, registry, deliverable_baseline, snapshot_label, snapshot_active = reporting_snapshot_state()

    with st.container(border=True):
        snap_text, snap_action, snap_clear = st.columns([3, 1, 1], vertical_alignment="center")
        with snap_text:
            st.markdown("**Reporting state**")
            if snapshot_active:
                st.success(f"Frozen snapshot active — {snapshot_label}")
                st.caption("All deliverable previews and exports below use this frozen project state.")
            else:
                st.info("Using live project state")
                st.caption("Freeze a reporting snapshot when you want every deliverable to keep exactly the same figures.")
        with snap_action:
            if st.button(
                "Refresh snapshot" if snapshot_active else "Freeze snapshot",
                type="primary",
                width="stretch",
                key="db_freeze_reporting_snapshot",
            ):
                freeze_reporting_snapshot()
                st.session_state[S_FLASH] = "Reporting snapshot saved."
                st.rerun()
        with snap_clear:
            if st.button(
                "Use live data",
                width="stretch",
                disabled=not snapshot_active,
                key="db_clear_reporting_snapshot",
            ):
                clear_reporting_snapshot()
                st.session_state[S_FLASH] = "Deliverables returned to the live project state."
                st.rerun()

    scope_mode = st.radio(
        "Deliverables scope",
        ["All companies", "One company"],
        horizontal=True,
        key="db_analysis_scope",
    )

    selected_company_id = None
    scope_label = "All companies"
    analysis_qa = deliverable_qa.copy()
    analysis_registry = registry.copy()

    if scope_mode == "One company":
        available = registry.loc[
            registry["Company ID"].astype(str).isin(set(deliverable_qa["Company ID"].astype(str)))
        ].copy()
        if available.empty:
            st.warning("No company-linked records are available in this reporting state.")
            st.stop()
        company_ids = available["Company ID"].astype(str).tolist()
        active_id = str(st.session_state.get(S_ACTIVE_COMPANY, "")).strip()
        selected_index = company_ids.index(active_id) if active_id in company_ids else 0
        selected_company_id = st.selectbox(
            "Company",
            company_ids,
            index=selected_index,
            format_func=lambda company_id: company_label(
                available.loc[available["Company ID"].eq(company_id)].iloc[0]
            ),
            key="db_analysis_company",
        )
        company_row = available.loc[available["Company ID"].eq(selected_company_id)].iloc[0]
        scope_label = company_row["Management/Owner"]
        analysis_qa = deliverable_qa.loc[
            deliverable_qa["Company ID"].astype(str).eq(selected_company_id)
        ].copy()
        analysis_registry = available.loc[
            available["Company ID"].astype(str).eq(selected_company_id)
        ].copy()

    coverage_matrix = closeout_research_coverage_matrix(
        analysis_qa, analysis_registry, baseline=deliverable_baseline
    )
    company_count_metric = len(analysis_registry) if not analysis_registry.empty else int(
        analysis_qa["Company ID"].astype(str).replace("", pd.NA).dropna().nunique()
    )
    existing_metric = int(analysis_qa["Directory Discovery Status"].eq("Existing Source Record").sum())
    discovered_metric = int(analysis_qa["Directory Discovery Status"].eq("Newly Discovered").sum())
    changed_metric = int(coverage_matrix["Changed Existing Records"].sum()) if not coverage_matrix.empty else 0
    needs_review_metric = int(deliverable_needs_review_mask(analysis_qa).sum())

    metric_row = st.columns(5)
    metric_row[0].metric("Companies", f"{company_count_metric:,}")
    metric_row[1].metric("Research properties", f"{len(analysis_qa):,}")
    metric_row[2].metric("Existing matches", f"{existing_metric:,}")
    metric_row[3].metric("New discoveries", f"{discovered_metric:,}")
    metric_row[4].metric("Needs review", f"{needs_review_metric:,}")

    with smart_expander("Deliverables map", expanded=False):
        st.dataframe(project_deliverables_table(), width="stretch", hide_index=True)

    project_slug = safe_filename(
        safe_text(st.session_state.get(S_PROJECT_NAME, "Datablix project")) or "Datablix project"
    )
    scope_slug = safe_filename(scope_label)
    package_bytes = deliverables_package_bytes(
        analysis_qa,
        analysis_registry,
        scope_label=scope_label,
        baseline=deliverable_baseline,
    )
    st.download_button(
        "Download All Deliverables — ZIP",
        data=package_bytes,
        file_name=f"{project_slug}_{scope_slug}_deliverables.zip",
        mime="application/zip",
        type="primary",
        width="stretch",
        key="db_download_all_deliverables",
    )
    st.caption("The ZIP uses the same reporting state as every individual export below.")

    analysis_tabs = st.tabs([
        "Research dataset",
        "Research coverage",
        "Verification",
        "Profiles",
        "Recommendations",
        "Methodology",
        "Final matrix",
    ])

    with analysis_tabs[0]:
        st.subheader("1. Research Dataset")
        dataset = deliverable_research_dataset(analysis_qa)
        final_directory_dataset = deliverable_final_directory_dataset(analysis_qa)
        st.caption(
            "Full Research preserves every researched property, including records already in Starting Data, "
            "new discoveries, rented properties, exclusions, duplicates, and records needing follow-up. "
            "Status fields explain each record's outcome."
        )
        if dataset.empty:
            st.info("No researched records are available in this reporting state.")
        else:
            st.dataframe(dataset, width="stretch", hide_index=True, height=560)

        with smart_expander(
            "Final Directory View",
            count=len(final_directory_dataset),
            expanded=False,
        ):
            st.caption(
                "This is the separate filtered view based on the current directory-eligibility rules. "
                "It does not remove rows from Full Research."
            )
            if final_directory_dataset.empty:
                st.info("No records currently meet the directory-view rules.")
            else:
                st.dataframe(final_directory_dataset, width="stretch", hide_index=True, height=500)

        d1_export_cols = st.columns(2)
        with d1_export_cols[0]:
            st.download_button(
                "Download Research Dataset — Excel",
                data=research_dataset_workbook(analysis_qa),
                file_name=f"{project_slug}_{scope_slug}_01_research_dataset.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                width="stretch",
                type="primary",
                key="db_deliverable_01_dataset_xlsx",
            )
        with d1_export_cols[1]:
            st.download_button(
                "Download Research Dataset — CSV",
                data=csv_bytes(dataset),
                file_name=f"{project_slug}_{scope_slug}_01_research_dataset.csv",
                mime="text/csv",
                width="stretch",
                key="db_deliverable_01_dataset_csv",
            )

    with analysis_tabs[1]:
        st.subheader("2. Organization Research Summary")
        st.caption(
            "This summary starts with every researched record by company, including records already in Starting Data, discoveries, exclusions, duplicates, rented properties, and review items. Reconciliation views remain secondary."
        )
        full_company_summary = full_company_research_summary(analysis_qa, analysis_registry)
        st.markdown("#### All research by company")
        st.dataframe(
            full_company_summary,
            width="stretch",
            hide_index=True,
            height=min(700, 90 + 36 * max(len(full_company_summary), 1)),
        )
        st.markdown("#### Starting Data reconciliation")
        coverage_totals = coverage_matrix_totals(coverage_matrix)
        total_row_1 = st.columns(4)
        total_row_1[0].metric("Total source records", f"{coverage_totals['Source Records']:,}")
        total_row_1[1].metric("Total research records", f"{coverage_totals['Research Records']:,}")
        total_row_1[2].metric("Existing matches", f"{coverage_totals['Existing Matches']:,}")
        total_row_1[3].metric("New discoveries", f"{coverage_totals['New Discoveries']:,}")

        total_row_2 = st.columns(3)
        total_row_2[0].metric("Changed existing records", f"{coverage_totals['Changed Existing Records']:,}")
        total_row_2[1].metric("Discoveries submitted", f"{coverage_totals['Discoveries Submitted']:,}")
        total_row_2[2].metric("Needs review", f"{coverage_totals['Needs Review']:,}")

        st.markdown("##### Starting Data geographic scope")
        st.caption(
            "Source Records always includes every original Starting Data row. "
            "Geographic scope is reported separately so out-of-scope records remain visible in the audit trail."
        )
        source_scope = starting_data_geographic_scope_table(
            deliverable_baseline,
            analysis_registry,
        )
        scope_status = (
            source_scope["Scope Status"].astype(str)
            if not source_scope.empty and "Scope Status" in source_scope.columns
            else pd.Series(dtype="string")
        )
        scope_inside = int(scope_status.eq("Inside City of Ottawa").sum())
        scope_outside = int(scope_status.eq("Outside City of Ottawa").sum())
        scope_review = int(scope_status.eq("Needs Geographic Review").sum())

        scope_cols = st.columns(3)
        scope_cols[0].metric("In Ottawa scope", f"{scope_inside:,}")
        scope_cols[1].metric("Out of scope", f"{scope_outside:,}")
        scope_cols[2].metric("Needs scope review", f"{scope_review:,}")

        outside_records = (
            source_scope.loc[source_scope["Scope Status"].eq("Outside City of Ottawa")].copy()
            if not source_scope.empty
            else pd.DataFrame()
        )
        with smart_expander(
            "Out-of-scope Starting Data records",
            count=len(outside_records),
            expanded=False,
        ):
            if outside_records.empty:
                st.success("No Starting Data records are currently classified outside the City of Ottawa.")
            else:
                st.caption(
                    "These rows remain included in Total Source Records. They are shown separately only to document scope assessment."
                )
                st.dataframe(
                    outside_records,
                    width="stretch",
                    hide_index=True,
                    height=min(460, 90 + 36 * max(len(outside_records), 1)),
                )

        review_records = (
            source_scope.loc[source_scope["Scope Status"].eq("Needs Geographic Review")].copy()
            if not source_scope.empty
            else pd.DataFrame()
        )
        if not review_records.empty:
            with smart_expander(
                "Starting Data records needing geographic review",
                count=len(review_records),
                expanded=False,
            ):
                st.dataframe(
                    review_records,
                    width="stretch",
                    hide_index=True,
                    height=min(460, 90 + 36 * max(len(review_records), 1)),
                )

        coverage_display = coverage_matrix_with_total_row(coverage_matrix)
        coverage_styler = coverage_display.style.apply(
            lambda row: [
                "font-weight:700;background-color:#F3F4F6"
                if safe_text(row.get("Company", "")) == "TOTAL"
                else ""
                for _ in row
            ],
            axis=1,
        )
        st.dataframe(
            coverage_styler,
            width="stretch",
            hide_index=True,
            height=min(740, 90 + 36 * max(len(coverage_display), 1)),
        )
        discoveries = discovery_submission_summary(analysis_qa)
        with smart_expander("New discoveries and external submission tracking", count=len(discoveries), expanded=False):
            if discoveries.empty:
                st.info("No records in this scope are currently classified as Newly Discovered.")
            else:
                st.dataframe(discoveries, width="stretch", hide_index=True, height=420)
        st.download_button(
            "Download Organization Research Summary — Excel",
            data=organization_research_summary_workbook(
                analysis_qa, analysis_registry, baseline=deliverable_baseline
            ),
            file_name=f"{project_slug}_{scope_slug}_02_organization_research_summary.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            key="db_deliverable_02_research_summary",
        )

    with analysis_tabs[4]:
        st.subheader("5. Structure & Search Recommendations")
        availability_matrix = field_availability_matrix(
            analysis_qa, analysis_registry, include_all=True
        )
        full_field_findings = field_availability_summary(analysis_qa, include_all=True)
        field_findings = field_availability_summary(analysis_qa)
        recommendations = directory_recommendations_with_coverage(analysis_qa)

        st.caption(
            "Full-research availability is shown first. Directory-candidate findings and recommendations are secondary views for final platform design."
        )
        st.markdown("#### Full research field availability")
        st.dataframe(
            availability_matrix,
            width="stretch",
            hide_index=True,
            height=min(620, 90 + 36 * max(len(availability_matrix), 1)),
        )
        st.markdown("#### Full research findings")
        st.dataframe(full_field_findings, width="stretch", hide_index=True, height=520)
        st.markdown("#### Directory-candidate findings and recommendations")
        st.dataframe(recommendations, width="stretch", hide_index=True, height=680)
        st.caption(
            "Recommendations follow observed coverage. The Rationale column can be refined independently without changing the underlying research data."
        )
        st.download_button(
            "Download Structure & Search Recommendations — Excel",
            data=recommendations_workbook(analysis_qa, analysis_registry),
            file_name=f"{project_slug}_{scope_slug}_05_structure_search_recommendations.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            key="db_deliverable_05_recommendations",
        )

    with analysis_tabs[2]:
        st.subheader("3. Source & Verification Tracker")
        followup = verification_followup_summary(analysis_qa)
        tracker = source_verification_tracker(analysis_qa)
        status_summary = full_research_status_summary(analysis_qa)
        st.caption(
            "The tracker preserves source and verification information for every researched record. The follow-up summary below is a separate active-candidate view."
        )
        st.markdown("#### Full research status")
        st.dataframe(status_summary, width="stretch", hide_index=True)
        with smart_expander("Full record-level source and verification tracker", count=len(tracker), expanded=True):
            st.dataframe(
                tracker,
                width="stretch",
                hide_index=True,
                height=560,
                column_config={
                    "Source URL": st.column_config.LinkColumn("Source URL", width="large"),
                },
            )
        st.markdown("#### Active follow-up summary")
        st.dataframe(followup, width="stretch", hide_index=True)
        st.download_button(
            "Download Source & Verification Tracker — Excel",
            data=verification_tracker_workbook(analysis_qa),
            file_name=f"{project_slug}_{scope_slug}_03_source_verification_tracker.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
            key="db_deliverable_03_verification",
        )

    with analysis_tabs[3]:
        st.subheader("4. Draft Record Profiles")
        profiles = draft_profiles_table(analysis_qa)
        final_profile_candidates = final_profile_candidates_table(analysis_qa)
        st.caption(
            "Profiles are generated for every researched property so the research history is preserved. A separate filtered candidate view identifies records suitable for final profile use."
        )
        if profiles.empty:
            st.info("No records are available for profile drafting in this scope.")
        else:
            profile_ids = profiles["Record ID"].astype(str).tolist()
            selected_profile_id = st.selectbox(
                "Preview profile",
                profile_ids,
                format_func=lambda record_id: (
                    f"{profiles.loc[profiles['Record ID'].astype(str).eq(record_id), 'Building Name'].iloc[0]} "
                    f"— {profiles.loc[profiles['Record ID'].astype(str).eq(record_id), 'Street Address'].iloc[0]}"
                ),
                key="db_profile_record_id",
            )
            profile_text = profiles.loc[
                profiles["Record ID"].astype(str).eq(selected_profile_id), "Profile Draft"
            ].iloc[0]
            st.code(profile_text, language="markdown")
        with smart_expander(
            "Filtered final profile candidates",
            count=len(final_profile_candidates),
            expanded=False,
        ):
            if final_profile_candidates.empty:
                st.info("No records currently meet the filtered profile-candidate rules.")
            else:
                st.dataframe(
                    final_profile_candidates.drop(columns=["Profile Draft"], errors="ignore"),
                    width="stretch",
                    hide_index=True,
                    height=420,
                )
        st.download_button(
            "Download Draft Record Profiles — HTML",
            data=profiles_html(analysis_qa),
            file_name=f"{project_slug}_{scope_slug}_04_draft_record_profiles.html",
            mime="text/html",
            width="stretch",
            key="db_deliverable_04_profiles",
        )

    with analysis_tabs[5]:
        st.subheader("6. Methodology & Limitations")
        method_report = methodology_and_limitations_report(analysis_qa, scope_label)
        st.caption(
            "The methodology is grounded in the complete research state; filtered final-use views are downstream outputs."
        )
        st.markdown("#### Full research status")
        st.dataframe(full_research_status_summary(analysis_qa), width="stretch", hide_index=True)
        st.markdown("#### Methodology and limitations")
        st.dataframe(method_report, width="stretch", hide_index=True)
        st.download_button(
            "Download Methodology & Limitations — HTML",
            data=methodology_html(analysis_qa, scope_label),
            file_name=f"{project_slug}_{scope_slug}_06_methodology_limitations.html",
            mime="text/html",
            width="stretch",
            key="db_deliverable_06_methodology",
        )

    with analysis_tabs[6]:
        st.subheader("7. Final Project Data Matrix")
        final_matrix = final_project_data_matrix(
            analysis_qa,
            analysis_registry,
            scope_label=scope_label,
            baseline=deliverable_baseline,
        )
        st.dataframe(final_matrix, width="stretch", hide_index=True, height=700)

        with smart_expander("Copy-ready presentation summary", expanded=False):
            st.code(
                presentation_summary_text(
                    analysis_qa,
                    analysis_registry,
                    scope_label,
                    baseline=deliverable_baseline,
                ),
                language="markdown",
            )

        final_html = project_closeout_report_html(
            analysis_qa,
            analysis_registry,
            scope_label=scope_label,
            baseline=deliverable_baseline,
        )
        final_export_cols = st.columns(2)
        with final_export_cols[0]:
            st.download_button(
                "Download Final Project Data Matrix — HTML",
                data=final_html,
                file_name=f"{project_slug}_{scope_slug}_07_final_project_data_matrix.html",
                mime="text/html",
                width="stretch",
                key="db_deliverable_07_matrix_html",
            )
        with final_export_cols[1]:
            st.download_button(
                "Download Final Project Data Matrix — CSV",
                data=csv_bytes(final_matrix),
                file_name=f"{project_slug}_{scope_slug}_07_final_project_data_matrix.csv",
                mime="text/csv",
                width="stretch",
                key="db_deliverable_07_matrix_csv",
            )
        st.caption(
            "This matrix starts from the full research population and keeps filtered final-use metrics explicitly separate. Narrative refinement can happen later without changing the underlying evidence."
        )

# -----------------------------
# Downloads
# -----------------------------
elif section == "Downloads":
    render_page_heading(
        "EXPORT",
        "Export approved records",
        "Choose the company or project scope, confirm which reviewed records to include, select the columns, preview the exact CSV, then download it.",
    )
    st.caption(f"Workspace build: {DATABLIX_BUILD}")

    registry = normalize_company_registry(st.session_state.get(S_COMPANIES))
    available = registry.loc[
        registry["Company ID"].astype(str).isin(set(qa["Company ID"].astype(str)))
    ].copy() if not registry.empty else registry

    st.subheader("1. Choose export scope")
    scope_options = ["One company", "Entire project"]
    if st.session_state.get("db_export_scope_mode") not in scope_options:
        st.session_state["db_export_scope_mode"] = "One company" if not available.empty else "Entire project"
    export_scope_mode = st.radio(
        "Scope",
        scope_options,
        horizontal=True,
        key="db_export_scope_mode",
        help="Use One company for the company you just reviewed. Use Entire project only when you intentionally want records from every company.",
    )

    export_company_id = None
    export_scope_label = str(st.session_state.get(S_PROJECT_NAME, "Datablix project")).strip() or "Datablix project"
    scope_qa = qa.copy()

    if export_scope_mode == "One company":
        if available.empty:
            st.warning("No company-linked records are available to export yet.")
            st.stop()
        company_ids = available["Company ID"].astype(str).tolist()
        active_id = str(st.session_state.get(S_ACTIVE_COMPANY, "")).strip()
        current_export_company = str(st.session_state.get("db_export_company", "")).strip()
        if current_export_company not in company_ids:
            st.session_state["db_export_company"] = active_id if active_id in company_ids else company_ids[0]
        export_company_id = st.selectbox(
            "Company",
            company_ids,
            format_func=lambda company_id: company_label(
                available.loc[available["Company ID"].eq(company_id)].iloc[0]
            ),
            key="db_export_company",
        )
        company_row = available.loc[available["Company ID"].eq(export_company_id)].iloc[0]
        export_scope_label = str(company_row["Management/Owner"]).strip() or export_company_id
        scope_qa = qa.loc[qa["Company ID"].astype(str).eq(str(export_company_id))].copy()

    approved_count = int(approved_for_export_mask(scope_qa).sum())
    excluded_count = int(scope_qa["Record Decision"].eq("Remove").sum())
    still_reviewing = int((~approved_for_export_mask(scope_qa) & ~scope_qa["Record Decision"].eq("Remove")).sum())

    scope_metrics = st.columns(4)
    scope_metrics[0].metric("Records in scope", f"{len(scope_qa):,}")
    scope_metrics[1].metric("Approved for Export", f"{approved_count:,}")
    scope_metrics[2].metric("Still in review", f"{still_reviewing:,}")
    scope_metrics[3].metric("Excluded", f"{excluded_count:,}")

    st.subheader("2. Choose records")
    record_options = ["Approved for Export", "All records"]
    if st.session_state.get("db_custom_export_scope") not in record_options:
        st.session_state["db_custom_export_scope"] = "Approved for Export"
    export_record_mode = st.radio(
        "Records to include",
        record_options,
        horizontal=True,
        key="db_custom_export_scope",
        help="Approved for Export means Completed + Verified + Keep with no critical data blocker.",
    )
    export_source = (
        scope_qa.loc[approved_for_export_mask(scope_qa)].copy()
        if export_record_mode == "Approved for Export"
        else scope_qa.copy()
    )

    if export_record_mode == "Approved for Export":
        if approved_count:
            st.success(
                f"{approved_count:,} of {len(scope_qa):,} record(s) in {export_scope_label} are approved for export."
            )
        else:
            st.warning(
                "No records in this scope are approved for export yet. Return to Review & Quality and complete the records you want to deliver."
            )

    st.subheader("3. Choose columns")

    # Start the custom export with the exact directory-entry form fields.
    # The default mirrors the final entry form while allowing any Datablix
    # research, QA or audit column to be added when needed.
    listing_view = listing_export(export_source)
    export_view = listing_view.copy()
    for column in export_source.columns:
        if column not in export_view.columns:
            export_view[column] = export_source[column]

    exportable_columns = list(export_view.columns)
    default_columns = [
        column for column in LISTING_COLUMNS
        if column in exportable_columns
    ]

    export_defaults_version = "listing_columns_v23"
    if st.session_state.get("db_export_defaults_version") != export_defaults_version:
        st.session_state["db_custom_export_columns"] = default_columns
        st.session_state["db_export_defaults_version"] = export_defaults_version

    stored_selection = st.session_state.get("db_custom_export_columns", default_columns)
    if not isinstance(stored_selection, list):
        stored_selection = default_columns
    stored_selection = [
        column for column in stored_selection if column in exportable_columns
    ]
    st.session_state["db_custom_export_columns"] = stored_selection

    with smart_expander(
        "Customize export columns",
        count=len(stored_selection),
        status="selected",
        expanded=False,
    ):
        export_controls = st.columns(2)
        if export_controls[0].button(
            "Select all columns",
            width="stretch",
            key="db_export_select_all",
        ):
            st.session_state["db_custom_export_columns"] = exportable_columns
            st.rerun()
        if export_controls[1].button(
            "Clear selection",
            width="stretch",
            key="db_export_clear_columns",
        ):
            st.session_state["db_custom_export_columns"] = []
            st.rerun()

        selected_columns = st.multiselect(
            "Columns to include",
            options=exportable_columns,
            key="db_custom_export_columns",
            help="The CSV will contain exactly these columns in the order shown here.",
        )

    if export_source.empty:
        st.info("There are no records to preview or download for the selected export choice.")
    elif not selected_columns:
        st.warning("Choose at least one column to create the CSV.")
    else:
        export_table = export_view[selected_columns].copy()
        preview_metrics = st.columns(2)
        preview_metrics[0].metric("Rows to download", f"{len(export_table):,}")
        preview_metrics[1].metric("Selected columns", f"{len(selected_columns):,}")

        with smart_expander(
            "Preview exact CSV",
            count=len(export_table),
            status=f"rows × {len(selected_columns):,} columns",
            expanded=False,
        ):
            st.caption(
                "This preview uses the exact selected rows and column order that will appear in the CSV."
            )
            st.dataframe(
                export_table.head(250),
                width="stretch",
                hide_index=True,
                height=500,
            )

        st.subheader("4. Download")
        scope_filename = safe_filename(export_scope_label)
        suffix = "approved" if export_record_mode == "Approved for Export" else "all_records"
        export_filename = f"{scope_filename}_{suffix}_selected_columns.csv"
        button_label = (
            f"Download {len(export_table):,} approved record(s) — CSV"
            if export_record_mode == "Approved for Export"
            else f"Download {len(export_table):,} record(s) — CSV"
        )
        st.download_button(
            button_label,
            data=csv_bytes(export_table),
            file_name=export_filename,
            mime="text/csv",
            type="primary",
            width="stretch",
            key="db_download_custom_export",
        )

        assistant_records = scope_qa.loc[
            approved_for_export_mask(scope_qa)
        ].copy()
        if not assistant_records.empty:
            assistant_records = assistant_records.sort_values(
                ["Management/Owner", "Building Name", "Street Address"],
                kind="stable",
            ).reset_index(drop=True)

        entered_count = int(
            assistant_records["Directory Entry Status"].eq("Entered").sum()
        ) if not assistant_records.empty else 0
        correction_count = int(
            assistant_records["Directory Entry Status"].eq("Needs Correction").sum()
        ) if not assistant_records.empty else 0
        remaining_count = max(len(assistant_records) - entered_count, 0)

        with smart_expander(
            "External Submission Assistant",
            count=remaining_count,
            status="remaining",
            expanded=remaining_count > 0,
        ):
            st.caption(
                "Use this after review. Copy each field into the external submission form, "
                "submit the building, then mark the record Entered so Datablix tracks what remains."
            )

            if assistant_records.empty:
                st.info(
                    "No approved records are available for external submission in this scope yet."
                )
            else:
                entry_metrics = st.columns(4)
                entry_metrics[0].metric("Approved", f"{len(assistant_records):,}")
                entry_metrics[1].metric("Entered", f"{entered_count:,}")
                entry_metrics[2].metric("Remaining", f"{remaining_count:,}")
                entry_metrics[3].metric("Needs correction", f"{correction_count:,}")

                scope_token = hashlib.sha256(
                    f"{export_scope_mode}|{export_company_id or 'project'}".encode("utf-8")
                ).hexdigest()[:10]
                entry_index_key = f"db_directory_entry_index_{scope_token}"
                try:
                    entry_index = int(st.session_state.get(entry_index_key, 0))
                except (TypeError, ValueError):
                    entry_index = 0
                entry_index = max(0, min(entry_index, len(assistant_records) - 1))
                st.session_state[entry_index_key] = entry_index

                nav_left, nav_middle, nav_right = st.columns([1, 2, 1])
                if nav_left.button(
                    "Previous record",
                    width="stretch",
                    disabled=entry_index <= 0,
                    key=f"db_entry_previous_{scope_token}",
                ):
                    st.session_state[entry_index_key] = max(entry_index - 1, 0)
                    st.rerun()
                nav_middle.markdown(
                    f"<div style='text-align:center;padding-top:.55rem'><strong>"
                    f"Record {entry_index + 1:,} of {len(assistant_records):,}"
                    f"</strong></div>",
                    unsafe_allow_html=True,
                )
                if nav_right.button(
                    "Next record",
                    width="stretch",
                    disabled=entry_index >= len(assistant_records) - 1,
                    key=f"db_entry_next_{scope_token}",
                ):
                    st.session_state[entry_index_key] = min(
                        entry_index + 1, len(assistant_records) - 1
                    )
                    st.rerun()

                entry_row = assistant_records.iloc[entry_index]
                record_id = str(entry_row.get("Record ID", "") or "").strip()
                entry_label = directory_entry_record_label(entry_row)
                current_entry_status = str(
                    entry_row.get("Directory Entry Status", "Not Entered")
                    or "Not Entered"
                ).strip()
                if current_entry_status not in DIRECTORY_ENTRY_STATUSES:
                    current_entry_status = "Not Entered"

                with st.container(border=True):
                    st.markdown(f"#### {escape(entry_label)}")
                    if current_entry_status == "Entered":
                        st.success("Directory Entry Status: Entered")
                    elif current_entry_status == "Needs Correction":
                        st.warning("Directory Entry Status: Needs Correction")
                    else:
                        st.info("Directory Entry Status: Not Entered")

                    st.caption(
                        "Each code box has a copy control. The fields are shown in the same order as the directory form."
                    )

                    listing_values = []
                    for listing_label, source_field in LISTING_FIELD_MAP:
                        value = (
                            formatted_location(entry_row)
                            if source_field is None
                            else entry_row.get(source_field, "")
                        )
                        clean_value = _excel_display_value(value)
                        listing_values.append((listing_label, clean_value))
                        st.markdown(f"**{listing_label}**")
                        st.code(clean_value or " ", language=None)

                    st.markdown("**Copy full record**")
                    st.caption(
                        "This tab-separated line follows the same 10-field order and is useful for spreadsheets or backup notes."
                    )
                    st.code(
                        "\t".join(value for _, value in listing_values),
                        language=None,
                    )

                    status_actions = st.columns(3)
                    if status_actions[0].button(
                        "Mark Entered & Next",
                        type="primary",
                        width="stretch",
                        key=f"db_entry_mark_entered_{scope_token}_{record_id}",
                    ):
                        if update_directory_entry_status(record_id, "Entered"):
                            if entry_index < len(assistant_records) - 1:
                                st.session_state[entry_index_key] = entry_index + 1
                            st.session_state[S_FLASH] = (
                                f"Marked {entry_label} as Entered."
                            )
                            st.rerun()
                    if status_actions[1].button(
                        "Needs Correction",
                        width="stretch",
                        key=f"db_entry_mark_correction_{scope_token}_{record_id}",
                    ):
                        if update_directory_entry_status(record_id, "Needs Correction"):
                            st.session_state[S_FLASH] = (
                                f"Marked {entry_label} as needing correction."
                            )
                            st.rerun()
                    if status_actions[2].button(
                        "Reset to Not Entered",
                        width="stretch",
                        key=f"db_entry_reset_{scope_token}_{record_id}",
                    ):
                        if update_directory_entry_status(record_id, "Not Entered"):
                            st.session_state[S_FLASH] = (
                                f"Reset {entry_label} to Not Entered."
                            )
                            st.rerun()


# Persist the latest completed state after every Streamlit rerun.
autosave_current_project()
