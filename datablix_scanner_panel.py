from __future__ import annotations

import copy
import hashlib
import io
import json
import os
import re
import unicodedata
from dataclasses import asdict
from datetime import date, datetime, timezone
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, Side

from full_site_scanner import ScanOptions, ScanReport, WebsiteScanError, scan_website


# Change this key if your existing Datablix working dataframe uses another
# st.session_state name.
WORKING_DATA_KEY = "working_df"

SCANNER_BUILD = "Current Inventory + Rich Extraction 2026.07.23-r2"
CHECKPOINT_DIRECTORY = Path(
    os.environ.get("DATABLIX_CHECKPOINT_DIRECTORY", "/tmp/datablix_checkpoints")
)

COMPANY_SCAN_STATE_STORE = "_db_company_scan_states"
ACTIVE_SCAN_COMPANY_KEY = "_db_active_scan_company"


def _scanner_session_keys() -> list[str]:
    """Return transient scanner keys that must stay isolated by company."""
    return [
        key for key in st.session_state
        if (
            key.startswith("website_scan_")
            or key.startswith("full_scan_")
            or key == "confirm_clear_full_scan"
        )
        and key not in {COMPANY_SCAN_STATE_STORE, ACTIVE_SCAN_COMPANY_KEY}
    ]


def _safe_session_copy(value):
    try:
        return copy.deepcopy(value)
    except Exception:
        return value


def _switch_company_scan_state(company_id: str, default_website: str = "") -> None:
    """Save the previous company's scanner UI and restore the selected company."""
    company_id = str(company_id or "").strip()
    if not company_id:
        return

    previous_company = str(
        st.session_state.get(ACTIVE_SCAN_COMPANY_KEY, "")
    ).strip()
    if previous_company == company_id:
        return

    store = st.session_state.get(COMPANY_SCAN_STATE_STORE)
    if not isinstance(store, dict):
        store = {}

    transient_keys = _scanner_session_keys()
    if previous_company:
        store[previous_company] = {
            key: _safe_session_copy(st.session_state[key])
            for key in transient_keys
            if key != "full_scan_submit"  # Button state must never be restored.
        }

    for key in transient_keys:
        st.session_state.pop(key, None)

    saved_state = store.get(company_id, {})
    if isinstance(saved_state, dict):
        for key, value in saved_state.items():
            st.session_state[key] = _safe_session_copy(value)

    if not str(st.session_state.get("full_scan_website_url", "")).strip():
        st.session_state["full_scan_website_url"] = str(default_website or "").strip()

    st.session_state[COMPANY_SCAN_STATE_STORE] = store
    st.session_state[ACTIVE_SCAN_COMPANY_KEY] = company_id


def _normalized_scan_url(value: str) -> str:
    return str(value or "").strip().lower().rstrip("/")


def _checkpoint_path(website_url: str) -> Path | None:
    normalized = _normalized_scan_url(website_url)
    if not normalized:
        return None
    digest = hashlib.sha256(normalized.encode("utf-8")).hexdigest()[:24]
    return CHECKPOINT_DIRECTORY / f"{digest}.json"


def _write_durable_checkpoint(
    report,
    website_url: str,
    scope: str,
    *,
    is_final: bool = False,
) -> Path | None:
    path = _checkpoint_path(website_url)
    if path is None:
        return None

    CHECKPOINT_DIRECTORY.mkdir(parents=True, exist_ok=True)
    payload = {
        "scanner_build": SCANNER_BUILD,
        "website_url": website_url,
        "scope": scope,
        "saved_at_utc": datetime.now(timezone.utc).isoformat(),
        "is_final": bool(is_final),
        "report": report.as_dict(),
    }

    temporary_path = path.with_suffix(".tmp")
    temporary_path.write_text(
        json.dumps(payload, ensure_ascii=False),
        encoding="utf-8",
    )
    os.replace(temporary_path, path)
    return path


def _read_durable_checkpoint(website_url: str) -> tuple[ScanReport, dict] | None:
    path = _checkpoint_path(website_url)
    if path is None or not path.exists():
        return None

    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
        if payload.get("scanner_build") != SCANNER_BUILD:
            # Old scan results can contain the exact extraction and inventory
            # behaviours this build is replacing. Never mix them with new code.
            try:
                path.unlink(missing_ok=True)
            except OSError:
                pass
            return None
        report_data = payload.get("report", {})
        report = ScanReport.from_dict(report_data)
    except (OSError, ValueError, TypeError, KeyError):
        return None

    if not report.pages:
        return None
    return report, payload


def _delete_durable_checkpoint(website_url: str) -> None:
    path = _checkpoint_path(website_url)
    if path is None:
        return
    try:
        path.unlink(missing_ok=True)
    except OSError:
        pass


DIRECTORY_FIELD_MAP = {
    "company_id": "Company ID",
    "scan_id": "Scan ID",
    "building_name": "Building Name",
    "management_owner": "Management/Owner",
    "street_address": "Street Address",
    "address_line_2": "Address Line 2",
    "city": "City",
    "province": "Province",
    "postal_code": "Postal Code",
    "country": "Country",
    "phone": "Phone",
    "primary_email": "Primary Email",
    "website": "Website",
    "number_of_apartments": "Number of Apartments",
    "amenities": "Amenities",
    "building_classification": "Building Classification",
    "inventory_status": "Current Inventory Status",
    "inventory_evidence": "Inventory Evidence",
    "found_on_city_page": "Found on City/Portfolio Page",
    "found_on_html_sitemap": "Found on HTML Sitemap",
    "found_on_xml_sitemap": "Found on XML Sitemap",
    "exclusion_reason": "Inventory Exclusion Reason",
    "source_url": "Source URL",
    "source_page_title": "Source Page Title",
    "extraction_method": "Extraction Method",
    "confidence": "Detection Confidence",
    "evidence": "Supporting Evidence",
    "ontario_scope_status": "Ontario Scope Status",
    "ontario_scope_reason": "Ontario Scope Reason",
    "scan_start_url": "Scan Start URL",
}


ONTARIO_SCOPE_CONFIRMED = "Confirmed Ontario"
ONTARIO_SCOPE_LIKELY = "Likely Ontario — Review"
ONTARIO_SCOPE_UNCLEAR = "Location Unclear"
ONTARIO_SCOPE_OUTSIDE = "Outside Ontario"

ONTARIO_APPROVABLE_STATUSES = {
    ONTARIO_SCOPE_CONFIRMED,
    ONTARIO_SCOPE_LIKELY,
}

INVENTORY_STATUS_CURRENT = "Current inventory"
INVENTORY_STATUS_REVIEW = "Review inventory"
INVENTORY_STATUS_EXCLUDED = "Excluded — not in current inventory"


def _inventory_eligible_mask(frame: pd.DataFrame) -> pd.Series:
    if "inventory_status" not in frame.columns:
        return pd.Series(True, index=frame.index)
    return ~frame["inventory_status"].fillna(INVENTORY_STATUS_REVIEW).eq(
        INVENTORY_STATUS_EXCLUDED
    )

ONTARIO_POSTAL_PREFIXES = set("KLMNP")

CANADIAN_PROVINCE_ALIASES = {
    "alberta": "AB", "ab": "AB",
    "british columbia": "BC", "bc": "BC",
    "manitoba": "MB", "mb": "MB",
    "new brunswick": "NB", "nb": "NB",
    "newfoundland and labrador": "NL", "newfoundland": "NL",
    "labrador": "NL", "nl": "NL",
    "nova scotia": "NS", "ns": "NS",
    "northwest territories": "NT", "nt": "NT",
    "nunavut": "NU", "nu": "NU",
    "ontario": "ON", "on": "ON",
    "prince edward island": "PE", "pei": "PE", "pe": "PE",
    "quebec": "QC", "québec": "QC", "qc": "QC",
    "saskatchewan": "SK", "sk": "SK",
    "yukon": "YT", "yt": "YT",
}

# Municipality names are supporting evidence only. Province and postal-code
# evidence take precedence whenever they are available.
ONTARIO_MUNICIPALITIES = {
    "ajax", "alexandria", "alliston", "almonte", "amherstburg", "ancaster",
    "arnprior", "aurora", "aylmer", "bancroft", "barrie", "beamsville",
    "belle river", "belleville", "blind river", "bolton", "bracebridge",
    "bradford", "brampton", "brantford", "brighton", "brockville",
    "burlington", "caledon", "caledonia", "cambridge", "campbellford",
    "carleton place", "casselman", "chatham", "chelmsford", "clarington",
    "cobourg", "cochrane", "collingwood", "cornwall", "deep river", "delhi",
    "dryden", "dunnville", "east gwillimbury", "elliot lake", "elmira",
    "elora", "embrun", "erin", "espanola", "essex", "etobicoke", "fergus",
    "fort erie", "fort frances", "gananoque", "georgetown", "geraldton",
    "gloucester", "goderich", "grand bend", "gravenhurst", "greater napanee",
    "greater sudbury", "grimsby", "guelph", "haliburton", "hamilton",
    "hanover", "harriston", "hawkesbury", "hearst", "huntsville",
    "ingersoll", "iroquois falls", "kanata", "kapuskasing", "kawartha lakes",
    "kemptville", "kenora", "keswick", "kincardine", "kingston",
    "kirkland lake", "kitchener", "lasalle", "leamington", "lindsay",
    "listowel", "london", "markham", "midland", "milton", "mississauga",
    "mitchell", "napanee", "nepean", "new hamburg", "new liskeard",
    "newmarket", "niagara falls", "niagara on the lake", "north bay",
    "north york", "oakville", "orangeville", "orillia", "oshawa", "ottawa",
    "owen sound", "paris", "parry sound", "pembroke", "perth", "petawawa",
    "peterborough", "petrolia", "pickering", "port colborne", "port elgin",
    "port hope", "prescott", "prince edward county", "renfrew",
    "richmond hill", "rockland", "sarnia", "sault ste marie", "scarborough",
    "simcoe", "smiths falls", "south porcupine", "st catharines",
    "st marys", "st thomas", "stittsville", "stratford", "strathroy",
    "sturgeon falls", "sudbury", "tecumseh", "temiskaming shores",
    "thunder bay", "tillsonburg", "timmins", "toronto", "trenton",
    "uxbridge", "vaughan", "walkerton", "wallaceburg", "waterdown",
    "waterloo", "welland", "whitby", "wiarton", "windsor", "wingham",
    "woodstock",
}

CANADIAN_POSTAL_PATTERN = re.compile(
    r"^[ABCEGHJ-NPRSTVXY]\d[ABCEGHJ-NPRSTV-Z]\d[ABCEGHJ-NPRSTV-Z]\d$",
    re.IGNORECASE,
)


def _normalize_location_text(value) -> str:
    if value is None or pd.isna(value):
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(
        character for character in text
        if not unicodedata.combining(character)
    )
    text = text.lower().replace("&", " and ")
    return re.sub(r"[^a-z0-9]+", " ", text).strip()


def _normalized_postal_code(value) -> str:
    return re.sub(r"[^A-Za-z0-9]", "", _clean_value(value)).upper()


def _province_code(value) -> str:
    return CANADIAN_PROVINCE_ALIASES.get(_normalize_location_text(value), "")


def _country_is_canada(value) -> bool:
    normalized = _normalize_location_text(value)
    return not normalized or normalized in {"canada", "ca", "can"}


def _classify_ontario_scope(row: pd.Series) -> tuple[str, str]:
    province_text = _clean_value(row.get("province"))
    province_code = _province_code(province_text)
    city = _normalize_location_text(row.get("city"))
    street = _clean_value(row.get("street_address"))
    postal = _normalized_postal_code(row.get("postal_code"))
    country = _clean_value(row.get("country"))

    source_context = " ".join(
        _normalize_location_text(row.get(field))
        for field in ["source_page_title", "evidence", "source_url"]
    )

    postal_is_canadian = bool(CANADIAN_POSTAL_PATTERN.fullmatch(postal))
    postal_is_ontario = (
        postal_is_canadian
        and bool(postal)
        and postal[0] in ONTARIO_POSTAL_PREFIXES
    )
    city_is_ontario = city in ONTARIO_MUNICIPALITIES
    source_mentions_ontario = bool(re.search(r"\bontario\b", source_context))

    if not _country_is_canada(country):
        return ONTARIO_SCOPE_OUTSIDE, f"Country is recorded as {country}."

    if province_code and province_code != "ON":
        if postal_is_ontario:
            return (
                ONTARIO_SCOPE_UNCLEAR,
                "Province conflicts with an Ontario postal-code pattern.",
            )
        return ONTARIO_SCOPE_OUTSIDE, f"Province is recorded as {province_text}."

    if province_code == "ON":
        details = ["Province is Ontario"]
        if street:
            details.append("a street address is present")
        if city:
            details.append(f"city is {_clean_value(row.get('city'))}")
        return ONTARIO_SCOPE_CONFIRMED, "; ".join(details) + "."

    if postal_is_ontario:
        return (
            ONTARIO_SCOPE_CONFIRMED,
            f"Postal code {_clean_value(row.get('postal_code'))} follows an Ontario pattern.",
        )

    if city_is_ontario and street:
        return (
            ONTARIO_SCOPE_LIKELY,
            f"{_clean_value(row.get('city'))} is recognized as an Ontario municipality and a street address is present.",
        )

    if city_is_ontario:
        return (
            ONTARIO_SCOPE_LIKELY,
            f"{_clean_value(row.get('city'))} is recognized as an Ontario municipality; confirm the street or province.",
        )

    if source_mentions_ontario and street:
        return (
            ONTARIO_SCOPE_LIKELY,
            "The source mentions Ontario and a street address is present; confirm the city or postal code.",
        )

    if source_mentions_ontario:
        return (
            ONTARIO_SCOPE_LIKELY,
            "The source mentions Ontario; confirm the city and street address.",
        )

    if street or city or postal or province_text:
        return (
            ONTARIO_SCOPE_UNCLEAR,
            "The available location details do not yet confirm Ontario.",
        )

    return (
        ONTARIO_SCOPE_UNCLEAR,
        "No province, Ontario postal code, recognized Ontario municipality, or street evidence was detected.",
    )


def _apply_ontario_scope(frame: pd.DataFrame) -> pd.DataFrame:
    scoped = frame.copy()
    if scoped.empty:
        scoped["ontario_scope_status"] = pd.Series(dtype="object")
        scoped["ontario_scope_reason"] = pd.Series(dtype="object")
        return scoped

    classifications = scoped.apply(_classify_ontario_scope, axis=1)
    scoped["ontario_scope_status"] = classifications.map(lambda result: result[0])
    scoped["ontario_scope_reason"] = classifications.map(lambda result: result[1])

    if "approved" not in scoped.columns:
        scoped["approved"] = False

    outside_or_unclear = scoped["ontario_scope_status"].isin(
        {ONTARIO_SCOPE_OUTSIDE, ONTARIO_SCOPE_UNCLEAR}
    )
    scoped.loc[outside_or_unclear, "approved"] = False
    return scoped


def _ontario_eligible_mask(frame: pd.DataFrame) -> pd.Series:
    scoped = (
        frame
        if "ontario_scope_status" in frame.columns
        else _apply_ontario_scope(frame)
    )
    return scoped["ontario_scope_status"].isin(ONTARIO_APPROVABLE_STATUSES)


def _apply_candidate_status(frame: pd.DataFrame) -> pd.DataFrame:
    """Keep scanner approval distinct from final Datablix verification."""
    out = frame.copy()
    if "approved" not in out.columns:
        out["approved"] = False
    requested = out["approved"].fillna(False).astype(bool)
    location_ok = _ontario_eligible_mask(out)
    inventory_ok = _inventory_eligible_mask(out)
    eligible = location_ok & inventory_ok

    status = pd.Series("Needs candidate review", index=out.index, dtype="object")
    status.loc[~inventory_ok] = "Excluded from current inventory"
    status.loc[inventory_ok & ~location_ok] = "Approval blocked — location"
    status.loc[requested & eligible] = "Approved for project"
    out["candidate_status"] = status
    return out



def _records_dataframe(report) -> pd.DataFrame:
    frame = pd.DataFrame([asdict(record) for record in report.records])
    if frame.empty:
        return pd.DataFrame(
            columns=[
                "approved", "building_name", "management_owner",
                "street_address",
                "address_line_2", "city", "province", "postal_code", "country",
                "phone", "primary_email", "website", "number_of_apartments",
                "amenities", "building_classification",
                "inventory_status", "inventory_evidence",
                "found_on_city_page", "found_on_html_sitemap",
                "found_on_xml_sitemap", "exclusion_reason",
                "source_url", "source_page_title",
                "extraction_method", "confidence",
                "review_status", "candidate_status", "evidence",
                "ontario_scope_status", "ontario_scope_reason",
            ]
        )
    frame["confidence"] = pd.to_numeric(frame["confidence"], errors="coerce").fillna(0.0)
    return _apply_ontario_scope(frame)


def _pages_dataframe(report) -> pd.DataFrame:
    return pd.DataFrame([asdict(page) for page in report.pages])


SCAN_LISTING_FIELDS = [
    ("Apartment Building Name", "building_name"),
    ("Street Address", "street_address"),
    ("City and Postal Code", None),
    ("Building Classification", "building_classification"),
    ("Number of Apartments", "number_of_apartments"),
    ("Apartment Building Management/Owner", "management_owner"),
    ("Phone Number", "phone"),
    ("Email Contact", "primary_email"),
    ("WebSite", "website"),
]

SCAN_ADDITIONAL_FIELDS = [
    ("Address Line 2", "address_line_2"),
    ("Amenities", "amenities"),
    ("Current Inventory Status", "inventory_status"),
    ("Inventory Evidence", "inventory_evidence"),
    ("Found on City/Portfolio Page", "found_on_city_page"),
    ("Found on HTML Sitemap", "found_on_html_sitemap"),
    ("Found on XML Sitemap", "found_on_xml_sitemap"),
    ("Inventory Exclusion Reason", "exclusion_reason"),
    ("Country", "country"),
    ("Official Source URL", "source_url"),
    ("Source Page Title", "source_page_title"),
    ("Extraction Method", "extraction_method"),
    ("Confidence", "confidence"),
    ("Ontario Scope Status", "ontario_scope_status"),
    ("Ontario Scope Reason", "ontario_scope_reason"),
    ("Evidence", "evidence"),
]


def _clean_value(value) -> str:
    if value is None or pd.isna(value):
        return ""
    return str(value).strip()


def _scan_location(row: pd.Series) -> str:
    city = _clean_value(row.get("city"))
    province = _clean_value(row.get("province"))
    postal = _clean_value(row.get("postal_code"))
    tail = " ".join(value for value in [province, postal] if value)
    return f"{city}, {tail}" if city and tail else city or tail


def _approved_listing_table(frame: pd.DataFrame) -> pd.DataFrame:
    output = pd.DataFrame(index=frame.index)
    for label, source_field in SCAN_LISTING_FIELDS:
        output[label] = (
            frame.apply(_scan_location, axis=1)
            if source_field is None
            else frame.get(source_field, "")
        )
    for label, source_field in SCAN_ADDITIONAL_FIELDS:
        output[label] = frame.get(source_field, "")
    return output.reset_index(drop=True)


def _write_approved_listing_blocks(ws, approved: pd.DataFrame) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    ws.merge_cells("A1:B1")
    ws["A1"] = "Create a listing for each Apartment Building as per sample below"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(wrap_text=True)

    row_number = 3
    required_fields = [label for label, _ in SCAN_LISTING_FIELDS]
    for listing_number, (_, row) in enumerate(approved.iterrows(), start=1):
        name = _clean_value(row.get("Apartment Building Name")) or f"Listing {listing_number}"
        ws.merge_cells(start_row=row_number, start_column=1, end_row=row_number, end_column=2)
        ws.cell(
            row=row_number,
            column=1,
            value=f"Apartment Building {listing_number}: {name}",
        ).font = Font(bold=True, size=12)
        row_number += 1

        for field_name, raw_value in row.items():
            value = _clean_value(raw_value)
            if field_name not in required_fields and not value:
                continue
            field_cell = ws.cell(row=row_number, column=1, value=field_name)
            value_cell = ws.cell(row=row_number, column=2, value=value)
            field_cell.font = Font(bold=True)
            field_cell.border = border
            value_cell.border = border
            field_cell.alignment = Alignment(wrap_text=True, vertical="top")
            value_cell.alignment = Alignment(wrap_text=True, vertical="top")
            if field_name in {"WebSite", "Official Source URL"} and value.startswith(("http://", "https://")):
                value_cell.hyperlink = value
                value_cell.style = "Hyperlink"
            elif field_name == "Email Contact" and value:
                value_cell.hyperlink = f"mailto:{value}"
                value_cell.style = "Hyperlink"
            row_number += 1
        row_number += 2

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 75
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False


def _excel_bytes(records_df: pd.DataFrame, pages_df: pd.DataFrame, report) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        approved_mask = records_df.get(
            "approved",
            pd.Series(False, index=records_df.index),
        ).fillna(False)
        approved = records_df.loc[
            approved_mask & _ontario_eligible_mask(records_df)
        ].copy()
        approved_listings = _approved_listing_table(approved)
        ws = writer.book.create_sheet("Approved Listings")
        _write_approved_listing_blocks(ws, approved_listings)
        records_df.to_excel(writer, sheet_name="Record Candidates", index=False)
        records_df.loc[
            records_df["ontario_scope_status"].eq(ONTARIO_SCOPE_OUTSIDE)
        ].to_excel(writer, sheet_name="Outside Ontario", index=False)
        records_df.loc[
            records_df["ontario_scope_status"].eq(ONTARIO_SCOPE_UNCLEAR)
        ].to_excel(writer, sheet_name="Location Review", index=False)
        if "inventory_status" in records_df.columns:
            records_df.loc[
                records_df["inventory_status"].eq(INVENTORY_STATUS_EXCLUDED)
            ].to_excel(writer, sheet_name="Excluded Legacy URLs", index=False)
        pages_df.to_excel(writer, sheet_name="Pages Scanned", index=False)
        pd.DataFrame({"Blocked URL": report.blocked_urls}).to_excel(
            writer,
            sheet_name="Blocked",
            index=False,
        )
        pd.DataFrame({"Error": report.errors}).to_excel(
            writer,
            sheet_name="Errors",
            index=False,
        )
    return output.getvalue()



SCAN_HISTORY_COLUMNS = [
    "Scan ID", "Company ID", "Management/Owner", "Start URL", "Coverage",
    "Started At UTC", "Completed At UTC", "Completion Reason",
    "Pages Scanned", "Candidates Detected", "Candidates Approved",
    "Records Added", "Duplicates Skipped", "Blocked URLs", "Scan Errors",
    "Recovered Partial", "Last Updated UTC",
]


def _scan_id_for_report(report, company_id: str, website_url: str) -> str:
    existing = str(st.session_state.get("website_scan_id", "")).strip()
    if existing:
        return existing

    started = str(getattr(report, "started_at_utc", "") or "").strip()
    source = "|".join([
        str(company_id or "").strip(),
        _normalized_scan_url(website_url),
        started,
    ])
    digest = hashlib.sha256(source.encode("utf-8")).hexdigest()[:8].upper()
    compact_time = re.sub(r"[^0-9]", "", started)[:14]
    if not compact_time:
        compact_time = datetime.now(timezone.utc).strftime("%Y%m%d%H%M%S")
    scan_id = f"SCAN-{compact_time}-{digest}"
    st.session_state["website_scan_id"] = scan_id
    return scan_id


def _attach_scan_context(
    frame: pd.DataFrame,
    *,
    scan_id: str,
    company_id: str,
    company_name: str,
    website_url: str,
) -> pd.DataFrame:
    contextual = frame.copy()
    contextual["scan_id"] = scan_id
    contextual["company_id"] = str(company_id or "").strip()
    contextual["assigned_company"] = str(company_name or "").strip()
    contextual["scan_start_url"] = str(website_url or "").strip()

    if "management_owner" not in contextual.columns:
        contextual["management_owner"] = ""
    owner_blank = contextual["management_owner"].apply(_clean_value).eq("")
    contextual.loc[owner_blank, "management_owner"] = str(company_name or "").strip()
    return contextual


def _history_frame(key: str, columns: list[str] | None = None) -> pd.DataFrame:
    value = st.session_state.get(key)
    if not isinstance(value, pd.DataFrame):
        value = pd.DataFrame(columns=columns or [])
    value = value.copy()
    if columns:
        for column in columns:
            if column not in value.columns:
                value[column] = pd.NA
    return value


def _replace_scan_rows(existing: pd.DataFrame, new_rows: pd.DataFrame, scan_id: str) -> pd.DataFrame:
    if "Scan ID" in existing.columns:
        existing = existing.loc[
            ~existing["Scan ID"].fillna("").astype(str).eq(str(scan_id))
        ].copy()
    if new_rows is None or new_rows.empty:
        return existing.reset_index(drop=True)

    all_columns = list(existing.columns)
    all_columns.extend(
        column for column in new_rows.columns
        if column not in all_columns
    )
    return pd.concat(
        [
            existing.reindex(columns=all_columns),
            new_rows.reindex(columns=all_columns),
        ],
        ignore_index=True,
        sort=False,
    )


def _persist_scan_evidence(
    *,
    report,
    records_df: pd.DataFrame,
    pages_df: pd.DataFrame,
    scan_id: str,
    company_id: str,
    company_name: str,
    website_url: str,
    scope: str,
    history_key: str,
    candidates_key: str,
    pages_key: str,
    added_count: int | None = None,
    duplicate_count: int | None = None,
) -> None:
    approval_mask = records_df.get(
        "approved",
        pd.Series(False, index=records_df.index),
    ).fillna(False)
    eligible_mask = _ontario_eligible_mask(records_df)
    approved_count = int((approval_mask & eligible_mask).sum())

    history = _history_frame(history_key, SCAN_HISTORY_COLUMNS)
    prior_added = 0
    prior_duplicates = 0
    if not history.empty and "Scan ID" in history.columns:
        prior = history.loc[history["Scan ID"].astype(str).eq(scan_id)]
        if not prior.empty:
            prior_added_value = pd.to_numeric(
                prior.iloc[-1].get("Records Added", 0),
                errors="coerce",
            )
            prior_duplicate_value = pd.to_numeric(
                prior.iloc[-1].get("Duplicates Skipped", 0),
                errors="coerce",
            )
            prior_added = (
                0 if pd.isna(prior_added_value) else int(prior_added_value)
            )
            prior_duplicates = (
                0
                if pd.isna(prior_duplicate_value)
                else int(prior_duplicate_value)
            )

    summary = pd.DataFrame([{
        "Scan ID": scan_id,
        "Company ID": str(company_id or "").strip(),
        "Management/Owner": str(company_name or "").strip(),
        "Start URL": str(website_url or "").strip(),
        "Coverage": str(scope or "").strip(),
        "Started At UTC": str(getattr(report, "started_at_utc", "") or ""),
        "Completed At UTC": str(getattr(report, "completed_at_utc", "") or ""),
        "Completion Reason": str(getattr(report, "completion_reason", "") or ""),
        "Pages Scanned": len(getattr(report, "pages", []) or []),
        "Candidates Detected": len(records_df),
        "Candidates Approved": approved_count,
        "Records Added": prior_added if added_count is None else int(added_count),
        "Duplicates Skipped": (
            prior_duplicates if duplicate_count is None else int(duplicate_count)
        ),
        "Blocked URLs": len(getattr(report, "blocked_urls", []) or []),
        "Scan Errors": len(getattr(report, "errors", []) or []),
        "Recovered Partial": bool(
            st.session_state.get("website_scan_recovered_partial", False)
        ),
        "Last Updated UTC": datetime.now(timezone.utc).isoformat(),
    }])
    st.session_state[history_key] = _replace_scan_rows(history, summary, scan_id)

    candidates = records_df.copy().rename(columns={
        "scan_id": "Scan ID",
        "company_id": "Company ID",
        "assigned_company": "Assigned Company",
        "scan_start_url": "Scan Start URL",
    })
    candidates.columns = [
        column
        if column in {
            "Scan ID", "Company ID", "Assigned Company", "Scan Start URL"
        }
        else str(column).replace("_", " ").title()
        for column in candidates.columns
    ]
    candidates["Scan ID"] = scan_id
    candidates["Company ID"] = str(company_id or "").strip()
    candidates["Management/Owner"] = str(company_name or "").strip()
    candidates["Candidate Outcome"] = "Not selected"
    if "Candidate Status" in candidates.columns:
        candidates["Candidate Outcome"] = candidates["Candidate Status"].fillna(
            "Not selected"
        )
    elif "Approved" in candidates.columns:
        candidates.loc[candidates["Approved"].fillna(False), "Candidate Outcome"] = (
            "Approved for project"
        )
    candidate_history = _history_frame(candidates_key)
    st.session_state[candidates_key] = _replace_scan_rows(
        candidate_history,
        candidates,
        scan_id,
    )

    pages = pages_df.copy()
    pages.columns = [
        str(column).replace("_", " ").title()
        for column in pages.columns
    ]
    pages["Scan ID"] = scan_id
    pages["Company ID"] = str(company_id or "").strip()
    pages["Management/Owner"] = str(company_name or "").strip()
    page_history = _history_frame(pages_key)
    st.session_state[pages_key] = _replace_scan_rows(
        page_history,
        pages,
        scan_id,
    )



def _merge_into_working_data(
    approved: pd.DataFrame,
    working_data_key: str,
    company_id: str,
    company_name: str,
) -> tuple[int, int]:
    mapped = approved.rename(columns=DIRECTORY_FIELD_MAP)
    available_mapped_columns = [
        destination
        for source, destination in DIRECTORY_FIELD_MAP.items()
        if source in approved.columns
    ]
    mapped = mapped[available_mapped_columns].copy()

    # The active company is the authoritative parent of every approved record.
    # Do not infer it later from whichever company happens to be selected.
    mapped["Company ID"] = str(company_id or "").strip()
    if "Management/Owner" not in mapped.columns:
        mapped["Management/Owner"] = str(company_name or "").strip()
    else:
        blank_owner = mapped["Management/Owner"].apply(_clean_value).eq("")
        mapped.loc[blank_owner, "Management/Owner"] = str(
            company_name or ""
        ).strip()

    mapped["Date Researched"] = date.today().isoformat()
    mapped["Research Status"] = "Ready for Review"
    mapped["Source Status"] = "Active"
    mapped["Verification Status"] = "Needs Review"
    mapped["Record Decision"] = "Undecided"
    mapped["Missing Information"] = (
        "Website-scanned rental property candidate. Confirm every extracted "
        "detail and supporting source before final use."
    )

    existing = st.session_state.get(working_data_key)
    if existing is None or not isinstance(existing, pd.DataFrame):
        st.session_state[working_data_key] = mapped.reset_index(drop=True)
        return len(mapped), 0

    existing = existing.copy()
    for column in mapped.columns:
        if column not in existing.columns:
            existing[column] = ""
    for column in existing.columns:
        if column not in mapped.columns:
            mapped[column] = ""
    mapped = mapped[existing.columns]

    def key_frame(frame: pd.DataFrame) -> pd.Series:
        company = frame.get(
            "Company ID",
            pd.Series("", index=frame.index),
        ).fillna("")
        address = frame.get(
            "Street Address",
            pd.Series("", index=frame.index),
        ).fillna("")
        name = frame.get(
            "Building Name",
            pd.Series("", index=frame.index),
        ).fillna("")
        source = frame.get(
            "Source URL",
            pd.Series("", index=frame.index),
        ).fillna("")
        return (
            company.astype(str).str.lower().str.replace(
                r"[^a-z0-9]", "", regex=True
            )
            + "|"
            + name.astype(str).str.lower().str.replace(
                r"[^a-z0-9]", "", regex=True
            )
            + "|"
            + address.astype(str).str.lower().str.replace(
                r"[^a-z0-9]", "", regex=True
            )
            + "|"
            + source.astype(str).str.lower().str.replace(
                r"[^a-z0-9]", "", regex=True
            )
        )

    existing_keys = set(key_frame(existing))
    mapped_keys = key_frame(mapped)
    new_rows = mapped.loc[~mapped_keys.isin(existing_keys)].copy()
    duplicates = len(mapped) - len(new_rows)
    st.session_state[working_data_key] = pd.concat(
        [existing, new_rows],
        ignore_index=True,
    )
    return len(new_rows), duplicates



def _clear_stale_scanner_session_if_needed() -> bool:
    """Remove in-memory results produced by an older scanner build."""
    has_results = any(
        key in st.session_state
        for key in (
            "website_scan_report",
            "website_scan_records",
            "website_scan_checkpoint_report",
        )
    )
    stored_build = str(st.session_state.get("website_scan_build", "")).strip()
    if not has_results or stored_build == SCANNER_BUILD:
        return False

    for key in list(st.session_state):
        if (
            key.startswith("website_scan_")
            or key.startswith("full_scan_record_editor_")
        ):
            st.session_state.pop(key, None)
    st.session_state["website_scan_build"] = SCANNER_BUILD
    return True


def render_website_scanner_panel(
    working_data_key: str = WORKING_DATA_KEY,
    *,
    active_company_id: str = "",
    active_company_name: str = "",
    active_company_website: str = "",
    scan_history_key: str = "db_scan_history",
    scan_candidates_key: str = "db_scan_candidates_history",
    scan_pages_key: str = "db_scan_pages_history",
) -> dict | None:
    result: dict | None = None
    active_company_id = str(active_company_id or "").strip()
    active_company_name = str(active_company_name or "").strip()
    active_company_website = str(active_company_website or "").strip()
    company_context_ready = bool(active_company_id and active_company_name)

    # The app shell already renders the shared Collect → Review → Verify →
    # Download process bar. Keep only the scanner page heading here so the
    # workflow controls do not appear twice.
    st.markdown(
        """
        <section class="db-page-head" aria-label="Scan a rental property website">
            <div class="db-eyebrow">COLLECT</div>
            <h2>Scan a rental property website</h2>
            <p>Search permitted public pages for listing details, contacts, building classifications, apartment counts, and supporting evidence. Only approved candidates move into the workspace.</p>
        </section>
        """,
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="db-guidance"><strong>Ontario-only research scope.</strong>'
        '<span>Use this optional scanner to cross-check website coverage and possible omissions. Only Ontario properties can be approved, added, or exported; confirmed non-Ontario records remain in the scan log.</span></div>',
        unsafe_allow_html=True,
    )
    st.markdown(
        '<div class="db-guidance"><strong>Approval and verification are separate.</strong>'
        '<span>Approve a candidate to add it to the workspace; complete human verification in Review records.</span></div>',
        unsafe_allow_html=True,
    )

    if not company_context_ready:
        st.warning(
            "Select or add the company in the Master project area before "
            "starting a scan. This prevents findings from being attached to "
            "the wrong organization."
        )
        return None

    _switch_company_scan_state(active_company_id, active_company_website)
    stale_scan_cleared = _clear_stale_scanner_session_if_needed()
    if stale_scan_cleared:
        st.info(
            "Older scanner results were cleared because the inventory and field-extraction rules changed. "
            "Run the company website again to use the corrected logic."
        )

    saved_scan_company_id = str(
        st.session_state.get("website_scan_company_id", "")
    ).strip()
    has_current_scan = (
        st.session_state.get("website_scan_report") is not None
        or st.session_state.get("website_scan_records") is not None
    )
    company_context_mismatch = bool(
        has_current_scan
        and saved_scan_company_id
        and saved_scan_company_id != active_company_id
    )

    st.success(
        f"Current company: **{active_company_name}** "
        f"({active_company_id}). Approved records and scan evidence will "
        "remain attached to this company."
    )
    if active_company_website:
        st.caption(f"Registered website: {active_company_website}")

    if company_context_mismatch:
        st.warning(
            "These restored scan results carry a different company identifier. "
            "Clear them before starting a new scan for this company."
        )

    st.markdown("### Choose website")
    website_url = st.text_input(
        "Website address",
        placeholder="https://examplepropertycompany.ca",
        help="Use the main public website of the property owner or management company.",
        key="full_scan_website_url",
    )

    coverage_settings = {
        "Quick check": (25, 3),
        "Recommended": (500, 12),
        "Extended scan": (2_000, 20),
        "Custom": (100, 5),
    }

    # Reset older preset names safely after an app update.
    if st.session_state.get("full_scan_scope") not in coverage_settings:
        st.session_state["full_scan_scope"] = "Recommended"

    def _select_extended_scan() -> None:
        st.session_state["full_scan_scope"] = "Extended scan"

    scope = st.session_state["full_scan_scope"]

    with st.expander("Change coverage", expanded=False):
        scope = st.radio(
            "Coverage",
            options=list(coverage_settings),
            help=(
                "Recommended is suitable for most websites. Quick check is for testing. "
                "Extended scan is a follow-up option. Custom is for unusual cases."
            ),
            key="full_scan_scope",
        )

    max_pages, max_depth = coverage_settings[scope]

    if scope == "Recommended":
        st.markdown(
            "**Recommended scan**  \n"
            "Up to **500 pages** · **12 link levels** · Sitemaps and related "
            "subdomains included"
        )
        st.caption(
            "Start here for broad, efficient coverage. Datablix will tell you when "
            "an extended scan may be useful."
        )
    elif scope == "Extended scan":
        st.markdown(
            "**Extended scan selected**  \n"
            "Up to **2,000 pages** · **20 link levels** · Sitemaps and related "
            "subdomains included"
        )
        st.caption(
            "Use this after the recommended scan reaches its limit or when known "
            "listings still appear to be missing."
        )
    elif scope == "Quick check":
        st.markdown(
            "**Quick check selected**  \n"
            "Up to **25 pages** · **3 link levels**"
        )
        st.caption("Use this only to confirm that a website can be scanned.")
    else:
        st.markdown("**Custom coverage selected**")
        st.caption("Set manual limits for unusual websites or troubleshooting.")

    with st.expander("Advanced scan settings", expanded=scope == "Custom"):
        if scope == "Custom":
            custom_col1, custom_col2 = st.columns(2)
            max_pages = custom_col1.number_input(
                "Maximum pages",
                min_value=1,
                max_value=2_000,
                value=100,
                step=25,
                help="The scan stops at this limit even when the site has more pages.",
                key="full_scan_max_pages_custom",
            )
            max_depth = custom_col2.number_input(
                "Maximum link depth",
                min_value=0,
                max_value=20,
                value=5,
                step=1,
                help="How many link levels the scanner follows from the starting page.",
                key="full_scan_max_depth_custom",
            )
        else:
            st.caption(
                f"Current coverage: up to **{max_pages:,} pages** and "
                f"**{max_depth} link levels**."
            )

        advanced_col1, advanced_col2 = st.columns(2)
        with advanced_col1:
            render_label = st.selectbox(
                "Page rendering",
                options=[
                    "Automatic",
                    "HTML only",
                    "Always render JavaScript",
                ],
                help=(
                    "Automatic reads ordinary HTML first and opens a browser only "
                    "when a page appears to require JavaScript."
                ),
                key="full_scan_render_mode",
            )
            delay = st.number_input(
                "Delay between requests (seconds)",
                min_value=0.1,
                max_value=30.0,
                value=0.75,
                step=0.25,
                help=(
                    "A longer delay is gentler on the website and can reduce "
                    "temporary blocking during large scans."
                ),
                key="full_scan_delay",
            )

        with advanced_col2:
            if scope == "Custom":
                use_sitemaps = st.checkbox(
                    "Use sitemaps and current listing pages",
                    value=True,
                    key="full_scan_sitemaps_custom",
                )
                include_subdomains = st.checkbox(
                    "Include subdomains",
                    value=False,
                    help="Turn this on when listings live on a related subdomain.",
                    key="full_scan_subdomains_custom",
                )
            else:
                use_sitemaps = True
                include_subdomains = scope in {"Recommended", "Extended scan"}
                discovery_note = (
                    "Sitemaps, current listing pages, and related subdomains are included automatically."
                    if include_subdomains
                    else "Sitemaps and current listing pages are included; related subdomains are not followed."
                )
                st.caption(discovery_note)

            follow_queries = st.checkbox(
                "Follow query-string pages",
                value=False,
                help=(
                    "Enable only when parameters such as ?page=2 or ?property=123 "
                    "lead to unique listing pages."
                ),
                key="full_scan_queries",
            )
            st.checkbox(
                "Respect robots.txt",
                value=True,
                disabled=True,
                help="Always on. The scanner never visits pages a site asks crawlers to skip.",
                key="full_scan_robots",
            )

    acknowledgement = st.checkbox(
        "I am scanning permitted public pages and will review every finding before use.",
        key="full_scan_acknowledgement",
    )

    missing_requirements = []
    if not website_url.strip():
        missing_requirements.append("enter the company website")
    if not acknowledgement:
        missing_requirements.append("confirm the public-page acknowledgement")
    if company_context_mismatch:
        missing_requirements.append("clear the mismatched restored scan")

    submitted = st.button(
        f"Start scan for {active_company_name}",
        type="primary",
        width="stretch",
        disabled=bool(missing_requirements),
        key="full_scan_submit",
    )
    if missing_requirements:
        st.caption("Before starting: " + "; ".join(missing_requirements) + ".")
    else:
        st.caption(
            f"Ready to scan {website_url.strip()} for {active_company_name}."
        )

    if submitted:
        st.session_state["website_scan_build"] = SCANNER_BUILD
        st.session_state["website_scan_active"] = True
        st.session_state["website_scan_company_id"] = active_company_id
        st.session_state["website_scan_company_name"] = active_company_name
        st.session_state["website_scan_company_website"] = active_company_website
        st.session_state.pop("website_scan_id", None)
        st.session_state["website_scan_start_url"] = website_url
        st.session_state["website_scan_last_progress"] = {
            "pages_processed": 0,
            "records_found": 0,
            "blocked_count": 0,
            "error_count": 0,
            "current_url": website_url,
            "max_pages": int(max_pages),
        }
        st.session_state.pop("website_scan_report", None)
        st.session_state.pop("website_scan_records", None)
        st.session_state.pop("website_scan_checkpoint_report", None)
        st.session_state.pop("website_scan_recovered_partial", None)

        render_mode = {
            "Automatic": "auto",
            "HTML only": "html",
            "Always render JavaScript": "javascript",
        }[render_label]

        options = ScanOptions(
            max_pages=int(max_pages),
            max_depth=int(max_depth),
            request_delay_seconds=float(delay),
            render_mode=render_mode,
            use_sitemaps=use_sitemaps,
            include_subdomains=include_subdomains,
            follow_query_strings=follow_queries,
            obey_robots_txt=True,
            checkpoint_every_pages=10,
        )

        status = st.status("Starting website scan…", expanded=True)
        progress = st.progress(0.0)
        current_page = st.empty()
        counters = st.empty()

        def update_progress(update: dict) -> None:
            st.session_state["website_scan_last_progress"] = dict(update)
            processed = update.get("pages_processed", 0)
            maximum = max(update.get("max_pages", 1), 1)
            progress.progress(min(processed / maximum, 1.0))
            current_page.caption(update.get("current_url", ""))
            counters.write(
                f"Pages processed: **{processed}** · "
                f"Candidates: **{update.get('records_found', 0)}** · "
                f"Blocked: **{update.get('blocked_count', 0)}** · "
                f"Errors: **{update.get('error_count', 0)}**"
            )

        def save_checkpoint(partial_report) -> None:
            # Keep an in-session copy for fast recovery and an atomic JSON file
            # for recovery after a Streamlit rerun or browser reconnection.
            st.session_state["website_scan_checkpoint_report"] = partial_report
            st.session_state["website_scan_checkpoint_pages"] = len(
                partial_report.pages
            )
            checkpoint_path = _write_durable_checkpoint(
                partial_report,
                website_url,
                scope,
                is_final=False,
            )
            if checkpoint_path is not None:
                st.session_state["website_scan_checkpoint_path"] = str(
                    checkpoint_path
                )

        try:
            report = scan_website(
                website_url=website_url,
                options=options,
                progress_callback=update_progress,
                checkpoint_callback=save_checkpoint,
            )
        except WebsiteScanError as exc:
            st.session_state["website_scan_active"] = False
            st.session_state["website_scan_stop_message"] = str(exc)
            status.update(label="The scan could not start", state="error", expanded=True)
            st.error(str(exc))
        except Exception as exc:
            st.session_state["website_scan_active"] = False
            st.session_state["website_scan_stop_message"] = str(exc)
            status.update(
                label="The scan stopped unexpectedly",
                state="error",
                expanded=True,
            )
            st.error(
                "The scan stopped because one page contained an unsupported or "
                "malformed link. The corrected scanner now skips that link and "
                "continues. Review the technical details only if the issue repeats."
            )
            with st.expander("Technical details"):
                st.code(f"{type(exc).__name__}: {exc}")

            recovered = _read_durable_checkpoint(website_url)
            if recovered is not None:
                recovered_report, checkpoint_meta = recovered
                recovered_report.completion_reason = "interrupted_recovered"
                recovered_records = _records_dataframe(recovered_report)
                st.session_state["website_scan_build"] = SCANNER_BUILD
                st.session_state["website_scan_report"] = recovered_report
                st.session_state["website_scan_records"] = recovered_records
                st.session_state["website_scan_scope"] = checkpoint_meta.get(
                    "scope", scope
                )
                st.session_state["website_scan_page_limit_reached"] = False
                st.session_state["website_scan_recovered_partial"] = True
                st.warning(
                    f"Recovered {len(recovered_report.pages):,} completed pages "
                    f"and {len(recovered_records):,} candidate(s) from the latest "
                    "durable checkpoint."
                )
        else:
            st.session_state["website_scan_active"] = False
            progress.progress(1.0)

            completion_reason = getattr(report, "completion_reason", "")
            page_limit_reached = (
                completion_reason == "page_limit"
                or len(report.pages) >= int(max_pages)
            )

            if completion_reason == "failure_limit":
                visible_message = (
                    f"Scan stopped after repeated page failures: "
                    f"{len(report.pages)} pages processed and "
                    f"{len(report.records)} unique candidates found."
                )
                status.update(
                    label=visible_message,
                    state="error",
                    expanded=True,
                )
                st.error(
                    visible_message
                    + " Review the Errors section before trying again."
                )
            elif page_limit_reached:
                visible_message = (
                    f"Page limit reached: {len(report.pages)} pages processed and "
                    f"{len(report.records)} unique candidates found."
                )
                status.update(
                    label=visible_message,
                    state="complete",
                    expanded=True,
                )
                st.warning(
                    visible_message
                    + " Additional permitted pages may remain."
                )
            else:
                visible_message = (
                    f"Coverage complete: {len(report.pages)} eligible pages processed "
                    f"and {len(report.records)} unique candidates found."
                )
                status.update(
                    label=visible_message,
                    state="complete",
                    expanded=True,
                )
                st.success(
                    visible_message
                    + " The scan stopped because no more eligible pages were available."
                )

            st.session_state["website_scan_stop_message"] = visible_message
            st.session_state["website_scan_build"] = SCANNER_BUILD
            st.session_state["website_scan_report"] = report
            st.session_state["website_scan_records"] = _records_dataframe(report)
            st.session_state["website_scan_scope"] = scope
            st.session_state["website_scan_page_limit_reached"] = page_limit_reached
            final_checkpoint_path = _write_durable_checkpoint(
                report,
                website_url,
                scope,
                is_final=True,
            )
            if final_checkpoint_path is not None:
                st.session_state["website_scan_checkpoint_path"] = str(
                    final_checkpoint_path
                )
            st.session_state.pop("website_scan_checkpoint_report", None)
            st.session_state.pop("website_scan_checkpoint_pages", None)
            st.session_state.pop("website_scan_recovered_partial", None)

    report = st.session_state.get("website_scan_report")
    records_df = st.session_state.get("website_scan_records")

    if report is None or not isinstance(records_df, pd.DataFrame):
        checkpoint_report = st.session_state.get(
            "website_scan_checkpoint_report"
        )
        checkpoint_meta = {}
        recovery_url = (
            website_url
            or st.session_state.get("website_scan_start_url", "")
        )

        if checkpoint_report is None:
            durable_recovery = _read_durable_checkpoint(recovery_url)
            if durable_recovery is not None:
                checkpoint_report, checkpoint_meta = durable_recovery

        last_progress = st.session_state.get("website_scan_last_progress", {})
        scan_active = bool(st.session_state.get("website_scan_active", False))
        processed = int(last_progress.get("pages_processed", 0) or 0)

        if checkpoint_report is not None and getattr(
            checkpoint_report, "pages", None
        ):
            report = checkpoint_report
            report.completion_reason = "interrupted_recovered"
            records_df = _records_dataframe(report)

            st.session_state["website_scan_active"] = False
            st.session_state["website_scan_build"] = SCANNER_BUILD
            st.session_state["website_scan_report"] = report
            st.session_state["website_scan_records"] = records_df
            st.session_state["website_scan_scope"] = checkpoint_meta.get(
                "scope", scope
            )
            st.session_state["website_scan_page_limit_reached"] = False
            st.session_state["website_scan_recovered_partial"] = True

            st.warning(
                f"Partial scan recovered: {len(report.pages):,} completed pages "
                f"and {len(records_df):,} candidate(s) are available for review. "
                "The latest durable checkpoint was restored."
            )
        elif scan_active and processed > 0:
            # This covers sessions started with an older Datablix version that
            # did not yet save checkpoints. Reset the stale flag so the same
            # notice does not remain indefinitely.
            st.session_state["website_scan_active"] = False
            st.warning(
                f"The earlier scan ended after approximately {processed:,} pages "
                "before a durable checkpoint was available. Start the scan once "
                "more. Future runs now save findings every 10 pages."
            )
            return
        else:
            st.caption(
                "Scan results will appear here for review. Nothing is added to the "
                "workspace automatically."
            )
            return

    scan_company_id = str(
        st.session_state.get("website_scan_company_id", active_company_id)
        or active_company_id
    ).strip()
    scan_company_name = str(
        st.session_state.get("website_scan_company_name", active_company_name)
        or active_company_name
    ).strip()
    scan_start_url = str(
        getattr(report, "start_url", "")
        or st.session_state.get("website_scan_start_url", website_url)
        or website_url
    ).strip()
    scan_id = _scan_id_for_report(
        report,
        scan_company_id,
        scan_start_url,
    )

    records_df = _apply_ontario_scope(records_df)
    records_df = _attach_scan_context(
        records_df,
        scan_id=scan_id,
        company_id=scan_company_id,
        company_name=scan_company_name,
        website_url=scan_start_url,
    )
    records_df = _apply_candidate_status(records_df)
    st.session_state["website_scan_records"] = records_df

    if (
        active_company_id
        and scan_company_id
        and active_company_id != scan_company_id
    ):
        st.warning(
            f"These results belong to **{scan_company_name}** "
            f"({scan_company_id}), not the newly selected company. Clear the "
            "current scan before starting work for another company."
        )

    if st.session_state.get("website_scan_recovered_partial", False):
        st.info(
            "You are reviewing recovered partial results. They can be approved, "
            "exported, or cleared normally. Run the website again later only when "
            "you need to search beyond the recovered pages."
        )

    last_scan_scope = st.session_state.get("website_scan_scope", "")
    last_scan_reached_limit = bool(
        st.session_state.get("website_scan_page_limit_reached", False)
    )

    if last_scan_scope == "Recommended" and last_scan_reached_limit:
        st.warning(
            "The recommended scan reached its 500-page limit. Some permitted pages "
            "may remain, so an extended scan is the appropriate next step."
        )
        st.button(
            "Select extended scan",
            type="primary",
            on_click=_select_extended_scan,
            key="select_extended_scan_after_limit",
        )
    elif last_scan_scope == "Recommended":
        st.success(
            "Recommended coverage completed without reaching its page limit. "
            "An extended scan is not needed unless a known listing is missing."
        )

    pages_df = _pages_dataframe(report)
    successful_pages = (
        int((pages_df.get("outcome") == "Scanned").sum())
        if not pages_df.empty
        else 0
    )
    rendered_pages = int(
        pages_df.get(
            "rendered_with_javascript",
            pd.Series(dtype=bool),
        ).fillna(False).sum()
    )

    st.divider()
    st.markdown("### Review detected listings")
    st.caption(
        "Confirm the required listing fields first, then inspect additional findings and source evidence. "
        "A blank value means the page did not provide enough information to confirm it."
    )

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("Pages scanned", successful_pages)
    m2.metric("Candidates", len(records_df))
    m3.metric("JavaScript pages", rendered_pages)
    m4.metric("Blocked or failed", len(report.blocked_urls) + len(report.errors))

    scope_counts = records_df["ontario_scope_status"].value_counts()
    s1, s2, s3, s4 = st.columns(4)
    s1.metric("Confirmed Ontario", int(scope_counts.get(ONTARIO_SCOPE_CONFIRMED, 0)))
    s2.metric("Likely Ontario — review", int(scope_counts.get(ONTARIO_SCOPE_LIKELY, 0)))
    s3.metric("Location unclear", int(scope_counts.get(ONTARIO_SCOPE_UNCLEAR, 0)))
    s4.metric("Outside Ontario", int(scope_counts.get(ONTARIO_SCOPE_OUTSIDE, 0)))
    st.caption(
        "Only Confirmed Ontario and human-approved Likely Ontario candidates can "
        "move into the selected company records. Location Unclear and Outside Ontario records are "
        "kept for traceability but excluded automatically."
    )

    if "inventory_status" in records_df.columns:
        inventory_counts = records_df["inventory_status"].value_counts()
        i1, i2, i3 = st.columns(3)
        i1.metric("Current inventory", int(inventory_counts.get(INVENTORY_STATUS_CURRENT, 0)))
        i2.metric("Inventory review", int(inventory_counts.get(INVENTORY_STATUS_REVIEW, 0)))
        i3.metric("Legacy / excluded", int(inventory_counts.get(INVENTORY_STATUS_EXCLUDED, 0)))
        st.caption(
            "A dedicated property URL is not enough by itself. Buildings linked from a current "
            "HTML sitemap or current city/portfolio page are treated as current. Legacy property "
            "URLs stay in the audit trail but cannot be approved."
        )

    with st.expander("Manage current scan results"):
        st.caption("Clearing removes this scan report and its review edits from the current session.")
        confirm_clear = st.checkbox(
            "Clear the current scan results",
            key="confirm_clear_full_scan",
        )
        if st.button(
            "Clear results",
            disabled=not confirm_clear,
            width="stretch",
            key="clear_full_scan",
        ):
            _delete_durable_checkpoint(
                st.session_state.get("website_scan_start_url", website_url)
            )
            keys_to_clear = [
                key for key in list(st.session_state)
                if key.startswith("full_scan_record_editor_")
            ]
            keys_to_clear.extend([
                "website_scan_report",
                "website_scan_records",
                "website_scan_active",
                "website_scan_last_progress",
                "website_scan_stop_message",
                "website_scan_checkpoint_report",
                "website_scan_checkpoint_pages",
                "website_scan_checkpoint_path",
                "website_scan_recovered_partial",
                "website_scan_start_url",
                "website_scan_scope",
                "website_scan_page_limit_reached",
                "website_scan_id",
                "website_scan_company_id",
                "website_scan_company_name",
                "website_scan_company_website",
                "full_scan_review_focus",
                "confirm_clear_full_scan",
            ])
            for key in keys_to_clear:
                st.session_state.pop(key, None)
            st.rerun()

    required_review_columns = [
        "approved",
        "ontario_scope_status",
        "ontario_scope_reason",
        "building_name",
        "street_address",
        "city",
        "province",
        "postal_code",
        "building_classification",
        "number_of_apartments",
        "amenities",
        "inventory_status",
        "inventory_evidence",
        "management_owner",
        "phone",
        "primary_email",
        "website",
        "source_url",
        "confidence",
    ]
    additional_review_columns = [
        "approved",
        "ontario_scope_status",
        "ontario_scope_reason",
        "inventory_status",
        "inventory_evidence",
        "found_on_city_page",
        "found_on_html_sitemap",
        "found_on_xml_sitemap",
        "exclusion_reason",
        "amenities",
        "address_line_2",
        "country",
        "source_page_title",
        "extraction_method",
        "candidate_status",
        "evidence",
        "source_url",
        "confidence",
    ]
    preferred_all_order = [
        "approved", "scan_id", "company_id", "assigned_company",
        "building_name", "management_owner", "street_address",
        "address_line_2", "city", "province", "postal_code", "country",
        "phone", "primary_email", "website", "number_of_apartments",
        "amenities", "building_classification",
        "inventory_status", "inventory_evidence",
        "found_on_city_page", "found_on_html_sitemap", "found_on_xml_sitemap",
        "exclusion_reason", "source_url", "source_page_title",
        "extraction_method", "confidence", "candidate_status", "evidence",
    ]
    all_review_columns = [column for column in preferred_all_order if column in records_df.columns]
    all_review_columns.extend(
        column for column in records_df.columns if column not in all_review_columns
    )

    for column in set(required_review_columns + additional_review_columns + all_review_columns):
        if column not in records_df.columns:
            records_df[column] = False if column == "approved" else ""

    review_views = {
        "Required listing information": required_review_columns,
        "Additional findings and evidence": additional_review_columns,
        "All detected fields": all_review_columns,
    }
    review_focus = st.selectbox(
        "Review focus",
        list(review_views),
        help="Use a focused view to reduce horizontal scrolling. Every detected field remains available in All detected fields.",
        key="full_scan_review_focus",
    )
    visible_review_columns = [
        column for column in review_views[review_focus] if column in records_df.columns
    ]

    editor_key = "full_scan_record_editor_" + review_focus.lower().replace(" ", "_").replace("/", "_")
    edited_review = st.data_editor(
        records_df[visible_review_columns].copy(),
        key=editor_key,
        width="stretch",
        hide_index=True,
        num_rows="fixed",
        disabled=[
            column for column in [
                "confidence", "source_page_title", "extraction_method",
                "candidate_status", "ontario_scope_status", "ontario_scope_reason",
                "inventory_status", "inventory_evidence", "found_on_city_page",
                "found_on_html_sitemap", "found_on_xml_sitemap", "exclusion_reason"
            ] if column in visible_review_columns
        ],
        column_config={
            "approved": st.column_config.CheckboxColumn(
                "Approve candidate",
                default=False,
                help=(
                    "Approve only after checking the source and confirming the "
                    "candidate is in Ontario."
                ),
            ),
            "ontario_scope_status": st.column_config.TextColumn(
                "Ontario Scope Status",
                width="medium",
                help=(
                    "Confirmed Ontario may be approved. Likely Ontario requires "
                    "human review. Unclear and outside-Ontario records are excluded."
                ),
            ),
            "ontario_scope_reason": st.column_config.TextColumn(
                "Ontario Scope Reason",
                width="large",
            ),
            "building_name": st.column_config.TextColumn("Apartment Building Name", width="large"),
            "management_owner": st.column_config.TextColumn("Management / Owner", width="large"),
            "street_address": st.column_config.TextColumn("Street Address"),
            "address_line_2": st.column_config.TextColumn("Address Line 2"),
            "city": st.column_config.TextColumn("City"),
            "province": st.column_config.TextColumn("Province"),
            "postal_code": st.column_config.TextColumn("Postal Code"),
            "amenities": st.column_config.TextColumn("Amenities", width="large"),
            "inventory_status": st.column_config.TextColumn(
                "Current Inventory Status", width="medium"
            ),
            "inventory_evidence": st.column_config.TextColumn(
                "Inventory Evidence", width="large"
            ),
            "found_on_city_page": st.column_config.CheckboxColumn(
                "On City/Portfolio Page"
            ),
            "found_on_html_sitemap": st.column_config.CheckboxColumn(
                "On HTML Sitemap"
            ),
            "found_on_xml_sitemap": st.column_config.CheckboxColumn(
                "On XML Sitemap"
            ),
            "exclusion_reason": st.column_config.TextColumn(
                "Inventory Exclusion Reason", width="large"
            ),
            "country": st.column_config.TextColumn("Country"),
            "phone": st.column_config.TextColumn("Phone Number"),
            "primary_email": st.column_config.TextColumn("Email Contact", width="large"),
            "website": st.column_config.LinkColumn("Website", width="large"),
            "number_of_apartments": st.column_config.TextColumn("Number of Apartments"),
            "building_classification": st.column_config.TextColumn(
                "Building Classification",
                help="For example: High Rise, Low Rise, Townhome, or Duplex.",
            ),
            "source_url": st.column_config.LinkColumn("Official Source URL", width="large"),
            "source_page_title": st.column_config.TextColumn("Source Page Title", width="large"),
            "extraction_method": st.column_config.TextColumn("Extraction Method"),
            "confidence": st.column_config.ProgressColumn(
                "Detection Confidence",
                min_value=0.0,
                max_value=1.0,
                format="percent",
                help="Use confidence to prioritize review; it is not proof that a value is correct.",
            ),
            "candidate_status": st.column_config.TextColumn(
                "Scanner Candidate Status",
                help="Scanner approval only. Final human verification is completed in Review records.",
            ),
            "evidence": st.column_config.TextColumn("Supporting Evidence", width="large"),
        },
    )

    updated_records = records_df.copy()
    for column in visible_review_columns:
        updated_records.loc[edited_review.index, column] = edited_review[column]

    updated_records = _apply_candidate_status(_apply_ontario_scope(updated_records))
    st.session_state["website_scan_records"] = updated_records

    requested_approval = updated_records["approved"].fillna(False)
    ontario_eligible = _ontario_eligible_mask(updated_records)
    inventory_eligible = _inventory_eligible_mask(updated_records)
    eligible_approval = ontario_eligible & inventory_eligible
    blocked_location_count = int((requested_approval & ~ontario_eligible).sum())
    blocked_inventory_count = int((requested_approval & ontario_eligible & ~inventory_eligible).sum())

    approved = updated_records.loc[
        requested_approval & eligible_approval
    ].copy()
    _persist_scan_evidence(
        report=report,
        records_df=updated_records,
        pages_df=pages_df,
        scan_id=scan_id,
        company_id=scan_company_id,
        company_name=scan_company_name,
        website_url=scan_start_url,
        scope=st.session_state.get("website_scan_scope", scope),
        history_key=scan_history_key,
        candidates_key=scan_candidates_key,
        pages_key=scan_pages_key,
    )

    if blocked_location_count:
        st.warning(
            f"{blocked_location_count} selected candidate(s) were not approved because "
            "their location is unclear or confirmed outside Ontario. Update the city, "
            "province, street address, or postal code, then review again."
        )
    if blocked_inventory_count:
        st.warning(
            f"{blocked_inventory_count} selected candidate(s) were not approved because "
            "their dedicated property page is not supported by the current HTML sitemap "
            "or current city/portfolio pages. They remain in the audit trail as legacy/excluded URLs."
        )

    with st.container(border=True):
        action_left, action_right = st.columns([2, 1], vertical_alignment="center")
        with action_left:
            st.subheader("Add approved records to project")
            st.write(
                f"{len(approved):,} Ontario-eligible candidate(s) are approved at the scanner stage. "
                "When added to the project, they begin final record verification in Review records; "
                "scanner approval itself is not the final human verification decision."
            )
        with action_right:
            add_approved = st.button(
                "Add approved records to project",
                type="primary",
                disabled=approved.empty or not scan_company_id,
                width="stretch",
                key="add_approved_scan_records",
            )
    if add_approved:
        added, duplicates = _merge_into_working_data(
            approved,
            working_data_key=working_data_key,
            company_id=scan_company_id,
            company_name=scan_company_name,
        )
        _persist_scan_evidence(
            report=report,
            records_df=updated_records,
            pages_df=pages_df,
            scan_id=scan_id,
            company_id=scan_company_id,
            company_name=scan_company_name,
            website_url=scan_start_url,
            scope=st.session_state.get("website_scan_scope", scope),
            history_key=scan_history_key,
            candidates_key=scan_candidates_key,
            pages_key=scan_pages_key,
            added_count=added,
            duplicate_count=duplicates,
        )
        result = {
            "scan_id": scan_id,
            "company_id": scan_company_id,
            "company_name": scan_company_name,
            "start_url": scan_start_url,
            "added": added,
            "duplicates": duplicates,
        }
        st.success(
            f"Added {added} record(s) to the master project for "
            f"{scan_company_name}. Skipped {duplicates} record(s) already "
            "saved for the same company, source, building name, and address."
        )

    with st.expander("Evidence, scan log, and downloads"):
        st.caption(
            "Keep the source page, extraction details, confidence, and supporting text with the research trail."
        )
        evidence_columns = [
            "inventory_status",
            "inventory_evidence",
            "found_on_city_page",
            "found_on_html_sitemap",
            "found_on_xml_sitemap",
            "exclusion_reason",
            "ontario_scope_status",
            "ontario_scope_reason",
            "building_name",
            "amenities",
            "building_classification",
            "number_of_apartments",
            "source_url",
            "source_page_title",
            "extraction_method",
            "confidence",
            "evidence",
        ]
        available_evidence = [
            column for column in evidence_columns if column in updated_records.columns
        ]
        if available_evidence:
            st.dataframe(
                updated_records[available_evidence],
                width="stretch",
                hide_index=True,
            )

        csv_data = _approved_listing_table(approved).to_csv(index=False).encode("utf-8-sig")
        excel_data = _excel_bytes(updated_records, pages_df, report)
        d1, d2, d3 = st.columns(3)
        d1.download_button(
            "Download approved listings — CSV",
            data=csv_data,
            file_name="approved_website_records.csv",
            mime="text/csv",
            disabled=approved.empty,
            width="stretch",
        )
        d2.download_button(
            "Download complete scan workbook",
            data=excel_data,
            file_name="website_scan_results.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            width="stretch",
        )
        reviewed_records_json = json.loads(
            updated_records.to_json(orient="records", force_ascii=False)
        )
        d3.download_button(
            "Download raw scan data — JSON",
            data=json.dumps(
                {
                    "research_scope": "Ontario only",
                    "scan_report": report.as_dict(),
                    "reviewed_records": reviewed_records_json,
                },
                indent=2,
                ensure_ascii=False,
            ),
            file_name="website_scan_report_ontario.json",
            mime="application/json",
            width="stretch",
        )

        st.write("**Pages scanned**")
        st.dataframe(pages_df, width="stretch", hide_index=True)
        if report.blocked_urls:
            st.write("**Skipped: blocked by robots.txt or scan policy**")
            st.dataframe(
                pd.DataFrame({"URL": report.blocked_urls}),
                width="stretch",
                hide_index=True,
            )
        if report.errors:
            st.write("**Errors during the scan**")
            st.dataframe(
                pd.DataFrame({"Error": report.errors}),
                width="stretch",
                hide_index=True,
            )

    return result
